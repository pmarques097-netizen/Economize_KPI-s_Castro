
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import psycopg2
import psycopg2.extras
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
import io
import base64
import sqlite3
import hashlib
import hmac
import time
import calendar
from sqlalchemy import create_engine, text
from urllib.parse import quote_plus

from pathlib import Path
import json
from datetime import datetime, date, timedelta
import unicodedata
import re
import traceback
import shutil
import gzip
import subprocess
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_EXPORT_OK = True
    EXCEL_EXPORT_ERROR = ""
except Exception as exc:
    Workbook = None
    EXCEL_EXPORT_OK = False
    EXCEL_EXPORT_ERROR = str(exc)

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    PDF_EXPORT_OK = True
    PDF_EXPORT_ERROR = ""
except Exception as exc:
    PDF_EXPORT_OK = False
    PDF_EXPORT_ERROR = str(exc)

# Compatibilidade com o módulo de exportação reconstruído.
REPORTLAB_DISPONIVEL = PDF_EXPORT_OK

st.set_page_config(
    page_title="Rede Economize | KPI Comercial",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# =========================================================
# MODO DE EXECUÇÃO
# =========================================================
MODO_APLICACAO = "VIEWER_STREAMLIT"
MODO_VIEWER = MODO_APLICACAO == "VIEWER_STREAMLIT"
MODO_ATUALIZADOR = MODO_APLICACAO == "ATUALIZADOR_LOCAL"

# =========================================================
# CONTROLE DE ACESSO — REDE ECONOMIZE
# =========================================================
_USUARIOS_SISTEMA = {
    # Administradores com visão corporativa completa.
    "paulo": "233fd8bd85b5d4fa7012d66a8d147b8790c1fcfe378210b9a703328e48c55278",
    "vanderlei": "233fd8bd85b5d4fa7012d66a8d147b8790c1fcfe378210b9a703328e48c55278",
    "ubiratan": "233fd8bd85b5d4fa7012d66a8d147b8790c1fcfe378210b9a703328e48c55278",
}

# Gerentes fixos do Viewer. Mesmo sendo contas internas, nunca recebem visão
# corporativa. O campo gerente é usado para filtrar lojas, cards, tabelas e exportações.
_USUARIOS_GERENTES_SISTEMA = {
    "fabio": {
        "senha_hash": "9cc1365d13c8bdba348b9d6ca8ab41672fbb095f464b66ce5fe7c389ea8b783c",
        "nome": "Fábio",
        "gerente": "Fábio",
    },
    "lanila": {
        "senha_hash": "37e2a5b868d2437b76aa9351b0c4850b078504299fd3727f2e82749cf8f0bc2a",
        "nome": "Lanila",
        "gerente": "Lanila",
    },
}
ACESSOS_COMPRADORES_FILE = Path("config") / "usuarios_compradores.json"


def _normalizar_usuario_acesso(valor):
    texto = unicodedata.normalize("NFKD", str(valor or "").strip().casefold())
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^a-z0-9._-]+", ".", texto)
    texto = re.sub(r"\.+", ".", texto).strip(".")
    return texto


def _hash_senha_acesso(senha):
    return hashlib.sha256(str(senha or "").encode("utf-8")).hexdigest()


def _carregar_acessos_compradores():
    try:
        if ACESSOS_COMPRADORES_FILE.exists():
            dados = json.loads(ACESSOS_COMPRADORES_FILE.read_text(encoding="utf-8"))
            if isinstance(dados, list):
                return dados
    except Exception:
        pass
    return []


def _salvar_acessos_compradores(dados):
    ACESSOS_COMPRADORES_FILE.parent.mkdir(parents=True, exist_ok=True)
    tmp = ACESSOS_COMPRADORES_FILE.with_suffix(".tmp")
    tmp.write_text(json.dumps(dados, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(ACESSOS_COMPRADORES_FILE)


def _dados_acesso_usuario(usuario, senha):
    usuario_norm = _normalizar_usuario_acesso(usuario)
    senha_hash = _hash_senha_acesso(senha)
    hash_admin = _USUARIOS_SISTEMA.get(usuario_norm)
    if hash_admin and hmac.compare_digest(senha_hash, hash_admin):
        return {"autorizado": True, "usuario": usuario_norm, "nome": usuario_norm.title(), "perfil": "Administrador", "comprador": "", "vendedor": "", "gerente": ""}

    gerente_sistema = _USUARIOS_GERENTES_SISTEMA.get(usuario_norm)
    if gerente_sistema:
        esperado = str(gerente_sistema.get("senha_hash", ""))
        if esperado and hmac.compare_digest(senha_hash, esperado):
            gerente_nome = str(gerente_sistema.get("gerente", "") or "").strip()
            return {
                "autorizado": True,
                "usuario": usuario_norm,
                "nome": str(gerente_sistema.get("nome", "") or gerente_nome or usuario_norm.title()),
                "perfil": "Gerente",
                "comprador": "",
                "vendedor": "",
                "gerente": gerente_nome,
            }

    for item in _carregar_acessos_compradores():
        if _normalizar_usuario_acesso(item.get("usuario")) != usuario_norm:
            continue
        if not bool(item.get("ativo", True)):
            continue
        esperado = str(item.get("senha_hash", ""))
        if not esperado or not hmac.compare_digest(senha_hash, esperado):
            continue
        perfil = str(item.get("perfil") or "Comprador").strip().title()
        if perfil not in {"Comprador", "Vendedor", "Gerente"}:
            perfil = "Comprador"
        comprador = str(item.get("comprador", "")).strip()
        vendedor = str(item.get("vendedor", "")).strip()
        gerente = str(item.get("gerente", "")).strip()
        if perfil == "Comprador":
            nome = comprador
        elif perfil == "Vendedor":
            nome = vendedor
        else:
            nome = gerente
        return {
            "autorizado": True,
            "usuario": usuario_norm,
            "nome": nome or usuario_norm.title(),
            "perfil": perfil,
            "comprador": comprador if perfil == "Comprador" else "",
            "vendedor": vendedor if perfil == "Vendedor" else "",
            "gerente": gerente if perfil == "Gerente" else "",
        }

    return {"autorizado": False, "usuario": usuario_norm, "nome": "", "perfil": "", "comprador": "", "vendedor": "", "gerente": ""}


def _renderizar_login_sistema():
    if st.session_state.get("_usuario_autenticado"):
        return

    st.markdown("""
    <style>
    [data-testid="stSidebar"]{display:none!important;}
    [data-testid="collapsedControl"]{display:none!important;}
    .login-card{max-width:520px;margin:7vh auto 18px;padding:30px 34px;border-radius:20px;border:1px solid #294b65;background:linear-gradient(145deg,#0d2032,#081724);box-shadow:0 24px 70px rgba(0,0,0,.36)}
    .login-card h1{border:0!important;padding:0!important;margin:0 0 8px!important;font-size:30px!important}
    .login-card p{color:#9fb4c5!important;margin:3px 0!important}
    </style>
    <div class="login-card"><h1>Rede Economize</h1><p>KPI Comercial • Enterprise Edition</p><p>Acesso restrito a usuários autorizados.</p></div>
    """, unsafe_allow_html=True)

    a,b,c = st.columns([1.25,1,1.25])
    with b:
        with st.form("login_rede_economize"):
            usuario = st.text_input("Usuário")
            senha = st.text_input("Senha", type="password")
            entrar = st.form_submit_button("Entrar", type="primary", use_container_width=True)
        if entrar:
            dados = _dados_acesso_usuario(usuario, senha)
            if dados["autorizado"]:
                st.session_state["_usuario_autenticado"] = dados["usuario"]
                st.session_state["_usuario_nome_exibicao"] = dados["nome"]
                st.session_state["_usuario_perfil"] = dados["perfil"]
                st.session_state["_usuario_comprador"] = dados.get("comprador", "")
                st.session_state["_usuario_vendedor"] = dados.get("vendedor", "")
                st.session_state["_usuario_gerente"] = dados.get("gerente", "")
                st.rerun()
            else:
                st.error("Usuário ou senha inválidos.")
    st.stop()


def _renderizar_usuario_logado():
    nome = st.session_state.get("_usuario_nome_exibicao", "")
    perfil = st.session_state.get("_usuario_perfil", "")
    if not nome:
        return
    st.sidebar.markdown("---")
    st.sidebar.caption("USUÁRIO CONECTADO")
    st.sidebar.markdown(f"**👤 {nome}**")
    st.sidebar.caption(f"Perfil: {perfil}")
    if st.sidebar.button("↩️ Sair", use_container_width=True, key="logout_rede"):
        for chave in ["_usuario_autenticado","_usuario_nome_exibicao","_usuario_perfil","_usuario_comprador","_usuario_vendedor","_usuario_gerente"]:
            st.session_state.pop(chave, None)
        st.rerun()


def _renderizar_gestao_acessos_compradores():
    if st.session_state.get("_usuario_perfil") != "Administrador":
        return
    st.markdown("### 🔐 Acessos dos Compradores")
    st.caption("Gerencie usuário, senha e status de acesso dos compradores. Senhas são armazenadas somente por hash.")
    try:
        compradores = sorted(lista_compradores_ativos(), key=str.casefold)
    except Exception:
        compradores = []
    acessos = _carregar_acessos_compradores()
    if acessos:
        tabela = pd.DataFrame([{"Comprador":x.get("comprador",""),"Usuário":x.get("usuario",""),"Status":"Ativo" if x.get("ativo",True) else "Bloqueado"} for x in acessos])
        dataframe_br(tabela, use_container_width=True, hide_index=True)
    if not compradores:
        st.info("Cadastre um comprador ativo antes de criar o acesso.")
        return
    comprador_sel = st.selectbox("Comprador para acesso", compradores, key="acesso_comprador_sel")
    atual = next((x for x in acessos if str(x.get("comprador","")).casefold()==str(comprador_sel).casefold()), {})
    c1,c2 = st.columns(2)
    with c1:
        usuario = st.text_input("Usuário de acesso", value=str(atual.get("usuario", "") or _normalizar_usuario_acesso(comprador_sel)), key="acesso_usuario")
    with c2:
        status = st.selectbox("Status", ["Ativo","Bloqueado"], index=0 if atual.get("ativo",True) else 1, key="acesso_status")
    senha = st.text_input("Nova senha", type="password", placeholder="Deixe em branco para manter a senha atual", key="acesso_senha")
    if st.button("💾 Salvar acesso do comprador", type="primary", use_container_width=True, key="salvar_acesso_comprador"):
        usuario_norm = _normalizar_usuario_acesso(usuario)
        if not usuario_norm:
            st.error("Informe um usuário válido.")
        elif not atual and not senha:
            st.error("Informe uma senha para o novo acesso.")
        elif senha and len(str(senha)) < 6:
            st.error("A senha deve ter pelo menos 6 caracteres.")
        elif usuario_norm in _USUARIOS_SISTEMA or usuario_norm in _USUARIOS_GERENTES_SISTEMA:
            st.error("Esse usuário pertence a uma conta interna do sistema.")
        else:
            conflito = next((x for x in acessos if _normalizar_usuario_acesso(x.get("usuario"))==usuario_norm and str(x.get("comprador","")).casefold()!=str(comprador_sel).casefold()), None)
            if conflito:
                st.error("Esse usuário já pertence a outro comprador.")
            else:
                registro = dict(atual) if atual else {}
                registro.update({"comprador": comprador_sel, "usuario": usuario_norm, "ativo": status=="Ativo", "atualizado_em": datetime.now().strftime("%d/%m/%Y %H:%M:%S")})
                if senha:
                    registro["senha_hash"] = _hash_senha_acesso(senha)
                novos = [x for x in acessos if str(x.get("comprador","")).casefold()!=str(comprador_sel).casefold()]
                novos.append(registro)
                _salvar_acessos_compradores(novos)
                st.success("Acesso atualizado.")
                st.rerun()

def _perfil_logado():
    return str(st.session_state.get("_usuario_perfil", "Administrador") or "Administrador").strip()

def _escopo_usuario_logado():
    perfil = _perfil_logado()
    if perfil == "Comprador":
        return str(st.session_state.get("_usuario_comprador", "") or "").strip()
    if perfil == "Vendedor":
        return str(st.session_state.get("_usuario_vendedor", "") or "").strip()
    if perfil == "Gerente":
        return str(st.session_state.get("_usuario_gerente", "") or "").strip()
    return ""

def _norm_escopo(valor):
    texto = unicodedata.normalize("NFKD", str(valor or "").strip().casefold())
    return "".join(c for c in texto if not unicodedata.combining(c))

def _filtrar_df_por_usuario_logado(df):
    if not isinstance(df, pd.DataFrame):
        return df
    perfil = _perfil_logado()
    if perfil == "Administrador":
        return df

    escopo = _escopo_usuario_logado()
    if not escopo:
        return df.iloc[0:0].copy()

    if perfil == "Comprador":
        candidatos = [
            "Comprador", "comprador", "Comprador Responsável",
            "comprador_responsavel", "Nome Comprador", "nome_comprador"
        ]
    elif perfil == "Gerente":
        candidatos = [
            "Gerente", "gerente", "Gerente Comercial", "gerente_comercial",
            "Nome Gerente", "nome_gerente"
        ]
    else:
        candidatos = [
            "Vendedor", "vendedor", "Nome Vendedor", "nome_vendedor",
            "Vendedor Responsável", "vendedor_responsavel"
        ]
    coluna = next((c for c in candidatos if c in df.columns), None)
    if coluna is None:
        return df.iloc[0:0].copy()

    alvo = _norm_escopo(escopo)
    return df.loc[df[coluna].map(_norm_escopo).eq(alvo)].copy()

def _filtrar_objeto_por_usuario_logado(obj):
    if isinstance(obj, pd.DataFrame):
        return _filtrar_df_por_usuario_logado(obj)
    if isinstance(obj, dict):
        if _perfil_logado() == "Administrador":
            return obj
        alvo = _norm_escopo(_escopo_usuario_logado())
        return {k: v for k, v in obj.items() if _norm_escopo(k) == alvo}
    return obj

def _filtrar_gerente_privado(df, coluna=None):
    """Filtra DataFrame estritamente para o gerente autenticado.

    Em perfil Gerente, nunca retorna dados de outro gerente. Para Administrador,
    preserva o DataFrame recebido.
    """
    if not isinstance(df, pd.DataFrame):
        return df
    if _perfil_logado() != "Gerente":
        return df
    gerente = _escopo_usuario_logado()
    if not gerente:
        return df.iloc[0:0].copy()
    candidatos = ([coluna] if coluna else []) + [
        "Gerente Comercial", "Gerente", "gerente", "gerente_comercial",
        "Nome Gerente", "nome_gerente"
    ]
    col = next((c for c in candidatos if c and c in df.columns), None)
    if col is None:
        return df.iloc[0:0].copy()
    alvo = _norm_escopo(gerente)
    return df.loc[df[col].map(_norm_escopo).eq(alvo)].copy()


def _menu_permitido_por_perfil():
    perfil = _perfil_logado()
    if perfil == "Comprador":
        return ["📌 Meu Resumo","📊 Realizados","🎯 Métricas Destaque","📈 Resultado Métricas","📋 Resultados dos KPI's","🏆 Prêmio Comprador","💰 Prêmio por KPI","🌟 Portal de Premiação","🧾 Holerite da Premiação"]
    if perfil == "Vendedor":
        return ["📊 Realizados","🌟 Portal de Premiação","🧾 Holerite da Premiação"]
    if perfil == "Gerente":
        # Perfil gerencial estritamente privado: somente visões do próprio gerente.
        # Não exibe portais/rankings consolidados da empresa.
        return ["📌 Meu Resumo","👔 Holerite do Gerente Comercial"]
    return None



# ============================================================
# AJUSTE GLOBAL — VALORES COMPLETOS NOS CARDS
# ============================================================
st.markdown(
    """
    <style id="eirox-card-values-full">
    /* Streamlit metric: nunca cortar com ... */
    [data-testid="stMetricValue"],
    [data-testid="stMetricValue"] > div,
    [data-testid="stMetricValue"] p,
    [data-testid="stMetricValue"] span {
        white-space: nowrap !important;
        overflow: visible !important;
        text-overflow: clip !important;
        max-width: none !important;
    }

    [data-testid="stMetricValue"] {
        font-size: clamp(1.18rem, 1.62vw, 2rem) !important;
        line-height: 1.12 !important;
        letter-spacing: -0.035em !important;
    }

    [data-testid="stMetricValue"] > div,
    [data-testid="stMetricValue"] p {
        font-size: inherit !important;
        line-height: inherit !important;
        letter-spacing: inherit !important;
    }

    /* Cards customizados usados no projeto */
    .hp-card-value,
    .rec-card-value,
    .premium-box .value,
    .status-card-kpi .value,
    .meta-card .value,
    .kpi-card .value,
    .metric-value,
    .card-value {
        white-space: nowrap !important;
        overflow: visible !important;
        text-overflow: clip !important;
        max-width: 100% !important;
        font-size: clamp(1.15rem, 1.55vw, 1.95rem) !important;
        line-height: 1.12 !important;
        letter-spacing: -0.035em !important;
    }

    /* Evita que o container interno aplique ellipsis */
    [data-testid="stMetric"] {
        overflow: visible !important;
        min-width: 0 !important;
    }

    [data-testid="stMetric"] > div {
        overflow: visible !important;
        min-width: 0 !important;
    }

    /* Em notebooks / resoluções médias, reduz antes de cortar */
    @media (max-width: 1500px) {
        [data-testid="stMetricValue"] {
            font-size: clamp(1.05rem, 1.45vw, 1.72rem) !important;
        }
        .hp-card-value,
        .rec-card-value,
        .premium-box .value,
        .status-card-kpi .value,
        .meta-card .value,
        .kpi-card .value,
        .metric-value,
        .card-value {
            font-size: clamp(1.02rem, 1.40vw, 1.68rem) !important;
        }
    }

    @media (max-width: 1180px) {
        [data-testid="stMetricValue"] {
            font-size: 1.08rem !important;
        }
        .hp-card-value,
        .rec-card-value,
        .premium-box .value,
        .status-card-kpi .value,
        .meta-card .value,
        .kpi-card .value,
        .metric-value,
        .card-value {
            font-size: 1.04rem !important;
        }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

_renderizar_login_sistema()


def _revalidar_perfil_conta_interna():
    """Impede que contas gerenciais antigas permaneçam como Administrador na sessão."""
    usuario = _normalizar_usuario_acesso(st.session_state.get("_usuario_autenticado", ""))
    gerente_sistema = _USUARIOS_GERENTES_SISTEMA.get(usuario)
    if gerente_sistema:
        gerente_nome = str(gerente_sistema.get("gerente", "") or "").strip()
        st.session_state["_usuario_nome_exibicao"] = str(
            gerente_sistema.get("nome", "") or gerente_nome or usuario.title()
        )
        st.session_state["_usuario_perfil"] = "Gerente"
        st.session_state["_usuario_comprador"] = ""
        st.session_state["_usuario_vendedor"] = ""
        st.session_state["_usuario_gerente"] = gerente_nome


_revalidar_perfil_conta_interna()

_perfil_validado_seguro = _perfil_logado()
if _perfil_validado_seguro not in {"Administrador", "Comprador", "Vendedor", "Gerente"}:
    st.error("Perfil de acesso inválido.")
    st.stop()


st.markdown(
    """
    <style>
    .valor-negativo {
        color: #ff4d4f !important;
        font-weight: 700 !important;
    }
    .status-card-kpi {
        border: 1px solid #254b6b;
        border-radius: 14px;
        padding: 14px 16px;
        min-height: 108px;
        background: #0d2032;
    }
    .status-card-kpi .titulo {
        font-size: 13px;
        font-weight: 700;
        color: #ffffff;
    }
    .status-card-kpi .valor {
        font-size: 29px;
        line-height: 1.2;
        margin-top: 8px;
        color: #ffffff;
        font-variant-numeric: tabular-nums;
    }
    .status-card-kpi .status {
        display: inline-block;
        margin-top: 8px;
        padding: 3px 8px;
        border-radius: 999px;
        background: rgba(46, 204, 113, .18);
        color: #2ecc71;
        font-size: 12px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# =========================================================
# VISUAL PREMIUM — HOLERITES DE PREMIAÇÃO
# =========================================================
st.markdown(
    """
    <style>
    .hp-title-wrap{display:flex;justify-content:space-between;align-items:flex-start;gap:16px;margin:2px 0 14px}
    .hp-title{font-size:30px;font-weight:850;line-height:1.05;color:#fff;margin:0}.hp-subtitle{color:#a9b9c8;font-size:14px;margin-top:6px}.hp-update{color:#a9b9c8;font-size:12px;white-space:nowrap;padding-top:6px}
    .hp-card{position:relative;overflow:hidden;background:linear-gradient(145deg,#10263a,#0b1c2d);border:1px solid #284a65;border-radius:15px;padding:15px 16px;min-height:118px;box-shadow:0 8px 22px rgba(0,0,0,.18)}
    .hp-card-top{display:flex;gap:12px;align-items:center}.hp-icon{width:42px;height:42px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:21px;font-weight:800;flex:0 0 42px}.hp-card-label{font-size:14px;font-weight:750;color:#fff}.hp-card-value{font-size:27px;font-weight:850;color:#fff;margin-top:6px;line-height:1.05;font-variant-numeric:tabular-nums}.hp-card-note{font-size:12px;color:#b6c5d2;margin-top:8px}
    .hp-red{color:#ff4d4f!important}.hp-green{color:#20c66b!important}.hp-gold{color:#f7b500!important}.hp-blue-icon{background:#2868d7}.hp-green-icon{background:#159653}.hp-purple-icon{background:#6a35b5}.hp-red-icon{background:#df2929}.hp-gold-icon{background:#cc8a00}
    .hp-section-title{font-size:24px;font-weight:850;color:#fff;margin:20px 0 12px}.hp-panel{background:linear-gradient(145deg,#0d2032,#091725);border:1px solid #29475e;border-radius:15px;padding:14px 15px;height:100%}.hp-panel-title{font-size:16px;font-weight:800;color:#fff;margin-bottom:12px}
    .hp-table-wrap{overflow:auto;border:1px solid #29475e;border-radius:14px;background:#07131f}table.hp-table{border-collapse:collapse;width:100%;min-width:980px;color:#fff;font-size:13px}.hp-table th{background:#1a202d;color:#c7d2dc;text-align:left;padding:11px 10px;border-right:1px solid #304153;border-bottom:1px solid #304153;font-weight:600;white-space:nowrap}.hp-table td{padding:11px 10px;border-right:1px solid #213445;border-bottom:1px solid #213445;white-space:nowrap;font-variant-numeric:tabular-nums}.hp-table tr:last-child td{font-weight:800;background:#0d2435}
    .hp-progress{height:9px;background:#213443;border-radius:99px;overflow:hidden;display:inline-block;width:135px;vertical-align:middle;margin-left:8px}.hp-progress>span{display:block;height:100%;border-radius:99px}.hp-badge{display:inline-flex;align-items:center;gap:6px;padding:4px 8px;border-radius:99px;font-weight:700;font-size:12px}.hp-badge-green{background:rgba(25,185,100,.16);color:#24d178}.hp-badge-yellow{background:rgba(247,181,0,.15);color:#ffc21c}.hp-badge-red{background:rgba(255,77,79,.15);color:#ff5c5f}.hp-footnote{text-align:center;color:#aab8c5;font-size:12px;margin:14px 0 6px}
    @media(max-width:1100px){.hp-title{font-size:25px}.hp-card-value{font-size:22px}.hp-update{display:none}}
    </style>
    """,
    unsafe_allow_html=True,
)



# =========================================================
# VISUAL PREMIUM GLOBAL — TODAS AS TELAS
# =========================================================
st.markdown(
    """
    <style>
    :root{
      --re-bg:#06111e;--re-bg2:#081827;--re-panel:#0d2032;--re-panel2:#10263a;
      --re-border:#294b65;--re-border-soft:#1e3a50;--re-text:#f7fbff;
      --re-muted:#a9bac8;--re-cyan:#43d7e8;--re-blue:#2e7bdc;--re-green:#20c66b;
      --re-yellow:#f7b500;--re-red:#ff4d4f;--re-radius:14px;
    }
    html,body,[data-testid="stAppViewContainer"],.stApp{
      background:radial-gradient(circle at 92% 4%,rgba(22,88,124,.18),transparent 25%),linear-gradient(180deg,#06111e 0%,#071522 100%)!important;
      color:var(--re-text)!important;
    }
    [data-testid="stHeader"]{background:rgba(5,15,25,.90)!important;border-bottom:1px solid rgba(67,215,232,.12)!important;backdrop-filter:blur(10px)}
    [data-testid="stMainBlockContainer"]{max-width:1680px!important;padding-top:1.2rem!important;padding-bottom:3rem!important}

    /* Sidebar */
    [data-testid="stSidebar"]{background:linear-gradient(180deg,#071727 0%,#091b2c 100%)!important;border-right:1px solid #203e54!important}
    [data-testid="stSidebarContent"]{padding-top:.6rem!important}
    [data-testid="stSidebar"] hr{border-color:#27445a!important}
    [data-testid="stSidebar"] label,[data-testid="stSidebar"] p{color:#d6e1ea!important}
    [data-testid="stSidebar"] [role="radiogroup"] label{
      border-radius:10px!important;padding:7px 9px!important;margin:1px 0!important;transition:.18s ease!important;
    }
    [data-testid="stSidebar"] [role="radiogroup"] label:hover{background:#102b42!important}
    [data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked){background:linear-gradient(90deg,#1b4f83,#183d63)!important;box-shadow:inset 3px 0 0 var(--re-cyan)!important}

    /* Titles and section rhythm */
    h1,h2,h3,h4{color:#fff!important;letter-spacing:-.02em!important}
    h1{font-size:2rem!important;font-weight:850!important;border-bottom:1px solid #24445b;padding-bottom:.7rem;margin-bottom:1rem!important}
    h2{font-size:1.55rem!important;font-weight:820!important;margin-top:1.25rem!important}
    h3{font-size:1.2rem!important;font-weight:800!important}
    .section-title{border-radius:12px!important;border:1px solid #34536a!important;box-shadow:0 8px 24px rgba(0,0,0,.16)!important}

    /* Metrics/cards */
    [data-testid="stMetric"]{
      background:linear-gradient(145deg,var(--re-panel2),var(--re-panel))!important;
      border:1px solid var(--re-border)!important;border-radius:var(--re-radius)!important;
      padding:15px 16px!important;min-height:112px!important;box-shadow:0 8px 24px rgba(0,0,0,.18)!important;
    }
    [data-testid="stMetricLabel"]{color:#c8d5df!important;font-weight:700!important}
    [data-testid="stMetricValue"]{color:#fff!important;font-weight:850!important;font-variant-numeric:tabular-nums!important}
    [data-testid="stMetricDelta"]{font-weight:750!important}
    .meta-card,.status-card-kpi{background:linear-gradient(145deg,var(--re-panel2),var(--re-panel))!important;border-color:var(--re-border)!important;box-shadow:0 8px 24px rgba(0,0,0,.17)!important}

    /* Inputs */
    [data-baseweb="select"]>div,[data-testid="stTextInput"] input,[data-testid="stNumberInput"] input,[data-testid="stDateInput"] input,[data-testid="stTextArea"] textarea{
      background:#081827!important;color:#fff!important;border:1px solid #31546e!important;border-radius:9px!important;
    }
    [data-baseweb="select"]>div:hover,[data-testid="stTextInput"] input:focus,[data-testid="stNumberInput"] input:focus,[data-testid="stDateInput"] input:focus,[data-testid="stTextArea"] textarea:focus{
      border-color:var(--re-cyan)!important;box-shadow:0 0 0 1px rgba(67,215,232,.25)!important;
    }
    [data-testid="stMultiSelect"] span{border-radius:8px!important}

    /* Buttons */
    .stButton>button,.stDownloadButton>button{
      border-radius:9px!important;border:1px solid #2c668b!important;background:linear-gradient(180deg,#12324b,#0d2437)!important;color:#fff!important;font-weight:750!important;min-height:40px!important;transition:.18s ease!important;
    }
    .stButton>button:hover,.stDownloadButton>button:hover{border-color:var(--re-cyan)!important;background:linear-gradient(180deg,#174565,#12344e)!important;transform:translateY(-1px);box-shadow:0 7px 18px rgba(0,0,0,.22)!important}
    .stButton>button[kind="primary"]{background:linear-gradient(90deg,#126cba,#1685cf)!important;border-color:#36b8e6!important}

    /* Tabs */
    [data-baseweb="tab-list"]{gap:6px!important;border-bottom:1px solid #28475d!important}
    [data-baseweb="tab"]{background:#0b1c2b!important;border:1px solid #28475d!important;border-radius:9px 9px 0 0!important;padding:9px 15px!important;color:#b9c8d4!important}
    [data-baseweb="tab"][aria-selected="true"]{background:#12324b!important;color:#fff!important;border-bottom-color:var(--re-cyan)!important;box-shadow:inset 0 -3px 0 var(--re-cyan)!important}

    /* Dataframes and native tables */
    [data-testid="stDataFrame"], [data-testid="stTable"]{
      border:1px solid var(--re-border)!important;border-radius:13px!important;overflow:hidden!important;background:#07131f!important;box-shadow:0 7px 22px rgba(0,0,0,.14)!important;
    }
    [data-testid="stDataFrame"] [role="columnheader"]{background:#1a202d!important;color:#cbd8e2!important;font-weight:700!important}
    [data-testid="stDataFrame"] [role="gridcell"]{border-color:#21394c!important;font-variant-numeric:tabular-nums!important}
    table{border-collapse:separate!important;border-spacing:0!important}
    table thead th{background:#1a202d!important;color:#cbd8e2!important;border-color:#304153!important}
    table tbody td{border-color:#213445!important;font-variant-numeric:tabular-nums!important}

    /* Plotly and containers */
    [data-testid="stPlotlyChart"]{background:linear-gradient(145deg,#0c1e2e,#081724)!important;border:1px solid var(--re-border-soft)!important;border-radius:14px!important;padding:8px!important;box-shadow:0 8px 22px rgba(0,0,0,.14)!important}
    [data-testid="stVerticalBlockBorderWrapper"]{border-radius:14px!important;border-color:var(--re-border)!important;background:rgba(10,28,43,.55)!important}

    /* Expanders */
    [data-testid="stExpander"]{border:1px solid var(--re-border)!important;border-radius:11px!important;background:#0a1b2a!important;overflow:hidden!important}
    [data-testid="stExpander"] summary{font-weight:750!important;color:#fff!important;background:#0d2234!important}

    /* Alerts */
    [data-testid="stAlert"]{border-radius:11px!important;border:1px solid rgba(255,255,255,.10)!important;box-shadow:0 6px 18px rgba(0,0,0,.12)!important}

    /* Progress */
    [data-testid="stProgress"]>div>div{background:linear-gradient(90deg,#22b5d2,#36d1dc)!important;border-radius:99px!important}
    [data-testid="stProgress"]>div{background:#1a3446!important;border-radius:99px!important}

    /* Dividers, captions and links */
    hr{border-color:#29475e!important}
    .stCaptionContainer,p,small{color:#b5c3cf}
    a{color:#57ddec!important}

    /* Universal status and negative presentation */
    .valor-negativo,.negative-value{color:var(--re-red)!important;font-weight:800!important}
    .positive-value{color:var(--re-green)!important;font-weight:800!important}

    /* Scrollbars */
    *{scrollbar-width:thin;scrollbar-color:#31536c #07131f}
    *::-webkit-scrollbar{width:9px;height:9px}*::-webkit-scrollbar-track{background:#07131f}*::-webkit-scrollbar-thumb{background:#31536c;border-radius:99px}

    @media(max-width:1000px){[data-testid="stMainBlockContainer"]{padding-left:.8rem!important;padding-right:.8rem!important}h1{font-size:1.65rem!important}}
    </style>
    """,
    unsafe_allow_html=True,
)

# =========================================================
# ENTERPRISE PRO — ACABAMENTO GLOBAL E EXPERIÊNCIA
# =========================================================
st.markdown(
    """
    <style>
    /* Tipografia, ritmo e legibilidade */
    html, body, [class*="css"] {
        font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont,
                     "Segoe UI", Arial, sans-serif !important;
    }
    [data-testid="stMainBlockContainer"] {
        max-width: 1720px !important;
        padding: 1.05rem 1.45rem 3.5rem !important;
    }
    p, label, .stCaptionContainer { line-height: 1.45 !important; }

    /* Cabeçalho da aplicação */
    [data-testid="stHeader"] {
        height: 3rem !important;
        background: rgba(5, 15, 25, .82) !important;
        border-bottom: 1px solid rgba(67, 215, 232, .13) !important;
        backdrop-filter: blur(16px) saturate(135%) !important;
    }

    /* Sidebar mais organizada */
    [data-testid="stSidebar"] { box-shadow: 12px 0 34px rgba(0,0,0,.18) !important; }
    [data-testid="stSidebarContent"] { padding: .7rem .75rem 1.5rem !important; }
    [data-testid="stSidebar"] [role="radiogroup"] { gap: 2px !important; }
    [data-testid="stSidebar"] [role="radiogroup"] label {
        min-height: 37px !important;
        border: 1px solid transparent !important;
        font-size: .91rem !important;
        transition: background .16s ease, border-color .16s ease, transform .16s ease !important;
    }
    [data-testid="stSidebar"] [role="radiogroup"] label:hover {
        border-color: rgba(67,215,232,.22) !important;
        transform: translateX(2px);
    }
    [data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked) {
        border-color: rgba(67,215,232,.38) !important;
        box-shadow: inset 3px 0 0 #43d7e8, 0 5px 16px rgba(0,0,0,.14) !important;
    }

    /* Cards consistentes */
    [data-testid="stMetric"], .meta-card, .status-card-kpi, .hp-card, .premium-box {
        transition: transform .16s ease, border-color .16s ease, box-shadow .16s ease !important;
    }
    [data-testid="stMetric"]:hover, .meta-card:hover, .status-card-kpi:hover, .hp-card:hover {
        transform: translateY(-2px);
        border-color: rgba(67,215,232,.48) !important;
        box-shadow: 0 14px 34px rgba(0,0,0,.26) !important;
    }
    [data-testid="stMetricLabel"] p {
        text-transform: uppercase !important;
        letter-spacing: .055em !important;
        font-size: .74rem !important;
    }
    [data-testid="stMetricValue"] { letter-spacing: -.025em !important; }

    /* Formulários e filtros */
    [data-baseweb="select"] > div,
    [data-testid="stTextInput"] input,
    [data-testid="stNumberInput"] input,
    [data-testid="stDateInput"] input,
    [data-testid="stTextArea"] textarea {
        min-height: 42px !important;
        box-shadow: inset 0 1px 0 rgba(255,255,255,.025) !important;
    }
    [data-testid="stWidgetLabel"] p { font-weight: 700 !important; color: #d5e1ea !important; }

    /* Botões com hierarquia clara */
    .stButton > button, .stDownloadButton > button {
        min-height: 42px !important;
        letter-spacing: .01em !important;
        box-shadow: 0 5px 14px rgba(0,0,0,.14) !important;
    }
    .stButton > button:active, .stDownloadButton > button:active { transform: translateY(0) !important; }
    button:focus-visible, input:focus-visible, textarea:focus-visible {
        outline: 2px solid #43d7e8 !important; outline-offset: 2px !important;
    }

    /* Tabelas profissionais */
    [data-testid="stDataFrame"] {
        border-radius: 14px !important;
        box-shadow: 0 10px 28px rgba(0,0,0,.18) !important;
    }
    [data-testid="stDataFrame"] [role="columnheader"] {
        text-transform: uppercase !important;
        letter-spacing: .035em !important;
        font-size: .75rem !important;
    }
    [data-testid="stDataFrame"] [role="gridcell"] { font-size: .86rem !important; }

    /* Gráficos */
    [data-testid="stPlotlyChart"] {
        padding: 10px 12px 6px !important;
        box-shadow: 0 10px 28px rgba(0,0,0,.16) !important;
    }

    /* Tabs */
    [data-baseweb="tab"] { min-height: 42px !important; font-weight: 750 !important; }

    /* Alertas e expansores */
    [data-testid="stAlert"] { border-left-width: 4px !important; }
    [data-testid="stExpander"] summary { min-height: 44px !important; }

    /* Rodapé corporativo */
    .eirox-footer {
        margin: 30px 0 4px !important;
        padding: 16px 12px !important;
        border-top: 1px solid #1d3a50 !important;
        color: #7f98ac !important;
        font-size: 11px !important;
        text-transform: uppercase;
        letter-spacing: .10em !important;
    }

    /* Responsividade */
    @media (max-width: 900px) {
        [data-testid="stMainBlockContainer"] { padding: .7rem .7rem 2.5rem !important; }
        [data-testid="stHorizontalBlock"] { gap: .55rem !important; }
        .hp-title-wrap, .brand-row, .premium-box { flex-direction: column !important; align-items: flex-start !important; }
        .hp-card-value { font-size: 21px !important; }
    }

    /* Impressão: neutraliza efeitos interativos */
    @media print {
        * { animation: none !important; transition: none !important; }
        [data-testid="stMetric"], .meta-card, .status-card-kpi, .hp-card { transform: none !important; box-shadow: none !important; }
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# =========================================================
# RESUMO EXECUTIVO CORPORATIVO — ANÁLISE COMERCIAL
# =========================================================
st.markdown(
    """
    <style>
    .rec-wrap{margin:20px 0 14px}
    .rec-hero{position:relative;overflow:hidden;border:1px solid #2d536d;border-radius:18px;padding:22px 24px;background:linear-gradient(135deg,#102b42 0%,#0a1d2d 58%,#071522 100%);box-shadow:0 16px 38px rgba(0,0,0,.24)}
    .rec-hero:after{content:"";position:absolute;right:-70px;top:-90px;width:260px;height:260px;border-radius:50%;background:radial-gradient(circle,rgba(67,215,232,.22),transparent 68%)}
    .rec-kicker{color:#55ddeb;font-size:12px;font-weight:850;letter-spacing:.14em;text-transform:uppercase;margin-bottom:8px}
    .rec-title{font-size:29px;line-height:1.08;font-weight:900;color:#fff;letter-spacing:-.025em;margin:0}
    .rec-subtitle{font-size:14px;color:#b7c8d5;margin-top:8px;max-width:920px}
    .rec-meta{display:flex;flex-wrap:wrap;gap:8px;margin-top:16px}
    .rec-chip{display:inline-flex;align-items:center;gap:7px;padding:6px 10px;border-radius:999px;border:1px solid #365d76;background:rgba(6,17,30,.58);color:#dceaf3;font-size:12px;font-weight:700}
    .rec-grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px;margin:14px 0 18px}
    .rec-card{position:relative;overflow:hidden;border:1px solid #294d66;border-radius:15px;background:linear-gradient(145deg,#10263a,#0b1d2d);padding:15px 16px;min-height:122px;box-shadow:0 9px 24px rgba(0,0,0,.18)}
    .rec-card:before{content:"";position:absolute;left:0;top:0;bottom:0;width:4px;background:var(--rec-accent,#43d7e8)}
    .rec-card-label{color:#aebfcb;font-size:12px;font-weight:800;text-transform:uppercase;letter-spacing:.06em}
    .rec-card-value{color:#fff;font-size:25px;font-weight:900;margin-top:8px;line-height:1.1;font-variant-numeric:tabular-nums}
    .rec-card-note{color:#b9c8d3;font-size:12px;margin-top:9px}
    .rec-card.good{--rec-accent:#20c66b}.rec-card.warn{--rec-accent:#f7b500}.rec-card.bad{--rec-accent:#ff4d4f}.rec-card.info{--rec-accent:#43d7e8}
    .rec-score{display:flex;align-items:center;gap:9px}.rec-score-dot{width:11px;height:11px;border-radius:50%;background:var(--rec-accent,#43d7e8);box-shadow:0 0 0 5px rgba(67,215,232,.16)}
    .rec-report{border:1px solid #2b4d65;border-radius:17px;background:linear-gradient(180deg,#0b1e2f,#081724);padding:20px 22px;margin:8px 0 18px;box-shadow:0 12px 30px rgba(0,0,0,.18)}
    .rec-report h2{font-size:18px!important;text-transform:uppercase;letter-spacing:.045em;color:#55ddeb!important;border-bottom:1px solid #28475d;padding-bottom:9px;margin-top:22px!important;margin-bottom:12px!important}
    .rec-report h2:first-child{margin-top:0!important}
    .rec-report h3{font-size:17px!important;color:#fff!important}
    .rec-report p,.rec-report li{font-size:14px;line-height:1.72;color:#d3dee7!important}
    .rec-report strong{color:#fff}
    .rec-report ul{padding-left:20px}
    .rec-report li{margin:6px 0}
    .rec-actions{border:1px solid #284a62;border-radius:14px;background:#0a1c2b;padding:14px 15px;margin-top:12px}
    .rec-section-label{font-size:13px;font-weight:850;color:#55ddeb;text-transform:uppercase;letter-spacing:.08em;margin-bottom:9px}
    .rec-insight-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px;margin:12px 0 18px}
    .rec-insight{border:1px solid #294d66;border-radius:15px;background:linear-gradient(145deg,#0d2234,#081724);padding:16px 17px;min-height:150px}
    .rec-insight.good{border-left:4px solid #20c66b}.rec-insight.warn{border-left:4px solid #f7b500}
    .rec-insight-title{font-size:14px;font-weight:900;color:#fff;text-transform:uppercase;letter-spacing:.055em;margin-bottom:10px}
    .rec-insight ul{margin:0;padding-left:18px}.rec-insight li{color:#d4e0e8;font-size:13px;line-height:1.55;margin:5px 0}
    .rec-diagnosis{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:10px;margin:12px 0 18px}
    .rec-diagnosis-item{border:1px solid #294d66;border-radius:13px;background:#0a1c2b;padding:13px 14px}
    .rec-diagnosis-name{color:#aebfcb;font-size:12px;font-weight:800;text-transform:uppercase;letter-spacing:.05em}
    .rec-diagnosis-status{font-size:17px;color:#fff;font-weight:900;margin-top:6px}
    .rec-diagnosis-note{color:#9eb0bd;font-size:11px;margin-top:5px}
    .rec-priority-table{width:100%;border-collapse:separate!important;border-spacing:0;border:1px solid #294d66;border-radius:14px;overflow:hidden;margin:10px 0 18px}
    .rec-priority-table th{background:#152a3b!important;color:#d7e5ee!important;padding:10px!important;font-size:12px;text-transform:uppercase;letter-spacing:.04em}
    .rec-priority-table td{background:#091a29;color:#d4e0e8;padding:11px 10px!important;border-top:1px solid #21394c;font-size:13px;vertical-align:top}
    .rec-projection{border:1px solid #315b75;border-radius:15px;background:linear-gradient(135deg,rgba(31,91,126,.28),rgba(9,26,41,.88));padding:16px 18px;margin:12px 0 18px}
    .rec-projection-value{font-size:22px;font-weight:900;color:#fff;margin-top:4px}
    @media(max-width:1100px){.rec-grid{grid-template-columns:repeat(2,minmax(0,1fr))}.rec-diagnosis{grid-template-columns:repeat(2,minmax(0,1fr))}}
    @media(max-width:650px){.rec-grid,.rec-insight-grid,.rec-diagnosis{grid-template-columns:1fr}.rec-hero{padding:18px}.rec-title{font-size:23px}.rec-card-value{font-size:22px}}
    @media print{.rec-hero,.rec-card,.rec-report{break-inside:avoid!important;page-break-inside:avoid!important}.rec-grid{grid-template-columns:repeat(4,1fr)!important}.rec-actions{display:none!important}}
    </style>
    """,
    unsafe_allow_html=True,
)

# =========================================================
# PERFORMANCE / CACHE
# =========================================================

def _arquivo_token(*caminhos):
    """Token leve para invalidar caches quando arquivos persistentes mudarem."""
    partes = []
    for caminho in caminhos:
        try:
            path = Path(caminho)
            stat = path.stat()
            partes.append(f"{path}:{stat.st_mtime_ns}:{stat.st_size}")
        except Exception:
            partes.append(f"{caminho}:0:0")
    return "|".join(partes)

def _limpar_cache_dados():
    try:
        st.cache_data.clear()
    except Exception:
        pass


def _registrar_atualizacao_dados(fonte="Banco de Dados", periodo="", registros=0):
    """Invalida todas as visões após qualquer alteração persistente.

    O marcador também permite que os cards sejam recalculados mesmo quando
    o SQLite estiver usando WAL e o arquivo principal não mudar de data.
    """
    marcador = Path("data") / "ultima_atualizacao_dados.json"
    marcador.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "fonte": str(fonte or "Banco de Dados"),
        "periodo": str(periodo or ""),
        "registros": int(registros or 0),
        "atualizado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f"),
        "nonce": time.time_ns(),
    }
    try:
        marcador.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception:
        pass

    _limpar_cache_dados()
    try:
        st.cache_resource.clear()
    except Exception:
        pass

    # Remove qualquer visão, agregado ou exportação construído antes da carga.
    # A limpeza por prefixo protege também telas novas que forem adicionadas depois.
    prefixos_invalidar = (
        "_dados_", "_chave_", "_cache_", "_visao_", "_resumo_",
        "_dashboard_", "_holerite_", "_premiacao_", "_export_",
    )
    chaves_fixas = {
        "_dados_visoes", "_chave_visoes", "_export_pdf_bytes",
        "_export_excel_bytes", "_export_nome_base",
    }
    for chave in list(st.session_state.keys()):
        if chave in chaves_fixas or str(chave).startswith(prefixos_invalidar):
            st.session_state.pop(chave, None)

    st.session_state["_ultima_atualizacao_dados"] = payload
    st.session_state["_forcar_recalculo_visoes"] = payload["nonce"]
    return payload


def _mensagem_atualizacao_pendente():
    mensagem = st.session_state.pop("_flash_atualizacao_dados", None)
    if mensagem:
        st.success(mensagem)


# =========================================================
# DADOS DEMONSTRATIVOS
# =========================================================

COMPRADORES = ["Paulo", "Francieli", "Sebastião"]

REALIZADOS = pd.DataFrame([
    ["Paulo", 1083974.52, 25.7, 961221.38, 88.7, 856101.24, 1.00, 69104.34, 8.1, 269908.06, 31.5, 460514.54, 53.8, 56574.30, 6.6, 13281.02, 1.2, 1003613.79, 104.4],
    ["Francieli", 1037338.03, 24.6, 708983.83, 68.3, 1155715.96, 2.00, 104317.28, 9.0, 278891.20, 24.1, 595751.47, 51.5, 176756.01, 15.3, 32132.06, 3.1, 807929.11, 114.0],
    ["Sebastião", 2102632.93, 49.8, 1239845.83, 59.0, 1951037.63, 2.00, 687151.14, 35.2, 661165.12, 33.9, 508011.09, 26.0, 94710.28, 4.9, 25934.34, 1.2, 1615224.91, 130.3],
], columns=[
    "Comprador", "Faturamento Total Atual", "Rep. Faturamento",
    "CMV mês Atual", "Rep. CMV", "Estoque Total", "Fator Cobertura",
    "Estoque Curva A", "Rep. Curva A", "Estoque Curva B", "Rep. Curva B",
    "Estoque Curva C", "Rep. Curva C", "Estoque Curva D", "Rep. Curva D",
    "Ruptura Ativa", "Ruptura %", "Entradas CUSTO", "Reposição CMV %"
])

METAS = pd.DataFrame([
    ["Paulo", 1093500.00, 27.0, 969571.19, 88.7, 1090767.59, 1.13, 218153.52, 20.0, 327230.28, 30.0, 436307.04, 40.0, 109076.76, 10.0, 54675.00, 5.0, 959875.48, 99.0],
    ["Francieli", 972000.00, 24.0, 664261.17, 68.3, 1252606.78, 1.89, 250521.36, 20.0, 375782.03, 30.0, 501042.71, 40.0, 125260.68, 10.0, 48600.00, 5.0, 657618.56, 99.0],
    ["Sebastião", 1984500.00, 49.0, 1170070.14, 59.0, 1880469.86, 1.61, 376093.97, 20.0, 564140.96, 30.0, 752187.95, 40.0, 188046.99, 10.0, 99225.00, 5.0, 1158369.44, 99.0],
], columns=[
    "Comprador", "Faturamento Total META", "Rep. Faturamento",
    "CMV mês META", "Rep. CMV", "Estoque Total META", "Fator Cobertura",
    "Estoque Curva A", "Rep. Curva A", "Estoque Curva B", "Rep. Curva B",
    "Estoque Curva C", "Rep. Curva C", "Estoque Curva D", "Rep. Curva D",
    "Ruptura Ativa", "Ruptura %", "Entradas CUSTO", "Reposição CMV %"
])

RESULTADO = pd.DataFrame([
    ["Paulo", 9525.48, 1.3, 8349.81, 0.0, 234666.36, 23.4, 149049.18, 11.9, 57322.22, -1.5, -24207.50, -13.8, 52502.46, 3.4, 41393.98, 3.8, -43738.31, -5.4],
    ["Francieli", -65338.03, -0.6, -44722.66, 0.0, 96890.82, 25.6, 146204.08, 11.0, 96890.83, 5.9, -94708.75, -11.5, -51495.33, -5.3, 16467.94, 1.9, -150310.55, -15.0],
    ["Sebastião", -118132.93, -0.8, -69775.69, 0.0, -70567.77, 3.4, -311057.17, -15.2, -97024.16, -3.9, 244176.85, 14.0, 93336.71, 5.1, 73290.66, 3.8, -456855.47, -31.3],
], columns=REALIZADOS.columns)

PREMIO = pd.DataFrame([
    ["Paulo", 28.51, 95.0, 30.00, 100.0, 95.66, 95.7, 96.65, 64.4, 149.61, 99.7, 88.11, 88.1, 88.50, 88.5, 86.01, 43.0, 139.58, 99.7],
    ["Francieli", 92.09, 102.3, 89.99, 100.0, 294.49, 98.2, 314.52, 69.9, 432.78, 96.2, 274.99, 91.7, 215.92, 72.0, 513.14, 85.5, 410.41, 97.7],
    ["Sebastião", 60.95, 101.6, 59.99, 100.0, 199.91, 100.0, 126.27, 42.1, 294.96, 98.3, 175.63, 87.8, 147.04, 73.5, 173.01, 43.3, 252.05, 90.0],
], columns=[
    "Comprador", "Faturamento Prêmio", "Faturamento Realizado",
    "CMV Prêmio", "CMV Realizado", "Estoque Total Prêmio", "Estoque Total Realizado",
    "Curva A Prêmio", "Curva A Realizado", "Curva B Prêmio", "Curva B Realizado",
    "Curva C Prêmio", "Curva C Realizado", "Curva D Prêmio", "Curva D Realizado",
    "Ruptura Prêmio", "Ruptura Realizado", "Entradas Prêmio", "Entradas Realizado"
])

PREMIO_KPI = pd.DataFrame([
    ["Faturamento", 3.0, 90.00, 102.3, 92.09],
    ["CMV", 3.0, 90.00, 100.0, 89.99],
    ["Fator Cobertura", 10.0, 300.00, 98.2, 294.49],
    ["Estoque Curva A", 15.0, 450.00, 69.9, 314.52],
    ["Estoque Curva B", 15.0, 450.00, 96.2, 432.78],
    ["Estoque Curva C", 10.0, 300.00, 91.7, 274.99],
    ["Estoque Curva D", 10.0, 300.00, 72.0, 215.92],
    ["Ruptura Ativa", 20.0, 600.00, 85.5, 513.14],
    ["Reposição CMV", 14.0, 420.00, 97.7, 410.41],
], columns=["KPI", "Peso sobre a meta", "Prêmio por KPI atingível", "Atingimento %", "Prêmio Atingido"])

# =========================================================
# FUNÇÕES
# =========================================================

def _numero_base(v, casas=2):
    """Formata número no padrão pt-BR."""
    if v is None or pd.isna(v):
        return ""
    try:
        numero = float(v)
    except (TypeError, ValueError):
        return str(v)

    formato = f"{numero:,.{int(casas)}f}"
    return (
        formato
        .replace(",", "\u0000")
        .replace(".", ",")
        .replace("\u0000", ".")
    )


def moeda(v):
    """Valor monetário sem o prefixo, sempre com duas casas."""
    if v is None or pd.isna(v):
        return ""
    return _numero_base(v, 2)


def moeda_real(v):
    """Valor monetário: R$ 100,00 ou R$ -100,00."""
    if v is None or pd.isna(v):
        return ""
    return f"R$ {_numero_base(v, 2)}"


def percentual(v):
    """Percentual: 100,00%."""
    if v is None or pd.isna(v):
        return ""
    return f"{_numero_base(v, 2)}%"


def numero_inteiro(v):
    """Número inteiro: 1.000."""
    if v is None or pd.isna(v):
        return ""
    try:
        return _numero_base(round(float(v)), 0)
    except (TypeError, ValueError):
        return str(v)


def numero_decimal(v, casas=2):
    """Número decimal sem símbolo monetário."""
    return _numero_base(v, casas)


def br_num(v, casas=2):
    """Compatibilidade com chamadas antigas."""
    return numero_decimal(v, casas)


def _nome_normalizado(nome):
    return str(nome).strip().casefold()


def _coluna_percentual(nome):
    texto = _nome_normalizado(nome)

    # Campos de valor que contêm as palavras CMV ou Margem não são percentuais.
    # Somente versões explicitamente percentuais, como "(%)", são tratadas
    # como percentual.
    campos_monetarios_exatos = {
        "cmv mês meta",
        "cmv mes meta",
        "cmv mês atual",
        "cmv mes atual",
        "margem bruta meta",
        "margem bruta atual",
        "meta margem bruta",
        "meta margem bruta (r$)",
        "margem bruta (r$)",
    }
    if texto in campos_monetarios_exatos:
        return False

    if "%" in texto or "percent" in texto:
        return True

    return any(chave in texto for chave in [
        "atingimento", "participação", "participacao",
        "reposição", "reposicao", "peso", "representatividade",
        "margem (%)", "cmv (%)", "cobertura (%)"
    ])


def _coluna_inteira(nome):
    texto = _nome_normalizado(nome)
    return any(chave in texto for chave in [
        "quantidade", "qtd", "itens", "item", "registros", "registro",
        "dias", "dia", "lojas", "loja", "produtos", "produto",
        "não mapeados", "nao mapeados", "código", "codigo",
        "parcela", "posição", "posicao"
    ])


def _coluna_monetaria(nome):
    texto = _nome_normalizado(nome)
    if _coluna_percentual(nome) or _coluna_inteira(nome):
        return False

    # Qualquer coluna marcada explicitamente como moeda deve sempre receber
    # o padrão brasileiro, por exemplo: Meta (R$), Realizado (R$) e Saldo R$.
    if "r$" in texto or "(r$)" in texto:
        return True

    campos_monetarios_exatos = {
        "cmv mês meta",
        "cmv mes meta",
        "cmv mês atual",
        "cmv mes atual",
        "margem bruta meta",
        "margem bruta atual",
        "meta margem bruta",
        "meta margem bruta (r$)",
        "margem bruta (r$)",
    }
    if texto in campos_monetarios_exatos:
        return True

    return any(chave in texto for chave in [
        "venda", "faturamento", "custo", "estoque", "entrada", "compra",
        "lucro", "cmv", "margem bruta", "meta mês", "meta mes",
        "meta venda", "prêmio", "premio", "pagamento",
        "contas a pagar", "ruptura", "necessidade", "valor",
        "entrega (r$)", "realizado", "saldo", "documento",
        "meta (r$)", "meta r$", "parcela máxima", "parcela maxima",
        "parcela conquistada", "valor perdido"
    ])



def cor_valor(v):
    try:
        numero = float(v)
    except (TypeError, ValueError):
        return "#ffffff"
    return "#ff4d4f" if numero < 0 else "#ffffff"


def formatar_valor_grafico(v, tipo):
    if tipo == "percentual":
        return percentual(v)
    if tipo == "moeda":
        return moeda_real(v)
    if tipo == "inteiro":
        return numero_inteiro(v)
    return numero_decimal(v, 2)


def detectar_tipo_grafico(fig):
    textos = []
    try:
        textos.append(str(fig.layout.xaxis.title.text or ""))
    except Exception:
        pass
    try:
        textos.append(str(fig.layout.yaxis.title.text or ""))
    except Exception:
        pass
    try:
        textos.append(str(fig.layout.title.text or ""))
    except Exception:
        pass
    texto = " ".join(textos).casefold()

    if any(chave in texto for chave in [
        "percent", "%", "cmv", "margem", "reposição", "reposicao",
        "atingimento", "participação", "participacao"
    ]):
        return "percentual"
    if any(chave in texto for chave in [
        "r$", "valor", "venda", "faturamento", "custo", "estoque",
        "entrada", "lucro", "pagamento", "ruptura", "contas a pagar"
    ]):
        return "moeda"
    if any(chave in texto for chave in [
        "itens", "quantidade", "registros", "dias"
    ]):
        return "inteiro"
    return "numero"


def aplicar_formato_grafico(fig, tipo=None):
    """Padroniza eixo, rótulo, tooltip e negativos dos traces."""
    tipo = tipo or detectar_tipo_grafico(fig)
    try:
        fig.update_layout(separators=",.")

        if tipo == "percentual":
            fig.update_xaxes(ticksuffix="%", tickformat=",.2f")
            fig.update_yaxes(ticksuffix="%", tickformat=",.2f")
        elif tipo == "moeda":
            fig.update_xaxes(tickprefix="R$ ", tickformat=",.2f")
            fig.update_yaxes(tickprefix="R$ ", tickformat=",.2f")
        elif tipo == "inteiro":
            fig.update_xaxes(tickformat=",.0f")
            fig.update_yaxes(tickformat=",.0f")

        fig.update_xaxes(exponentformat="none", separatethousands=True)
        fig.update_yaxes(exponentformat="none", separatethousands=True)

        for trace in fig.data:
            valores = None
            if getattr(trace, "orientation", None) == "h":
                valores = getattr(trace, "x", None)
            else:
                valores = getattr(trace, "y", None)

            if valores is None:
                continue

            textos = []
            cores = []
            for valor in valores:
                try:
                    numero = float(valor)
                    textos.append(formatar_valor_grafico(numero, tipo))
                    cores.append("#ff4d4f" if numero < 0 else "#ffffff")
                except (TypeError, ValueError):
                    textos.append(str(valor))
                    cores.append("#ffffff")

            try:
                trace.text = textos
                trace.texttemplate = "%{text}"
                trace.textfont = dict(color=cores)
                trace.hovertemplate = "%{fullData.name}<br>%{text}<extra></extra>"
                trace.cliponaxis = False
            except Exception:
                pass
    except Exception:
        pass
    return fig


def _markdown_executivo_seguro(texto):
    """Evita que valores como R$ 1.000,00 sejam interpretados como LaTeX.

    A substituição é somente para exibição no Streamlit. O texto original
    permanece limpo para download, histórico e demais exportações.
    """
    if texto is None:
        return ""
    saida = str(texto)
    # O Markdown usa $ como delimitador matemático. Em valores brasileiros,
    # escapa apenas o cifrão que compõe o prefixo monetário R$.
    saida = re.sub(r"R\s*\$", lambda m: m.group(0).replace("$", r"\$"), saida)
    return saida


def card_status_base(titulo, registros):
    registrar_card_exportacao(titulo, numero_inteiro(registros))
    st.markdown(
        f"""
        <div class="status-card-kpi">
            <div class="titulo">{titulo}</div>
            <div class="valor">{numero_inteiro(registros)}</div>
            <div class="status">↑ registros salvos</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def contar_registros_cache(tabela, periodo=None):
    """Conta os registros reais no SQLite para o período selecionado."""
    try:
        with conexao_cache() as con:
            existe = con.execute(
                "SELECT 1 FROM sqlite_master "
                "WHERE type='table' AND name=?",
                (tabela,),
            ).fetchone()
            if not existe:
                return 0

            colunas = {
                linha[1]
                for linha in con.execute(
                    f'PRAGMA table_info("{tabela}")'
                ).fetchall()
            }
            if periodo and "periodo_referencia" in colunas:
                return int(
                    con.execute(
                        f'SELECT COUNT(*) FROM "{tabela}" '
                        'WHERE periodo_referencia=?',
                        (str(periodo),),
                    ).fetchone()[0]
                )
            return int(
                con.execute(
                    f'SELECT COUNT(*) FROM "{tabela}"'
                ).fetchone()[0]
            )
    except Exception:
        return 0



# =========================================================
# EXPORTAÇÃO UNIVERSAL DA TELA ATUAL
# =========================================================

def iniciar_contexto_exportacao(visao_atual, periodo_atual):
    st.session_state["_export_visao"] = str(visao_atual)
    st.session_state["_export_periodo"] = str(periodo_atual)
    st.session_state["_export_tabelas"] = []
    st.session_state["_export_cards"] = []
    st.session_state["_export_graficos"] = []
    st.session_state.pop("_export_pdf_bytes", None)
    st.session_state.pop("_export_excel_bytes", None)
    st.session_state.pop("_export_nome_base", None)


def registrar_tabela_exportacao(dados, titulo=None):
    if not isinstance(dados, pd.DataFrame):
        return
    try:
        numero = len(st.session_state.get("_export_tabelas", [])) + 1
        st.session_state.setdefault("_export_tabelas", []).append({
            "titulo": titulo or f"Tabela {numero}",
            "dados": dados.copy(),
        })
    except Exception:
        pass


def registrar_card_exportacao(titulo, valor):
    try:
        st.session_state.setdefault("_export_cards", []).append({
            "titulo": str(titulo),
            "valor": str(valor),
        })
    except Exception:
        pass


def _lista_plotly(valor):
    if valor is None:
        return []
    try:
        return list(valor)
    except Exception:
        return []


def registrar_grafico_exportacao(fig):
    try:
        titulo = "Gráfico"
        try:
            titulo = str(fig.layout.title.text or titulo)
        except Exception:
            pass
        registros = []
        for trace in fig.data:
            nome = str(getattr(trace, "name", "") or "Série")
            horizontal = getattr(trace, "orientation", None) == "h"
            categorias = _lista_plotly(
                getattr(trace, "y", None) if horizontal else getattr(trace, "x", None)
            )
            valores = _lista_plotly(
                getattr(trace, "x", None) if horizontal else getattr(trace, "y", None)
            )
            tamanho = max(len(categorias), len(valores))
            for i in range(tamanho):
                registros.append({
                    "Série": nome,
                    "Categoria": categorias[i] if i < len(categorias) else "",
                    "Valor": valores[i] if i < len(valores) else None,
                })
        if registros:
            st.session_state.setdefault("_export_graficos", []).append({
                "titulo": titulo,
                "dados": pd.DataFrame(registros),
            })
    except Exception:
        pass


def _sanitizar_nome_arquivo(texto):
    texto = unicodedata.normalize("NFKD", str(texto))
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^A-Za-z0-9_-]+", "_", texto).strip("_")
    return texto or "exportacao"


def _nome_aba_excel(nome, usados):
    nome = re.sub(r'[:\\/?*\[\]]', " ", str(nome)).strip()
    nome = re.sub(r"\s+", " ", nome) or "Dados"
    base = nome[:31]
    candidato = base
    contador = 2
    while candidato.casefold() in usados:
        sufixo = f" {contador}"
        candidato = base[:31-len(sufixo)] + sufixo
        contador += 1
    usados.add(candidato.casefold())
    return candidato


def _tipo_coluna_exportacao(nome):
    if _coluna_percentual(nome):
        return "percentual"
    if _coluna_monetaria(nome):
        return "moeda"
    if _coluna_inteira(nome):
        return "inteiro"
    texto = _nome_normalizado(nome)
    if any(chave in texto for chave in ["data", "competência", "competencia", "período", "periodo"]):
        return "data"
    return "geral"


def gerar_excel_tela(visao_atual, periodo_atual, cards, tabelas, graficos):
    if not EXCEL_EXPORT_OK:
        raise RuntimeError(f"Biblioteca openpyxl não instalada: {EXCEL_EXPORT_ERROR}")
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    azul, ciano, branco, vermelho = "0D2032", "43D7E8", "FFFFFF", "FF4D4F"
    borda = Side(style="thin", color="35536D")
    ws["A1"] = "REDE ECONOMIZE - KPI COMERCIAL"
    ws["A2"] = str(visao_atual)
    ws["A3"] = f"Período: {periodo_atual}"
    ws["A4"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")
    for celula in ("A1", "A2"):
        ws[celula].fill = PatternFill("solid", fgColor=azul)
        ws[celula].font = Font(color=branco, bold=True, size=14)
        ws[celula].alignment = Alignment(horizontal="center")
    linha = 6
    if cards:
        ws.cell(linha, 1, "Indicador")
        ws.cell(linha, 2, "Valor")
        for celula in ws[linha]:
            celula.fill = PatternFill("solid", fgColor=ciano)
            celula.font = Font(bold=True)
        linha += 1
        for item in cards:
            ws.cell(linha, 1, item["titulo"])
            ws.cell(linha, 2, item["valor"])
            linha += 1
    usados = {ws.title.casefold()}
    for indice, item in enumerate(list(tabelas) + list(graficos), start=1):
        df = item.get("dados")
        if not isinstance(df, pd.DataFrame):
            continue
        aba = wb.create_sheet(_nome_aba_excel(item.get("titulo") or f"Dados {indice}", usados))
        aba.freeze_panes = "A2"
        for col_idx, coluna in enumerate(df.columns, start=1):
            celula = aba.cell(1, col_idx, str(coluna))
            celula.fill = PatternFill("solid", fgColor=azul)
            celula.font = Font(color=branco, bold=True)
            celula.alignment = Alignment(horizontal="center")
            celula.border = Border(left=borda, right=borda, top=borda, bottom=borda)
        for row_idx, valores in enumerate(df.itertuples(index=False, name=None), start=2):
            for col_idx, valor in enumerate(valores, start=1):
                coluna = df.columns[col_idx - 1]
                celula = aba.cell(row_idx, col_idx)
                if pd.isna(valor):
                    valor = None
                elif isinstance(valor, np.generic):
                    valor = valor.item()
                elif isinstance(valor, pd.Timestamp):
                    valor = valor.to_pydatetime()
                celula.value = valor
                tipo = _tipo_coluna_exportacao(coluna)
                if tipo == "moeda" and isinstance(valor, (int, float)):
                    celula.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
                elif tipo == "percentual" and isinstance(valor, (int, float)):
                    celula.number_format = '0.00"%";[Red]-0.00"%"'
                elif tipo == "inteiro" and isinstance(valor, (int, float)):
                    celula.number_format = '#,##0;[Red]-#,##0'
                elif isinstance(valor, float):
                    celula.number_format = '#,##0.00;[Red]-#,##0.00'
                elif tipo == "data" and isinstance(valor, (datetime, date)):
                    celula.number_format = "dd/mm/yyyy"
                if isinstance(valor, (int, float)) and valor < 0:
                    celula.font = Font(color=vermelho)
                celula.border = Border(left=borda, right=borda, top=borda, bottom=borda)
                celula.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx, coluna in enumerate(df.columns, start=1):
            max_len = len(str(coluna))
            for valor in df.iloc[:500, col_idx - 1]:
                if not pd.isna(valor):
                    max_len = max(max_len, len(str(valor)))
            aba.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 38)
        aba.auto_filter.ref = aba.dimensions
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 25
    ws.sheet_view.showGridLines = False
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _texto_pdf(valor, coluna):
    if pd.isna(valor):
        return ""
    tipo = _tipo_coluna_exportacao(coluna)
    if isinstance(valor, (int, float, np.number)):
        if tipo == "moeda": return moeda_real(valor)
        if tipo == "percentual": return percentual(valor)
        if tipo == "inteiro": return numero_inteiro(valor)
        return numero_decimal(valor, 2)
    if isinstance(valor, (datetime, date, pd.Timestamp)):
        return valor.strftime("%d/%m/%Y")
    return str(valor)


def gerar_pdf_tela(visao_atual, periodo_atual, cards, tabelas, graficos):
    if not PDF_EXPORT_OK:
        raise RuntimeError(f"Biblioteca reportlab não instalada: {PDF_EXPORT_ERROR}")
    buffer = io.BytesIO()
    pagina = landscape(A4)
    doc = SimpleDocTemplate(buffer, pagesize=pagina, leftMargin=22, rightMargin=22,
        topMargin=22, bottomMargin=22, title=f"{visao_atual} - {periodo_atual}", author="Rede Economize")
    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle("TituloKPI", parent=estilos["Title"], fontName="Helvetica-Bold",
        fontSize=17, leading=21, textColor=colors.HexColor("#0D2032"), alignment=TA_CENTER, spaceAfter=8)
    estilo_sub = ParagraphStyle("SubKPI", parent=estilos["Normal"], fontSize=9, leading=12,
        alignment=TA_CENTER, textColor=colors.HexColor("#35536D"), spaceAfter=12)
    estilo_secao = ParagraphStyle("SecaoKPI", parent=estilos["Heading2"], fontName="Helvetica-Bold",
        fontSize=11, leading=14, textColor=colors.HexColor("#0D2032"), spaceBefore=8, spaceAfter=6)
    estilo_normal = ParagraphStyle("NormalKPI", parent=estilos["Normal"], fontSize=7, leading=9)
    estilo_negativo = ParagraphStyle("NegativoKPI", parent=estilo_normal, textColor=colors.HexColor("#FF4D4F"))
    elementos = [Paragraph("REDE ECONOMIZE - KPI COMERCIAL", estilo_titulo),
        Paragraph(f"{visao_atual} | Período: {periodo_atual} | Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", estilo_sub)]
    if cards:
        elementos.append(Paragraph("Indicadores", estilo_secao))
        dados_cards = [["Indicador", "Valor"]] + [[i["titulo"], i["valor"]] for i in cards]
        tabela_cards = Table(dados_cards, colWidths=[300, 180], repeatRows=1)
        tabela_cards.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#0D2032")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 8), ("GRID", (0,0), (-1,-1), .35, colors.HexColor("#35536D")),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#EDF3F8")]),
            ("VALIGN", (0,0), (-1,-1), "TOP")]))
        elementos.extend([tabela_cards, Spacer(1,10)])
    conjuntos = list(tabelas) + list(graficos)
    if not conjuntos:
        elementos.append(Paragraph("Esta tela não possui tabelas ou séries de gráfico para exportação.", estilo_normal))
    limite_pdf = 2000
    for indice, item in enumerate(conjuntos, start=1):
        df = item.get("dados")
        if not isinstance(df, pd.DataFrame): continue
        elementos.append(Paragraph(str(item.get("titulo") or f"Dados {indice}"), estilo_secao))
        df_pdf = df.head(limite_pdf).copy()
        if len(df) > limite_pdf:
            elementos.extend([Paragraph(f"PDF limitado aos primeiros {numero_inteiro(limite_pdf)} de {numero_inteiro(len(df))} registros. O Excel contém a base completa.", estilo_normal), Spacer(1,4)])
        colunas = list(df_pdf.columns)
        if not colunas: continue
        largura = max(45, min(110, (pagina[0]-44)/max(len(colunas),1)))
        dados_pdf = [[Paragraph(str(c), estilo_normal) for c in colunas]]
        for valores in df_pdf.itertuples(index=False, name=None):
            linha_pdf = []
            for col_idx, valor in enumerate(valores):
                estilo = estilo_negativo if isinstance(valor, (int,float,np.number)) and valor < 0 else estilo_normal
                linha_pdf.append(Paragraph(_texto_pdf(valor, colunas[col_idx]), estilo))
            dados_pdf.append(linha_pdf)
        tabela = Table(dados_pdf, colWidths=[largura]*len(colunas), repeatRows=1, hAlign="LEFT")
        tabela.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#0D2032")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("GRID", (0,0), (-1,-1), .3, colors.HexColor("#35536D")),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#EDF3F8")]),
            ("VALIGN", (0,0), (-1,-1), "TOP"), ("LEFTPADDING", (0,0), (-1,-1), 3),
            ("RIGHTPADDING", (0,0), (-1,-1), 3), ("TOPPADDING", (0,0), (-1,-1), 3),
            ("BOTTOMPADDING", (0,0), (-1,-1), 3)]))
        elementos.extend([tabela, Spacer(1,10)])
    def rodape(canvas, documento):
        canvas.saveState(); canvas.setFont("Helvetica",7); canvas.setFillColor(colors.HexColor("#35536D"))
        canvas.drawString(22,10,f"Rede Economize - {visao_atual}")
        canvas.drawRightString(pagina[0]-22,10,f"Página {documento.page}"); canvas.restoreState()
    doc.build(elementos, onFirstPage=rodape, onLaterPages=rodape)
    return buffer.getvalue()


def renderizar_botao_pdf_identico(visao_atual, periodo_atual):
    """Abre a impressão nativa preservando visual, cores, cards e gráficos."""
    nome_documento = (
        _sanitizar_nome_arquivo(visao_atual)
        + "_"
        + _sanitizar_nome_arquivo(periodo_atual)
    )
    html = f"""
    <html>
    <head>
      <style>
        html, body {{ margin: 0; padding: 0; background: transparent; }}
        button {{
          width: 100%; height: 42px; border: 1px solid #2c668b;
          border-radius: 8px; background: #0d2032; color: white;
          font-family: Arial, sans-serif; font-size: 14px; font-weight: 700;
          cursor: pointer;
        }}
        button:hover {{ background: #143653; border-color: #43d7e8; }}
      </style>
    </head>
    <body>
      <button onclick="exportarPDF()">🖨️ Exportar PDF idêntico à tela</button>
      <script>
        function exportarPDF() {{
          const w = window.parent;
          const d = w.document;
          const tituloAnterior = d.title;
          d.title = {nome_documento!r};

          let estilo = d.getElementById('rede-economize-pdf-visual');
          if (!estilo) {{
            estilo = d.createElement('style');
            estilo.id = 'rede-economize-pdf-visual';
            estilo.innerHTML = `
              @media print {{
                @page {{ size: A4 landscape; margin: 7mm; }}
                * {{
                  -webkit-print-color-adjust: exact !important;
                  print-color-adjust: exact !important;
                }}
                html, body, .stApp,
                [data-testid="stAppViewContainer"],
                [data-testid="stMain"] {{
                  background: #06111e !important;
                  overflow: visible !important;
                  height: auto !important;
                  min-height: 0 !important;
                }}
                [data-testid="stHeader"],
                [data-testid="stToolbar"],
                [data-testid="stDecoration"],
                [data-testid="stStatusWidget"],
                [data-testid="stBottomBlockContainer"],
                .stDeployButton,
                footer {{ display: none !important; }}

                [data-testid="stSidebar"] {{
                  position: static !important;
                  transform: none !important;
                  min-width: 225px !important;
                  width: 225px !important;
                  height: auto !important;
                  overflow: visible !important;
                }}
                [data-testid="stSidebarContent"] {{
                  height: auto !important;
                  overflow: visible !important;
                }}
                .main .block-container,
                [data-testid="stMainBlockContainer"] {{
                  max-width: none !important;
                  width: 100% !important;
                  padding: 5mm !important;
                  overflow: visible !important;
                }}
                [data-testid="stVerticalBlock"] {{
                  position: static !important;
                  transform: none !important;
                  overflow: visible !important;
                  height: auto !important;
                }}
                [data-testid="stHorizontalBlock"] {{
                  display: grid !important;
                  grid-template-columns: repeat(auto-fit, minmax(240px, 1fr)) !important;
                  gap: 6mm !important;
                  align-items: start !important;
                  overflow: visible !important;
                  height: auto !important;
                }}
                [data-testid="stHorizontalBlock"]:has([data-testid="stPlotlyChart"]) {{
                  display: block !important;
                }}
                [data-testid="stHorizontalBlock"]:has([data-testid="stPlotlyChart"]) > div {{
                  width: 100% !important;
                  max-width: 100% !important;
                  margin-bottom: 8mm !important;
                  clear: both !important;
                }}
                [data-testid="stDataFrame"],
                [data-testid="stMetric"],
                [data-testid="stExpanderDetails"],
                .status-card-kpi,
                .meta-card {{
                  break-inside: avoid !important;
                  page-break-inside: avoid !important;
                }}
                [data-testid="stDataFrame"] {{
                  height: auto !important;
                  max-height: none !important;
                  overflow: visible !important;
                }}
                [data-testid="stDataFrame"] > div,
                [data-testid="stDataFrame"] iframe {{
                  height: auto !important;
                  max-height: none !important;
                }}
                [data-testid="stPlotlyChart"] {{
                  position: relative !important;
                  display: block !important;
                  width: 100% !important;
                  max-width: 100% !important;
                  height: 112mm !important;
                  min-height: 112mm !important;
                  margin: 0 0 8mm 0 !important;
                  clear: both !important;
                  overflow: hidden !important;
                  break-inside: avoid !important;
                  page-break-inside: avoid !important;
                }}
                [data-testid="stPlotlyChart"] .js-plotly-plot,
                [data-testid="stPlotlyChart"] .plot-container,
                [data-testid="stPlotlyChart"] .svg-container {{
                  position: relative !important;
                  width: 100% !important;
                  max-width: 100% !important;
                  height: 108mm !important;
                  min-height: 108mm !important;
                  inset: auto !important;
                  transform: none !important;
                }}
                details {{ display: block !important; }}
                details > summary {{ display: none !important; }}
                iframe[title="streamlit_component"] {{ display: none !important; }}
                button, .stButton, .stDownloadButton {{ display: none !important; }}
                a {{ color: inherit !important; text-decoration: none !important; }}
              }}
            `;
            d.head.appendChild(estilo);
          }}

          const restaurar = () => {{
            d.title = tituloAnterior;
            w.removeEventListener('afterprint', restaurar);
          }};
          w.addEventListener('afterprint', restaurar);
          w.focus();
          setTimeout(() => w.print(), 350);
        }}
      </script>
    </body>
    </html>
    """
    components.html(html, height=48, scrolling=False)


def renderizar_exportacao_tela():
    """Módulo único de exportação, sem expanders aninhados."""
    visao_atual = st.session_state.get("_export_visao", "Tela")
    periodo_atual = st.session_state.get("_export_periodo", "")
    tabelas = st.session_state.get("_export_tabelas", [])
    cards = st.session_state.get("_export_cards", [])
    graficos = st.session_state.get("_export_graficos", [])

    st.markdown("---")
    st.markdown("### 📤 Exportar esta tela")
    st.caption(
        "Escolha o formato. A geração ocorre somente quando solicitada, "
        "preservando a performance da navegação."
    )

    aba_visual, aba_analitico, aba_excel = st.tabs([
        "🖨️ PDF visual",
        "📄 PDF analítico",
        "📊 Excel completo",
    ])

    nome_base_padrao = (
        _sanitizar_nome_arquivo(visao_atual)
        + "_"
        + _sanitizar_nome_arquivo(periodo_atual)
    )

    with aba_visual:
        st.markdown("#### PDF visual - igual à tela")
        st.caption(
            "Abre a impressão do navegador preservando o tema, cards, "
            "gráficos e tabelas exibidos."
        )
        componentes_html = """
        <script>
        async function carregarHtml2Canvas(topWindow) {
            if (topWindow.html2canvas) return topWindow.html2canvas;
            await new Promise((resolve, reject) => {
                const existente = topWindow.document.getElementById('rede-html2canvas');
                if (existente) {
                    existente.addEventListener('load', resolve, {once:true});
                    existente.addEventListener('error', reject, {once:true});
                    if (topWindow.html2canvas) resolve();
                    return;
                }
                const script = topWindow.document.createElement('script');
                script.id = 'rede-html2canvas';
                script.src = 'https://cdn.jsdelivr.net/npm/html2canvas@1.4.1/dist/html2canvas.min.js';
                script.onload = resolve;
                script.onerror = () => reject(new Error('Não foi possível carregar o mecanismo de captura.'));
                topWindow.document.head.appendChild(script);
            });
            return topWindow.html2canvas;
        }

        function localizarInicioExportacao(mainOriginal) {
            const elementos = Array.from(mainOriginal.querySelectorAll('h1,h2,h3,h4,p,div,span'));
            const alvo = elementos.find(el => {
                const texto = (el.textContent || '').trim().replace(/\\s+/g, ' ');
                return texto === '📤 Exportar esta tela' || texto === 'Exportar esta tela';
            });
            return alvo ? alvo.closest('[data-testid="stElementContainer"]') : null;
        }

        function localizarContextoStreamlit() {
            const candidatos = [];
            let atual = window;

            // O components.html roda dentro de iframe. No Streamlit Cloud a árvore
            // pode ganhar níveis extras, então não assumimos mais window.top.
            for (let i = 0; i < 8; i += 1) {
                try {
                    if (atual && !candidatos.includes(atual)) candidatos.push(atual);
                    if (!atual.parent || atual.parent === atual) break;
                    atual = atual.parent;
                } catch (e) {
                    break;
                }
            }
            try {
                if (window.top && !candidatos.includes(window.top)) candidatos.push(window.top);
            } catch (e) {}

            const seletores = [
                '[data-testid="stMainBlockContainer"]',
                '[data-testid="stMain"] [data-testid="stMainBlockContainer"]',
                '.main .block-container',
                '.stMain .block-container',
                '[data-testid="stMain"]',
                'section.main'
            ];

            for (const win of candidatos) {
                try {
                    const doc = win.document;
                    if (!doc || !doc.documentElement) continue;
                    for (const seletor of seletores) {
                        const main = doc.querySelector(seletor);
                        if (main) return { win, doc, main, seletor };
                    }
                } catch (e) {
                    // Ignora somente frames sem acesso e continua procurando.
                }
            }
            return null;
        }

        async function imprimirTelaKPI() {
            const contexto = localizarContextoStreamlit();
            if (!contexto) {
                alert('Não foi possível localizar a área principal do Streamlit para gerar o PDF. Atualize a página e tente novamente.');
                return;
            }

            const topWindow = contexto.win;
            const topDocument = contexto.doc;
            const mainOriginal = contexto.main;

            const botao = document.getElementById('btnImprimirKPI');
            const textoOriginal = botao.textContent;
            botao.disabled = true;
            botao.textContent = '⏳ Preparando PDF idêntico à apresentação...';

            const printWindow = topWindow.open('', '_blank');
            if (!printWindow) {
                botao.disabled = false;
                botao.textContent = textoOriginal;
                alert('Permita pop-ups para este endereço e tente novamente.');
                return;
            }

            printWindow.document.write(`
                <!doctype html><html lang="pt-BR"><head><meta charset="utf-8">
                <title>Rede Economize - KPI Comercial</title>
                <style>
                    html,body{margin:0;background:#06111e;color:#fff;font-family:Arial,sans-serif}
                    .loading{padding:34px;text-align:center}.loading h2{margin-bottom:8px}
                </style></head><body><div class="loading">
                <h2>Preparando apresentação para PDF...</h2>
                <p>Aguarde a captura completa dos gráficos e tabelas.</p></div></body></html>`);
            printWindow.document.close();

            const restauracoes = [];
            try {
                const html2canvas = await carregarHtml2Canvas(topWindow);

                // Oculta somente componentes de operação, sem reorganizar a apresentação.
                const ocultarTemporariamente = (el) => {
                    if (!el) return;
                    const anterior = el.style.cssText;
                    restauracoes.push(() => { el.style.cssText = anterior; });
                    el.style.setProperty('display', 'none', 'important');
                };

                topDocument.querySelectorAll(
                    '[data-testid="stHeader"], [data-testid="stToolbar"], [data-testid="stDecoration"], '
                    + '[data-testid="stStatusWidget"], [data-testid="stSidebar"], .stDeployButton, footer'
                ).forEach(ocultarTemporariamente);

                // Esconde o bloco de exportação para ele não aparecer dentro do próprio PDF.
                const inicioExportacao = localizarInicioExportacao(mainOriginal);
                if (inicioExportacao && inicioExportacao.parentElement) {
                    let atual = inicioExportacao;
                    while (atual) {
                        ocultarTemporariamente(atual);
                        atual = atual.nextElementSibling;
                    }
                }

                // Esconde controles, mas mantém os textos, cards, gráficos e tabelas no lugar.
                mainOriginal.querySelectorAll(
                    'button, [data-testid="stFileUploader"], [data-testid="stDownloadButton"], '
                    + '[data-testid="stButton"], iframe[title="streamlit_component"]'
                ).forEach(ocultarTemporariamente);

                // Remove barras de ferramentas dos gráficos.
                mainOriginal.querySelectorAll('.modebar').forEach(ocultarTemporariamente);

                // Garante que todos os gráficos terminem o redimensionamento antes da captura.
                if (topWindow.Plotly && topWindow.Plotly.Plots) {
                    mainOriginal.querySelectorAll('.js-plotly-plot').forEach(g => {
                        try { topWindow.Plotly.Plots.resize(g); } catch (e) {}
                    });
                }
                await new Promise(resolve => setTimeout(resolve, 900));

                const largura = Math.ceil(mainOriginal.getBoundingClientRect().width);
                const altura = Math.ceil(Math.max(mainOriginal.scrollHeight, mainOriginal.getBoundingClientRect().height));

                const canvas = await html2canvas(mainOriginal, {
                    backgroundColor: '#06111e',
                    scale: 2,
                    useCORS: true,
                    allowTaint: true,
                    logging: false,
                    width: largura,
                    height: altura,
                    windowWidth: Math.max(topDocument.documentElement.clientWidth, largura),
                    windowHeight: Math.max(topDocument.documentElement.clientHeight, altura),
                    scrollX: 0,
                    scrollY: -topWindow.scrollY,
                    imageTimeout: 20000,
                    removeContainer: true,
                    onclone: (documentoClonado) => {
                        // html2canvas 1.4.1 não interpreta color(), color-mix(),
                        // oklch(), lab() e lch(). A apresentação usa somente
                        // cores RGB/HEX durante a captura.
                        const estiloCompatibilidade = documentoClonado.createElement('style');
                        estiloCompatibilidade.textContent = `
                            .rec-score-dot {
                                box-shadow: 0 0 0 5px rgba(67,215,232,.16) !important;
                            }
                            * {
                                color-scheme: dark !important;
                            }
                        `;
                        documentoClonado.head.appendChild(estiloCompatibilidade);

                        const propriedadesCor = [
                            'color', 'backgroundColor', 'borderTopColor',
                            'borderRightColor', 'borderBottomColor', 'borderLeftColor',
                            'outlineColor', 'textDecorationColor', 'columnRuleColor'
                        ];
                        const normalizarCor = (valor) => {
                            const texto = String(valor || '').trim();
                            if (!texto || !/(?:color\(|color-mix\(|oklch\(|lab\(|lch\()/i.test(texto)) {
                                return texto;
                            }
                            const canvasCor = documentoClonado.createElement('canvas');
                            const contextoCor = canvasCor.getContext('2d');
                            try {
                                contextoCor.fillStyle = '#000000';
                                contextoCor.fillStyle = texto;
                                const convertido = contextoCor.fillStyle;
                                return /(?:color\(|color-mix\(|oklch\(|lab\(|lch\()/i.test(convertido)
                                    ? '#ffffff' : convertido;
                            } catch (e) {
                                return '#ffffff';
                            }
                        };

                        documentoClonado.querySelectorAll('*').forEach((elemento) => {
                            let calculado;
                            try { calculado = documentoClonado.defaultView.getComputedStyle(elemento); }
                            catch (e) { return; }
                            propriedadesCor.forEach((prop) => {
                                const valor = calculado[prop];
                                const seguro = normalizarCor(valor);
                                if (seguro && seguro !== valor) {
                                    try { elemento.style[prop] = seguro; } catch (e) {}
                                }
                            });
                            const sombra = String(calculado.boxShadow || '');
                            if (/(?:color\(|color-mix\(|oklch\(|lab\(|lch\()/i.test(sombra)) {
                                elemento.style.boxShadow = 'none';
                            }
                            const fundo = String(calculado.backgroundImage || '');
                            if (/(?:color\(|color-mix\(|oklch\(|lab\(|lch\()/i.test(fundo)) {
                                elemento.style.backgroundImage = 'none';
                            }
                        });
                    }
                });

                // A4 paisagem com margens de 7 mm: área útil aproximada de 283 x 196 mm.
                // As quebras são calculadas nos limites reais dos componentes da tela,
                // evitando cortar cards, gráficos, tabelas, títulos e painéis ao meio.
                const proporcaoPagina = 283 / 196;
                const alturaPaginaFonte = Math.floor(canvas.width / proporcaoPagina);
                const escalaCaptura = canvas.width / Math.max(largura, 1);
                const raizRect = mainOriginal.getBoundingClientRect();

                const seletoresBloco = [
                    '[data-testid="stElementContainer"]',
                    '[data-testid="stVerticalBlockBorderWrapper"]',
                    '[data-testid="stPlotlyChart"]',
                    '[data-testid="stDataFrame"]',
                    '.hp-panel', '.hp-table-wrap', '.rec-panel', '.rec-hero',
                    '.rec-actions', '.rec-trend', '.premium-box', '.section-title'
                ].join(',');

                const intervalosProtegidos = [];
                const candidatosQuebra = new Set([0, canvas.height]);

                mainOriginal.querySelectorAll(seletoresBloco).forEach((elemento) => {
                    if (!elemento || elemento.offsetParent === null) return;
                    const r = elemento.getBoundingClientRect();
                    const topo = Math.max(0, Math.round((r.top - raizRect.top) * escalaCaptura));
                    const fundo = Math.min(canvas.height, Math.round((r.bottom - raizRect.top) * escalaCaptura));
                    if (fundo <= topo + 4) return;
                    intervalosProtegidos.push({topo, fundo});
                    candidatosQuebra.add(topo);
                    candidatosQuebra.add(fundo);
                });

                const candidatos = Array.from(candidatosQuebra)
                    .filter(v => Number.isFinite(v) && v >= 0 && v <= canvas.height)
                    .sort((a,b) => a-b);

                function dentroDeBloco(posicao) {
                    return intervalosProtegidos.find(i => posicao > i.topo + 8 && posicao < i.fundo - 8);
                }

                function proximaQuebra(inicio) {
                    const limiteIdeal = Math.min(canvas.height, inicio + alturaPaginaFonte);
                    if (limiteIdeal >= canvas.height) return canvas.height;

                    // Prioriza uma quebra entre 62% e 100% da página, sempre no
                    // último limite de componente disponível antes do fim da folha.
                    const minimoAceitavel = inicio + Math.floor(alturaPaginaFonte * 0.62);
                    const opcoes = candidatos.filter(v => v > minimoAceitavel && v <= limiteIdeal - 10);
                    if (opcoes.length) return opcoes[opcoes.length - 1];

                    // Se o limite cair dentro de um componente grande, tenta quebrar
                    // antes dele. Componentes maiores que uma página podem ser divididos.
                    const bloco = dentroDeBloco(limiteIdeal);
                    if (bloco && bloco.topo > inicio + Math.floor(alturaPaginaFonte * 0.40)) {
                        return bloco.topo;
                    }

                    return limiteIdeal;
                }

                const faixas = [];
                let inicioPagina = 0;
                let seguranca = 0;
                while (inicioPagina < canvas.height && seguranca < 500) {
                    let fimPagina = proximaQuebra(inicioPagina);
                    if (fimPagina <= inicioPagina + 20) {
                        fimPagina = Math.min(canvas.height, inicioPagina + alturaPaginaFonte);
                    }
                    faixas.push({inicio: inicioPagina, fim: fimPagina});
                    inicioPagina = fimPagina;
                    seguranca += 1;
                }

                const paginas = [];
                faixas.forEach((faixa) => {
                    const alturaFatia = Math.max(1, faixa.fim - faixa.inicio);
                    const paginaCanvas = topDocument.createElement('canvas');
                    paginaCanvas.width = canvas.width;
                    paginaCanvas.height = alturaFatia;
                    const ctx = paginaCanvas.getContext('2d');
                    ctx.fillStyle = '#06111e';
                    ctx.fillRect(0, 0, paginaCanvas.width, paginaCanvas.height);
                    ctx.drawImage(
                        canvas,
                        0, faixa.inicio, canvas.width, alturaFatia,
                        0, 0, canvas.width, alturaFatia
                    );
                    paginas.push(paginaCanvas.toDataURL('image/png', 1.0));
                });

                const paginasHtml = paginas.map((src, indice) => `
                    <section class="pagina ${indice === paginas.length - 1 ? 'ultima' : ''}">
                        <div class="conteudo-pagina">
                            <img src="${src}" alt="Página ${indice + 1} do relatório">
                        </div>
                        <footer class="rodape-pagina">
                            <span>REDE ECONOMIZE • KPI COMERCIAL</span>
                            <span>Página ${indice + 1} de ${paginas.length}</span>
                        </footer>
                    </section>`).join('');

                printWindow.document.open();
                printWindow.document.write(`
                    <!doctype html><html lang="pt-BR"><head><meta charset="utf-8">
                    <title>Rede Economize - KPI Comercial</title>
                    <style>
                        @page { size:A4 landscape; margin:7mm; }
                        * { box-sizing:border-box; -webkit-print-color-adjust:exact !important; print-color-adjust:exact !important; }
                        html,body { margin:0; padding:0; background:#06111e; }
                        .pagina {
                            width:283mm; height:196mm; margin:0 auto;
                            display:flex; flex-direction:column; align-items:stretch; justify-content:flex-start;
                            overflow:hidden; background:#06111e;
                            break-after:page; page-break-after:always;
                        }
                        .pagina.ultima { break-after:auto; page-break-after:auto; }
                        .conteudo-pagina {
                            width:100%; height:188mm; display:flex; align-items:flex-start;
                            justify-content:center; overflow:hidden; background:#06111e;
                        }
                        .pagina img {
                            display:block; max-width:100%; max-height:188mm;
                            width:auto; height:auto; object-fit:contain; object-position:top center; margin:0;
                        }
                        .rodape-pagina {
                            height:8mm; padding:1.6mm 2mm 0; border-top:0.25mm solid #24455f;
                            display:flex; align-items:flex-start; justify-content:space-between;
                            color:#8fa8ba; font:700 8pt Arial,sans-serif; letter-spacing:.25px;
                            background:#06111e;
                        }
                        @media screen {
                            body { padding:12px; }
                            .pagina { margin-bottom:12px; box-shadow:0 0 0 1px #24455f; }
                        }
                    </style></head><body>${paginasHtml}</body></html>`);
                printWindow.document.close();
                await new Promise(resolve => setTimeout(resolve, 1000));
                printWindow.focus();
                printWindow.print();
            } catch (erro) {
                printWindow.document.body.innerHTML = `
                    <div style="padding:32px;background:#06111e;color:white;font-family:Arial">
                        <h2>Não foi possível capturar a apresentação.</h2>
                        <p>${String(erro)}</p>
                        <p>A captura encontrou um estilo de cor incompatível. Utilize esta versão corrigida e tente novamente.</p>
                    </div>`;
            } finally {
                restauracoes.reverse().forEach(restaurar => {
                    try { restaurar(); } catch (e) {}
                });
                botao.disabled = false;
                botao.textContent = textoOriginal;
            }
        }
        </script>
        <button id="btnImprimirKPI" style="
            width:100%; border:1px solid #35536d; border-radius:8px;
            padding:12px 16px; background:#0d2032; color:#ffffff;
            font-size:15px; font-weight:700; cursor:pointer;">
            🖨️ Abrir apresentação idêntica para salvar como PDF
        </button>
        <script>
            document.getElementById('btnImprimirKPI').addEventListener('click', imprimirTelaKPI);
        </script>
        """
        components.html(componentes_html, height=58)
        st.info(
            "Na janela de impressão, selecione **Salvar como PDF**, "
            "orientação **Paisagem** e ative **Gráficos de plano de fundo**."
        )

    with aba_analitico:
        st.markdown("#### PDF analítico")
        st.caption(
            "Relatório estruturado com indicadores, tabelas e dados das "
            "séries dos gráficos."
        )
        if st.button(
            "Gerar PDF analítico", use_container_width=True,
            key=f"gerar_pdf_analitico_{_sanitizar_nome_arquivo(visao_atual)}",
        ):
            if not REPORTLAB_DISPONIVEL:
                st.error(
                    "A biblioteca ReportLab não está instalada. Execute o "
                    "iniciador do projeto novamente para instalar as dependências."
                )
            else:
                with st.spinner("Preparando PDF analítico..."):
                    st.session_state["_export_pdf_bytes"] = gerar_pdf_tela(
                        visao_atual, periodo_atual, cards, tabelas, graficos
                    )
                    st.session_state["_export_nome_base"] = nome_base_padrao
                st.success("PDF analítico preparado.")
        if st.session_state.get("_export_pdf_bytes"):
            st.download_button(
                "⬇️ Baixar PDF analítico",
                data=st.session_state["_export_pdf_bytes"],
                file_name=f"{st.session_state.get('_export_nome_base', nome_base_padrao)}.pdf",
                mime="application/pdf", use_container_width=True,
                key=f"download_pdf_analitico_{nome_base_padrao}",
            )
        else:
            st.button(
                "⬇️ Baixar PDF analítico", disabled=True,
                use_container_width=True,
                key=f"pdf_analitico_desabilitado_{nome_base_padrao}",
            )

    with aba_excel:
        st.markdown("#### Excel completo")
        st.caption("Contém os registros capturados na tela em abas separadas.")
        if st.button(
            "Gerar Excel desta tela", use_container_width=True,
            key=f"gerar_excel_{_sanitizar_nome_arquivo(visao_atual)}",
        ):
            with st.spinner("Preparando Excel..."):
                st.session_state["_export_excel_bytes"] = gerar_excel_tela(
                    visao_atual, periodo_atual, cards, tabelas, graficos
                )
                st.session_state["_export_nome_base"] = nome_base_padrao
            st.success("Excel preparado.")
        if st.session_state.get("_export_excel_bytes"):
            st.download_button(
                "⬇️ Baixar Excel",
                data=st.session_state["_export_excel_bytes"],
                file_name=f"{st.session_state.get('_export_nome_base', nome_base_padrao)}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"download_excel_{nome_base_padrao}",
            )
        else:
            st.button(
                "⬇️ Baixar Excel", disabled=True, use_container_width=True,
                key=f"excel_desabilitado_{nome_base_padrao}",
            )


def dataframe_br(dados, *args, **kwargs):
    """Exibe tabelas no padrão brasileiro sem alterar os dados salvos."""
    export_title = kwargs.pop("export_title", None)
    if isinstance(dados, pd.DataFrame):
        registrar_tabela_exportacao(dados, export_title)
    if not isinstance(dados, pd.DataFrame):
        return st.dataframe(dados, *args, **kwargs)

    exibicao = dados.copy()

    indices_percentuais = {
        "cmv geral", "margem contribuição", "margem contribuicao",
        "margem (%)", "cmv (%)", "reposição cmv", "reposicao cmv",
        "atingimento", "participação", "participacao"
    }
    indices_monetarios = {
        "venda geral", "custo médio geral", "custo medio geral",
        "entrada geral", "lucro bruto geral", "pagamento de fornecedor",
        "custo médio - pagamento de fornecedor (caixa)",
        "custo medio - pagamento de fornecedor (caixa)",
        "custo médio geral - entrada geral (competência)",
        "custo medio geral - entrada geral (competencia)",
        "contas a pagar fornecedor total", "estoque mês", "estoque mes",
        "estoque - contas a pagar"
    }
    indices_inteiros = {
        "itens", "quantidade", "registros", "dias"
    }

    # Matrizes com o indicador no índice.
    if len(exibicao.index):
        for indice in exibicao.index:
            nome_indice = _nome_normalizado(indice)
            if nome_indice in indices_percentuais:
                for coluna in exibicao.columns:
                    valor = exibicao.loc[indice, coluna]
                    if isinstance(valor, (int, float, np.number)) and not pd.isna(valor):
                        exibicao.loc[indice, coluna] = percentual(valor)
            elif nome_indice in indices_monetarios:
                for coluna in exibicao.columns:
                    valor = exibicao.loc[indice, coluna]
                    if isinstance(valor, (int, float, np.number)) and not pd.isna(valor):
                        exibicao.loc[indice, coluna] = moeda_real(valor)
            elif nome_indice in indices_inteiros:
                for coluna in exibicao.columns:
                    valor = exibicao.loc[indice, coluna]
                    if isinstance(valor, (int, float, np.number)) and not pd.isna(valor):
                        exibicao.loc[indice, coluna] = numero_inteiro(valor)

    # Tabelas convencionais com o indicador no nome da coluna.
    for coluna in exibicao.columns:
        if not pd.api.types.is_numeric_dtype(exibicao[coluna]):
            continue
        if _coluna_percentual(coluna):
            exibicao[coluna] = exibicao[coluna].map(percentual)
        elif _coluna_monetaria(coluna):
            exibicao[coluna] = exibicao[coluna].map(moeda_real)
        elif _coluna_inteira(coluna):
            exibicao[coluna] = exibicao[coluna].map(numero_inteiro)
        else:
            # Evita a exibição técnica do pandas, como 914607.200000.
            # Colunas numéricas sem unidade explícita ficam com duas casas,
            # separadores brasileiros e sem símbolo indevido.
            exibicao[coluna] = exibicao[coluna].map(lambda v: numero_decimal(v, 2))

    try:
        styler = exibicao.style
        for coluna in dados.columns:
            if coluna not in exibicao.columns:
                continue
            if not pd.api.types.is_numeric_dtype(dados[coluna]):
                continue
            valores_originais = pd.to_numeric(
                dados[coluna], errors="coerce"
            ).reset_index(drop=True)
            styler = styler.apply(
                lambda serie, valores=valores_originais: [
                    "color:#ff4d4f;font-weight:700"
                    if (
                        i < len(valores)
                        and pd.notna(valores.iloc[i])
                        and float(valores.iloc[i]) < 0
                    )
                    else ""
                    for i in range(len(serie))
                ],
                subset=[coluna],
            )
        return st.dataframe(styler, *args, **kwargs)
    except Exception:
        return st.dataframe(exibicao, *args, **kwargs)


def preparar_tabela(df):
    """Prepara tabelas de resultado no padrão visual brasileiro."""
    out = df.copy()
    for col in out.columns:
        if col in ["Comprador", "KPI"]:
            continue
        if not pd.api.types.is_numeric_dtype(out[col]):
            continue
        if _coluna_percentual(col):
            out[col] = out[col].map(percentual)
        elif _coluna_monetaria(col):
            out[col] = out[col].map(moeda_real)
        elif _coluna_inteira(col):
            out[col] = out[col].map(numero_inteiro)
        else:
            out[col] = out[col].map(lambda v: numero_decimal(v, 2))
    return out


def plotly_chart_br(fig, *args, **kwargs):
    """Exibe gráfico com padrão brasileiro em todos os elementos."""
    registrar_grafico_exportacao(fig)
    tipo = kwargs.pop("tipo", None)
    fig = aplicar_formato_grafico(fig, tipo=tipo)
    config = dict(kwargs.pop("config", {}) or {})
    config.setdefault("displayModeBar", False)
    config.setdefault("locale", "pt-BR")
    return st.plotly_chart(fig, *args, config=config, **kwargs)



def html_valor_negativo(valor, texto):
    try:
        negativo = float(valor) < 0
    except (TypeError, ValueError):
        negativo = False
    if negativo:
        return (
            '<span class="valor-negativo">'
            + str(texto)
            + '</span>'
        )
    return str(texto)


def card_meta(titulo, linhas, destaque=False):
    for rotulo, valor in linhas:
        registrar_card_exportacao(f"{titulo} - {rotulo}", valor)
    classe = "meta-card premium" if destaque else "meta-card"
    html = [f"<div class='{classe}'><div class='meta-card-title'>{titulo}</div>"]
    for rotulo, valor in linhas:
        html.append(f"<div class='meta-line'><span>{rotulo}</span><strong>{valor}</strong></div>")
    html.append("</div>")
    st.markdown("".join(html), unsafe_allow_html=True)

def section(texto, cls):
    st.markdown(f"<div class='section-title {cls}'>{texto}</div>", unsafe_allow_html=True)


# =========================================================
# CONFIGURAÇÃO, PERÍODO E HISTÓRICO DAS METAS
# =========================================================

DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)
LOG_DIR = DATA_DIR / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)

METAS_FILE = DATA_DIR / "metas_gestor.json"
HISTORICO_FILE = DATA_DIR / "historico_metas.json"

METAS_PADRAO = {
    "id_meta": "META-2026-07",
    "periodo_referencia": "2026-05",
    "data_inicio": "2026-07-01",
    "data_fim": "2026-07-31",
    "descricao": "Metas comerciais do período",
    "status": "Ativa",
    "meta_venda_total_mes": 4050000.00,
    "meta_cmv_mes": 67.40,
    "fator_reducao_cmv": 0.01,
    "fator_cobertura": 1.50,
    "meta_ruptura": 5.00,
    "meta_reposicao": 99.00,
    "curva_a": 20.00,
    "curva_b": 30.00,
    "curva_c": 40.00,
    "curva_d": 10.00,
    "rep_paulo": 27.00,
    "rep_francieli": 24.00,
    "rep_sebastiao": 49.00,
    "peso_faturamento": 3.00,
    "peso_cmv": 3.00,
    "peso_fator_cobertura": 10.00,
    "peso_curva_a": 15.00,
    "peso_curva_b": 15.00,
    "peso_curva_c": 10.00,
    "peso_curva_d": 10.00,
    "peso_ruptura": 20.00,
    "peso_reposicao": 14.00,
    "valor_premio_total": 3000.00,
    "usuario_cadastro": "Gestor",
    "data_cadastro": "",
    "ultima_atualizacao": "",
}

@st.cache_data(show_spinner=False)
def carregar_historico():
    if HISTORICO_FILE.exists():
        try:
            dados = json.loads(HISTORICO_FILE.read_text(encoding="utf-8"))
            return dados if isinstance(dados, list) else []
        except Exception:
            return []
    return []

def salvar_historico(lista):
    HISTORICO_FILE.write_text(
        json.dumps(lista, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

@st.cache_data(show_spinner=False)
def carregar_metas():
    if METAS_FILE.exists():
        try:
            dados = json.loads(METAS_FILE.read_text(encoding="utf-8"))
            return {**METAS_PADRAO, **dados}
        except Exception:
            return METAS_PADRAO.copy()
    return METAS_PADRAO.copy()

def salvar_metas(dados, registrar_historico=True):
    agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    if not dados.get("data_cadastro"):
        dados["data_cadastro"] = agora
    dados["ultima_atualizacao"] = agora

    METAS_FILE.write_text(
        json.dumps(dados, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

    if registrar_historico:
        historico = carregar_historico()
        historico = [h for h in historico if h.get("id_meta") != dados.get("id_meta")]
        historico.append(dados.copy())
        historico = sorted(historico, key=lambda x: x.get("periodo_referencia", ""), reverse=True)
        salvar_historico(historico)
    _limpar_cache_dados()


def data_br(valor):
    try:
        return datetime.strptime(valor, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return valor or "-"

METAS_GESTOR = carregar_metas()
HISTORICO_METAS = carregar_historico()


# =========================================================
# METAS DE LOJA: FATURAMENTO E MARGEM BRUTA
# =========================================================

METAS_LOJAS_FILE = DATA_DIR / "metas_lojas.json"
HISTORICO_METAS_LOJAS_FILE = DATA_DIR / "historico_metas_lojas.json"

METAS_LOJAS_PADRAO = [
    {
        "periodo_referencia": "2026-05",
        "regional_loja": "Filial 1 (Matriz)",
        "gerente": "Fábio",
        "meta_mes": 508547.37,
        "meta_margem_bruta_valor": 172906.11,
        "meta_margem_bruta_pct": 34.0,
        "representatividade_entrega_pct": 25.4,
        "representatividade_entrega_valor": 129251.71,
        "status": "Ativa",
    },
    {
        "periodo_referencia": "2026-05",
        "regional_loja": "Filial 3",
        "gerente": "Lanila",
        "meta_mes": 914607.20,
        "meta_margem_bruta_valor": 310966.45,
        "meta_margem_bruta_pct": 34.0,
        "representatividade_entrega_pct": 4.7,
        "representatividade_entrega_valor": 42718.64,
        "status": "Ativa",
    },
]

def carregar_metas_lojas():
    if METAS_LOJAS_FILE.exists():
        try:
            dados = json.loads(METAS_LOJAS_FILE.read_text(encoding="utf-8"))
            if isinstance(dados, list):
                return dados
        except Exception:
            pass
    METAS_LOJAS_FILE.write_text(
        json.dumps(METAS_LOJAS_PADRAO, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return [dict(x) for x in METAS_LOJAS_PADRAO]

def carregar_historico_metas_lojas():
    if HISTORICO_METAS_LOJAS_FILE.exists():
        try:
            dados = json.loads(HISTORICO_METAS_LOJAS_FILE.read_text(encoding="utf-8"))
            return dados if isinstance(dados, list) else []
        except Exception:
            return []
    return []


def _replicar_lista_metas_maio_2026(arquivo, campo_chave):
    """Cria metas ausentes de 2026 usando maio/2026 como modelo.

    Nunca sobrescreve uma competência já existente.
    """
    if not arquivo.exists():
        return 0

    try:
        conteudo = json.loads(arquivo.read_text(encoding="utf-8"))
    except Exception:
        return 0

    lista = conteudo.get("metas", []) if isinstance(conteudo, dict) else conteudo
    if not isinstance(lista, list):
        return 0

    def periodo(item):
        return str(
            item.get("periodo_referencia")
            or item.get("competencia")
            or item.get("periodo")
            or ""
        )[:7]

    origem = [item for item in lista if periodo(item) == "2026-05"]
    if not origem:
        return 0

    existentes = {
        (periodo(item), str(item.get(campo_chave, "")).strip().upper())
        for item in lista
    }

    criadas = 0
    for mes in range(1, 13):
        competencia = f"2026-{mes:02d}"
        for modelo in origem:
            chave = (
                competencia,
                str(modelo.get(campo_chave, "")).strip().upper(),
            )
            if chave in existentes:
                continue

            nova = dict(modelo)
            nova["periodo_referencia"] = competencia
            nova["competencia"] = competencia

            primeiro_dia = date(2026, mes, 1)
            proximo = date(2027, 1, 1) if mes == 12 else date(2026, mes + 1, 1)
            ultimo_dia = proximo - timedelta(days=1)
            nova["data_inicio"] = primeiro_dia.isoformat()
            nova["data_fim"] = ultimo_dia.isoformat()
            nova["origem"] = "Replicada da meta de maio/2026"
            nova["atualizado_em"] = datetime.now().isoformat(timespec="seconds")

            lista.append(nova)
            existentes.add(chave)
            criadas += 1

    if criadas:
        if isinstance(conteudo, dict):
            conteudo["metas"] = lista
            conteudo["ultima_replicacao_maio_2026"] = datetime.now().isoformat(
                timespec="seconds"
            )
            saida = conteudo
        else:
            saida = lista

        arquivo.write_text(
            json.dumps(saida, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    return criadas


def preencher_metas_2026_com_maio():
    lojas = _replicar_lista_metas_maio_2026(
        DATA_DIR / "metas_lojas.json", "loja"
    )
    compradores = _replicar_lista_metas_maio_2026(
        DATA_DIR / "metas_compradores.json", "comprador"
    )
    return lojas, compradores


# Preenchimento inicial sem sobrescrever alterações já realizadas.
try:
    preencher_metas_2026_com_maio()
except Exception:
    pass


def salvar_metas_lojas(lista, usuario="Gestor"):
    agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    dados_limpos = []
    for item in lista:
        registro = dict(item)
        registro["ultima_atualizacao"] = agora
        registro["usuario_atualizacao"] = usuario or "Gestor"
        dados_limpos.append(registro)
    METAS_LOJAS_FILE.write_text(
        json.dumps(dados_limpos, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    historico = carregar_historico_metas_lojas()
    historico.append({
        "data_hora": agora,
        "usuario": usuario or "Gestor",
        "registros": dados_limpos,
    })
    HISTORICO_METAS_LOJAS_FILE.write_text(
        json.dumps(historico[-100:], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

def dataframe_metas_lojas(periodo=None):
    dados = carregar_metas_lojas()
    df = pd.DataFrame(dados)
    colunas = [
        "periodo_referencia", "regional_loja", "gerente", "meta_mes",
        "meta_margem_bruta_valor", "meta_margem_bruta_pct",
        "representatividade_entrega_pct", "representatividade_entrega_valor", "status"
    ]
    for col in colunas:
        if col not in df.columns:
            df[col] = "" if col in ["periodo_referencia", "regional_loja", "gerente", "status"] else 0.0
    if periodo and "periodo_referencia" in df.columns:
        filtrado = df[df["periodo_referencia"].astype(str) == str(periodo)].copy()
        if not filtrado.empty:
            df = filtrado
    return df[colunas].copy()



@st.cache_data(ttl=600, show_spinner=False, max_entries=24)
def carregar_realizado_filiais_ceo(periodo, token_cache):
    """Agrega faturamento e margem bruta por filial diretamente no SQLite."""
    colunas = ["numero_loja", "loja", "faturamento_atual", "margem_bruta_atual"]
    try:
        with conexao_cache() as con:
            tabelas = {linha[0] for linha in con.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()}
            if "base_vendas" not in tabelas:
                return pd.DataFrame(columns=colunas)
            estrutura = {linha[1] for linha in con.execute(
                'PRAGMA table_info("base_vendas")'
            ).fetchall()}
            if "periodo_referencia" not in estrutura:
                return pd.DataFrame(columns=colunas)
            expr_venda = "COALESCE(CAST(valortotal AS REAL), 0)" if "valortotal" in estrutura else "0"
            if "lucro" in estrutura:
                expr_lucro = "COALESCE(CAST(lucro AS REAL), 0)"
            elif "custo" in estrutura and "valortotal" in estrutura:
                expr_lucro = "COALESCE(CAST(valortotal AS REAL), 0) - COALESCE(CAST(custo AS REAL), 0)"
            else:
                expr_lucro = "0"
            numero = "COALESCE(CAST(numero_loja AS TEXT), '')" if "numero_loja" in estrutura else "''"
            nome = "COALESCE(CAST(loja AS TEXT), '')" if "loja" in estrutura else "''"
            consulta = f"""
                SELECT {numero} AS numero_loja, {nome} AS loja,
                       SUM({expr_venda}) AS faturamento_atual,
                       SUM({expr_lucro}) AS margem_bruta_atual
                FROM base_vendas
                WHERE periodo_referencia = ?
                GROUP BY {numero}, {nome}
                ORDER BY {numero}, {nome}
            """
            return pd.read_sql_query(consulta, con, params=(str(periodo),))
    except Exception:
        return pd.DataFrame(columns=colunas)


def montar_quadro_filiais_ceo(periodo):
    """Cruza metas cadastradas por filial com o realizado da base de vendas."""
    metas_filiais = dataframe_metas_lojas(periodo).copy()
    if metas_filiais.empty:
        return pd.DataFrame()

    realizado = carregar_realizado_filiais_ceo(periodo, _arquivo_token(CACHE_DB_FILE)).copy()

    def chave_numero(valor):
        numeros = re.findall(r"\d+", str(valor or "").strip())
        return str(int(numeros[-1])).zfill(2) if numeros else ""

    metas_filiais["chave_filial"] = metas_filiais["regional_loja"].map(chave_numero)
    if realizado.empty:
        realizado = pd.DataFrame(columns=["numero_loja", "loja", "faturamento_atual", "margem_bruta_atual"])
    realizado["chave_filial"] = realizado["numero_loja"].map(chave_numero)

    mapa_romano = {" I":"01", " II":"02", " III":"03", " IV":"04", " V":"05", " VI":"06", " VII":"07", " VIII":"08", " IX":"09", " X":"10"}
    for indice in realizado[realizado["chave_filial"].eq("")].index:
        nome_loja = " " + str(realizado.at[indice, "loja"]).strip().upper()
        for romano, numero in sorted(mapa_romano.items(), key=lambda item: len(item[0]), reverse=True):
            if nome_loja.endswith(romano):
                realizado.at[indice, "chave_filial"] = numero
                break

    realizado = realizado[["chave_filial", "loja", "faturamento_atual", "margem_bruta_atual"]].copy()
    quadro = metas_filiais.merge(realizado, on="chave_filial", how="left")
    for coluna in ["meta_mes", "meta_margem_bruta_valor", "meta_margem_bruta_pct", "faturamento_atual", "margem_bruta_atual"]:
        quadro[coluna] = pd.to_numeric(quadro.get(coluna, 0), errors="coerce").fillna(0.0)

    quadro["Atingimento Faturamento (%)"] = quadro.apply(
        lambda linha: linha["faturamento_atual"] / linha["meta_mes"] * 100 if linha["meta_mes"] else 0.0, axis=1)
    quadro["Atingimento Margem Bruta (%)"] = quadro.apply(
        lambda linha: linha["margem_bruta_atual"] / linha["meta_margem_bruta_valor"] * 100 if linha["meta_margem_bruta_valor"] else 0.0, axis=1)
    quadro["Margem Bruta Atual (%)"] = quadro.apply(
        lambda linha: linha["margem_bruta_atual"] / linha["faturamento_atual"] * 100 if linha["faturamento_atual"] else 0.0, axis=1)

    quadro["Filial"] = quadro["regional_loja"].fillna("")
    quadro["Gerente"] = quadro["gerente"].fillna("")
    quadro["Faturamento Total META"] = quadro["meta_mes"]
    quadro["Faturamento Total Atual"] = quadro["faturamento_atual"]
    quadro["Margem Bruta META"] = quadro["meta_margem_bruta_valor"]
    quadro["Margem Bruta Atual"] = quadro["margem_bruta_atual"]
    quadro["Margem Bruta META (%)"] = quadro["meta_margem_bruta_pct"]
    return quadro[["Filial", "Gerente", "Faturamento Total META", "Faturamento Total Atual", "Atingimento Faturamento (%)", "Margem Bruta META", "Margem Bruta Atual", "Atingimento Margem Bruta (%)", "Margem Bruta META (%)", "Margem Bruta Atual (%)"]].sort_values(["Filial", "Gerente"]).reset_index(drop=True)


# =========================================================
# IMPORTAÇÃO TEMPORÁRIA DA RUPTURA
# =========================================================

RUPTURA_FILE = DATA_DIR / "ruptura_importada.csv"
RUPTURA_META_FILE = DATA_DIR / "ruptura_importacao_meta.json"

COLUNAS_RUPTURA_PADRAO = [
    "Loja",
    "Comprador",
    "Classificação 3º Nível",
    "Código Interno",
    "EAN",
    "Produto",
    "Ruptura Ativa",
    "Data Referência",
]

def normalizar_nome_coluna(nome):
    return (
        str(nome).strip()
        .replace("\n", " ")
        .replace("  ", " ")
    )

def detectar_coluna(df, alternativas):
    mapa = {normalizar_nome_coluna(c).lower(): c for c in df.columns}
    for alternativa in alternativas:
        chave = alternativa.lower()
        if chave in mapa:
            return mapa[chave]
    return None

def preparar_ruptura_importada(df):
    df = df.copy()
    df.columns = [normalizar_nome_coluna(c) for c in df.columns]

    col_loja = detectar_coluna(df, ["loja", "numero_loja", "filial"])
    col_comprador = detectar_coluna(df, ["comprador", "comprador responsável", "comprador responsavel"])
    col_classificacao = detectar_coluna(df, ["classificação 3º nível", "classificacao 3º nivel", "classificacao 3 nivel", "classificação"])
    col_codigo = detectar_coluna(df, ["código interno", "codigo interno", "cod_interno", "produtoid"])
    col_ean = detectar_coluna(df, ["ean", "código de barras", "codigo de barras", "codigobarras"])
    col_produto = detectar_coluna(df, ["produto", "descrição", "descricao"])
    col_ruptura = detectar_coluna(
        df, [
            "valor ruptura", "ruptura venda", "ruptura ativa",
            "ruptura", "ruptura_r$", "valor_ruptura", "ruptura_venda"
        ]
    )
    col_data = detectar_coluna(df, ["data referência", "data referencia", "data", "competência", "competencia"])

    if col_ruptura is None:
        raise ValueError(
            "Não foi encontrada a coluna obrigatória de ruptura. "
            "Use uma coluna chamada 'Ruptura Ativa' ou 'Ruptura'."
        )

    saida = pd.DataFrame()
    saida["Loja"] = df[col_loja] if col_loja else ""
    saida["Comprador"] = df[col_comprador] if col_comprador else ""
    saida["Classificação 3º Nível"] = df[col_classificacao] if col_classificacao else ""
    saida["Código Interno"] = df[col_codigo] if col_codigo else ""
    saida["EAN"] = df[col_ean] if col_ean else ""
    saida["Produto"] = df[col_produto] if col_produto else ""
    saida["Ruptura Ativa"] = pd.to_numeric(
        df[col_ruptura].astype(str)
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False),
        errors="coerce"
    ).fillna(0)

    if col_data:
        saida["Data Referência"] = pd.to_datetime(df[col_data], errors="coerce", dayfirst=True)
    else:
        saida["Data Referência"] = pd.Timestamp.today().normalize()

    saida["Data Referência"] = saida["Data Referência"].dt.strftime("%Y-%m-%d")
    return saida

def carregar_ruptura_importada():
    if RUPTURA_FILE.exists():
        try:
            return pd.read_csv(RUPTURA_FILE, sep=";", encoding="utf-8-sig")
        except Exception:
            return pd.DataFrame(columns=COLUNAS_RUPTURA_PADRAO)
    return pd.DataFrame(columns=COLUNAS_RUPTURA_PADRAO)

def salvar_ruptura_importada(df, nome_arquivo):
    df.to_csv(RUPTURA_FILE, sep=";", index=False, encoding="utf-8-sig")
    meta = {
        "arquivo_origem": nome_arquivo,
        "data_importacao": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "linhas": int(len(df)),
        "valor_total_ruptura": float(df["Ruptura Ativa"].sum()),
    }
    RUPTURA_META_FILE.write_text(
        json.dumps(meta, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

def carregar_meta_ruptura():
    if RUPTURA_META_FILE.exists():
        try:
            return json.loads(RUPTURA_META_FILE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}

RUPTURA_IMPORTADA = carregar_ruptura_importada()
META_RUPTURA_IMPORTADA = carregar_meta_ruptura()


def carregar_logo_economize():
    caminho = Path("assets/logo_rede_economize.png")
    if caminho.exists():
        return base64.b64encode(caminho.read_bytes()).decode("utf-8")
    return ""

LOGO_ECONOMIZE_B64 = carregar_logo_economize()


# =========================================================
# BANCO DE DADOS E ATUALIZAÇÕES MENSAIS
# =========================================================

CONFIG_DIR = Path("config")
SQL_DIR = Path("sql")
DATA_DIR = Path("data")
CONFIG_DIR.mkdir(exist_ok=True)
SQL_DIR.mkdir(exist_ok=True)
DATA_DIR.mkdir(exist_ok=True)

DB_CONFIG_FILE = CONFIG_DIR / "database.json"
CACHE_DB_FILE = DATA_DIR / "kpis_mensal.sqlite"
CACHE_DB_GZ_FILE = DATA_DIR / "kpis_mensal.sqlite.gz"

# Viewer: o banco publicado nunca é aberto diretamente no checkout do Streamlit.
# Cada instância trabalha com uma cópia runtime validada em /tmp.
VIEWER_RUNTIME_DIR = Path("/tmp/eirox_kpi_viewer")
VIEWER_RUNTIME_DB_FILE = VIEWER_RUNTIME_DIR / "kpis_mensal.sqlite"
VIEWER_RUNTIME_META_FILE = VIEWER_RUNTIME_DIR / "kpis_mensal.runtime.json"

def _assinatura_sqlite_publicado():
    origem = CACHE_DB_GZ_FILE if CACHE_DB_GZ_FILE.exists() else CACHE_DB_FILE
    if not origem.exists():
        return origem, "ausente"
    stat = origem.stat()
    return origem, f"{origem.name}|{stat.st_size}|{stat.st_mtime_ns}"

def _validar_sqlite_arquivo(caminho):
    con = sqlite3.connect(str(caminho), timeout=60)
    try:
        check = con.execute("PRAGMA quick_check").fetchone()
        if not check or str(check[0]).strip().lower() != "ok":
            raise RuntimeError(f"SQLite publicado inválido: {check}")
    finally:
        con.close()

def _preparar_sqlite_publicado(forcar=False):
    """Materializa o SQLite publicado em /tmp, valida e publica atomicamente.

    No Viewer o arquivo dentro de data/ é somente uma origem publicada. A navegação
    lê exclusivamente a cópia runtime, evitando troca do arquivo sob consultas ativas.
    """
    if not MODO_VIEWER:
        return CACHE_DB_FILE

    origem, assinatura = _assinatura_sqlite_publicado()
    if assinatura == "ausente":
        raise FileNotFoundError(
            f"Base publicada não encontrada: {CACHE_DB_GZ_FILE} nem {CACHE_DB_FILE}"
        )

    VIEWER_RUNTIME_DIR.mkdir(parents=True, exist_ok=True)
    if not forcar and VIEWER_RUNTIME_DB_FILE.exists() and VIEWER_RUNTIME_META_FILE.exists():
        try:
            meta = json.loads(VIEWER_RUNTIME_META_FILE.read_text(encoding="utf-8"))
            if meta.get("assinatura") == assinatura and VIEWER_RUNTIME_DB_FILE.stat().st_size > 0:
                return VIEWER_RUNTIME_DB_FILE
        except Exception:
            pass

    nonce = f"{os.getpid()}_{time.time_ns()}" if "os" in globals() else str(time.time_ns())
    tmp = VIEWER_RUNTIME_DIR / f"kpis_mensal.{nonce}.tmp.sqlite"
    tmp_meta = VIEWER_RUNTIME_DIR / f"kpis_mensal.{nonce}.tmp.json"
    tmp.unlink(missing_ok=True)
    tmp_meta.unlink(missing_ok=True)
    try:
        if origem == CACHE_DB_GZ_FILE:
            with gzip.open(origem, "rb") as entrada, tmp.open("wb") as destino:
                shutil.copyfileobj(entrada, destino, length=1024 * 1024)
        else:
            shutil.copy2(origem, tmp)

        _validar_sqlite_arquivo(tmp)
        tmp.replace(VIEWER_RUNTIME_DB_FILE)
        tmp_meta.write_text(
            json.dumps({
                "assinatura": assinatura,
                "origem": str(origem),
                "preparado_em": datetime.now().isoformat(timespec="seconds"),
            }, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        tmp_meta.replace(VIEWER_RUNTIME_META_FILE)
        return VIEWER_RUNTIME_DB_FILE
    finally:
        tmp.unlink(missing_ok=True)
        tmp_meta.unlink(missing_ok=True)

def _arquivo_sqlite_leitura():
    return _preparar_sqlite_publicado() if MODO_VIEWER else CACHE_DB_FILE

_preparar_sqlite_publicado()

CACHE_DB_ORIGIN_FILE = DATA_DIR / "origem_banco_cache.json"

def _identificador_config_banco(cfg):
    """Identificador sem senha para separar caches de bancos diferentes."""
    partes = [
        str(cfg.get("host", "")).strip().casefold(),
        str(cfg.get("porta", 5432)).strip(),
        str(cfg.get("banco", "")).strip().casefold(),
        str(cfg.get("usuario", "")).strip().casefold(),
    ]
    return hashlib.sha256("|".join(partes).encode("utf-8")).hexdigest()

def _registrar_origem_cache_banco(cfg, identidade=None):
    payload = {
        "identificador": _identificador_config_banco(cfg),
        "host": str(cfg.get("host", "")),
        "porta": int(cfg.get("porta", 5432) or 5432),
        "banco": str(cfg.get("banco", "")),
        "usuario": str(cfg.get("usuario", "")),
        "identidade_real": identidade or {},
        "atualizado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f"),
        "nonce": time.time_ns(),
    }
    CACHE_DB_ORIGIN_FILE.parent.mkdir(parents=True, exist_ok=True)
    CACHE_DB_ORIGIN_FILE.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return payload

def _origem_cache_corresponde(cfg):
    if not CACHE_DB_ORIGIN_FILE.exists():
        return False
    try:
        dados = json.loads(CACHE_DB_ORIGIN_FILE.read_text(encoding="utf-8"))
        return dados.get("identificador") == _identificador_config_banco(cfg)
    except Exception:
        return False


def preparar_cache_para_banco_atual(cfg):
    """Impede mistura de dados entre bancos diferentes na mesma pasta."""
    identidade = validar_identidade_banco(cfg)
    origem_igual = _origem_cache_corresponde(cfg)
    if origem_igual:
        return identidade, False

    # Preserva uma cópia técnica antes de limpar o cache operacional antigo.
    try:
        if CACHE_DB_FILE.exists():
            backup_dir = DATA_DIR / "backups_troca_banco"
            backup_dir.mkdir(parents=True, exist_ok=True)
            nome = datetime.now().strftime("kpis_mensal_%Y%m%d_%H%M%S.sqlite")
            shutil.copy2(CACHE_DB_FILE, backup_dir / nome)
    except Exception:
        pass

    limpar_dados_operacionais()
    # Camada financeira de Entradas não faz parte da função antiga de limpeza.
    try:
        with sqlite3.connect(CACHE_DB_FILE, timeout=30) as con:
            for tabela in ["base_entradas_financeira"]:
                try:
                    con.execute(f'DELETE FROM "{tabela}"')
                except sqlite3.OperationalError:
                    pass
            con.commit()
    except Exception:
        pass

    st.session_state.pop("_dados_visoes", None)
    st.session_state.pop("_chave_visoes", None)
    _limpar_cache_dados()
    return identidade, True


@st.cache_data(ttl=60, show_spinner=False, max_entries=24)
def carregar_totais_cards_diretos(periodo, token_dados=None):
    """Totais executivos diretamente do SQLite, sem depender do mapa de compradores."""
    periodo = str(periodo or "")[:7]
    totais = {
        "faturamento": 0.0, "cmv": 0.0, "estoque": 0.0,
        "entradas": 0.0, "ruptura": 0.0, "reposicao": 0.0,
    }
    # No Viewer, a leitura deve usar exclusivamente a cópia runtime validada.
    # Não testar CACHE_DB_FILE diretamente, pois ele é apenas a origem publicada.
    try:
        arquivo_leitura = _arquivo_sqlite_leitura()
        if not Path(arquivo_leitura).exists():
            return totais
    except Exception:
        return totais

    def cols(con, tabela):
        try:
            return {r[1] for r in con.execute(f'PRAGMA table_info("{tabela}")')}
        except Exception:
            return set()

    with conexao_cache() as con:
        tabelas = {r[0] for r in con.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        )}

        if "base_vendas" in tabelas:
            c = cols(con, "base_vendas")
            if "periodo_referencia" in c:
                valor = 'COALESCE(SUM(CAST("valortotal" AS REAL)),0)' if "valortotal" in c else '0'
                if "custo" in c and "quantidade" in c:
                    cmv = 'COALESCE(SUM(CAST("custo" AS REAL)*CAST("quantidade" AS REAL)),0)'
                elif "custototal" in c:
                    cmv = 'COALESCE(SUM(CAST("custototal" AS REAL)),0)'
                else:
                    cmv = '0'
                row = con.execute(
                    f'SELECT {valor}, {cmv} FROM "base_vendas" WHERE periodo_referencia=?',
                    (periodo,),
                ).fetchone()
                totais["faturamento"] = float(row[0] or 0)
                totais["cmv"] = float(row[1] or 0)

        if "base_estoque" in tabelas:
            c = cols(con, "base_estoque")
            if "periodo_referencia" in c:
                qtd = next((x for x in ["estoque", "saldo_estoque", "quantidade_estoque"] if x in c), None)
                cmed = next((x for x in ["custo_medio_atual", "custo_medio", "customedio"] if x in c), None)
                cunit = next((x for x in ["custo_unit_atual", "custo", "custo_unitario"] if x in c), None)
                if qtd:
                    if cmed and cunit:
                        custo = f'CASE WHEN COALESCE(CAST("{cmed}" AS REAL),0)>0 THEN CAST("{cmed}" AS REAL) ELSE COALESCE(CAST("{cunit}" AS REAL),0) END'
                    elif cmed:
                        custo = f'COALESCE(CAST("{cmed}" AS REAL),0)'
                    elif cunit:
                        custo = f'COALESCE(CAST("{cunit}" AS REAL),0)'
                    else:
                        custo = '0'
                    row = con.execute(
                        f'SELECT COALESCE(SUM(CAST("{qtd}" AS REAL)*({custo})),0) FROM "base_estoque" WHERE periodo_referencia=?',
                        (periodo,),
                    ).fetchone()
                    totais["estoque"] = float(row[0] or 0)

        if "base_entradas_financeira" in tabelas:
            c = cols(con, "base_entradas_financeira")
            if "periodo_referencia" in c and "valor_nf_total" in c:
                row = con.execute(
                    'SELECT COALESCE(SUM(CAST("valor_nf_total" AS REAL)),0) FROM "base_entradas_financeira" WHERE periodo_referencia=?',
                    (periodo,),
                ).fetchone()
                totais["entradas"] = float(row[0] or 0)
        elif "base_entradas" in tabelas:
            c = cols(con, "base_entradas")
            campo = next((x for x in ["valor_nf_total", "entrada_custo_total"] if x in c), None)
            if "periodo_referencia" in c and campo:
                row = con.execute(
                    f'SELECT COALESCE(SUM(CAST("{campo}" AS REAL)),0) FROM "base_entradas" WHERE periodo_referencia=?',
                    (periodo,),
                ).fetchone()
                totais["entradas"] = float(row[0] or 0)

    try:
        rup = carregar_ruptura_auto(periodo, _arquivo_token(RUPTURA_AUTO_DB, RUPTURA_AUTO_CONTROLE))
        if isinstance(rup, pd.DataFrame) and not rup.empty:
            # A importação automática grava o valor oficial como "Valor Ruptura".
            # Mantemos aliases para bases antigas e arquivos com outros cabeçalhos.
            aliases_ruptura = [
                "Valor Ruptura", "valor_ruptura",
                "Ruptura Venda", "ruptura_venda",
                "Ruptura Ativa", "ruptura_ativa",
                "Ruptura", "ruptura",
            ]
            campo = next((c for c in aliases_ruptura if c in rup.columns), None)
            if campo:
                totais["ruptura"] = float(
                    pd.to_numeric(rup[campo], errors="coerce").fillna(0).sum()
                )
    except Exception:
        pass

    totais["reposicao"] = (totais["entradas"] / totais["cmv"] * 100.0) if totais["cmv"] else 0.0
    return totais
ANALISE_COMERCIAL_CONFIG_FILE = CONFIG_DIR / "analise_comercial.json"
PLANO_CONTAS_PAGAMENTO_PADRAO = "Resultado > 03.1 - Despesas Operacionais > 2-CUSTOS VARIAVEIS > 1-C.M.V / DUPL. PAGAS"
PLANOS_CONTAS_CATALOGO_FILE = CONFIG_DIR / "planos_contas_catalogo.json"

@st.cache_data(show_spinner=False)
def carregar_config_analise_comercial():
    padrao = {
        "plano_contas_padrao": PLANO_CONTAS_PAGAMENTO_PADRAO,
        "planos_contas_selecionados": [PLANO_CONTAS_PAGAMENTO_PADRAO],
        "planos_adicionais": [],
    }
    if ANALISE_COMERCIAL_CONFIG_FILE.exists():
        try:
            dados = json.loads(ANALISE_COMERCIAL_CONFIG_FILE.read_text(encoding="utf-8"))
            if isinstance(dados, dict):
                padrao.update(dados)
        except Exception:
            pass
    return padrao

def salvar_config_analise_comercial(dados):
    ANALISE_COMERCIAL_CONFIG_FILE.write_text(
        json.dumps(dados, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    _limpar_cache_dados()

@st.cache_data(show_spinner=False)
def carregar_catalogo_planos_contas():
    if PLANOS_CONTAS_CATALOGO_FILE.exists():
        try:
            dados = json.loads(
                PLANOS_CONTAS_CATALOGO_FILE.read_text(encoding="utf-8")
            )
            planos = dados.get("planos", []) if isinstance(dados, dict) else []
            return sorted({
                str(plano).strip()
                for plano in planos
                if str(plano).strip()
            })
        except Exception:
            pass
    return []


def _numero_br(valor):
    """Converte valores brasileiros ou numéricos para float."""
    if valor is None or pd.isna(valor):
        return 0.0
    if isinstance(valor, (int, float, np.number)):
        return float(valor)
    texto = str(valor).strip().replace("R$", "").strip()
    if not texto:
        return 0.0
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except Exception:
        return 0.0


def normalizar_contas_pagar_df(df):
    """Adapta a saída oficial do DBeaver ao formato interno do projeto."""
    base = df.copy()
    mapa = {normalizar_nome_coluna(c).casefold(): c for c in base.columns}

    def coluna(*nomes):
        for nome in nomes:
            achada = mapa.get(normalizar_nome_coluna(nome).casefold())
            if achada is not None:
                return achada
        return None

    col_status = coluna("Status", "status")
    col_venc = coluna("Data de Vencimento", "data_vencimento")
    col_pag = coluna("Data Pagamento", "data_pagamento")
    col_valor = coluna("Valor Documento", "valor_documento", "Valor")
    col_fornecedor = coluna("Credor", "fornecedor")
    col_unidade = coluna("Unidade", "unidade")
    col_apelido = coluna("Apelido Un. Neg.", "apelido_unidade")
    col_plano = coluna("Plano de Contas", "plano_contas")
    col_doc = coluna("Número Documento", "numero_documento")

    saida = pd.DataFrame(index=base.index)
    saida["status"] = base[col_status].astype(str) if col_status else ""
    saida["data_vencimento"] = (
        pd.to_datetime(base[col_venc], errors="coerce", dayfirst=True)
        if col_venc else pd.NaT
    )
    saida["data_pagamento"] = (
        pd.to_datetime(base[col_pag], errors="coerce", dayfirst=True)
        if col_pag else pd.NaT
    )
    saida["valor_documento"] = (
        base[col_valor].map(_numero_br) if col_valor else 0.0
    )
    saida["fornecedor"] = base[col_fornecedor].astype(str) if col_fornecedor else ""
    saida["unidade"] = base[col_unidade].astype(str) if col_unidade else ""
    saida["apelido_unidade"] = base[col_apelido].astype(str) if col_apelido else ""
    saida["plano_contas"] = base[col_plano].astype(str).str.strip() if col_plano else ""
    saida["numero_documento"] = base[col_doc].astype(str) if col_doc else ""

    paga = saida["data_pagamento"].notna() | saida["status"].str.casefold().eq("paga")
    pendente = saida["status"].str.casefold().eq("pendente")
    saida["valor_pago"] = saida["valor_documento"].where(paga, 0.0)
    saida["saldo_aberto"] = saida["valor_documento"].where(pendente, 0.0)

    saida["data_vencimento"] = saida["data_vencimento"].dt.strftime("%Y-%m-%d")
    saida["data_pagamento"] = saida["data_pagamento"].dt.strftime("%Y-%m-%d")
    return saida


FONTES_BANCO = {
    "vendas": {
        "titulo": "Vendas",
        "arquivo_sql": SQL_DIR / "vendas.sql",
        "tabela_cache": "base_vendas",
    },
    "estoque": {
        "titulo": "Estoque",
        "arquivo_sql": SQL_DIR / "estoque.sql",
        "tabela_cache": "base_estoque",
    },
    "entradas": {
        "titulo": "Entradas",
        "arquivo_sql": SQL_DIR / "entradas.sql",
        "tabela_cache": "base_entradas",
    },
    "contas_pagar": {
        "titulo": "Contas a Pagar",
        "arquivo_sql": SQL_DIR / "contas_pagar.sql",
        "tabela_cache": "base_contas_pagar",
    },
}

DB_CONFIG_PADRAO = {
    "tipo": "PostgreSQL",
    "host": "",
    "porta": 5432,
    "banco": "",
    "usuario": "",
    "senha": "",
    "sslmode": "prefer",
    "salvar_senha": True,
    "ultima_validacao": "",
}

def carregar_config_banco():
    if DB_CONFIG_FILE.exists():
        try:
            dados = json.loads(DB_CONFIG_FILE.read_text(encoding="utf-8"))
            return {**DB_CONFIG_PADRAO, **dados}
        except Exception:
            pass
    return DB_CONFIG_PADRAO.copy()

def salvar_config_banco(dados):
    """Salva a conexão desta máquina e invalida qualquer configuração anterior."""
    configuracao_anterior = carregar_config_banco()
    banco_mudou = (
        _identificador_config_banco(configuracao_anterior)
        != _identificador_config_banco(dados)
    )
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    temporario = DB_CONFIG_FILE.with_suffix(".json.tmp")
    temporario.write_text(
        json.dumps(dados, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    temporario.replace(DB_CONFIG_FILE)
    # Garante que nenhum dado/resultado memorizado continue ligado ao banco anterior.
    try:
        st.cache_data.clear()
    except Exception:
        pass
    try:
        st.cache_resource.clear()
    except Exception:
        pass
    if banco_mudou:
        st.session_state["_banco_alterado_pendente"] = True
        st.session_state.pop("_dados_visoes", None)
        st.session_state.pop("_chave_visoes", None)


def apagar_config_banco_local():
    """Remove somente host, banco, usuário e senha salvos nesta máquina."""
    if DB_CONFIG_FILE.exists():
        DB_CONFIG_FILE.unlink()
    try:
        st.cache_data.clear()
    except Exception:
        pass
    try:
        st.cache_resource.clear()
    except Exception:
        pass


def identificar_banco_real(cfg):
    """Retorna a identidade efetiva do PostgreSQL ao qual a conexão chegou."""
    engine = criar_engine_banco(cfg)
    try:
        with engine.connect() as conn:
            row = conn.execute(text("""
                SELECT
                    inet_server_addr()::text AS servidor,
                    inet_server_port() AS porta,
                    current_database() AS banco,
                    current_user AS usuario,
                    current_schema() AS esquema
            """)).mappings().one()
            return dict(row)
    finally:
        engine.dispose()


def validar_identidade_banco(cfg):
    """Evita atualizar um banco diferente daquele digitado na configuração."""
    identidade = identificar_banco_real(cfg)
    banco_esperado = str(cfg.get("banco") or cfg.get("database") or "").strip()
    banco_real = str(identidade.get("banco") or "").strip()
    if banco_esperado and banco_real.casefold() != banco_esperado.casefold():
        raise RuntimeError(
            f"Conexão direcionada ao banco '{banco_real}', mas a configuração solicita "
            f"'{banco_esperado}'. Atualização cancelada para proteger os dados."
        )
    return identidade

def montar_url_banco(cfg):
    usuario = quote_plus(str(cfg.get("usuario", "")))
    senha = quote_plus(str(cfg.get("senha", "")))
    host = cfg.get("host", "")
    porta = int(cfg.get("porta", 5432))
    banco = cfg.get("banco", "")
    sslmode = cfg.get("sslmode", "prefer")
    return f"postgresql+psycopg2://{usuario}:{senha}@{host}:{porta}/{banco}?sslmode={sslmode}"

def criar_engine_banco(cfg):
    return create_engine(
        montar_url_banco(cfg),
        pool_pre_ping=True,
        pool_recycle=1800,
        connect_args={"connect_timeout": 15},
    )

def garantir_sqls():
    exemplos = {
        "vendas.sql": """-- Use os parâmetros :data_inicio e :data_fim
-- Cole aqui o SQL validado de vendas.
-- Exemplo de filtro:
-- WHERE datahora_venda_final::date BETWEEN :data_inicio AND :data_fim
SELECT
    NULL::text AS loja,
    NULL::text AS classificacao_3_nivel,
    NULL::numeric AS valor_venda
WHERE 1 = 0;
""",
        "estoque.sql": """-- Use os parâmetros :data_inicio e :data_fim
-- Cole aqui o SQL validado de estoque.
-- Para estoque atual, os parâmetros podem ser ignorados.
SELECT
    NULL::text AS loja,
    NULL::text AS classificacao_3_nivel,
    NULL::numeric AS valor_estoque
WHERE 1 = 0;
""",
        "entradas.sql": """-- Use os parâmetros :data_inicio e :data_fim
-- Cole aqui o SQL validado de entradas.
-- Exemplo de filtro:
-- WHERE datahoraentrada::date BETWEEN :data_inicio AND :data_fim
SELECT
    NULL::text AS loja,
    NULL::text AS classificacao_3_nivel,
    NULL::numeric AS valor_entrada
WHERE 1 = 0;
""",
        "contas_pagar.sql": """-- A consulta oficial de Contas a Pagar
-- é armazenada em sql/contas_pagar.sql.
-- Use os parâmetros :data_inicio e :data_fim.
""",
    }
    for nome, conteudo in exemplos.items():
        caminho = SQL_DIR / nome
        if not caminho.exists() or not caminho.read_text(encoding="utf-8").strip():
            caminho.write_text(conteudo, encoding="utf-8")

def ler_sql(caminho):
    if caminho.exists():
        return caminho.read_text(encoding="utf-8")
    return ""


def diagnosticar_sql_fonte(fonte, sql):
    """Valida se a fonte possui uma consulta real antes da execução."""
    texto_sql = str(sql or "").strip()
    normalizado = re.sub(r"\s+", " ", texto_sql).casefold()

    if not texto_sql:
        return False, "O arquivo SQL está vazio."

    marcadores_modelo = [
        "where 1 = 0",
        "cole aqui o sql",
        "null::text as unidade",
        "null::numeric as valor_documento",
    ]
    if any(marcador in normalizado for marcador in marcadores_modelo):
        if fonte == "contas_pagar":
            return (
                False,
                "A consulta de Contas a Pagar ainda é apenas um modelo. "
                "Cole o SQL oficial no menu Banco de Dados > Editar SQL."
            )
        return False, "A consulta desta fonte ainda é apenas um modelo."

    if fonte == "contas_pagar":
        if "plano" not in normalizado or "valor" not in normalizado:
            return (
                False,
                "O SQL de Contas a Pagar deve retornar o Plano de Contas "
                "e o Valor Documento."
            )

    return True, "SQL configurado."


def salvar_sql(caminho, conteudo):
    caminho.write_text(conteudo.strip() + "\n", encoding="utf-8")

def conexao_cache():
    # Viewer usa exclusivamente a cópia runtime validada e imutável.
    if MODO_VIEWER:
        caminho = _arquivo_sqlite_leitura().resolve().as_posix()
        con = sqlite3.connect(
            f"file:{caminho}?mode=ro&immutable=1",
            uri=True,
            timeout=30,
        )
        try:
            con.execute("PRAGMA query_only=ON")
            con.execute("PRAGMA busy_timeout=30000")
            con.execute("PRAGMA cache_size=-64000")
        except Exception:
            pass
        return con

    con = sqlite3.connect(CACHE_DB_FILE, timeout=30)
    try:
        con.execute("PRAGMA journal_mode=WAL")
        con.execute("PRAGMA synchronous=NORMAL")
        con.execute("PRAGMA temp_store=MEMORY")
        con.execute("PRAGMA cache_size=-64000")
        con.execute("PRAGMA busy_timeout=30000")
        con.execute("PRAGMA mmap_size=268435456")
    except Exception:
        pass
    con.execute("""
        CREATE TABLE IF NOT EXISTS atualizacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fonte TEXT NOT NULL,
            periodo_referencia TEXT NOT NULL,
            data_inicio TEXT NOT NULL,
            data_fim TEXT NOT NULL,
            registros INTEGER NOT NULL,
            status TEXT NOT NULL,
            mensagem TEXT,
            atualizado_em TEXT NOT NULL
        )
    """)
    con.commit()
    return con


def reconstruir_posicoes_mensais(con, periodo=None):
    """Recalcula posições mensais preservando todas as bases existentes."""
    garantir_tabelas_analise(con)
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    existe_cp = con.execute(
        "SELECT 1 FROM sqlite_master "
        "WHERE type='table' AND name='base_contas_pagar'"
    ).fetchone()

    if existe_cp:
        filtro_cp = " WHERE periodo_referencia = ?" if periodo else ""
        params_cp = (periodo,) if periodo else ()
        linhas_cp = con.execute(
            """
            SELECT
                periodo_referencia,
                SUM(COALESCE(CAST(valor_documento AS REAL), 0))
            FROM base_contas_pagar
            """
            + filtro_cp
            + """
            GROUP BY periodo_referencia
            """,
            params_cp,
        ).fetchall()

        for competencia, valor_cp in linhas_cp:
            existe = con.execute(
                """
                SELECT 1
                FROM analise_posicao_resumo
                WHERE periodo_referencia = ?
                """,
                (competencia,),
            ).fetchone()

            if existe:
                con.execute(
                    """
                    UPDATE analise_posicao_resumo
                    SET contas_pagar = ?,
                        origem = 'Reconstrução',
                        atualizado_em = ?
                    WHERE periodo_referencia = ?
                    """,
                    (float(valor_cp or 0), agora, competencia),
                )
            else:
                con.execute(
                    """
                    INSERT INTO analise_posicao_resumo
                        (
                            periodo_referencia,
                            contas_pagar,
                            estoque,
                            origem,
                            atualizado_em
                        )
                    VALUES (?, ?, 0, 'Reconstrução', ?)
                    """,
                    (competencia, float(valor_cp or 0), agora),
                )

    existe_estoque = con.execute(
        "SELECT 1 FROM sqlite_master "
        "WHERE type='table' AND name='analise_estoque_resumo'"
    ).fetchone()

    if existe_estoque:
        filtro_est = " WHERE periodo_referencia = ?" if periodo else ""
        params_est = (periodo,) if periodo else ()
        linhas_est = con.execute(
            """
            SELECT
                periodo_referencia,
                COALESCE(estoque, 0)
            FROM analise_estoque_resumo
            """
            + filtro_est,
            params_est,
        ).fetchall()

        for competencia, valor_estoque in linhas_est:
            existe = con.execute(
                """
                SELECT 1
                FROM analise_posicao_resumo
                WHERE periodo_referencia = ?
                """,
                (competencia,),
            ).fetchone()

            if existe:
                con.execute(
                    """
                    UPDATE analise_posicao_resumo
                    SET estoque = ?,
                        origem = 'Reconstrução',
                        atualizado_em = ?
                    WHERE periodo_referencia = ?
                    """,
                    (float(valor_estoque or 0), agora, competencia),
                )
            else:
                con.execute(
                    """
                    INSERT INTO analise_posicao_resumo
                        (
                            periodo_referencia,
                            contas_pagar,
                            estoque,
                            origem,
                            atualizado_em
                        )
                    VALUES (?, 0, ?, 'Reconstrução', ?)
                    """,
                    (competencia, float(valor_estoque or 0), agora),
                )

    con.commit()




def diagnosticar_competencias_sqlite(con):
    """Retorna competências presentes nas bases brutas e nos resumos."""
    resultado = {}
    tabelas = [
        "base_vendas",
        "base_entradas",
        "base_entradas_financeira",
        "base_contas_pagar",
        "base_estoque",
        "analise_vendas_resumo",
        "analise_entradas_resumo",
        "analise_contas_resumo",
        "analise_estoque_resumo",
    ]
    existentes = {
        r[0] for r in con.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        ).fetchall()
    }
    for tabela in tabelas:
        if tabela not in existentes:
            resultado[tabela] = []
            continue
        colunas = {
            r[1] for r in con.execute(f'PRAGMA table_info("{tabela}")').fetchall()
        }
        if "periodo_referencia" not in colunas:
            resultado[tabela] = []
            continue
        rows = con.execute(
            f'SELECT DISTINCT TRIM(periodo_referencia) '
            f'FROM "{tabela}" '
            f'WHERE periodo_referencia IS NOT NULL '
            f'AND TRIM(periodo_referencia) <> "" '
            f'ORDER BY 1'
        ).fetchall()
        resultado[tabela] = [str(r[0]) for r in rows if r and r[0]]
    return resultado


def garantir_tabelas_analise(con):
    """Cria todas as tabelas de resumo necessárias antes de qualquer atualização."""
    con.execute("""
        CREATE TABLE IF NOT EXISTS analise_vendas_resumo (
            periodo_referencia TEXT NOT NULL,
            classificacao TEXT,
            curva TEXT,
            venda REAL NOT NULL DEFAULT 0,
            custo REAL NOT NULL DEFAULT 0,
            PRIMARY KEY (periodo_referencia, classificacao, curva)
        )
    """)

    con.execute("""
        CREATE TABLE IF NOT EXISTS analise_entradas_resumo (
            periodo_referencia TEXT NOT NULL,
            classificacao TEXT,
            entrada REAL NOT NULL DEFAULT 0,
            PRIMARY KEY (periodo_referencia, classificacao)
        )
    """)

    con.execute("""
        CREATE TABLE IF NOT EXISTS analise_entradas_financeiro_resumo (
            periodo_referencia TEXT NOT NULL PRIMARY KEY,
            compra REAL NOT NULL DEFAULT 0,
            notas INTEGER NOT NULL DEFAULT 0
        )
    """)

    con.execute("""
        CREATE TABLE IF NOT EXISTS analise_contas_resumo (
            periodo_referencia TEXT NOT NULL,
            plano_contas TEXT,
            pagamento REAL NOT NULL DEFAULT 0,
            PRIMARY KEY (periodo_referencia, plano_contas)
        )
    """)

    con.execute("""
        CREATE TABLE IF NOT EXISTS analise_estoque_resumo (
            periodo_referencia TEXT NOT NULL PRIMARY KEY,
            estoque REAL NOT NULL DEFAULT 0
        )
    """)

    con.execute("""
        CREATE TABLE IF NOT EXISTS analise_posicao_resumo (
            periodo_referencia TEXT NOT NULL PRIMARY KEY,
            contas_pagar REAL NOT NULL DEFAULT 0,
            estoque REAL NOT NULL DEFAULT 0,
            origem TEXT,
            atualizado_em TEXT
        )
    """)

    con.commit()


def atualizar_resumos_analise(con, fonte=None, periodo=None):
    garantir_tabelas_analise(con)
    """Mantém tabelas pequenas usadas pela Análise Comercial.

    A tela nunca precisa carregar as bases brutas. Os resumos são refeitos
    somente após atualização da fonte ou em uma reconstrução manual.
    """
    filtros_delete = " WHERE periodo_referencia = ?" if periodo else ""
    params = (periodo,) if periodo else ()

    if fonte in (None, "vendas"):
        con.execute("""
            CREATE TABLE IF NOT EXISTS analise_vendas_resumo (
                periodo_referencia TEXT NOT NULL,
                classificacao TEXT,
                venda REAL NOT NULL DEFAULT 0,
                custo REAL NOT NULL DEFAULT 0,
                PRIMARY KEY (periodo_referencia, classificacao)
            )
        """)
        con.execute("DELETE FROM analise_vendas_resumo" + filtros_delete, params)
        where = " WHERE periodo_referencia = ?" if periodo else ""
        con.execute("""
            INSERT OR REPLACE INTO analise_vendas_resumo
                (periodo_referencia, classificacao, venda, custo)
            SELECT
                TRIM(periodo_referencia) AS periodo_referencia,
                COALESCE(NULLIF(TRIM(classificacao_3_nivel), ''),
                         NULLIF(TRIM(classificacao_resumida), ''),
                         'SEM CLASSIFICACAO') AS classificacao,
                SUM(COALESCE(CAST(valortotal AS REAL), 0)) AS venda,
                SUM(
                    ABS(COALESCE(CAST(qtd_mov AS REAL), CAST(quantidade AS REAL), 0))
                    * COALESCE(CAST(custo AS REAL), 0)
                ) AS custo
            FROM base_vendas
        """ + (
            " WHERE TRIM(periodo_referencia) = ?"
            if periodo else
            " WHERE COALESCE(TRIM(periodo_referencia), '') <> ''"
        ) + """
            GROUP BY
                TRIM(periodo_referencia),
                COALESCE(NULLIF(TRIM(classificacao_3_nivel), ''),
                         NULLIF(TRIM(classificacao_resumida), ''),
                         'SEM CLASSIFICACAO')
        """, params)

    if fonte in (None, "entradas"):
        con.execute("""
            CREATE TABLE IF NOT EXISTS analise_entradas_resumo (
                periodo_referencia TEXT NOT NULL,
                classificacao TEXT,
                compra REAL NOT NULL DEFAULT 0,
                PRIMARY KEY (periodo_referencia, classificacao)
            )
        """)
        con.execute("DELETE FROM analise_entradas_resumo" + filtros_delete, params)
        where = " WHERE periodo_referencia = ?" if periodo else ""
        con.execute("""
            INSERT OR REPLACE INTO analise_entradas_resumo
                (periodo_referencia, classificacao, compra)
            SELECT
                TRIM(periodo_referencia) AS periodo_referencia,
                COALESCE(NULLIF(TRIM(classificacao_3_nivel), ''), 'SEM CLASSIFICACAO'),
                SUM(
                    COALESCE(
                        CAST(valor_nf_total AS REAL),
                        CAST(entrada_custo_total AS REAL),
                        0
                    )
                )
            FROM base_entradas
        """ + (
            " WHERE TRIM(periodo_referencia) = ?"
            if periodo else
            " WHERE COALESCE(TRIM(periodo_referencia), '') <> ''"
        ) + """
            GROUP BY
                TRIM(periodo_referencia),
                COALESCE(NULLIF(TRIM(classificacao_3_nivel), ''), 'SEM CLASSIFICACAO')
        """, params)

        con.execute("""
            CREATE TABLE IF NOT EXISTS analise_entradas_financeiro_resumo (
                periodo_referencia TEXT NOT NULL PRIMARY KEY,
                compra REAL NOT NULL DEFAULT 0,
                notas INTEGER NOT NULL DEFAULT 0
            )
        """)
        con.execute(
            "DELETE FROM analise_entradas_financeiro_resumo" + filtros_delete,
            params,
        )
        if con.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='base_entradas_financeira'"
        ).fetchone():
            con.execute("""
                INSERT OR REPLACE INTO analise_entradas_financeiro_resumo
                    (periodo_referencia, compra, notas)
                SELECT
                    periodo_referencia,
                    COALESCE(SUM(CAST(valor_nf_total AS REAL)), 0),
                    COUNT(DISTINCT nota_fiscal_id)
                FROM base_entradas_financeira
            """ + (
                " WHERE periodo_referencia = ?" if periodo else ""
            ) + """
                GROUP BY periodo_referencia
            """, params)

    if fonte in (None, "contas_pagar"):
        con.execute("""
            CREATE TABLE IF NOT EXISTS analise_contas_resumo (
                periodo_referencia TEXT NOT NULL,
                plano_contas TEXT,
                pagamento REAL NOT NULL DEFAULT 0,
                PRIMARY KEY (periodo_referencia, plano_contas)
            )
        """)
        con.execute("DELETE FROM analise_contas_resumo" + filtros_delete, params)

        con.execute("""
            INSERT OR REPLACE INTO analise_contas_resumo
                (periodo_referencia, plano_contas, pagamento)
            SELECT
                TRIM(periodo_referencia) AS periodo_referencia,
                COALESCE(NULLIF(TRIM(plano_contas), ''), 'SEM PLANO DE CONTAS'),
                SUM(COALESCE(CAST(valor_pago AS REAL), 0))
            FROM base_contas_pagar
        """ + (
            " WHERE TRIM(periodo_referencia) = ? "
            "AND COALESCE(CAST(valor_pago AS REAL), 0) <> 0"
            if periodo else
            " WHERE COALESCE(TRIM(periodo_referencia), '') <> '' "
            "AND COALESCE(CAST(valor_pago AS REAL), 0) <> 0"
        ) + """
            GROUP BY
                TRIM(periodo_referencia),
                COALESCE(NULLIF(TRIM(plano_contas), ''), 'SEM PLANO DE CONTAS')
        """, params)

        con.execute("""
            CREATE TABLE IF NOT EXISTS analise_posicao_resumo (
                periodo_referencia TEXT NOT NULL PRIMARY KEY,
                contas_pagar REAL NOT NULL DEFAULT 0,
                estoque REAL NOT NULL DEFAULT 0,
                origem TEXT,
                atualizado_em TEXT
            )
        """)

        if periodo:
            saldo_contas = con.execute(
                """
                SELECT SUM(COALESCE(CAST(valor_documento AS REAL), 0))
                FROM base_contas_pagar
                WHERE periodo_referencia = ?
                """,
                (periodo,),
            ).fetchone()[0] or 0

            con.execute(
                """
                INSERT INTO analise_posicao_resumo
                    (periodo_referencia, contas_pagar, estoque, origem, atualizado_em)
                VALUES (?, ?, 0, 'Banco de Dados', ?)
                ON CONFLICT(periodo_referencia) DO UPDATE SET
                    contas_pagar = excluded.contas_pagar,
                    origem = excluded.origem,
                    atualizado_em = excluded.atualizado_em
                """,
                (
                    periodo,
                    float(saldo_contas),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                ),
            )

    if fonte in (None, "estoque"):
        con.execute("""
            CREATE TABLE IF NOT EXISTS analise_estoque_resumo (
                periodo_referencia TEXT NOT NULL PRIMARY KEY,
                estoque REAL NOT NULL DEFAULT 0
            )
        """)
        con.execute("DELETE FROM analise_estoque_resumo" + filtros_delete, params)

        existe_estoque = con.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='base_estoque'"
        ).fetchone()
        if existe_estoque:
            colunas = {
                str(linha[1]).strip().lower(): str(linha[1])
                for linha in con.execute("PRAGMA table_info(base_estoque)").fetchall()
            }
            col_periodo = colunas.get("periodo_referencia")
            col_qtd = next(
                (colunas[c] for c in [
                    "estoque", "quantidade estoque", "saldo estoque", "qtd estoque"
                ] if c in colunas),
                None
            )
            col_custo_medio = next(
                (colunas[c] for c in [
                    "custo_medio_atual", "custo medio atual",
                    "custo médio atual", "customedio", "custo medio"
                ] if c in colunas),
                None
            )
            col_custo_unit = next(
                (colunas[c] for c in [
                    "custo_unit_atual", "custo unit atual",
                    "custo unitario atual", "custo", "custo_final_r"
                ] if c in colunas),
                None
            )
            col_valor = next(
                (colunas[c] for c in [
                    "valor_estoque", "valor estoque", "estoque total",
                    "estoque a custo", "estoque x custo medio"
                ] if c in colunas),
                None
            )

            if col_periodo and (col_valor or (col_qtd and (col_custo_medio or col_custo_unit))):
                where_est = f' WHERE "{col_periodo}" = ?' if periodo else ""
                if col_valor:
                    expr_valor = f'COALESCE(CAST("{col_valor}" AS REAL), 0)'
                else:
                    if col_custo_medio and col_custo_unit:
                        expr_custo = (
                            f'CASE WHEN COALESCE(CAST("{col_custo_medio}" AS REAL),0) > 0 '
                            f'THEN CAST("{col_custo_medio}" AS REAL) '
                            f'ELSE COALESCE(CAST("{col_custo_unit}" AS REAL),0) END'
                        )
                    elif col_custo_medio:
                        expr_custo = f'COALESCE(CAST("{col_custo_medio}" AS REAL),0)'
                    else:
                        expr_custo = f'COALESCE(CAST("{col_custo_unit}" AS REAL),0)'
                    expr_valor = (
                        f'COALESCE(CAST("{col_qtd}" AS REAL),0) * ({expr_custo})'
                    )

                con.execute(
                    f"""
                    INSERT OR REPLACE INTO analise_estoque_resumo
                        (periodo_referencia, estoque)
                    SELECT
                        "{col_periodo}",
                        SUM({expr_valor})
                    FROM base_estoque
                    {where_est}
                    GROUP BY "{col_periodo}"
                    """,
                    params
                )

                con.execute("""
                    CREATE TABLE IF NOT EXISTS analise_posicao_resumo (
                        periodo_referencia TEXT NOT NULL PRIMARY KEY,
                        contas_pagar REAL NOT NULL DEFAULT 0,
                        estoque REAL NOT NULL DEFAULT 0,
                        origem TEXT,
                        atualizado_em TEXT
                    )
                """)

                if periodo:
                    linha_estoque = con.execute(
                        """
                        SELECT COALESCE(estoque, 0)
                        FROM analise_estoque_resumo
                        WHERE periodo_referencia = ?
                        """,
                        (periodo,),
                    ).fetchone()
                    valor_estoque_posicao = (
                        float(linha_estoque[0])
                        if linha_estoque is not None
                        else 0.0
                    )
                    agora_posicao = datetime.now().strftime(
                        "%Y-%m-%d %H:%M:%S"
                    )
                    existe_posicao = con.execute(
                        """
                        SELECT 1
                        FROM analise_posicao_resumo
                        WHERE periodo_referencia = ?
                        """,
                        (periodo,),
                    ).fetchone()
                    if existe_posicao:
                        con.execute(
                            """
                            UPDATE analise_posicao_resumo
                            SET estoque = ?,
                                origem = 'Banco de Dados',
                                atualizado_em = ?
                            WHERE periodo_referencia = ?
                            """,
                            (
                                valor_estoque_posicao,
                                agora_posicao,
                                periodo,
                            ),
                        )
                    else:
                        con.execute(
                            """
                            INSERT INTO analise_posicao_resumo
                                (
                                    periodo_referencia,
                                    contas_pagar,
                                    estoque,
                                    origem,
                                    atualizado_em
                                )
                            VALUES (?, 0, ?, 'Banco de Dados', ?)
                            """,
                            (
                                periodo,
                                valor_estoque_posicao,
                                agora_posicao,
                            ),
                        )
    con.execute("""
        CREATE TABLE IF NOT EXISTS analise_posicao_resumo (
            periodo_referencia TEXT NOT NULL PRIMARY KEY,
            contas_pagar REAL NOT NULL DEFAULT 0,
            estoque REAL NOT NULL DEFAULT 0,
            origem TEXT,
            atualizado_em TEXT
        )
    """)

    con.commit()


def salvar_contas_pagar_independente(df, periodo, data_inicio, data_fim):
    """Fluxo exclusivo de consulta já normalizada até o cache e resumos."""
    obrigatorias = {
        "status", "data_vencimento", "data_pagamento",
        "valor_documento", "valor_pago", "saldo_aberto",
        "fornecedor", "unidade", "plano_contas",
    }
    faltantes = sorted(obrigatorias - set(df.columns))
    if faltantes:
        raise ValueError(
            "Campos obrigatórios ausentes: " + ", ".join(faltantes)
        )
    if df.empty:
        raise ValueError("A base normalizada está vazia.")

    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    base = df.copy()

    for coluna in ("valor_documento", "valor_pago", "saldo_aberto"):
        base[coluna] = pd.to_numeric(
            base[coluna], errors="coerce"
        ).fillna(0.0)

    for coluna in (
        "status", "data_vencimento", "data_pagamento",
        "fornecedor", "unidade", "apelido_unidade",
        "plano_contas", "numero_documento",
    ):
        if coluna not in base.columns:
            base[coluna] = ""
        base[coluna] = base[coluna].fillna("").astype(str)

    base["periodo_referencia"] = str(periodo)
    base["data_inicio_meta"] = str(data_inicio)
    base["data_fim_meta"] = str(data_fim)
    base["atualizado_em"] = agora

    destino = "base_contas_pagar"
    temporaria = "_tmp_contas_pagar_atualizacao"
    etapa = "abertura do SQLite"
    con = conexao_cache()

    try:
        etapa = "gravação temporária"
        con.execute(f'DROP TABLE IF EXISTS "{temporaria}"')
        con.commit()
        base.to_sql(temporaria, con, if_exists="replace", index=False)

        gravados_tmp = con.execute(
            f'SELECT COUNT(*) FROM "{temporaria}"'
        ).fetchone()[0]
        if int(gravados_tmp) != len(base):
            raise RuntimeError(
                f"Tabela temporária: esperado {len(base)}, "
                f"gravado {gravados_tmp}."
            )

        etapa = "adequação da tabela definitiva"
        existe = con.execute(
            "SELECT 1 FROM sqlite_master "
            "WHERE type='table' AND name=?",
            (destino,),
        ).fetchone()

        if not existe:
            con.execute(
                f'CREATE TABLE "{destino}" AS '
                f'SELECT * FROM "{temporaria}" WHERE 1=0'
            )
            con.commit()
        else:
            cols_destino = {
                r[1] for r in con.execute(
                    f'PRAGMA table_info("{destino}")'
                ).fetchall()
            }
            cols_tmp = [
                r[1] for r in con.execute(
                    f'PRAGMA table_info("{temporaria}")'
                ).fetchall()
            ]
            for coluna in cols_tmp:
                if coluna not in cols_destino:
                    nome_seguro = coluna.replace('"', '""')
                    con.execute(
                        f'ALTER TABLE "{destino}" '
                        f'ADD COLUMN "{nome_seguro}"'
                    )
            con.commit()

        cols_destino = {
            r[1] for r in con.execute(
                f'PRAGMA table_info("{destino}")'
            ).fetchall()
        }
        cols_tmp = [
            r[1] for r in con.execute(
                f'PRAGMA table_info("{temporaria}")'
            ).fetchall()
        ]
        cols = [c for c in cols_tmp if c in cols_destino]
        if not cols:
            raise RuntimeError("Nenhuma coluna compatível para inserção.")

        campos = ", ".join(
            '"' + c.replace('"', '""') + '"' for c in cols
        )

        etapa = "substituição atômica do mês"
        con.execute("BEGIN IMMEDIATE")
        con.execute(
            f'DELETE FROM "{destino}" '
            'WHERE periodo_referencia=?',
            (str(periodo),),
        )
        con.execute(
            f'INSERT INTO "{destino}" ({campos}) '
            f'SELECT {campos} FROM "{temporaria}"'
        )

        gravados = con.execute(
            f'SELECT COUNT(*) FROM "{destino}" '
            'WHERE periodo_referencia=?',
            (str(periodo),),
        ).fetchone()[0]
        if int(gravados) != len(base):
            raise RuntimeError(
                f"Tabela definitiva: esperado {len(base)}, "
                f"gravado {gravados}."
            )

        etapa = "criação dos resumos"
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS analise_contas_resumo (
                periodo_referencia TEXT NOT NULL,
                plano_contas TEXT NOT NULL,
                pagamento REAL NOT NULL DEFAULT 0,
                PRIMARY KEY (periodo_referencia, plano_contas)
            )
            """
        )
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS analise_posicao_resumo (
                periodo_referencia TEXT NOT NULL PRIMARY KEY,
                contas_pagar REAL NOT NULL DEFAULT 0,
                estoque REAL NOT NULL DEFAULT 0,
                origem TEXT,
                atualizado_em TEXT
            )
            """
        )

        con.execute(
            "DELETE FROM analise_contas_resumo "
            "WHERE periodo_referencia=?",
            (str(periodo),),
        )
        con.execute(
            """
            INSERT OR REPLACE INTO analise_contas_resumo
                (periodo_referencia, plano_contas, pagamento)
            SELECT
                periodo_referencia,
                COALESCE(
                    NULLIF(TRIM(plano_contas), ''),
                    'SEM PLANO DE CONTAS'
                ),
                SUM(COALESCE(CAST(valor_pago AS REAL), 0))
            FROM base_contas_pagar
            WHERE periodo_referencia=?
            GROUP BY
                periodo_referencia,
                COALESCE(
                    NULLIF(TRIM(plano_contas), ''),
                    'SEM PLANO DE CONTAS'
                )
            """,
            (str(periodo),),
        )

        # Total de documentos com vencimento na competência.
        # Pagamentos realizados continuam separados no indicador de caixa.
        saldo = con.execute(
            """
            SELECT COALESCE(
                SUM(COALESCE(CAST(valor_documento AS REAL), 0)), 0
            )
            FROM base_contas_pagar
            WHERE periodo_referencia=?
            """,
            (str(periodo),),
        ).fetchone()[0] or 0

        con.execute(
            """
            INSERT INTO analise_posicao_resumo
                (periodo_referencia, contas_pagar, estoque,
                 origem, atualizado_em)
            VALUES (?, ?, 0, 'Banco de Dados', ?)
            ON CONFLICT(periodo_referencia) DO UPDATE SET
                contas_pagar=excluded.contas_pagar,
                origem=excluded.origem,
                atualizado_em=excluded.atualizado_em
            """,
            (str(periodo), float(saldo), agora),
        )

        etapa = "histórico"
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS atualizacoes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fonte TEXT,
                periodo_referencia TEXT,
                data_inicio TEXT,
                data_fim TEXT,
                registros INTEGER,
                status TEXT,
                mensagem TEXT,
                atualizado_em TEXT
            )
            """
        )
        con.execute(
            """
            INSERT INTO atualizacoes
                (fonte, periodo_referencia, data_inicio, data_fim,
                 registros, status, mensagem, atualizado_em)
            VALUES (?, ?, ?, ?, ?, 'Sucesso', ?, ?)
            """,
            (
                "contas_pagar", str(periodo),
                str(data_inicio), str(data_fim),
                int(len(base)),
                "Rotina exclusiva concluída",
                agora,
            ),
        )

        con.execute(f'DROP TABLE IF EXISTS "{temporaria}"')
        con.commit()
        _limpar_cache_dados()
        st.cache_data.clear()
        return int(len(base))

    except Exception as erro:
        try:
            con.rollback()
        except Exception:
            pass
        try:
            con.execute(f'DROP TABLE IF EXISTS "{temporaria}"')
            con.commit()
        except Exception:
            pass

        LOG_DIR.mkdir(parents=True, exist_ok=True)
        log = LOG_DIR / "contas_pagar_gravacao_erro.txt"
        detalhe = (
            f"Etapa: {etapa}\n"
            f"Período: {periodo}\n"
            f"Registros: {len(base)}\n"
            f"Tipo: {type(erro).__name__}\n"
            f"Mensagem: {erro}\n\n"
            f"{traceback.format_exc()}"
        )
        log.write_text(detalhe, encoding="utf-8")
        raise RuntimeError(
            "Falha na atualização exclusiva de Contas a Pagar.\n"
            f"Etapa: {etapa}\nDetalhe: {erro}\nLog: {log}"
        ) from erro
    finally:
        con.close()


def salvar_snapshot_mensal(df, fonte, periodo, data_inicio, data_fim):
    with conexao_cache() as con_estrutura:
        garantir_tabelas_analise(con_estrutura)

    info = FONTES_BANCO[fonte]
    tabela = info["tabela_cache"]
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    base = df.copy()
    base["periodo_referencia"] = periodo
    base["data_inicio_meta"] = data_inicio
    base["data_fim_meta"] = data_fim
    base["atualizado_em"] = agora

    con = conexao_cache()
    try:
        # Remove somente o período que será atualizado.
        try:
            con.execute(
                f'DELETE FROM "{tabela}" WHERE periodo_referencia = ?',
                (periodo,)
            )
            con.commit()
        except sqlite3.OperationalError:
            pass

        base.to_sql(tabela, con, if_exists="append", index=False)
        # Atualiza somente o pequeno resumo da fonte/período alterado.
        atualizar_resumos_analise(con, fonte=fonte, periodo=periodo)
        con.execute(
            """
            INSERT INTO atualizacoes
            (fonte, periodo_referencia, data_inicio, data_fim, registros, status, mensagem, atualizado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                fonte, periodo, data_inicio, data_fim,
                len(base),
                "Sucesso",
                (
                    (
                        "Atualização concluída • "
                        f"Faturamento: {float(pd.to_numeric(base.get('valortotal', pd.Series(dtype=float)), errors='coerce').fillna(0).sum()):.2f} • "
                        f"Itens: {float(pd.to_numeric(base.get('quantidade', pd.Series(dtype=float)), errors='coerce').fillna(0).sum()):.0f}"
                    )
                    if fonte == "vendas"
                    else (
                        (
                            "Atualização concluída • "
                            f"Compra A7: {float(pd.to_numeric(base.get('valor_nf_total', base.get('entrada_custo_total', pd.Series(dtype=float))), errors='coerce').fillna(0).sum()):.2f} • "
                            f"Itens: {float(pd.to_numeric(base.get('quantidade_por_produto', pd.Series(dtype=float)), errors='coerce').fillna(0).sum()):.2f}"
                        )
                        if fonte == "entradas"
                        else "Atualização concluída"
                    )
                ),
                agora
            )
        )
        con.commit()
        _limpar_cache_dados()
    finally:
        con.close()

def limpar_dados_operacionais():
    """Remove bases, resumos e histórico técnico sem apagar cadastros e metas."""
    tabelas = [
        "base_vendas",
        "base_estoque",
        "base_entradas",
        "base_contas_pagar",
        "analise_vendas_resumo",
        "analise_entradas_resumo",
        "analise_contas_resumo",
        "analise_estoque_resumo",
        "analise_posicao_resumo",
        "atualizacoes",
    ]
    con = conexao_cache()
    try:
        for tabela in tabelas:
            try:
                con.execute(f'DELETE FROM "{tabela}"')
            except sqlite3.OperationalError:
                pass
        con.commit()
        try:
            con.execute("VACUUM")
        except Exception:
            pass
    finally:
        con.close()

    _limpar_cache_dados()
    st.cache_data.clear()


def _config_psycopg2(cfg):
    return {
        "host": cfg.get("host") or cfg.get("servidor"),
        "port": int(cfg.get("port") or cfg.get("porta") or 5432),
        "dbname": cfg.get("database") or cfg.get("banco"),
        "user": cfg.get("user") or cfg.get("usuario"),
        "password": cfg.get("password") or cfg.get("senha"),
        "connect_timeout": 20,
        "application_name": "Rede Economize KPI Comercial",
    }


def _gravar_log_contas_pagar(texto):
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    caminho = LOG_DIR / "contas_pagar_erros.log"
    with caminho.open("a", encoding="utf-8") as arquivo:
        arquivo.write(
            f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]\n{texto}\n"
        )
    return caminho


def executar_contas_pagar_psycopg2(cfg, sql, data_inicio, data_fim, periodo):
    """Executa exatamente o SQL validado no DBeaver via psycopg2 direto."""
    inicio = datetime.strptime(str(data_inicio)[:10], "%Y-%m-%d").date()
    fim = datetime.strptime(str(data_fim)[:10], "%Y-%m-%d").date()
    sql_final = str(sql)
    sql_final, qtd_inicio = re.subn(
        r"DATE\s+'\d{4}-\d{2}-\d{2}'\s+AS\s+data_inicial",
        f"DATE '{inicio.isoformat()}' AS data_inicial",
        sql_final, count=1, flags=re.IGNORECASE,
    )
    sql_final, qtd_fim = re.subn(
        r"DATE\s+'\d{4}-\d{2}-\d{2}'\s+AS\s+data_final",
        f"DATE '{fim.isoformat()}' AS data_final",
        sql_final, count=1, flags=re.IGNORECASE,
    )
    if qtd_inicio != 1 or qtd_fim != 1:
        raise RuntimeError("As datas do CTE parametros não foram encontradas no SQL oficial.")

    LOG_DIR.mkdir(parents=True, exist_ok=True)
    arquivo_sql = LOG_DIR / "contas_pagar_sql_executado.sql"
    arquivo_erro = LOG_DIR / "contas_pagar_erro.txt"
    arquivo_sql.write_text(sql_final, encoding="utf-8")

    parametros = {
        "host": cfg.get("host"),
        "port": int(cfg.get("porta", 5432)),
        "dbname": cfg.get("banco"),
        "user": cfg.get("usuario"),
        "password": cfg.get("senha"),
        "connect_timeout": 20,
        "application_name": "Rede Economize KPI Comercial",
        "options": "-c statement_timeout=600000",
    }
    sslmode = str(cfg.get("sslmode", "prefer") or "prefer").strip()
    if sslmode:
        parametros["sslmode"] = sslmode

    conexao = None
    cursor = None
    try:
        conexao = psycopg2.connect(**parametros)
        conexao.autocommit = False
        cursor = conexao.cursor()
        cursor.execute("SELECT current_database(), current_user, current_schema(), current_setting('search_path')")
        ambiente = cursor.fetchone()
        cursor.execute(sql_final)
        if cursor.description is None:
            raise RuntimeError("O SQL executou, mas não retornou colunas de resultado.")
        colunas = [d.name for d in cursor.description]
        registros = cursor.fetchall()
        conexao.rollback()
        if not registros:
            raise ValueError(f"O SQL executou corretamente, mas retornou 0 registros entre {inicio} e {fim}.")
        df = pd.DataFrame.from_records(registros, columns=colunas)
        df.attrs["ambiente_postgresql"] = ambiente
        return df
    except Exception as erro:
        if conexao is not None:
            try: conexao.rollback()
            except Exception: pass
        detalhes=[f"Competência: {periodo}",f"Período SQL: {inicio} até {fim}",f"Tipo: {type(erro).__name__}",f"Mensagem: {erro}"]
        if getattr(erro,'pgcode',None): detalhes.append(f"SQLSTATE: {erro.pgcode}")
        if getattr(erro,'pgerror',None): detalhes.append(f"PostgreSQL: {erro.pgerror}")
        diag=getattr(erro,'diag',None)
        if diag is not None:
            for attr,label in [("severity","Severidade"),("message_primary","Mensagem principal"),("message_detail","Detalhe"),("message_hint","Sugestão"),("statement_position","Posição no SQL"),("context","Contexto"),("schema_name","Schema"),("table_name","Tabela"),("column_name","Coluna")]:
                val=getattr(diag,attr,None)
                if val: detalhes.append(f"{label}: {val}")
        detalhes += ["", "TRACEBACK:", traceback.format_exc()]
        texto="\n".join(detalhes)
        arquivo_erro.write_text(texto,encoding='utf-8')
        raise RuntimeError(texto) from erro
    finally:
        if cursor is not None:
            try: cursor.close()
            except Exception: pass
        if conexao is not None:
            try: conexao.close()
            except Exception: pass



def _gravar_log_entradas(texto):
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    caminho = LOG_DIR / "entradas_erros.log"
    with caminho.open("a", encoding="utf-8") as arquivo:
        arquivo.write(
            f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]\n{texto}\n"
        )
    return caminho


def executar_entradas_psycopg2(cfg, sql, data_inicio, data_fim, periodo):
    """Executa Entradas diretamente via psycopg2 e preserva o erro real do PostgreSQL."""
    inicio = datetime.strptime(str(data_inicio)[:10], "%Y-%m-%d").replace(
        hour=0, minute=0, second=0, microsecond=0
    )
    fim = datetime.strptime(str(data_fim)[:10], "%Y-%m-%d").replace(
        hour=23, minute=59, second=59, microsecond=999999
    )

    sql_psycopg2 = str(sql)
    sql_psycopg2 = sql_psycopg2.replace(":data_inicio", "%(data_inicio)s")
    sql_psycopg2 = sql_psycopg2.replace(":data_fim", "%(data_fim)s")
    sql_psycopg2 = sql_psycopg2.replace(
        ":periodo_referencia", "%(periodo_referencia)s"
    )

    conexao = None
    cursor = None
    inicio_execucao = time.perf_counter()
    try:
        conexao = psycopg2.connect(**_config_psycopg2(cfg))
        conexao.set_session(readonly=True, autocommit=False)
        cursor = conexao.cursor()
        cursor.execute(
            sql_psycopg2,
            {
                "data_inicio": inicio,
                "data_fim": fim,
                "periodo_referencia": str(periodo),
            },
        )
        colunas = [desc[0] for desc in cursor.description]
        linhas = cursor.fetchall()
        df = pd.DataFrame(linhas, columns=colunas)
        df.attrs["tempo_consulta_segundos"] = time.perf_counter() - inicio_execucao
        df.attrs["periodo_sql"] = (inicio.isoformat(), fim.isoformat())
        df.attrs["ambiente_postgresql"] = (
            cfg.get("banco") or cfg.get("database"),
            cfg.get("usuario") or cfg.get("user"),
        )
        conexao.rollback()
        return df
    except Exception as erro:
        try:
            if conexao is not None:
                conexao.rollback()
        except Exception:
            pass

        detalhes = [
            f"Fonte: Entradas",
            f"Competência: {periodo}",
            f"Período SQL: {inicio} até {fim}",
            f"Tipo: {type(erro).__name__}",
            f"Mensagem: {erro}",
        ]
        if getattr(erro, "pgcode", None):
            detalhes.append(f"SQLSTATE: {erro.pgcode}")
        if getattr(erro, "pgerror", None):
            detalhes.append(f"PostgreSQL: {erro.pgerror}")
        diag = getattr(erro, "diag", None)
        if diag is not None:
            for attr, label in [
                ("severity", "Severidade"),
                ("message_primary", "Mensagem principal"),
                ("message_detail", "Detalhe"),
                ("message_hint", "Sugestão"),
                ("statement_position", "Posição no SQL"),
                ("internal_position", "Posição interna"),
                ("context", "Contexto"),
                ("schema_name", "Schema"),
                ("table_name", "Tabela"),
                ("column_name", "Coluna"),
            ]:
                valor = getattr(diag, attr, None)
                if valor:
                    detalhes.append(f"{label}: {valor}")
        detalhes += ["", "TRACEBACK:", traceback.format_exc()]
        texto = "\n".join(detalhes)
        caminho = _gravar_log_entradas(texto)
        raise RuntimeError(
            f"Falha real do PostgreSQL na consulta de Entradas.\n{texto}\nLog: {caminho}"
        ) from erro
    finally:
        if cursor is not None:
            try:
                cursor.close()
            except Exception:
                pass
        if conexao is not None:
            try:
                conexao.close()
            except Exception:
                pass



def executar_entradas_reconstruidas(cfg, sql, data_inicio, data_fim, periodo):
    '''Executa o SQL original de Entradas pela conexão padrão do projeto.'''
    inicio_execucao = time.perf_counter()
    engine = None
    try:
        engine = criar_engine_banco(cfg)
        with engine.connect() as conn:
            df = executar_sql_com_fallback_datas(conn, sql, data_inicio, data_fim, periodo)
        if df is None or df.empty:
            raise ValueError('O SQL original de Entradas retornou 0 registros. A base anterior foi preservada.')
        if 'valor_nf_total' not in df.columns:
            df['valor_nf_total'] = pd.to_numeric(
                df.get('entrada_custo_total', pd.Series(index=df.index, dtype=float)),
                errors='coerce'
            ).fillna(0)
        df.attrs['tempo_consulta_segundos'] = time.perf_counter() - inicio_execucao
        df.attrs['metodo_execucao'] = 'SQLAlchemy + SQL original'
        return df
    except Exception as erro:
        detalhes = [
            'MÓDULO DE ENTRADAS RECONSTRUÍDO',
            f'Competência: {periodo}',
            f'Período: {data_inicio} até {data_fim}',
            'Método: SQLAlchemy + SQL original',
            f'Tipo: {type(erro).__name__}',
            f'Mensagem: {erro}',
        ]
        origem = getattr(erro, 'orig', None)
        if origem is not None:
            detalhes.append(f'Erro original do driver: {origem}')
            if getattr(origem, 'pgcode', None):
                detalhes.append(f'SQLSTATE: {origem.pgcode}')
        detalhes.extend(['', 'TRACEBACK:', traceback.format_exc()])
        texto_erro = '\n'.join(detalhes)
        caminho = _gravar_log_entradas(texto_erro)
        raise RuntimeError(f'{texto_erro}\nLog: {caminho}') from erro
    finally:
        if engine is not None:
            try:
                engine.dispose()
            except Exception:
                pass



def gerar_auditoria_regras_entradas(df, valor_referencia=0.0):
    """Compara fórmulas possíveis sem alterar os dados salvos."""
    if df is None or df.empty:
        raise ValueError("A consulta de Entradas não retornou registros para auditar.")

    base = df.copy()
    for coluna in [
        "valorunitario_bruto", "custo_bruto", "acrescimo_bruto",
        "quantidade_bruta", "quantidade_por_produto", "quant_embalagem",
        "entrada_custo_total", "custo_unit_r", "custo_final_r", "imposto_r",
    ]:
        if coluna not in base.columns:
            base[coluna] = 0.0
        base[coluna] = pd.to_numeric(base[coluna], errors="coerce").fillna(0.0)

    base["formula_valorunitario_x_quantidade"] = (
        base["valorunitario_bruto"] * base["quantidade_bruta"]
    )
    base["formula_custo_x_quantidade"] = (
        base["custo_bruto"] * base["quantidade_bruta"]
    )
    base["formula_valorunitario_menos_acrescimo_x_qtd"] = (
        (base["valorunitario_bruto"] - base["acrescimo_bruto"]) * base["quantidade_bruta"]
    )
    base["formula_custo_menos_acrescimo_x_qtd"] = (
        (base["custo_bruto"] - base["acrescimo_bruto"]) * base["quantidade_bruta"]
    )
    base["formula_custo_unitario_x_quantidade"] = (
        base["custo_final_r"] * base["quantidade_bruta"]
    )
    base["formula_valorunitario_unit_x_quantidade"] = (
        base["custo_unit_r"] * base["quantidade_bruta"]
    )

    formulas = [
        ("Atual do KPI: custo × quantidade", "formula_custo_x_quantidade"),
        ("Valor unitário × quantidade", "formula_valorunitario_x_quantidade"),
        ("(Valor unitário − acréscimo) × quantidade", "formula_valorunitario_menos_acrescimo_x_qtd"),
        ("(Custo − acréscimo) × quantidade", "formula_custo_menos_acrescimo_x_qtd"),
        ("Custo final unitário × quantidade", "formula_custo_unitario_x_quantidade"),
        ("Valor unitário por embalagem × quantidade", "formula_valorunitario_unit_x_quantidade"),
    ]

    resultados = []
    for nome, coluna in formulas:
        total = float(base[coluna].sum())
        resultados.append({
            "Regra": nome,
            "Linhas": int(len(base)),
            "Total": total,
            "Diferença para A7": total - float(valor_referencia or 0),
            "Distância absoluta": abs(total - float(valor_referencia or 0)),
            "Tipo": "Todas as linhas",
        })

    # A mesma comparação sem repetir o mesmo item de NF. Isso identifica
    # multiplicação causada por mais de uma classificação principal.
    chave_item = "item_nota_fiscal_id" if "item_nota_fiscal_id" in base.columns else None
    if chave_item:
        unicos = base.drop_duplicates(subset=[chave_item], keep="first").copy()
        for nome, coluna in formulas:
            total = float(unicos[coluna].sum())
            resultados.append({
                "Regra": nome,
                "Linhas": int(len(unicos)),
                "Total": total,
                "Diferença para A7": total - float(valor_referencia or 0),
                "Distância absoluta": abs(total - float(valor_referencia or 0)),
                "Tipo": "Itens únicos da NF",
            })

    resumo = pd.DataFrame(resultados).sort_values(
        ["Distância absoluta", "Tipo", "Regra"], ascending=[True, True, True]
    ).reset_index(drop=True)

    # Diagnóstico de duplicidade.
    duplicidade = pd.DataFrame()
    if chave_item:
        contagem = base.groupby(chave_item, dropna=False).size().reset_index(name="repetições")
        ids_dup = contagem.loc[contagem["repetições"] > 1, chave_item]
        if not ids_dup.empty:
            colunas = [c for c in [
                chave_item, "nota_fiscal_id", "numero_nf", "numero_loja",
                "cod_interno", "codigobarras", "descricao_embalagem",
                "classificacao_3_nivel", "cfop", "fornecedor",
                "quantidade_bruta", "valorunitario_bruto", "custo_bruto",
                "formula_custo_x_quantidade",
            ] if c in base.columns]
            duplicidade = base[base[chave_item].isin(ids_dup)][colunas].copy()
            duplicidade = duplicidade.sort_values([chave_item, "classificacao_3_nivel"] if "classificacao_3_nivel" in duplicidade.columns else [chave_item])

    # Totais por CFOP para encontrar operações incluídas/excluídas pelo A7.
    por_cfop = pd.DataFrame()
    if "cfop" in base.columns:
        por_cfop = (
            base.groupby("cfop", dropna=False)
            .agg(
                Linhas=("cfop", "size"),
                Itens=("quantidade_bruta", "sum"),
                Total_custo=("formula_custo_x_quantidade", "sum"),
                Total_valor_unitario=("formula_valorunitario_x_quantidade", "sum"),
            )
            .reset_index()
            .sort_values("Total_custo", ascending=False)
        )

    diagnostico = {
        "linhas": int(len(base)),
        "itens_unicos": int(base[chave_item].nunique()) if chave_item else int(len(base)),
        "linhas_duplicadas": int(len(base) - base[chave_item].nunique()) if chave_item else 0,
        "valor_referencia": float(valor_referencia or 0),
    }
    return resumo, duplicidade, por_cfop, diagnostico

def salvar_entradas_transacional(df, df_financeiro, periodo, data_inicio, data_fim):
    """Grava as camadas analítica e financeira na mesma transação SQLite."""
    if df is None or df.empty:
        raise ValueError("Entradas analíticas vazias; a base anterior foi preservada.")
    if df_financeiro is None or df_financeiro.empty:
        raise ValueError("Entradas financeiras vazias; a base anterior foi preservada.")

    tabela_analitica = "base_entradas"
    tabela_financeira = "base_entradas_financeira"
    sufixo = datetime.now().strftime("%Y%m%d%H%M%S%f")
    tmp_analitica = f"{tabela_analitica}_nova_{sufixo}"
    tmp_financeira = f"{tabela_financeira}_nova_{sufixo}"
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    analitica = df.copy()
    financeira = df_financeiro.copy()
    for base in (analitica, financeira):
        base["periodo_referencia"] = str(periodo)
        base["data_inicio_meta"] = str(data_inicio)
        base["data_fim_meta"] = str(data_fim)
        base["atualizado_em"] = agora

    # Mantém compatibilidade das telas analíticas antigas.
    if "valor_nf_total" not in analitica.columns:
        analitica["valor_nf_total"] = pd.to_numeric(
            analitica.get(
                "entrada_custo_total",
                pd.Series(index=analitica.index, dtype=float),
            ),
            errors="coerce",
        ).fillna(0)

    financeira["valor_nf_total"] = pd.to_numeric(
        financeira.get("valor_nf_total", pd.Series(index=financeira.index, dtype=float)),
        errors="coerce",
    ).fillna(0)
    financeira = financeira.drop_duplicates(subset=["nota_fiscal_id"], keep="first")

    valor_financeiro = float(financeira["valor_nf_total"].sum())
    valor_analitico = float(pd.to_numeric(
        analitica["valor_nf_total"], errors="coerce"
    ).fillna(0).sum())
    quantidade_analitica = float(pd.to_numeric(
        analitica.get(
            "quantidade_por_produto",
            pd.Series(index=analitica.index, dtype=float),
        ),
        errors="coerce",
    ).fillna(0).sum())

    con = conexao_cache()
    etapa = "início"
    try:
        etapa = "criação das tabelas temporárias"
        analitica.to_sql(tmp_analitica, con, if_exists="fail", index=False)
        financeira.to_sql(tmp_financeira, con, if_exists="fail", index=False)

        etapa = "validação das temporárias"
        val_an = con.execute(
            f'SELECT COUNT(*) FROM "{tmp_analitica}"'
        ).fetchone()
        val_fin = con.execute(
            f'SELECT COUNT(*), COALESCE(SUM(CAST(valor_nf_total AS REAL)),0), '
            f'COUNT(DISTINCT nota_fiscal_id) FROM "{tmp_financeira}"'
        ).fetchone()
        if int(val_an[0] or 0) != len(analitica):
            raise RuntimeError("Falha na validação da camada analítica temporária.")
        if int(val_fin[0] or 0) != len(financeira):
            raise RuntimeError("Falha na validação da camada financeira temporária.")
        if int(val_fin[2] or 0) != len(financeira):
            raise RuntimeError("Existem notas duplicadas na camada financeira.")
        if abs(float(val_fin[1] or 0) - valor_financeiro) > 0.01:
            raise RuntimeError("Falha na validação do valor financeiro temporário.")

        con.commit()
        etapa = "substituição atômica das duas camadas"
        con.execute("BEGIN IMMEDIATE")

        def substituir_periodo(tabela_destino, tabela_tmp):
            existe = con.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
                (tabela_destino,),
            ).fetchone()
            if not existe:
                con.execute(
                    f'CREATE TABLE "{tabela_destino}" AS '
                    f'SELECT * FROM "{tabela_tmp}" WHERE 0'
                )
            con.execute(
                f'DELETE FROM "{tabela_destino}" WHERE periodo_referencia = ?',
                (str(periodo),),
            )
            info_tmp = list(con.execute(f'PRAGMA table_info("{tabela_tmp}")'))
            cols_tmp = [r[1] for r in info_tmp]
            tipos_tmp = {r[1]: (r[2] or "TEXT") for r in info_tmp}
            cols_dest = [r[1] for r in con.execute(
                f'PRAGMA table_info("{tabela_destino}")'
            )]
            for coluna in cols_tmp:
                if coluna not in cols_dest:
                    con.execute(
                        f'ALTER TABLE "{tabela_destino}" ADD COLUMN '
                        f'"{coluna}" {tipos_tmp[coluna]}'
                    )
            cols_dest = [r[1] for r in con.execute(
                f'PRAGMA table_info("{tabela_destino}")'
            )]
            comuns = [c for c in cols_tmp if c in cols_dest]
            lista = ", ".join(f'"{c}"' for c in comuns)
            con.execute(
                f'INSERT INTO "{tabela_destino}" ({lista}) '
                f'SELECT {lista} FROM "{tabela_tmp}"'
            )

        substituir_periodo(tabela_analitica, tmp_analitica)
        substituir_periodo(tabela_financeira, tmp_financeira)

        etapa = "validação das tabelas definitivas"
        grav_an = con.execute(
            f'SELECT COUNT(*) FROM "{tabela_analitica}" '
            'WHERE periodo_referencia = ?',
            (str(periodo),),
        ).fetchone()
        grav_fin = con.execute(
            f'SELECT COUNT(*), COALESCE(SUM(CAST(valor_nf_total AS REAL)),0), '
            f'COUNT(DISTINCT nota_fiscal_id) FROM "{tabela_financeira}" '
            'WHERE periodo_referencia = ?',
            (str(periodo),),
        ).fetchone()
        if int(grav_an[0] or 0) != len(analitica):
            raise RuntimeError("A camada analítica definitiva ficou incompleta.")
        if int(grav_fin[0] or 0) != len(financeira):
            raise RuntimeError("A camada financeira definitiva ficou incompleta.")
        if int(grav_fin[2] or 0) != len(financeira):
            raise RuntimeError("A camada financeira definitiva contém notas duplicadas.")
        if abs(float(grav_fin[1] or 0) - valor_financeiro) > 0.01:
            raise RuntimeError("O total financeiro gravado diverge da consulta.")

        etapa = "atualização dos resumos"
        atualizar_resumos_analise(con, fonte="entradas", periodo=str(periodo))
        con.execute(
            """
            INSERT INTO atualizacoes
                (fonte, periodo_referencia, data_inicio, data_fim,
                 registros, status, mensagem, atualizado_em)
            VALUES (?, ?, ?, ?, ?, 'Sucesso', ?, ?)
            """,
            (
                "entradas",
                str(periodo),
                str(data_inicio),
                str(data_fim),
                len(analitica),
                (
                    f"Analítica: {len(analitica)} linhas / R$ {valor_analitico:.2f} • "
                    f"Financeira A7: {len(financeira)} notas / R$ {valor_financeiro:.2f}"
                ),
                agora,
            ),
        )
        con.execute(f'DROP TABLE IF EXISTS "{tmp_analitica}"')
        con.execute(f'DROP TABLE IF EXISTS "{tmp_financeira}"')
        con.commit()

        _limpar_cache_dados()
        st.cache_data.clear()
        st.session_state["_ultima_validacao_entradas"] = {
            "registros": len(analitica),
            "notas": len(financeira),
            "valor": valor_financeiro,
            "valor_analitico": valor_analitico,
            "quantidade": quantidade_analitica,
            "metodo": "Duas camadas: itens + nf.valortotal (A7)",
            "tempo_consulta": float(df.attrs.get("tempo_consulta_segundos", 0)),
        }
        return int(len(analitica))
    except Exception as erro:
        try:
            con.rollback()
        except Exception:
            pass
        for tabela_tmp in (tmp_analitica, tmp_financeira):
            try:
                con.execute(f'DROP TABLE IF EXISTS "{tabela_tmp}"')
                con.commit()
            except Exception:
                pass
        detalhe = (
            "MÓDULO DE ENTRADAS — FALHA NAS DUAS CAMADAS\n"
            f"Etapa: {etapa}\nPeríodo: {periodo}\n"
            f"Linhas analíticas: {len(analitica)}\n"
            f"Notas financeiras: {len(financeira)}\n"
            f"Valor financeiro: {valor_financeiro:.2f}\n"
            f"Tipo: {type(erro).__name__}\nMensagem: {erro}\n\n"
            f"{traceback.format_exc()}"
        )
        caminho = _gravar_log_entradas(detalhe)
        raise RuntimeError(f"{detalhe}\nLog: {caminho}") from erro
    finally:
        con.close()


def executar_sql_com_fallback_datas(conn, sql, data_inicio, data_fim, periodo):
    """Executa consultas incluindo integralmente o primeiro e o último dia.

    Antes, ``data_fim`` era enviada como uma data simples. No PostgreSQL,
    comparações como ``BETWEEN :data_inicio AND :data_fim`` convertiam o fim
    para 00:00:00 e excluíam praticamente todo o último dia do mês.
    """
    inicio_base = datetime.strptime(str(data_inicio)[:10], "%Y-%m-%d")
    fim_base = datetime.strptime(str(data_fim)[:10], "%Y-%m-%d")

    inicio_timestamp = inicio_base.replace(
        hour=0, minute=0, second=0, microsecond=0
    )
    fim_timestamp = fim_base.replace(
        hour=23, minute=59, second=59, microsecond=999999
    )

    return pd.read_sql_query(
        text(sql),
        conn,
        params={
            "data_inicio": inicio_timestamp,
            "data_fim": fim_timestamp,
            "periodo_referencia": str(periodo),
        },
    )



def importar_csv_anual_contas_pagar(arquivo):
    """Importa uma base anual e cria snapshots mensais pelo vencimento."""
    nome = str(getattr(arquivo, "name", "contas_pagar.csv"))
    arquivo.seek(0)
    try:
        bruto = pd.read_csv(
            arquivo,
            sep=None,
            engine="python",
            dtype=str,
            encoding="utf-8-sig",
        )
    except UnicodeDecodeError:
        arquivo.seek(0)
        bruto = pd.read_csv(
            arquivo,
            sep=None,
            engine="python",
            dtype=str,
            encoding="latin-1",
        )

    if bruto.empty:
        raise ValueError("O CSV de Contas a Pagar está vazio.")

    normalizado = normalizar_contas_pagar_df(bruto)
    vencimentos = pd.to_datetime(
        normalizado["data_vencimento"], errors="coerce"
    )
    normalizado = normalizado.loc[vencimentos.notna()].copy()
    vencimentos = vencimentos.loc[vencimentos.notna()]
    normalizado["periodo_referencia"] = vencimentos.dt.strftime("%Y-%m")

    if normalizado.empty:
        raise ValueError(
            "Não foi possível reconhecer a coluna Data de Vencimento no CSV."
        )

    resultados = []
    planos = sorted(
        p for p in normalizado["plano_contas"].dropna().astype(str).str.strip().unique()
        if p
    )

    for periodo, grupo in normalizado.groupby("periodo_referencia"):
        ano, mes = map(int, periodo.split("-"))
        import calendar
        inicio = f"{periodo}-01"
        fim = f"{periodo}-{calendar.monthrange(ano, mes)[1]:02d}"
        salvar_snapshot_mensal(
            grupo.drop(columns=["periodo_referencia"]),
            "contas_pagar",
            periodo,
            inicio,
            fim,
        )
        resultados.append({"Período": periodo, "Registros": len(grupo)})

    _gravar_log_contas_pagar(
        f"Importação CSV concluída: {nome}; "
        f"{len(normalizado)} registros; {len(resultados)} competências."
    )
    return pd.DataFrame(resultados), planos


def registrar_erro_atualizacao(fonte, periodo, data_inicio, data_fim, mensagem):
    con = conexao_cache()
    try:
        con.execute(
            """
            INSERT INTO atualizacoes
            (fonte, periodo_referencia, data_inicio, data_fim, registros, status, mensagem, atualizado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                fonte, periodo, data_inicio, data_fim,
                0, "Erro", str(mensagem)[:8000],
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
        )
        con.commit()
    finally:
        con.close()

def executar_atualizacao_fonte(fonte, cfg, periodo, data_inicio, data_fim):
    # Nunca confia em uma configuração mantida em memória: relê a conexão desta máquina.
    cfg = carregar_config_banco()
    validar_identidade_banco(cfg)
    info = FONTES_BANCO[fonte]
    sql = ler_sql(info["arquivo_sql"])

    if not str(sql).strip():
        raise ValueError(f"O SQL de {info['titulo']} está vazio.")

    engine = None

    try:
        if fonte == "contas_pagar":
            df = executar_contas_pagar_psycopg2(
                cfg,
                sql,
                data_inicio,
                data_fim,
                periodo,
            )
        elif fonte == "entradas":
            df = executar_entradas_reconstruidas(
                cfg,
                sql,
                data_inicio,
                data_fim,
                periodo,
            )
            sql_financeiro = ler_sql(SQL_DIR / "entradas_financeiras.sql")
            if not str(sql_financeiro).strip():
                raise ValueError("O SQL de Entradas Financeiras está vazio.")
            engine_financeiro = criar_engine_banco(cfg)
            try:
                with engine_financeiro.connect() as conn_financeira:
                    df_financeiro = executar_sql_com_fallback_datas(
                        conn_financeira,
                        sql_financeiro,
                        data_inicio,
                        data_fim,
                        periodo,
                    )
            finally:
                engine_financeiro.dispose()
            if df_financeiro is None or df_financeiro.empty:
                raise ValueError(
                    "A consulta financeira de Entradas retornou 0 notas. "
                    "A base anterior foi preservada."
                )
        else:
            engine = criar_engine_banco(cfg)
            with engine.connect() as conn:
                df = executar_sql_com_fallback_datas(
                    conn,
                    sql,
                    data_inicio,
                    data_fim,
                    periodo,
                )

        if fonte == "vendas" and not df.empty:
            # Diagnóstico do intervalo efetivamente retornado. Ajuda a impedir
            # que um mês seja considerado atualizado sem o último dia.
            coluna_data_venda = next(
                (
                    c for c in (
                        "datahora", "datahora_venda_final", "datahora_fechamento"
                    )
                    if c in df.columns
                ),
                None,
            )
            if coluna_data_venda:
                datas_venda = pd.to_datetime(
                    df[coluna_data_venda], errors="coerce"
                ).dropna()
                if not datas_venda.empty:
                    df.attrs["intervalo_vendas"] = {
                        "inicio": datas_venda.min().strftime("%Y-%m-%d %H:%M:%S"),
                        "fim": datas_venda.max().strftime("%Y-%m-%d %H:%M:%S"),
                    }

        if fonte == "contas_pagar":
            colunas_originais = [str(c) for c in df.columns]
            df = normalizar_contas_pagar_df(df)

            if df.empty:
                raise ValueError(
                    "O SQL retornou registros, mas a normalização resultou em base vazia. "
                    f"Colunas recebidas: {colunas_originais}"
                )

            if "plano_contas" not in df.columns:
                raise ValueError(
                    "A coluna Plano de Contas não foi reconhecida no resultado."
                )

        if fonte == "entradas" and df.empty:
            raise ValueError(
                "A consulta de Entradas foi executada, mas retornou 0 registros. "
                "A base anterior foi preservada e nenhuma substituição foi realizada."
            )

        if fonte == "contas_pagar":
            return salvar_contas_pagar_independente(
                df,
                periodo,
                data_inicio,
                data_fim,
            )
        if fonte == "entradas":
            return salvar_entradas_transacional(
                df,
                df_financeiro,
                periodo,
                data_inicio,
                data_fim,
            )

        salvar_snapshot_mensal(
            df,
            fonte,
            periodo,
            data_inicio,
            data_fim,
        )

        if fonte == "vendas":
            valor_consulta = float(
                pd.to_numeric(
                    df.get("valortotal", pd.Series(dtype=float)),
                    errors="coerce",
                ).fillna(0).sum()
            )
            quantidade_consulta = float(
                pd.to_numeric(
                    df.get("quantidade", pd.Series(dtype=float)),
                    errors="coerce",
                ).fillna(0).sum()
            )

            with conexao_cache() as con_validacao:
                gravado = con_validacao.execute(
                    """
                    SELECT
                        COUNT(*),
                        COALESCE(SUM(CAST(valortotal AS REAL)), 0),
                        COALESCE(SUM(CAST(quantidade AS REAL)), 0),
                        MIN(datahora_venda_final),
                        MAX(datahora_venda_final)
                    FROM base_vendas
                    WHERE periodo_referencia = ?
                    """,
                    (str(periodo),),
                ).fetchone()

            registros_gravados = int(gravado[0] or 0)
            valor_gravado = float(gravado[1] or 0)
            quantidade_gravada = float(gravado[2] or 0)

            if registros_gravados != len(df):
                raise RuntimeError(
                    "A validação da carga de Vendas falhou: "
                    f"consulta={len(df)} linhas e SQLite={registros_gravados} linhas."
                )
            if abs(valor_gravado - valor_consulta) > 0.01:
                raise RuntimeError(
                    "A validação do faturamento falhou: "
                    f"consulta={valor_consulta:.2f} e SQLite={valor_gravado:.2f}."
                )
            if abs(quantidade_gravada - quantidade_consulta) > 0.001:
                raise RuntimeError(
                    "A validação da quantidade vendida falhou: "
                    f"consulta={quantidade_consulta:.3f} e SQLite={quantidade_gravada:.3f}."
                )

            st.session_state["_ultima_validacao_vendas"] = {
                "periodo": str(periodo),
                "registros": registros_gravados,
                "quantidade": quantidade_gravada,
                "faturamento": valor_gravado,
                "data_inicial": str(gravado[3] or ""),
                "data_final": str(gravado[4] or ""),
            }

        return len(df)

    except Exception as erro:
        mensagem = str(erro)
        registrar_erro_atualizacao(
            fonte, periodo, data_inicio, data_fim, mensagem,
        )
        if fonte == "contas_pagar":
            raise
        raise RuntimeError(
            f"Falha na consulta de {info['titulo']}:\n{mensagem}"
        ) from erro

    finally:
        if engine is not None:
            engine.dispose()



def status_configuracao_fontes():
    resultado = {}
    for codigo, info in FONTES_BANCO.items():
        sql = ler_sql(info["arquivo_sql"])
        configurado, mensagem = diagnosticar_sql_fonte(codigo, sql)
        resultado[codigo] = {
            "configurado": configurado,
            "mensagem": mensagem,
        }
    return resultado


def historico_atualizacoes():
    con = conexao_cache()
    try:
        return pd.read_sql_query(
            "SELECT * FROM atualizacoes ORDER BY id DESC",
            con
        )
    finally:
        con.close()

def carregar_snapshot(fonte, periodo):
    tabela = FONTES_BANCO[fonte]["tabela_cache"]
    con = conexao_cache()
    try:
        try:
            return pd.read_sql_query(
                f'SELECT * FROM "{tabela}" WHERE periodo_referencia = ?',
                con,
                params=(periodo,)
            )
        except Exception:
            return pd.DataFrame()
    finally:
        con.close()


def garantir_indices_cache():
    """Índices mínimos e objetivos para acelerar consultas por competência."""
    con = conexao_cache()
    try:
        tabelas = ["base_vendas", "base_entradas", "base_estoque", "base_contas_pagar"]
        candidatas_class = [
            "classificacao_resumida", "classificacao_3_nivel",
            "classificacao_geral", "classificacao_principal", "classificacao"
        ]
        for tabela in tabelas:
            try:
                cols = {r[1] for r in con.execute(f'PRAGMA table_info("{tabela}")').fetchall()}
                if "periodo_referencia" not in cols:
                    continue
                con.execute(
                    f'CREATE INDEX IF NOT EXISTS "idx_{tabela}_periodo" '
                    f'ON "{tabela}" (periodo_referencia)'
                )
                col_class = next((c for c in candidatas_class if c in cols), None)
                if col_class:
                    con.execute(
                        f'CREATE INDEX IF NOT EXISTS "idx_{tabela}_periodo_class" '
                        f'ON "{tabela}" (periodo_referencia, "{col_class}")'
                    )
            except Exception:
                continue
        con.commit()
    finally:
        con.close()


garantir_sqls()
# PERFORMANCE: não criar índices sobre milhões de linhas durante a abertura.
# A manutenção de índices deve ocorrer somente após atualização manual do banco.
CONFIG_BANCO = carregar_config_banco()
if (
    str(CONFIG_BANCO.get("host", "")).strip()
    and not _origem_cache_corresponde(CONFIG_BANCO)
):
    st.session_state["_banco_alterado_pendente"] = True


# =========================================================
# RUPTURA AUTOMÁTICA POR PERÍODO DA META
# =========================================================

PASTA_RUPTURA_AUTO = Path("IMPORTAR_RUPTURA")
PASTA_RUPTURA_AUTO.mkdir(exist_ok=True)
RUPTURA_AUTO_DB = DATA_DIR / "ruptura_mensal.sqlite"
RUPTURA_AUTO_CONTROLE = DATA_DIR / "controle_ruptura_auto.json"

MAPA_COMPRADORES_RUPTURA = {
    "PRINCIPAL > NAO MED > BRINQUEDOS": "Francieli",
    "PRINCIPAL > NAO MED > PERFUMARIA": "Francieli",
    "PRINCIPAL > NAO MED > PERFUMARIA COMBATE": "Francieli",
    "PRINCIPAL > NAO MED > PERFUMARIA ENCALHADOS": "Francieli",
    "PRINCIPAL > NAO MED > CONVENIÊNCIA": "Francieli",
    "PRINCIPAL > NAO MED > SUPLEMENTOS": "Francieli",
    "PRINCIPAL > MED > ÉTICOS": "Paulo",
    "PRINCIPAL > NAO MED > MASCARA": "Sebastião",
    "PRINCIPAL > NAO MED > FRALDAS": "Sebastião",
    "PRINCIPAL > MED > GEN - SIM": "Sebastião",
    "PRINCIPAL > MED > MEDICAMENTO CURVA D (TOP 15)": "Sebastião",
    "PRINCIPAL > MED > NATURAIS": "Sebastião",
    "PRINCIPAL > MED > PRÓPRIOS": "Sebastião",
    "PRINCIPAL > NAO MED > HOSPITALARES": "Sebastião",
    "PRINCIPAL > NAO MED > LEITES": "Sebastião",
    "PRINCIPAL > NAO MED > VAREJO": "Sebastião",
}

def _hash_arquivo(caminho):
    h = hashlib.sha256()
    with open(caminho, "rb") as f:
        for bloco in iter(lambda: f.read(1024 * 1024), b""):
            h.update(bloco)
    return h.hexdigest()

def _numero(serie):
    if pd.api.types.is_numeric_dtype(serie):
        return pd.to_numeric(serie, errors="coerce").fillna(0)
    s = serie.astype(str).str.strip()
    tem_virgula = s.str.contains(",", regex=False)
    r = pd.Series(index=s.index, dtype="float64")
    r.loc[tem_virgula] = pd.to_numeric(
        s.loc[tem_virgula].str.replace(".", "", regex=False).str.replace(",", ".", regex=False),
        errors="coerce"
    )
    r.loc[~tem_virgula] = pd.to_numeric(s.loc[~tem_virgula], errors="coerce")
    return r.fillna(0)

def _comprador(classificacao):
    chave = " ".join(str(classificacao or "").strip().upper().split())
    mapa = {" ".join(k.upper().split()): v for k, v in MAPA_COMPRADORES_RUPTURA.items()}
    return mapa.get(chave, "Não mapeado")

def _arquivo_mais_recente():
    arquivos = []
    for padrao in ("*.xlsx", "*.xls", "*.csv"):
        arquivos.extend(PASTA_RUPTURA_AUTO.glob(padrao))
    arquivos = [p for p in arquivos if not p.name.startswith("~$")]
    return max(arquivos, key=lambda p: p.stat().st_mtime) if arquivos else None

def _ler_modelo_ruptura(caminho):
    if caminho.suffix.lower() == ".csv":
        try:
            bruto = pd.read_csv(caminho, sep=";", encoding="utf-8-sig")
        except Exception:
            bruto = pd.read_csv(caminho)
    else:
        bruto = pd.read_excel(caminho)

    bruto.columns = [str(c).strip() for c in bruto.columns]
    obrigatorias = [
        "Un. Neg.", "Apelido Un. Neg.", "Produto", "Fabricante",
        "Ruptura Venda", "Necessidade", "Estoque", "Custo Médio",
        "Curva Valor", "Cód. de Barras", "Classificação Principal"
    ]
    faltantes = [c for c in obrigatorias if c not in bruto.columns]
    if faltantes:
        raise ValueError("Colunas ausentes: " + ", ".join(faltantes))

    out = pd.DataFrame()
    out["Loja"] = bruto["Un. Neg."].astype(str).str.strip()
    out["Apelido Loja"] = bruto["Apelido Un. Neg."].astype(str).str.strip()
    out["Produto"] = bruto["Produto"].astype(str).str.strip()
    out["Fabricante"] = bruto["Fabricante"].astype(str).str.strip()
    out["EAN"] = bruto["Cód. de Barras"].astype(str).str.replace(r"\.0$", "", regex=True)
    out["Classificação Principal"] = bruto["Classificação Principal"].astype(str).str.strip()
    out["Comprador"] = out["Classificação Principal"].map(_comprador)
    out["Curva Valor"] = bruto["Curva Valor"].astype(str).str.strip()
    out["Curva Qtd."] = bruto["Curva Qtd."].astype(str).str.strip() if "Curva Qtd." in bruto else ""
    out["Ruptura Venda"] = _numero(bruto["Ruptura Venda"])
    out["Necessidade"] = _numero(bruto["Necessidade"])
    out["Estoque"] = _numero(bruto["Estoque"])
    out["Custo Médio"] = _numero(bruto["Custo Médio"])
    out["Valor Necessidade Custo"] = out["Necessidade"] * out["Custo Médio"]
    out["Valor Ruptura"] = out["Ruptura Venda"]
    out = out[(out["Valor Ruptura"] != 0) | (out["Necessidade"] != 0) | (out["Estoque"] != 0)]
    return out

def _con_ruptura():
    con = sqlite3.connect(RUPTURA_AUTO_DB)
    con.execute("""
        CREATE TABLE IF NOT EXISTS importacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            periodo_referencia TEXT,
            arquivo TEXT,
            hash_arquivo TEXT,
            registros INTEGER,
            valor_ruptura REAL,
            importado_em TEXT
        )
    """)
    con.commit()
    return con

def processar_ruptura_automatica(periodo=None, forcar=False):
    periodo = periodo or METAS_GESTOR.get("periodo_referencia", "")
    arquivo = _arquivo_mais_recente()
    if arquivo is None:
        return {"status": "sem_arquivo", "mensagem": "Nenhum arquivo na pasta IMPORTAR_RUPTURA."}

    hash_atual = _hash_arquivo(arquivo)
    controle = {}
    if RUPTURA_AUTO_CONTROLE.exists():
        try:
            controle = json.loads(RUPTURA_AUTO_CONTROLE.read_text(encoding="utf-8"))
        except Exception:
            controle = {}

    chave = f"{periodo}|{arquivo.name}"
    if not forcar and controle.get(chave) == hash_atual:
        return {"status": "sem_alteracao", "mensagem": "Arquivo já processado.", "arquivo": arquivo.name}

    df = _ler_modelo_ruptura(arquivo)
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    base = df.copy()
    base["periodo_referencia"] = periodo
    base["arquivo_origem"] = arquivo.name
    base["importado_em"] = agora

    con = _con_ruptura()
    try:
        try:
            con.execute("DELETE FROM ruptura_detalhe WHERE periodo_referencia = ?", (periodo,))
            con.commit()
        except Exception:
            pass
        base.to_sql("ruptura_detalhe", con, if_exists="append", index=False)
        con.execute(
            "INSERT INTO importacoes(periodo_referencia,arquivo,hash_arquivo,registros,valor_ruptura,importado_em) VALUES(?,?,?,?,?,?)",
            (periodo, arquivo.name, hash_atual, len(base), float(base["Valor Ruptura"].sum()), agora)
        )
        con.commit()
    finally:
        con.close()

    controle[chave] = hash_atual
    controle["ultima_importacao"] = {
        "periodo": periodo,
        "arquivo": arquivo.name,
        "registros": len(base),
        "valor_ruptura": float(base["Valor Ruptura"].sum()),
        "data": datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    }
    RUPTURA_AUTO_CONTROLE.write_text(json.dumps(controle, ensure_ascii=False, indent=2), encoding="utf-8")
    _registrar_atualizacao_dados("Ruptura", periodo, len(base))
    return {"status": "importado", **controle["ultima_importacao"]}

@st.cache_data(ttl=600, show_spinner=False)
def carregar_ruptura_auto(periodo, token_dados=None):
    con = _con_ruptura()
    try:
        try:
            base = pd.read_sql_query(
                "SELECT * FROM ruptura_detalhe WHERE periodo_referencia = ?",
                con, params=(periodo,)
            )
        except Exception:
            return pd.DataFrame()
    finally:
        con.close()

    # Recalcula sempre o comprador pela Classificação Principal.
    # Isso corrige também registros antigos que foram gravados como
    # "Não mapeado", sem exigir uma nova importação da planilha.
    if not base.empty:
        coluna_classificacao = next(
            (
                coluna for coluna in [
                    "Classificação Principal", "classificacao principal",
                    "classificacao_principal", "classificacao_geral",
                    "classificacao"
                ]
                if coluna in base.columns
            ),
            None
        )
        if coluna_classificacao:
            base["Comprador"] = base[coluna_classificacao].map(_mapear_comprador)
        elif "Comprador" not in base.columns:
            base["Comprador"] = "Não mapeado"

    return base

def historico_ruptura_auto():
    con = _con_ruptura()
    try:
        return pd.read_sql_query("SELECT * FROM importacoes ORDER BY id DESC", con)
    finally:
        con.close()

# Executa uma vez ao abrir e não repete se o arquivo não mudou.
RESULTADO_AUTO_RUPTURA = processar_ruptura_automatica()


# =========================================================
# GESTÃO EDITÁVEL DE COMPRADORES POR CLASSIFICAÇÃO
# =========================================================

MAPA_COMPRADORES_FILE = DATA_DIR / "mapa_compradores_editavel.json"

MAPA_COMPRADORES_PADRAO = [
    {"Área": "Higiene e Beleza", "Classificação Principal": "PRINCIPAL > NAO MED > BRINQUEDOS", "Comprador": "Francieli"},
    {"Área": "Higiene e Beleza", "Classificação Principal": "PRINCIPAL > NAO MED > PERFUMARIA", "Comprador": "Francieli"},
    {"Área": "Higiene e Beleza", "Classificação Principal": "PRINCIPAL > NAO MED > PERFUMARIA COMBATE", "Comprador": "Francieli"},
    {"Área": "Higiene e Beleza", "Classificação Principal": "PRINCIPAL > NAO MED > PERFUMARIA ENCALHADOS", "Comprador": "Francieli"},
    {"Área": "Diversos", "Classificação Principal": "PRINCIPAL > NAO MED > CONVENIÊNCIA", "Comprador": "Francieli"},
    {"Área": "Diversos", "Classificação Principal": "PRINCIPAL > NAO MED > SUPLEMENTOS", "Comprador": "Francieli"},
    {"Área": "Propagados", "Classificação Principal": "PRINCIPAL > MED > ÉTICOS", "Comprador": "Paulo"},
    {"Área": "Higiene e Beleza", "Classificação Principal": "PRINCIPAL > NAO MED > MASCARA", "Comprador": "Não mapeado"},
    {"Área": "Higiene e Beleza", "Classificação Principal": "PRINCIPAL > NAO MED > FRALDAS", "Comprador": "Não mapeado"},
    {"Área": "Diversos", "Classificação Principal": "PRINCIPAL > MED > GEN - SIM", "Comprador": "Não mapeado"},
    {"Área": "Diversos", "Classificação Principal": "PRINCIPAL > MED > MEDICAMENTO CURVA D (TOP 15)", "Comprador": "Não mapeado"},
    {"Área": "Diversos", "Classificação Principal": "PRINCIPAL > MED > NATURAIS", "Comprador": "Não mapeado"},
    {"Área": "Diversos", "Classificação Principal": "PRINCIPAL > MED > PRÓPRIOS", "Comprador": "Não mapeado"},
    {"Área": "Diversos", "Classificação Principal": "PRINCIPAL > NAO MED > HOSPITALARES", "Comprador": "Não mapeado"},
    {"Área": "Diversos", "Classificação Principal": "PRINCIPAL > NAO MED > LEITES", "Comprador": "Não mapeado"},
    {"Área": "Diversos", "Classificação Principal": "PRINCIPAL > NAO MED > VAREJO", "Comprador": "Não mapeado"},
]

def carregar_mapa_compradores_editavel():
    if MAPA_COMPRADORES_FILE.exists():
        try:
            dados = json.loads(MAPA_COMPRADORES_FILE.read_text(encoding="utf-8"))
            if isinstance(dados, list):
                return dados
        except Exception:
            pass
    MAPA_COMPRADORES_FILE.write_text(
        json.dumps(MAPA_COMPRADORES_PADRAO, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    return MAPA_COMPRADORES_PADRAO.copy()

def salvar_mapa_compradores_editavel(dados):
    MAPA_COMPRADORES_FILE.write_text(
        json.dumps(dados, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

def mapa_compradores_dict():
    dados = carregar_mapa_compradores_editavel()
    return {
        _normalizar_classificacao(item.get("Classificação Principal", "")): item.get("Comprador", "Não mapeado")
        for item in dados
        if item.get("Classificação Principal")
    }

MAPA_COMPRADORES_EDITAVEL = carregar_mapa_compradores_editavel()



# =========================================================
# AUDITORIA DE COBERTURA DAS CATEGORIAS POR COMPRADOR
# =========================================================

@st.cache_data(ttl=300, show_spinner=False, max_entries=12)
def carregar_auditoria_categorias(periodo_referencia, token_banco, token_mapa):
    """Consolida categorias reais das bases sem carregar as tabelas brutas."""
    colunas_saida = [
        "Classificação Principal", "Comprador atual", "Situação",
        "Produtos", "Itens vendidos", "Venda", "Estoque", "Entradas",
        "Fontes"
    ]
    if not CACHE_DB_FILE.exists():
        return pd.DataFrame(columns=colunas_saida)

    acumulado = {}

    def adicionar(classificacao, produtos=0, itens=0, venda=0, estoque=0, entradas=0, fonte=""):
        texto = str(classificacao or "").strip()
        if not texto:
            texto = "SEM CLASSIFICAÇÃO"
        chave = _normalizar_classificacao(texto) or "SEM CLASSIFICACAO"
        item = acumulado.setdefault(chave, {
            "classificacao": texto,
            "produtos": 0,
            "itens": 0.0,
            "venda": 0.0,
            "estoque": 0.0,
            "entradas": 0.0,
            "fontes": set(),
        })
        # Mantém o caminho mais completo quando a mesma chave aparecer em formatos diferentes.
        if len(texto) > len(item["classificacao"]):
            item["classificacao"] = texto
        item["produtos"] += int(produtos or 0)
        item["itens"] += float(itens or 0)
        item["venda"] += float(venda or 0)
        item["estoque"] += float(estoque or 0)
        item["entradas"] += float(entradas or 0)
        if fonte:
            item["fontes"].add(fonte)

    try:
        with conexao_cache() as con:
            tabelas = {
                r[0] for r in con.execute(
                    "SELECT name FROM sqlite_master WHERE type='table'"
                ).fetchall()
            }

            if "base_vendas" in tabelas:
                cols = {r[1] for r in con.execute('PRAGMA table_info("base_vendas")')}
                col_class = next((c for c in [
                    "classificacao_resumida", "classificacao_3_nivel",
                    "classificacao_principal", "classificacao_geral", "classificacao"
                ] if c in cols), None)
                col_prod = next((c for c in ["cod_interno", "codigobarras", "descricao"] if c in cols), None)
                if col_class:
                    prod_expr = f'COUNT(DISTINCT CAST("{col_prod}" AS TEXT))' if col_prod else '0'
                    itens_expr = 'SUM(COALESCE(CAST(quantidade AS REAL),0))' if 'quantidade' in cols else '0'
                    venda_expr = 'SUM(COALESCE(CAST(valortotal AS REAL),0))' if 'valortotal' in cols else '0'
                    periodo_where = 'WHERE periodo_referencia = ?' if 'periodo_referencia' in cols else ''
                    params = (str(periodo_referencia),) if periodo_where else ()
                    sql = f"""SELECT COALESCE(CAST("{col_class}" AS TEXT), ''), {prod_expr}, {itens_expr}, {venda_expr}
                              FROM base_vendas {periodo_where}
                              GROUP BY COALESCE(CAST("{col_class}" AS TEXT), '')"""
                    for classificacao, produtos, itens, venda in con.execute(sql, params):
                        adicionar(classificacao, produtos, itens, venda, fonte="Vendas")

            if "base_estoque" in tabelas:
                cols = {r[1] for r in con.execute('PRAGMA table_info("base_estoque")')}
                col_class = next((c for c in [
                    "classificacao_geral", "classificao", "classificacao_3_nivel",
                    "classificacao_principal", "classificacao"
                ] if c in cols), None)
                col_prod = next((c for c in ["cod_int", "cod_barras", "descricao"] if c in cols), None)
                if col_class:
                    prod_expr = f'COUNT(DISTINCT CAST("{col_prod}" AS TEXT))' if col_prod else '0'
                    estoque_expr = 'SUM(COALESCE(CAST(estoque AS REAL),0) * COALESCE(CAST(custo_medio_atual AS REAL), CAST(custo_unit_atual AS REAL),0))' if 'estoque' in cols else '0'
                    periodo_where = 'WHERE periodo_referencia = ?' if 'periodo_referencia' in cols else ''
                    params = (str(periodo_referencia),) if periodo_where else ()
                    sql = f"""SELECT COALESCE(CAST("{col_class}" AS TEXT), ''), {prod_expr}, {estoque_expr}
                              FROM base_estoque {periodo_where}
                              GROUP BY COALESCE(CAST("{col_class}" AS TEXT), '')"""
                    for classificacao, produtos, estoque in con.execute(sql, params):
                        adicionar(classificacao, produtos, estoque=estoque, fonte="Estoque")

            if "base_entradas" in tabelas:
                cols = {r[1] for r in con.execute('PRAGMA table_info("base_entradas")')}
                col_class = next((c for c in [
                    "classificacao_3_nivel", "classificacao_resumida",
                    "classificacao_principal", "classificacao_geral", "classificacao"
                ] if c in cols), None)
                col_prod = next((c for c in ["cod_interno", "codigobarras", "descricao_embalagem"] if c in cols), None)
                if col_class:
                    prod_expr = f'COUNT(DISTINCT CAST("{col_prod}" AS TEXT))' if col_prod else '0'
                    entrada_expr = 'SUM(COALESCE(CAST(entrada_custo_total AS REAL),0))' if 'entrada_custo_total' in cols else '0'
                    periodo_where = 'WHERE periodo_referencia = ?' if 'periodo_referencia' in cols else ''
                    params = (str(periodo_referencia),) if periodo_where else ()
                    sql = f"""SELECT COALESCE(CAST("{col_class}" AS TEXT), ''), {prod_expr}, {entrada_expr}
                              FROM base_entradas {periodo_where}
                              GROUP BY COALESCE(CAST("{col_class}" AS TEXT), '')"""
                    for classificacao, produtos, entradas in con.execute(sql, params):
                        adicionar(classificacao, produtos, entradas=entradas, fonte="Entradas")
    except Exception:
        return pd.DataFrame(columns=colunas_saida)

    mapa_bruto = carregar_mapa_compradores_editavel()
    ativos = {n.casefold(): n for n in lista_compradores_ativos()}
    por_chave = {}
    for registro in mapa_bruto:
        caminho = str(registro.get("Classificação Principal", "")).strip()
        comprador = str(registro.get("Comprador", "")).strip()
        chave = _normalizar_classificacao(caminho)
        if chave:
            por_chave.setdefault(chave, []).append(comprador)

    linhas = []
    for chave, item in acumulado.items():
        classificacao = item["classificacao"]
        compradores_exatos = [c for c in por_chave.get(chave, []) if c]
        compradores_unicos = sorted(set(compradores_exatos), key=str.casefold)
        comprador_mapeado = _mapear_comprador(classificacao)

        if classificacao == "SEM CLASSIFICAÇÃO":
            situacao = "Sem classificação"
            comprador_atual = ""
        elif len({c.casefold() for c in compradores_unicos}) > 1:
            situacao = "Duplicidade"
            comprador_atual = " / ".join(compradores_unicos)
        elif comprador_mapeado in ("", "Não mapeado", "Nao mapeado"):
            situacao = "Sem comprador"
            comprador_atual = ""
        elif comprador_mapeado.casefold() not in ativos:
            situacao = "Comprador inativo"
            comprador_atual = comprador_mapeado
        else:
            situacao = "OK"
            comprador_atual = ativos[comprador_mapeado.casefold()]

        linhas.append({
            "Classificação Principal": classificacao,
            "Comprador atual": comprador_atual,
            "Situação": situacao,
            "Produtos": int(item["produtos"]),
            "Itens vendidos": float(item["itens"]),
            "Venda": float(item["venda"]),
            "Estoque": float(item["estoque"]),
            "Entradas": float(item["entradas"]),
            "Fontes": ", ".join(sorted(item["fontes"])),
        })

    if not linhas:
        return pd.DataFrame(columns=colunas_saida)
    return pd.DataFrame(linhas).sort_values(
        ["Situação", "Venda", "Classificação Principal"],
        ascending=[True, False, True]
    ).reset_index(drop=True)


def salvar_vinculos_auditoria(classificacoes, comprador):
    """Substitui vínculos exatos das categorias selecionadas e preserva os demais."""
    selecionadas = {
        _normalizar_classificacao(c): str(c).strip()
        for c in classificacoes if str(c).strip()
    }
    mapa = carregar_mapa_compradores_editavel()
    novo = []
    for item in mapa:
        chave = _normalizar_classificacao(item.get("Classificação Principal", ""))
        if chave not in selecionadas:
            novo.append(item)
    for chave, texto in selecionadas.items():
        partes = [p.strip() for p in texto.split(">") if p.strip()]
        area = partes[1] if len(partes) > 1 else (partes[0] if partes else "")
        novo.append({
            "Área": area,
            "Classificação Principal": texto,
            "Comprador": str(comprador).strip(),
        })
    salvar_mapa_compradores_editavel(novo)
    st.cache_data.clear()
    return len(selecionadas)


# =========================================================
# CADASTRO EDITÁVEL DE COMPRADORES
# =========================================================

COMPRADORES_FILE = DATA_DIR / "cadastro_compradores.json"

COMPRADORES_PADRAO = [
    {"Comprador": "Geane", "Status": "Ativo"},
    {"Comprador": "Renato", "Status": "Ativo"},
]

def carregar_cadastro_compradores():
    """Carrega o cadastro persistente, aceitando inclusive uma lista vazia.

    Uma lista vazia significa que o usuário excluiu todos os compradores e não
    deve provocar restauração automática dos nomes padrão.
    """
    if COMPRADORES_FILE.exists():
        try:
            dados = json.loads(COMPRADORES_FILE.read_text(encoding="utf-8"))
            if isinstance(dados, list):
                return dados
        except Exception:
            pass
    COMPRADORES_FILE.parent.mkdir(parents=True, exist_ok=True)
    COMPRADORES_FILE.write_text(
        json.dumps(COMPRADORES_PADRAO, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    return [dict(item) for item in COMPRADORES_PADRAO]

def salvar_cadastro_compradores(dados):
    """Salva o cadastro de modo atômico e invalida as visões dependentes."""
    COMPRADORES_FILE.parent.mkdir(parents=True, exist_ok=True)
    temporario = COMPRADORES_FILE.with_suffix('.json.tmp')
    temporario.write_text(
        json.dumps(list(dados or []), ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    temporario.replace(COMPRADORES_FILE)
    _limpar_cache_dados()
    for chave in [
        "editor_cadastro_compradores",
        "_dados_visoes",
        "_chave_visoes",
        "_premiacao_analitica_cache",
    ]:
        st.session_state.pop(chave, None)

def lista_compradores_ativos():
    return [
        str(item.get("Comprador", "")).strip()
        for item in carregar_cadastro_compradores()
        if str(item.get("Status", "Ativo")).strip() == "Ativo"
        and str(item.get("Comprador", "")).strip()
    ]

COMPRADORES = lista_compradores_ativos()


def _conjunto_compradores_ativos():
    return {nome.casefold(): nome for nome in lista_compradores_ativos()}


def filtrar_dataframe_compradores_ativos(df):
    if not isinstance(df, pd.DataFrame) or df.empty or "Comprador" not in df.columns:
        return df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
    ativos = set(_conjunto_compradores_ativos())
    nomes = df["Comprador"].astype(str).str.strip().str.casefold()
    return df.loc[nomes.isin(ativos)].copy()


# =========================================================
# METAS EDITÁVEIS POR COMPRADOR E PERÍODO
# =========================================================

METAS_COMPRADORES_FILE = DATA_DIR / "metas_por_comprador.json"

def estrutura_meta_comprador_padrao(comprador, periodo, participacao_inicial=0.0):
    # Compradores e participações não são fixos no código. Para compradores
    # recém-identificados na base, a participação inicial pode ser calculada
    # pelo faturamento real do período e depois alterada na Gestão de Metas.
    participacao = float(participacao_inicial or 0.0)
    meta_venda = float(METAS_GESTOR.get("meta_venda_total_mes", 0)) * participacao / 100.0
    meta_cmv_pct = float(METAS_GESTOR.get("meta_cmv_mes", 0))
    meta_cmv = meta_venda * meta_cmv_pct / 100.0
    meta_estoque = meta_cmv * float(METAS_GESTOR.get("fator_cobertura", 0))

    return {
        "periodo_referencia": periodo,
        "comprador": comprador,
        "meta_venda": meta_venda,
        "participacao_venda_pct": participacao,
        "meta_cmv_pct": meta_cmv_pct,
        "meta_cmv_valor": meta_cmv,
        "fator_cobertura": float(METAS_GESTOR.get("fator_cobertura", 0)),
        "meta_estoque_total": meta_estoque,
        "meta_curva_a_pct": float(METAS_GESTOR.get("curva_a", 0)),
        "meta_curva_b_pct": float(METAS_GESTOR.get("curva_b", 0)),
        "meta_curva_c_pct": float(METAS_GESTOR.get("curva_c", 0)),
        "meta_curva_d_pct": float(METAS_GESTOR.get("curva_d", 0)),
        "meta_ruptura_pct": float(METAS_GESTOR.get("meta_ruptura", 0)),
        "meta_reposicao_pct": float(METAS_GESTOR.get("meta_reposicao", 0)),
        "valor_premio": float(METAS_GESTOR.get("valor_premio_total", 0)),
        "status": "Ativa",
        "ultima_atualizacao": "",
    }

def carregar_metas_por_comprador():
    if METAS_COMPRADORES_FILE.exists():
        try:
            dados = json.loads(METAS_COMPRADORES_FILE.read_text(encoding="utf-8"))
            if isinstance(dados, list):
                return dados
        except Exception:
            pass
    return []

def salvar_metas_por_comprador(dados):
    METAS_COMPRADORES_FILE.write_text(
        json.dumps(dados, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )


def _nome_comprador_valido(valor):
    texto = str(valor or "").strip()
    return texto and texto.upper() not in {
        "NÃO MAPEADO", "NAO MAPEADO", "NAN", "NONE", "NULL"
    }


def sincronizar_compradores_reconhecidos(nomes):
    """Inclui no cadastro compradores encontrados nas bases, sem apagar os existentes."""
    atuais = carregar_cadastro_compradores()
    conhecidos = {
        str(item.get("Comprador", "")).strip().casefold()
        for item in atuais if str(item.get("Comprador", "")).strip()
    }
    alterado = False
    for nome in sorted({str(x).strip() for x in nomes if _nome_comprador_valido(x)}):
        if nome.casefold() not in conhecidos:
            atuais.append({"Comprador": nome, "Status": "Ativo", "Origem": "Base de dados"})
            conhecidos.add(nome.casefold())
            alterado = True
    if alterado:
        salvar_cadastro_compradores(atuais)
    return atuais


def garantir_meta_comprador_periodo(comprador, periodo, participacao_inicial=0.0):
    dados = carregar_metas_por_comprador()
    for item in dados:
        if (
            str(item.get("periodo_referencia", "")) == str(periodo)
            and str(item.get("comprador", "")).strip().casefold() == str(comprador).strip().casefold()
        ):
            return item
    novo = estrutura_meta_comprador_padrao(comprador, periodo, participacao_inicial)
    dados.append(novo)
    salvar_metas_por_comprador(dados)
    return novo

def garantir_metas_compradores_periodo(periodo):
    dados = carregar_metas_por_comprador()
    ativos = lista_compradores_ativos()
    existentes = {
        (str(x.get("periodo_referencia", "")), str(x.get("comprador", "")))
        for x in dados
    }
    alterado = False

    for comprador in ativos:
        chave = (periodo, comprador)
        if chave not in existentes:
            dados.append(estrutura_meta_comprador_padrao(comprador, periodo))
            alterado = True

    if alterado:
        salvar_metas_por_comprador(dados)
    return dados

def obter_meta_comprador(comprador, periodo):
    dados = garantir_metas_compradores_periodo(periodo)
    for item in dados:
        if (
            str(item.get("periodo_referencia", "")) == str(periodo)
            and str(item.get("comprador", "")) == str(comprador)
        ):
            return item
    return estrutura_meta_comprador_padrao(comprador, periodo)

def atualizar_nome_comprador_metas(nome_antigo, nome_novo):
    dados = carregar_metas_por_comprador()
    alterado = False
    for item in dados:
        if str(item.get("comprador", "")).strip() == nome_antigo:
            item["comprador"] = nome_novo
            alterado = True
    if alterado:
        salvar_metas_por_comprador(dados)

garantir_metas_compradores_periodo(METAS_GESTOR.get("periodo_referencia", ""))

# Compradores inativos permanecem apenas no histórico.
_dados_metas_ativos = carregar_metas_por_comprador()
_alterou_status = False
_ativos_cf = set(_conjunto_compradores_ativos())
for _item_meta in _dados_metas_ativos:
    _nome_cf = str(_item_meta.get("comprador", "")).strip().casefold()
    if _nome_cf and _nome_cf not in _ativos_cf and str(_item_meta.get("status", "")).casefold() != "inativa":
        _item_meta["status"] = "Inativa"
        _item_meta["motivo_inativacao"] = "Comprador fora do cadastro ativo"
        _alterou_status = True
if _alterou_status:
    salvar_metas_por_comprador(_dados_metas_ativos)

REALIZADOS = filtrar_dataframe_compradores_ativos(REALIZADOS)
METAS = filtrar_dataframe_compradores_ativos(METAS)
RESULTADO = filtrar_dataframe_compradores_ativos(RESULTADO)
PREMIO = filtrar_dataframe_compradores_ativos(PREMIO)

REALIZADOS = _filtrar_df_por_usuario_logado(REALIZADOS)
METAS = _filtrar_df_por_usuario_logado(METAS)
RESULTADO = _filtrar_df_por_usuario_logado(RESULTADO)
PREMIO = _filtrar_df_por_usuario_logado(PREMIO)
PREMIO_KPI = _filtrar_objeto_por_usuario_logado(PREMIO_KPI)
if "RUPTURA_IMPORTADA" in globals() and isinstance(RUPTURA_IMPORTADA, pd.DataFrame):
    RUPTURA_IMPORTADA = _filtrar_df_por_usuario_logado(RUPTURA_IMPORTADA)


# =========================================================
# MOTOR DINÂMICO DAS VISÕES
# =========================================================

MAPA_CLASSIFICACAO_COMPRADOR = {
    "PRINCIPAL > NAO MED > BRINQUEDOS": "Francieli",
    "PRINCIPAL > NAO MED > PERFUMARIA": "Francieli",
    "PRINCIPAL > NAO MED > PERFUMARIA COMBATE": "Francieli",
    "PRINCIPAL > NAO MED > PERFUMARIA ENCALHADOS": "Francieli",
    "PRINCIPAL > NAO MED > CONVENIÊNCIA": "Francieli",
    "PRINCIPAL > NAO MED > SUPLEMENTOS": "Francieli",
    "PRINCIPAL > MED > ÉTICOS": "Paulo",
    "PRINCIPAL > NAO MED > MASCARA": "Não mapeado",
    "PRINCIPAL > NAO MED > FRALDAS": "Não mapeado",
    "PRINCIPAL > MED > GEN - SIM": "Não mapeado",
    "PRINCIPAL > MED > MEDICAMENTO CURVA D (TOP 15)": "Não mapeado",
    "PRINCIPAL > MED > NATURAIS": "Não mapeado",
    "PRINCIPAL > MED > PRÓPRIOS": "Não mapeado",
    "PRINCIPAL > NAO MED > HOSPITALARES": "Não mapeado",
    "PRINCIPAL > NAO MED > LEITES": "Não mapeado",
    "PRINCIPAL > NAO MED > VAREJO": "Não mapeado",
}

def _norm_coluna(valor):
    import unicodedata
    txt = unicodedata.normalize("NFKD", str(valor))
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    return " ".join(txt.lower().replace("_", " ").strip().split())

def _achar_coluna(df, candidatos):
    if df is None or df.empty:
        return None
    mapa = {_norm_coluna(c): c for c in df.columns}
    for candidato in candidatos:
        chave = _norm_coluna(candidato)
        if chave in mapa:
            return mapa[chave]
    for candidato in candidatos:
        chave = _norm_coluna(candidato)
        for norm, original in mapa.items():
            if chave in norm or norm in chave:
                return original
    return None

def _numero_df(df, candidatos):
    coluna = _achar_coluna(df, candidatos)
    if coluna is None:
        return pd.Series(0.0, index=df.index)
    serie = df[coluna]
    if pd.api.types.is_numeric_dtype(serie):
        return pd.to_numeric(serie, errors="coerce").fillna(0.0)
    texto = serie.astype(str).str.strip()
    tem_virgula = texto.str.contains(",", regex=False)
    resultado = pd.Series(index=texto.index, dtype="float64")
    resultado.loc[tem_virgula] = pd.to_numeric(
        texto.loc[tem_virgula]
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False),
        errors="coerce"
    )
    resultado.loc[~tem_virgula] = pd.to_numeric(texto.loc[~tem_virgula], errors="coerce")
    return resultado.fillna(0.0)

def _texto_df(df, candidatos, padrao=""):
    coluna = _achar_coluna(df, candidatos)
    if coluna is None:
        return pd.Series(padrao, index=df.index, dtype="object")
    return df[coluna].fillna(padrao).astype(str).str.strip()

def _normalizar_classificacao(valor):
    """Normaliza o caminho sem perder os níveis da classificação.

    Trata acentos, quebras de linha, espaços duplicados e diferenças de
    digitação entre PostgreSQL, Excel e o cadastro do comprador.
    """
    texto = str(valor or "").replace("\n", " ").replace("\r", " ")
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = texto.upper().strip()
    texto = re.sub(r"\s*>\s*", " > ", texto)
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip(" >")


def _mapa_compradores_ordenado():
    """Retorna somente mapeamentos de compradores ativos.

    O Cadastro de Compradores é a fonte oficial. Metas antigas, cache e
    mapeamentos históricos nunca reativam um comprador removido/inativo.
    """
    ativos = {nome.casefold(): nome for nome in lista_compradores_ativos()}
    itens = []
    for item in carregar_mapa_compradores_editavel():
        caminho = _normalizar_classificacao(
            item.get("Classificação Principal", "")
        )
        comprador_original = str(item.get("Comprador", "Não mapeado")).strip()
        comprador = ativos.get(comprador_original.casefold(), "")
        if caminho and comprador:
            itens.append((caminho, comprador))

    # Mais níveis e caminhos maiores primeiro. Isso impede que
    # PRINCIPAL > ETICOS capture antes de PRINCIPAL > ETICOS > ETICO LINEAR.
    itens.sort(
        key=lambda par: (par[0].count(" > "), len(par[0])),
        reverse=True,
    )
    return itens


def _mapear_comprador(valor):
    """Reconhece o comprador pelo caminho completo da classificação.

    Ordem de decisão:
    1. caminho completo exatamente igual ao cadastro;
    2. caminho cadastrado como ancestral mais específico;
    3. subclassificação recebida como ancestral de um único comprador;
    4. não mapeado quando houver ambiguidade.

    A regra nunca escolhe uma classificação genérica antes de uma
    subclassificação mais detalhada.
    """
    chave = _normalizar_classificacao(valor)
    if not chave:
        return "Não mapeado"

    itens = _mapa_compradores_ordenado()
    mapa_exato = {caminho: comprador for caminho, comprador in itens}

    # 1. A classificação completa é a fonte prioritária.
    if chave in mapa_exato:
        return mapa_exato[chave]

    # 2. Procura o ancestral cadastrado mais específico.
    ancestrais = [
        (caminho, comprador)
        for caminho, comprador in itens
        if chave.startswith(caminho + " > ")
    ]
    if ancestrais:
        maior_nivel = max(caminho.count(" > ") for caminho, _ in ancestrais)
        compradores = {
            comprador
            for caminho, comprador in ancestrais
            if caminho.count(" > ") == maior_nivel
        }
        if len(compradores) == 1:
            return next(iter(compradores))

    # 3. Quando a base vier resumida, aceita apenas se todas as
    # subclassificações abaixo dela pertencem ao mesmo comprador.
    descendentes = {
        comprador
        for caminho, comprador in itens
        if caminho.startswith(chave + " > ")
    }
    if len(descendentes) == 1:
        return next(iter(descendentes))

    return "Não mapeado"


def _atribuir_comprador(df):
    if df is None or df.empty:
        return pd.Series(dtype="object")

    # Prioriza sempre a classificação completa/subclassificada.
    classificacao = _texto_df(
        df,
        [
            "classificacao_resumida", "classificação resumida",
            "classificacao 3 nivel", "classificação 3º nível",
            "classificacao_3_nivel",
            "classificacao principal", "classificação principal",
            "classificacao_principal",
            "classificacao geral", "classificacao_geral",
            "classificacao",
        ],
        "",
    )

    # Recalcula pela classificação em todas as bases. Um comprador antigo
    # salvo no cache não pode prevalecer sobre um caminho atualizado.
    comprador_calculado = classificacao.map(_mapear_comprador)

    # Usa a coluna existente somente quando não existe classificação para
    # recalcular e o valor existente é realmente válido.
    comprador_existente = _texto_df(
        df, ["comprador", "comprador responsavel"], ""
    ).astype(str).str.strip()
    valido_existente = ~comprador_existente.str.upper().isin({
        "", "NÃO MAPEADO", "NAO MAPEADO", "NAN", "NONE", "NULL"
    })
    sem_classificacao = classificacao.astype(str).str.strip().eq("")
    resultado = comprador_calculado.where(
        ~(sem_classificacao & valido_existente),
        comprador_existente,
    )
    return resultado.replace("", "Não mapeado").fillna("Não mapeado")

def _snapshot_seguro(fonte, periodo):
    try:
        return carregar_snapshot(fonte, periodo)
    except Exception:
        return pd.DataFrame()

def _meta_comprador(comprador):
    periodo = METAS_GESTOR.get("periodo_referencia", "")
    individual = obter_meta_comprador(comprador, periodo)

    meta_venda = float(individual.get("meta_venda", 0))
    participacao = float(individual.get("participacao_venda_pct", 0)) / 100.0
    meta_cmv_pct = float(individual.get("meta_cmv_pct", 0))
    meta_cmv = float(individual.get("meta_cmv_valor", 0))
    if meta_cmv <= 0:
        meta_cmv = meta_venda * meta_cmv_pct / 100.0

    fator_cobertura = float(individual.get("fator_cobertura", 0))
    meta_estoque_total = float(individual.get("meta_estoque_total", 0))
    if meta_estoque_total <= 0:
        meta_estoque_total = meta_cmv * fator_cobertura

    return {
        "participacao": participacao,
        "meta_venda": meta_venda,
        "meta_cmv": meta_cmv,
        "meta_estoque_total": meta_estoque_total,
        "meta_curva_a": meta_estoque_total * float(individual.get("meta_curva_a_pct", 0)) / 100.0,
        "meta_curva_b": meta_estoque_total * float(individual.get("meta_curva_b_pct", 0)) / 100.0,
        "meta_curva_c": meta_estoque_total * float(individual.get("meta_curva_c_pct", 0)) / 100.0,
        "meta_curva_d": meta_estoque_total * float(individual.get("meta_curva_d_pct", 0)) / 100.0,
        "meta_ruptura": meta_venda * float(individual.get("meta_ruptura_pct", 0)) / 100.0,
        "meta_entradas": meta_cmv * float(individual.get("meta_reposicao_pct", 0)) / 100.0,
        "meta_cmv_pct": meta_cmv_pct,
        "fator_cobertura": fator_cobertura,
        "meta_curva_a_pct": float(individual.get("meta_curva_a_pct", 0)),
        "meta_curva_b_pct": float(individual.get("meta_curva_b_pct", 0)),
        "meta_curva_c_pct": float(individual.get("meta_curva_c_pct", 0)),
        "meta_curva_d_pct": float(individual.get("meta_curva_d_pct", 0)),
        "meta_ruptura_pct": float(individual.get("meta_ruptura_pct", 0)),
        "meta_reposicao_pct": float(individual.get("meta_reposicao_pct", 0)),
        "valor_premio": float(individual.get("valor_premio", METAS_GESTOR.get("valor_premio_total", 0))),
    }

def _atingimento_maior(real, meta):
    if meta <= 0:
        return 0.0
    return max(0.0, min(real / meta, 1.0)) * 100.0

def _atingimento_menor(real, meta):
    if real <= 0:
        return 100.0
    if meta <= 0:
        return 0.0
    return max(0.0, min(meta / real, 1.0)) * 100.0

def _agregar_vendas(df):
    saida = {}
    if df.empty:
        return saida
    base = df.copy()
    base["Comprador"] = _atribuir_comprador(base)
    valor = _numero_df(base, ["valortotal", "valor total", "faturamento", "valor venda", "receita"])
    custo_unit = _numero_df(base, ["custo", "custo unitario", "custo_unit_r", "cmv"])
    quantidade = _numero_df(base, ["quantidade", "qtd", "quantidade vendida"])
    custo_total_existente = _numero_df(base, ["custo total", "cmv total", "valor custo"])
    custo_total = custo_total_existente.where(custo_total_existente != 0, custo_unit * quantidade)
    base["_venda"] = valor
    base["_cmv"] = custo_total
    for comprador, grupo in base.groupby("Comprador"):
        saida[comprador] = {
            "faturamento": float(grupo["_venda"].sum()),
            "cmv": float(grupo["_cmv"].sum()),
        }
    return saida

def _agregar_entradas(df):
    saida = {}
    if df.empty:
        return saida
    base = df.copy()
    base["Comprador"] = _atribuir_comprador(base)
    valor = _numero_df(
        base,
        [
            "entrada custo total", "entradas custo", "valor entrada",
            "valor nf total", "custo total", "entrada_custo_total"
        ]
    )
    if float(valor.abs().sum()) == 0:
        custo = _numero_df(base, ["custo_final_r", "custo final", "custo"])
        qtd = _numero_df(base, ["quantidade_por_produto", "quantidade", "qtd"])
        valor = custo * qtd
    base["_entrada"] = valor
    for comprador, grupo in base.groupby("Comprador"):
        saida[comprador] = float(grupo["_entrada"].sum())
    return saida

def _agregar_estoque(df):
    saida = {}
    if df.empty:
        return saida

    base = df.copy()
    base["Comprador"] = _atribuir_comprador(base)

    curva = _texto_df(
        base,
        ["curva valor", "curva", "curva abc", "curva qtd", "cabc nome"],
        ""
    ).str.upper().str.strip()

    # Caso a consulta já traga o valor financeiro do estoque, utiliza diretamente.
    valor = _numero_df(
        base,
        [
            "estoque x custo medio", "valor estoque", "estoque a custo",
            "estoque total", "valor_estoque", "valor estoque atual"
        ]
    )

    # No novo script oficial, o valor é:
    # estoque × custo_medio_atual.
    # Se custo médio estiver zerado, usa custo_unit_atual.
    if float(valor.abs().sum()) == 0:
        qtd = _numero_df(
            base,
            ["estoque", "quantidade estoque", "saldo estoque", "qtd estoque"]
        )
        custo_medio = _numero_df(
            base,
            [
                "custo_medio_atual", "custo medio atual",
                "custo médio atual", "customedio", "custo medio"
            ]
        )
        custo_unitario = _numero_df(
            base,
            [
                "custo_unit_atual", "custo unit atual",
                "custo unitario atual", "custo", "custo_final_r"
            ]
        )
        custo_usado = custo_medio.where(custo_medio > 0, custo_unitario)
        valor = qtd * custo_usado

    base["_valor_estoque"] = valor
    base["_curva"] = curva

    for comprador, grupo in base.groupby("Comprador"):
        curvas = {}
        for letra in ["A", "B", "C", "D"]:
            mascara = grupo["_curva"].str.contains(
                rf"(^|[^A-Z]){letra}([^A-Z]|$)",
                regex=True,
                na=False
            )
            curvas[letra] = float(
                grupo.loc[mascara, "_valor_estoque"].sum()
            )

        saida[comprador] = {
            "total": float(grupo["_valor_estoque"].sum()),
            "curvas": curvas,
        }

    return saida

@st.cache_data(ttl=600, show_spinner=False)
def _agregar_ruptura(periodo, token_dados=None):
    try:
        base = carregar_ruptura_auto(periodo, token_dados)
    except Exception:
        base = pd.DataFrame()
    saida = {}
    if base.empty:
        return saida
    # Recalcula inclusive quando a coluna já existe, pois versões anteriores
    # gravavam o texto "Não mapeado" e impediam uma nova classificação.
    base["Comprador"] = _atribuir_comprador(base)
    valor_col = "Valor Ruptura" if "Valor Ruptura" in base.columns else _achar_coluna(base, ["ruptura ativa", "ruptura venda"])
    if isinstance(valor_col, str):
        base["_ruptura"] = pd.to_numeric(base[valor_col], errors="coerce").fillna(0)
    else:
        base["_ruptura"] = 0.0
    for comprador, grupo in base.groupby("Comprador"):
        saida[comprador] = float(grupo["_ruptura"].sum())
    return saida

def _colunas_tabela_cache(con, tabela):
    try:
        return {linha[1] for linha in con.execute(f'PRAGMA table_info("{tabela}")').fetchall()}
    except Exception:
        return set()


@st.cache_data(ttl=3600, show_spinner=False, max_entries=16)
def _resolver_periodo_realizado(periodo_solicitado, token_dados=None):
    """Retorna a competência realizada disponível no cache.

    Se o mês solicitado ainda não foi atualizado, utiliza o último mês
    disponível até a competência solicitada.
    """
    periodo_solicitado = str(periodo_solicitado or "")[:7]
    periodos = set()
    con = conexao_cache()
    try:
        for tabela in ["base_vendas", "base_entradas", "base_estoque"]:
            existe = con.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
                (tabela,),
            ).fetchone()
            if not existe:
                continue
            colunas = _colunas_tabela_cache(con, tabela)
            if "periodo_referencia" not in colunas:
                continue
            sql = (
                f'SELECT DISTINCT periodo_referencia FROM "{tabela}" '
                "WHERE periodo_referencia IS NOT NULL "
                "AND TRIM(periodo_referencia) <> ''"
            )
            for (valor,) in con.execute(sql).fetchall():
                texto = str(valor)[:7]
                if re.match(r"^\d{4}-\d{2}$", texto):
                    periodos.add(texto)
    finally:
        con.close()

    if periodo_solicitado in periodos:
        return periodo_solicitado
    anteriores = sorted(p for p in periodos if p <= periodo_solicitado)
    if anteriores:
        return anteriores[-1]
    return max(periodos) if periodos else periodo_solicitado


def _agregados_cache_rapidos(periodo):
    """Calcula os KPIs no SQLite sem carregar milhões de linhas na memória."""
    vendas, estoque, entradas, contas = {}, {}, {}, {}
    con = conexao_cache()
    try:
        cols = _colunas_tabela_cache(con, "base_vendas")
        if cols and "periodo_referencia" in cols:
            col_class = next((c for c in ["classificacao_resumida", "classificacao_3_nivel", "classificacao_geral", "classificacao_principal", "classificacao"] if c in cols), None)
            if col_class:
                valor_expr = 'COALESCE(SUM(CAST("valortotal" AS REAL)),0)' if "valortotal" in cols else '0'
                cmv_expr = 'COALESCE(SUM(CAST("custo" AS REAL) * CAST("quantidade" AS REAL)),0)' if "custo" in cols and "quantidade" in cols else '0'
                sql = f"SELECT COALESCE(\"{col_class}\", ''), {valor_expr}, {cmv_expr} FROM base_vendas WHERE periodo_referencia = ? GROUP BY COALESCE(\"{col_class}\", '')"
                for classificacao, faturamento, cmv in con.execute(sql, (periodo,)):
                    comprador = _mapear_comprador(classificacao)
                    item = vendas.setdefault(comprador, {"faturamento": 0.0, "cmv": 0.0})
                    item["faturamento"] += float(faturamento or 0)
                    item["cmv"] += float(cmv or 0)

        cols = _colunas_tabela_cache(con, "base_entradas")
        if cols and "periodo_referencia" in cols:
            col_class = next((c for c in ["classificacao_resumida", "classificacao_3_nivel", "classificacao_geral", "classificacao_principal", "classificacao"] if c in cols), None)
            if col_class:
                if "valor_nf_total" in cols:
                    valor_expr = 'COALESCE(SUM(CAST("valor_nf_total" AS REAL)),0)'
                elif "entrada_custo_total" in cols:
                    valor_expr = 'COALESCE(SUM(CAST("entrada_custo_total" AS REAL)),0)'
                elif "custo_final_r" in cols and "quantidade_por_produto" in cols:
                    valor_expr = 'COALESCE(SUM(CAST("custo_final_r" AS REAL) * CAST("quantidade_por_produto" AS REAL)),0)'
                else:
                    valor_expr = '0'
                sql = f"SELECT COALESCE(\"{col_class}\", ''), {valor_expr} FROM base_entradas WHERE periodo_referencia = ? GROUP BY COALESCE(\"{col_class}\", '')"
                for classificacao, valor in con.execute(sql, (periodo,)):
                    comprador = _mapear_comprador(classificacao)
                    entradas[comprador] = entradas.get(comprador, 0.0) + float(valor or 0)

        cols = _colunas_tabela_cache(con, "base_estoque")
        if cols and "periodo_referencia" in cols:
            col_class = next((c for c in ["classificacao_geral", "classificacao_3_nivel", "classificacao_principal", "classificacao"] if c in cols), None)
            col_curva = next((c for c in ["curva", "curva_abc", "curva_valor"] if c in cols), None)
            col_qtd = next((c for c in ["estoque", "saldo_estoque", "quantidade_estoque"] if c in cols), None)
            col_cmed = next((c for c in ["custo_medio_atual", "custo_medio", "customedio"] if c in cols), None)
            col_cunit = next((c for c in ["custo_unit_atual", "custo", "custo_unitario"] if c in cols), None)
            if col_class and col_qtd:
                curva_expr = f"COALESCE(\"{col_curva}\", '')" if col_curva else "''"
                if col_cmed and col_cunit:
                    custo_expr = f'CASE WHEN COALESCE(CAST(\"{col_cmed}\" AS REAL),0)>0 THEN CAST(\"{col_cmed}\" AS REAL) ELSE COALESCE(CAST(\"{col_cunit}\" AS REAL),0) END'
                elif col_cmed:
                    custo_expr = f'COALESCE(CAST(\"{col_cmed}\" AS REAL),0)'
                elif col_cunit:
                    custo_expr = f'COALESCE(CAST(\"{col_cunit}\" AS REAL),0)'
                else:
                    custo_expr = '0'
                sql = f"SELECT COALESCE(\"{col_class}\", ''), {curva_expr}, COALESCE(SUM(CAST(\"{col_qtd}\" AS REAL) * ({custo_expr})),0) FROM base_estoque WHERE periodo_referencia = ? GROUP BY COALESCE(\"{col_class}\", ''), {curva_expr}"
                for classificacao, curva, valor in con.execute(sql, (periodo,)):
                    comprador = _mapear_comprador(classificacao)
                    item = estoque.setdefault(comprador, {"total": 0.0, "curvas": {"A": 0.0, "B": 0.0, "C": 0.0, "D": 0.0}})
                    valor = float(valor or 0)
                    item["total"] += valor
                    curva_txt = str(curva or '').upper().strip()
                    for letra in ["A", "B", "C", "D"]:
                        if curva_txt == letra or curva_txt.startswith(letra + " ") or curva_txt.startswith(letra + "-"):
                            item["curvas"][letra] += valor
                            break
    except Exception as erro_cache:
        # Uma tabela ausente, coluna incompatível ou SQL inválido não pode
        # derrubar todo o painel. Mantém os agregados já calculados e segue.
        try:
            print(f"Aviso ao consolidar cache SQLite: {erro_cache}")
        except Exception:
            pass
    finally:
        con.close()
    return vendas, estoque, entradas, contas


@st.cache_data(ttl=3600, show_spinner=False, max_entries=16)
def construir_visoes_dinamicas(periodo, token_dados=None):
    periodo_realizado = _resolver_periodo_realizado(periodo, token_dados)
    vendas, estoque, entradas, contas = _agregados_cache_rapidos(periodo_realizado)
    ruptura = _agregar_ruptura(periodo_realizado, token_dados)

    linhas_real = []
    linhas_meta = []

    total_fat = sum(v.get("faturamento", 0) for v in vendas.values())

    # O Cadastro de Compradores ativos é a fonte oficial da interface.
    # Bases, metas antigas e mapeamentos históricos não podem reintroduzir
    # um comprador removido ou inativado.
    compradores_periodo = sorted(
        {nome for nome in lista_compradores_ativos() if _nome_comprador_valido(nome)},
        key=lambda x: x.casefold(),
    )

    # Novos compradores recebem inicialmente a participação real encontrada
    # na base. O percentual continua editável na Gestão de Metas.
    for nome in compradores_periodo:
        fat_nome = float(vendas.get(nome, {}).get("faturamento", 0.0))
        participacao_inicial = (fat_nome / total_fat * 100.0) if total_fat else 0.0
        garantir_meta_comprador_periodo(nome, periodo, participacao_inicial)

    for comprador in compradores_periodo:
        met = _meta_comprador(comprador)
        fat = vendas.get(comprador, {}).get("faturamento", 0.0)
        cmv = vendas.get(comprador, {}).get("cmv", 0.0)
        est = estoque.get(comprador, {})
        est_total = est.get("total", 0.0)
        curvas = est.get("curvas", {})
        ent = entradas.get(comprador, 0.0)
        rup = ruptura.get(comprador, 0.0)

        rep_fat = (fat / total_fat * 100.0) if total_fat else 0.0
        rep_cmv = (cmv / fat * 100.0) if fat else 0.0
        fator_cob = (est_total / cmv) if cmv else 0.0
        rep_a = (curvas.get("A", 0) / est_total * 100.0) if est_total else 0.0
        rep_b = (curvas.get("B", 0) / est_total * 100.0) if est_total else 0.0
        rep_c = (curvas.get("C", 0) / est_total * 100.0) if est_total else 0.0
        rep_d = (curvas.get("D", 0) / est_total * 100.0) if est_total else 0.0
        rup_pct = (rup / fat * 100.0) if fat else 0.0
        reposicao_pct = (ent / cmv * 100.0) if cmv else 0.0

        linhas_real.append([
            comprador, fat, rep_fat, cmv, rep_cmv, est_total, fator_cob,
            curvas.get("A", 0.0), rep_a,
            curvas.get("B", 0.0), rep_b,
            curvas.get("C", 0.0), rep_c,
            curvas.get("D", 0.0), rep_d,
            rup, rup_pct, ent, reposicao_pct
        ])

        linhas_meta.append([
            comprador,
            met["meta_venda"], met["participacao"] * 100,
            met["meta_cmv"], met["meta_cmv_pct"],
            met["meta_estoque_total"], met["fator_cobertura"],
            met["meta_curva_a"], met["meta_curva_a_pct"],
            met["meta_curva_b"], met["meta_curva_b_pct"],
            met["meta_curva_c"], met["meta_curva_c_pct"],
            met["meta_curva_d"], met["meta_curva_d_pct"],
            met["meta_ruptura"], met["meta_ruptura_pct"],
            met["meta_entradas"], met["meta_reposicao_pct"],
        ])

    colunas_real = [
        "Comprador", "Faturamento Total Atual", "Rep. Faturamento",
        "CMV mês Atual", "Rep. CMV", "Estoque Total", "Fator Cobertura",
        "Estoque Curva A", "Rep. Curva A", "Estoque Curva B", "Rep. Curva B",
        "Estoque Curva C", "Rep. Curva C", "Estoque Curva D", "Rep. Curva D",
        "Ruptura Ativa", "Ruptura %", "Entradas CUSTO", "Reposição CMV %"
    ]
    colunas_meta = [
        "Comprador", "Faturamento Total META", "Rep. Faturamento",
        "CMV mês META", "Rep. CMV", "Estoque Total META", "Fator Cobertura",
        "Estoque Curva A", "Rep. Curva A", "Estoque Curva B", "Rep. Curva B",
        "Estoque Curva C", "Rep. Curva C", "Estoque Curva D", "Rep. Curva D",
        "Ruptura Ativa", "Ruptura %", "Entradas CUSTO", "Reposição CMV %"
    ]

    real = pd.DataFrame(linhas_real, columns=colunas_real)
    metas = pd.DataFrame(linhas_meta, columns=colunas_meta)

    resultado = real.copy()
    pares = [
        ("Faturamento Total Atual", "Faturamento Total META"),
        ("CMV mês Atual", "CMV mês META"),
        ("Estoque Total", "Estoque Total META"),
        ("Estoque Curva A", "Estoque Curva A"),
        ("Estoque Curva B", "Estoque Curva B"),
        ("Estoque Curva C", "Estoque Curva C"),
        ("Estoque Curva D", "Estoque Curva D"),
        ("Ruptura Ativa", "Ruptura Ativa"),
        ("Entradas CUSTO", "Entradas CUSTO"),
    ]
    for atual, meta_col in pares:
        resultado[atual] = metas[meta_col] - real[atual]
    resultado["Rep. Faturamento"] = metas["Rep. Faturamento"] - real["Rep. Faturamento"]
    resultado["Rep. CMV"] = metas["Rep. CMV"] - real["Rep. CMV"]
    resultado["Fator Cobertura"] = metas["Fator Cobertura"] - real["Fator Cobertura"]
    for curva in ["A", "B", "C", "D"]:
        resultado[f"Rep. Curva {curva}"] = metas[f"Rep. Curva {curva}"] - real[f"Rep. Curva {curva}"]
    resultado["Ruptura %"] = metas["Ruptura %"] - real["Ruptura %"]
    resultado["Reposição CMV %"] = metas["Reposição CMV %"] - real["Reposição CMV %"]

    pesos = {
        "Faturamento": float(METAS_GESTOR.get("peso_faturamento", 0)),
        "CMV": float(METAS_GESTOR.get("peso_cmv", 0)),
        "Fator Cobertura": float(METAS_GESTOR.get("peso_fator_cobertura", 0)),
        "Estoque Curva A": float(METAS_GESTOR.get("peso_curva_a", 0)),
        "Estoque Curva B": float(METAS_GESTOR.get("peso_curva_b", 0)),
        "Estoque Curva C": float(METAS_GESTOR.get("peso_curva_c", 0)),
        "Estoque Curva D": float(METAS_GESTOR.get("peso_curva_d", 0)),
        "Ruptura Ativa": float(METAS_GESTOR.get("peso_ruptura", 0)),
        "Reposição CMV": float(METAS_GESTOR.get("peso_reposicao", 0)),
    }
    premio_total = float(METAS_GESTOR.get("valor_premio_total", 0))

    premio_linhas = []
    for i, comprador in enumerate(compradores_periodo):
        r = real.iloc[i]
        m = metas.iloc[i]
        ating = {
            "Faturamento": _atingimento_maior(r["Faturamento Total Atual"], m["Faturamento Total META"]),
            "CMV": _atingimento_menor(r["CMV mês Atual"], m["CMV mês META"]),
            "Fator Cobertura": _atingimento_menor(r["Fator Cobertura"], m["Fator Cobertura"]),
            "Estoque Curva A": _atingimento_maior(r["Estoque Curva A"], m["Estoque Curva A"]),
            "Estoque Curva B": _atingimento_maior(r["Estoque Curva B"], m["Estoque Curva B"]),
            "Estoque Curva C": _atingimento_maior(r["Estoque Curva C"], m["Estoque Curva C"]),
            "Estoque Curva D": _atingimento_maior(r["Estoque Curva D"], m["Estoque Curva D"]),
            "Ruptura Ativa": _atingimento_menor(r["Ruptura Ativa"], m["Ruptura Ativa"]),
            "Reposição CMV": _atingimento_maior(r["Reposição CMV %"], m["Reposição CMV %"]),
        }
        premio_linhas.append([
            comprador,
            premio_total * pesos["Faturamento"]/100 * ating["Faturamento"]/100, ating["Faturamento"],
            premio_total * pesos["CMV"]/100 * ating["CMV"]/100, ating["CMV"],
            premio_total * pesos["Fator Cobertura"]/100 * ating["Fator Cobertura"]/100, ating["Fator Cobertura"],
            premio_total * pesos["Estoque Curva A"]/100 * ating["Estoque Curva A"]/100, ating["Estoque Curva A"],
            premio_total * pesos["Estoque Curva B"]/100 * ating["Estoque Curva B"]/100, ating["Estoque Curva B"],
            premio_total * pesos["Estoque Curva C"]/100 * ating["Estoque Curva C"]/100, ating["Estoque Curva C"],
            premio_total * pesos["Estoque Curva D"]/100 * ating["Estoque Curva D"]/100, ating["Estoque Curva D"],
            premio_total * pesos["Ruptura Ativa"]/100 * ating["Ruptura Ativa"]/100, ating["Ruptura Ativa"],
            premio_total * pesos["Reposição CMV"]/100 * ating["Reposição CMV"]/100, ating["Reposição CMV"],
        ])

    premio = pd.DataFrame(premio_linhas, columns=[
        "Comprador", "Faturamento Prêmio", "Faturamento Realizado",
        "CMV Prêmio", "CMV Realizado", "Estoque Total Prêmio", "Estoque Total Realizado",
        "Curva A Prêmio", "Curva A Realizado", "Curva B Prêmio", "Curva B Realizado",
        "Curva C Prêmio", "Curva C Realizado", "Curva D Prêmio", "Curva D Realizado",
        "Ruptura Prêmio", "Ruptura Realizado", "Entradas Prêmio", "Entradas Realizado"
    ])

    # Referência segura para o quadro de prêmio por KPI.
    # Usa Francieli quando existir; caso contrário usa o primeiro comprador
    # disponível. Se não houver resultados, mantém os indicadores zerados.
    referencia = None
    if not premio.empty and "Comprador" in premio.columns:
        referencia_francieli = premio.loc[
            premio["Comprador"].astype(str).str.strip().str.casefold() == "francieli"
        ]
        if not referencia_francieli.empty:
            referencia = referencia_francieli.iloc[0]
        else:
            referencia = premio.iloc[0]
    premio_kpi = pd.DataFrame([
        ["Faturamento", pesos["Faturamento"], premio_total*pesos["Faturamento"]/100, referencia["Faturamento Realizado"] if referencia is not None else 0, referencia["Faturamento Prêmio"] if referencia is not None else 0],
        ["CMV", pesos["CMV"], premio_total*pesos["CMV"]/100, referencia["CMV Realizado"] if referencia is not None else 0, referencia["CMV Prêmio"] if referencia is not None else 0],
        ["Fator Cobertura", pesos["Fator Cobertura"], premio_total*pesos["Fator Cobertura"]/100, referencia["Estoque Total Realizado"] if referencia is not None else 0, referencia["Estoque Total Prêmio"] if referencia is not None else 0],
        ["Estoque Curva A", pesos["Estoque Curva A"], premio_total*pesos["Estoque Curva A"]/100, referencia["Curva A Realizado"] if referencia is not None else 0, referencia["Curva A Prêmio"] if referencia is not None else 0],
        ["Estoque Curva B", pesos["Estoque Curva B"], premio_total*pesos["Estoque Curva B"]/100, referencia["Curva B Realizado"] if referencia is not None else 0, referencia["Curva B Prêmio"] if referencia is not None else 0],
        ["Estoque Curva C", pesos["Estoque Curva C"], premio_total*pesos["Estoque Curva C"]/100, referencia["Curva C Realizado"] if referencia is not None else 0, referencia["Curva C Prêmio"] if referencia is not None else 0],
        ["Estoque Curva D", pesos["Estoque Curva D"], premio_total*pesos["Estoque Curva D"]/100, referencia["Curva D Realizado"] if referencia is not None else 0, referencia["Curva D Prêmio"] if referencia is not None else 0],
        ["Ruptura Ativa", pesos["Ruptura Ativa"], premio_total*pesos["Ruptura Ativa"]/100, referencia["Ruptura Realizado"] if referencia is not None else 0, referencia["Ruptura Prêmio"] if referencia is not None else 0],
        ["Reposição CMV", pesos["Reposição CMV"], premio_total*pesos["Reposição CMV"]/100, referencia["Entradas Realizado"] if referencia is not None else 0, referencia["Entradas Prêmio"] if referencia is not None else 0],
    ], columns=["KPI", "Peso sobre a meta", "Prêmio por KPI atingível", "Atingimento %", "Prêmio Atingido"])

    # As fontes acima já foram consolidadas diretamente no SQLite.
    # Por isso, o quadro de status deve usar as estruturas agregadas e não
    # DataFrames completos, evitando carregar milhões de linhas na memória.
    status_fontes = {
        "Vendas": contar_registros_cache(
            "base_vendas", periodo_realizado
        ),
        "Estoque": contar_registros_cache(
            "base_estoque", periodo_realizado
        ),
        "Entradas": contar_registros_cache(
            "base_entradas", periodo_realizado
        ),
        "Contas a Pagar": contar_registros_cache(
            "base_contas_pagar", periodo_realizado
        ),
        "Ruptura": len(carregar_ruptura_auto(periodo_realizado))
        if "carregar_ruptura_auto" in globals()
        else 0,
    }
    return real, metas, resultado, premio, premio_kpi, status_fontes, periodo_realizado


def reconstruir_visoes_imediatamente(periodo):
    """Reconstrói cards e gráficos a partir do SQLite recém-gravado."""
    try:
        construir_visoes_dinamicas.clear()
    except Exception:
        pass
    try:
        carregar_totais_cards_diretos.clear()
    except Exception:
        pass
    _limpar_cache_dados()
    token = _arquivo_token(
        CACHE_DB_FILE, Path(str(CACHE_DB_FILE) + "-wal"),
        Path(str(CACHE_DB_FILE) + "-shm"), RUPTURA_AUTO_DB,
        Path(str(RUPTURA_AUTO_DB) + "-wal"),
        Path(str(RUPTURA_AUTO_DB) + "-shm"),
        DATA_DIR / "ultima_atualizacao_dados.json", DB_CONFIG_FILE,
        CACHE_DB_ORIGIN_FILE, RUPTURA_AUTO_CONTROLE, METAS_FILE,
        COMPRADORES_FILE, MAPA_COMPRADORES_FILE,
    )
    dados = construir_visoes_dinamicas(str(periodo), token)
    st.session_state["_dados_visoes"] = dados
    st.session_state["_chave_visoes"] = f"{periodo}|{token}"
    return dados


# =========================================================
# PORTAL ANALÍTICO DE PREMIAÇÃO
# =========================================================

HIERARQUIA_PREMIACAO_FILE = DATA_DIR / "hierarquia_premiacao.json"


def _premio_colunas():
    return [c for c in PREMIO.columns if str(c).endswith(" Prêmio")]


def _premio_total_por_comprador():
    if PREMIO is None or PREMIO.empty:
        return pd.DataFrame(columns=["Comprador", "Prêmio Total"])
    base = PREMIO.copy()
    cols = _premio_colunas()
    if not cols:
        base["Prêmio Total"] = 0.0
    else:
        base["Prêmio Total"] = base[cols].apply(
            pd.to_numeric, errors="coerce"
        ).fillna(0).sum(axis=1)
    return base[["Comprador", "Prêmio Total"]].sort_values(
        "Prêmio Total", ascending=False
    ).reset_index(drop=True)


def preparar_tabela_premio_comprador(df):
    """Formata Prêmio em R$ e Realizado em percentual nas telas de premiação."""
    exib = df.copy()
    for coluna in exib.columns:
        nome = str(coluna)
        if nome == "Comprador":
            continue
        if nome.endswith(" Prêmio"):
            exib[coluna] = pd.to_numeric(exib[coluna], errors="coerce").map(moeda_real)
        elif nome.endswith(" Realizado"):
            exib[coluna] = pd.to_numeric(exib[coluna], errors="coerce").map(percentual)
    return exib


def preparar_holerite_exibicao(df):
    """Cria uma cópia apenas visual, sem alterar os tipos da base de cálculo."""
    if df is None or df.empty:
        return df

    exib = df.copy(deep=True)

    # As colunas abaixo receberão textos formatados. Convertê-las previamente
    # para object evita TypeError ao atribuir "R$ ..." em colunas float64.
    colunas_visuais = [
        "Meta", "Realizado", "Peso (%)", "Prêmio máximo",
        "Atingimento (%)", "Prêmio conquistado",
        "Saldo não conquistado",
    ]
    for coluna in colunas_visuais:
        if coluna in exib.columns:
            exib[coluna] = exib[coluna].astype("object")

    kpis_monetarios = {
        "Faturamento", "CMV", "Estoque Curva A", "Estoque Curva B",
        "Estoque Curva C", "Estoque Curva D", "Ruptura Ativa",
    }
    kpis_percentuais = {"Reposição CMV"}
    kpis_decimais = {"Fator Cobertura"}

    for idx, linha_original in df.iterrows():
        kpi = str(linha_original.get("KPI", ""))
        for coluna in ["Meta", "Realizado"]:
            if coluna not in exib.columns:
                continue
            valor = linha_original.get(coluna, 0)
            if kpi in kpis_monetarios:
                texto = moeda_real(valor)
            elif kpi in kpis_percentuais:
                texto = percentual(valor)
            elif kpi in kpis_decimais:
                texto = numero_decimal(valor, 2)
            else:
                texto = numero_decimal(valor, 2)
            exib.at[idx, coluna] = texto

        formatadores = {
            "Peso (%)": percentual,
            "Prêmio máximo": moeda_real,
            "Atingimento (%)": percentual,
            "Prêmio conquistado": moeda_real,
            "Saldo não conquistado": moeda_real,
        }
        for coluna, formatador in formatadores.items():
            if coluna in exib.columns:
                exib.at[idx, coluna] = formatador(linha_original.get(coluna, 0))

    return exib


def _kpis_holerite_comprador(nome_comprador):
    premio_row = PREMIO.loc[PREMIO["Comprador"] == nome_comprador]
    real_row = REALIZADOS.loc[REALIZADOS["Comprador"] == nome_comprador]
    meta_row = METAS.loc[METAS["Comprador"] == nome_comprador]
    if premio_row.empty or real_row.empty or meta_row.empty:
        return pd.DataFrame()
    p, r, m = premio_row.iloc[0], real_row.iloc[0], meta_row.iloc[0]
    valor_premio = float(_meta_comprador(nome_comprador).get("valor_premio", 0))
    pesos = {
        "Faturamento": float(METAS_GESTOR.get("peso_faturamento", 0)),
        "CMV": float(METAS_GESTOR.get("peso_cmv", 0)),
        "Fator Cobertura": float(METAS_GESTOR.get("peso_fator_cobertura", 0)),
        "Estoque Curva A": float(METAS_GESTOR.get("peso_curva_a", 0)),
        "Estoque Curva B": float(METAS_GESTOR.get("peso_curva_b", 0)),
        "Estoque Curva C": float(METAS_GESTOR.get("peso_curva_c", 0)),
        "Estoque Curva D": float(METAS_GESTOR.get("peso_curva_d", 0)),
        "Ruptura Ativa": float(METAS_GESTOR.get("peso_ruptura", 0)),
        "Reposição CMV": float(METAS_GESTOR.get("peso_reposicao", 0)),
    }
    regras = [
        ("Faturamento", "Faturamento Total META", "Faturamento Total Atual", "Faturamento Realizado", "Faturamento Prêmio", "Maior ou igual"),
        ("CMV", "CMV mês META", "CMV mês Atual", "CMV Realizado", "CMV Prêmio", "Menor ou igual"),
        ("Fator Cobertura", "Fator Cobertura", "Fator Cobertura", "Estoque Total Realizado", "Estoque Total Prêmio", "Menor ou igual"),
        ("Estoque Curva A", "Estoque Curva A", "Estoque Curva A", "Curva A Realizado", "Curva A Prêmio", "Maior ou igual"),
        ("Estoque Curva B", "Estoque Curva B", "Estoque Curva B", "Curva B Realizado", "Curva B Prêmio", "Maior ou igual"),
        ("Estoque Curva C", "Estoque Curva C", "Estoque Curva C", "Curva C Realizado", "Curva C Prêmio", "Maior ou igual"),
        ("Estoque Curva D", "Estoque Curva D", "Estoque Curva D", "Curva D Realizado", "Curva D Prêmio", "Maior ou igual"),
        ("Ruptura Ativa", "Ruptura Ativa", "Ruptura Ativa", "Ruptura Realizado", "Ruptura Prêmio", "Menor ou igual"),
        ("Reposição CMV", "Reposição CMV %", "Reposição CMV %", "Entradas Realizado", "Entradas Prêmio", "Maior ou igual"),
    ]
    linhas = []
    for kpi, col_meta, col_real, col_ating, col_premio, criterio in regras:
        meta_v = float(pd.to_numeric(pd.Series([m.get(col_meta, 0)]), errors="coerce").fillna(0).iloc[0])
        real_v = float(pd.to_numeric(pd.Series([r.get(col_real, 0)]), errors="coerce").fillna(0).iloc[0])
        ating = float(pd.to_numeric(pd.Series([p.get(col_ating, 0)]), errors="coerce").fillna(0).iloc[0])
        premio = float(pd.to_numeric(pd.Series([p.get(col_premio, 0)]), errors="coerce").fillna(0).iloc[0])
        peso = pesos.get(kpi, 0.0)
        maximo = valor_premio * peso / 100.0
        linhas.append({
            "KPI": kpi,
            "Critério": criterio,
            "Meta": meta_v,
            "Realizado": real_v,
            "Peso (%)": peso,
            "Prêmio máximo": maximo,
            "Atingimento (%)": ating,
            "Prêmio conquistado": premio,
            "Saldo não conquistado": max(0.0, maximo - premio),
        })
    return pd.DataFrame(linhas)


def _chave_loja_premiacao(valor):
    texto = str(valor or "").strip().upper()
    achou = re.search(r"\d+", texto)
    return achou.group(0).lstrip("0") or "0" if achou else texto


def carregar_hierarquia_premiacao():
    if HIERARQUIA_PREMIACAO_FILE.exists():
        try:
            dados = json.loads(HIERARQUIA_PREMIACAO_FILE.read_text(encoding="utf-8"))
            if isinstance(dados, list):
                return dados
        except Exception:
            pass
    registros = []
    lojas = dataframe_metas_lojas(PERIODO_DASHBOARD)
    if lojas.empty:
        lojas = dataframe_metas_lojas()
    for _, item in lojas.iterrows():
        registros.append({
            "Loja": str(item.get("regional_loja", "")).strip(),
            "Supervisor": "Não cadastrado",
            "Gerente": str(item.get("gerente", "Não cadastrado")).strip() or "Não cadastrado",
        })
    return registros


def salvar_hierarquia_premiacao(registros):
    HIERARQUIA_PREMIACAO_FILE.write_text(
        json.dumps(registros, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    _registrar_atualizacao_dados("Hierarquia de Premiação", PERIODO_DASHBOARD, len(registros))


@st.cache_data(ttl=600, show_spinner=False, max_entries=16)
def _vendas_loja_comprador_premiacao(periodo, token_dados=None):
    con = conexao_cache()
    try:
        cols = _colunas_tabela_cache(con, "base_vendas")
        if not cols or "periodo_referencia" not in cols:
            return pd.DataFrame(columns=["Loja", "Comprador", "Venda"])
        col_loja = next((c for c in ["numero_loja", "loja", "unidade_negocio", "unidade", "filial"] if c in cols), None)
        col_class = next((c for c in ["classificacao_resumida", "classificacao_3_nivel", "classificacao_geral", "classificacao_principal", "classificacao"] if c in cols), None)
        col_valor = next((c for c in ["valortotal", "valor_total", "faturamento", "valor_venda"] if c in cols), None)
        if not col_loja or not col_class or not col_valor:
            return pd.DataFrame(columns=["Loja", "Comprador", "Venda"])
        sql = f'''SELECT COALESCE(CAST("{col_loja}" AS TEXT),''), COALESCE(CAST("{col_class}" AS TEXT),''), COALESCE(SUM(CAST("{col_valor}" AS REAL)),0)
                  FROM base_vendas WHERE periodo_referencia=?
                  GROUP BY COALESCE(CAST("{col_loja}" AS TEXT),''), COALESCE(CAST("{col_class}" AS TEXT),'')'''
        linhas = []
        for loja, classificacao, venda in con.execute(sql, (periodo,)):
            linhas.append({"Loja": str(loja).strip() or "Loja não identificada", "Comprador": _mapear_comprador(classificacao), "Venda": float(venda or 0)})
        if not linhas:
            return pd.DataFrame(columns=["Loja", "Comprador", "Venda"])
        return pd.DataFrame(linhas).groupby(["Loja", "Comprador"], as_index=False)["Venda"].sum()
    except Exception:
        return pd.DataFrame(columns=["Loja", "Comprador", "Venda"])
    finally:
        con.close()


def _rateio_premiacao_loja(periodo):
    vendas = _vendas_loja_comprador_premiacao(periodo, _TOKEN_VISOES)
    premios = _premio_total_por_comprador()
    if vendas.empty or premios.empty:
        return pd.DataFrame(columns=["Loja", "Comprador", "Venda", "Participação no comprador (%)", "Prêmio rateado"])
    base = vendas.merge(premios, on="Comprador", how="inner")
    total_comp = base.groupby("Comprador")["Venda"].transform("sum")
    base["Participação no comprador (%)"] = np.where(total_comp > 0, base["Venda"] / total_comp * 100, 0)
    base["Prêmio rateado"] = np.where(total_comp > 0, base["Prêmio Total"] * base["Venda"] / total_comp, 0)
    return base[["Loja", "Comprador", "Venda", "Participação no comprador (%)", "Prêmio rateado"]].sort_values("Prêmio rateado", ascending=False)


def _hierarquia_rateio(periodo):
    rateio = _rateio_premiacao_loja(periodo)
    if rateio.empty:
        return rateio
    hier = pd.DataFrame(carregar_hierarquia_premiacao())
    if hier.empty:
        rateio["Supervisor"] = "Não cadastrado"
        rateio["Gerente"] = "Não cadastrado"
        return rateio
    for col in ["Loja", "Supervisor", "Gerente"]:
        if col not in hier.columns:
            hier[col] = "Não cadastrado"
    rateio["_chave"] = rateio["Loja"].map(_chave_loja_premiacao)
    hier["_chave"] = hier["Loja"].map(_chave_loja_premiacao)
    mapa = hier.drop_duplicates("_chave")[["_chave", "Supervisor", "Gerente"]]
    out = rateio.merge(mapa, on="_chave", how="left").drop(columns=["_chave"])
    out["Supervisor"] = out["Supervisor"].fillna("Não cadastrado").replace("", "Não cadastrado")
    out["Gerente"] = out["Gerente"].fillna("Não cadastrado").replace("", "Não cadastrado")
    return out



def _premio_maximo_por_comprador():
    """Retorna o prêmio máximo configurado para cada comprador."""
    linhas = []
    for nome in _premio_total_por_comprador().get("Comprador", pd.Series(dtype=str)).tolist():
        detalhe = _kpis_holerite_comprador(nome)
        linhas.append({
            "Comprador": nome,
            "Prêmio máximo comprador": float(
                pd.to_numeric(detalhe.get("Prêmio máximo", pd.Series(dtype=float)), errors="coerce")
                .fillna(0).sum()
            ) if not detalhe.empty else 0.0,
        })
    return pd.DataFrame(linhas)


def _holerite_lojas_por_meta(periodo):
    """Monta o holerite da loja usando as metas financeiras cadastradas."""
    quadro = montar_quadro_filiais_ceo(periodo).copy()
    if quadro.empty:
        return pd.DataFrame()

    rateio = _hierarquia_rateio(periodo).copy()
    premio_max = _premio_maximo_por_comprador()
    if not rateio.empty:
        rateio = rateio.merge(premio_max, on="Comprador", how="left")
        rateio["Prêmio máximo comprador"] = pd.to_numeric(
            rateio["Prêmio máximo comprador"], errors="coerce"
        ).fillna(0)
        rateio["Prêmio máximo rateado"] = (
            rateio["Prêmio máximo comprador"]
            * pd.to_numeric(rateio["Participação no comprador (%)"], errors="coerce").fillna(0)
            / 100.0
        )
        rateio["_chave_loja"] = rateio["Loja"].map(_chave_loja_premiacao)
        premios_loja = rateio.groupby("_chave_loja", as_index=False).agg(
            **{
                "Prêmio oficial rateado": ("Prêmio rateado", "sum"),
                "Prêmio máximo da loja": ("Prêmio máximo rateado", "sum"),
                "Compradores envolvidos": ("Comprador", "nunique"),
                "Supervisor": ("Supervisor", "first"),
                "Gerente hierarquia": ("Gerente", "first"),
            }
        )
    else:
        premios_loja = pd.DataFrame(columns=[
            "_chave_loja", "Prêmio oficial rateado", "Prêmio máximo da loja",
            "Compradores envolvidos", "Supervisor", "Gerente hierarquia"
        ])

    quadro["_chave_loja"] = quadro["Filial"].map(_chave_loja_premiacao)
    quadro = quadro.merge(premios_loja, on="_chave_loja", how="left")
    for col in [
        "Prêmio oficial rateado", "Prêmio máximo da loja", "Compradores envolvidos",
        "Faturamento Total META", "Faturamento Total Atual",
        "Margem Bruta META", "Margem Bruta Atual",
        "Atingimento Faturamento (%)", "Atingimento Margem Bruta (%)",
    ]:
        quadro[col] = pd.to_numeric(quadro.get(col, 0), errors="coerce").fillna(0)

    quadro["Supervisor"] = quadro.get("Supervisor", "Não cadastrado").fillna("Não cadastrado")
    quadro["Gerente Comercial"] = quadro["Gerente"].fillna("")
    vazio_gerente = quadro["Gerente Comercial"].astype(str).str.strip().eq("")
    quadro.loc[vazio_gerente, "Gerente Comercial"] = quadro.loc[vazio_gerente, "Gerente hierarquia"].fillna("Não cadastrado")

    # Regra transparente: 50% faturamento + 50% margem bruta, cada parcela limitada a 100%.
    quadro["Atingimento Geral (%)"] = (
        quadro["Atingimento Faturamento (%)"].clip(lower=0, upper=100) * 0.50
        + quadro["Atingimento Margem Bruta (%)"].clip(lower=0, upper=100) * 0.50
    )
    quadro["Prêmio conquistado pelas metas"] = (
        quadro["Prêmio máximo da loja"] * quadro["Atingimento Geral (%)"] / 100.0
    )
    quadro["Saldo não conquistado"] = (
        quadro["Prêmio máximo da loja"] - quadro["Prêmio conquistado pelas metas"]
    ).clip(lower=0)

    return quadro[[
        "Filial", "Supervisor", "Gerente Comercial",
        "Faturamento Total META", "Faturamento Total Atual", "Atingimento Faturamento (%)",
        "Margem Bruta META", "Margem Bruta Atual", "Atingimento Margem Bruta (%)",
        "Prêmio máximo da loja", "Prêmio conquistado pelas metas", "Saldo não conquistado",
        "Atingimento Geral (%)", "Prêmio oficial rateado", "Compradores envolvidos",
    ]].sort_values("Prêmio conquistado pelas metas", ascending=False).reset_index(drop=True)


def _holerite_gerentes_por_meta(periodo):
    lojas = _holerite_lojas_por_meta(periodo)
    if lojas.empty:
        return pd.DataFrame()
    return lojas.groupby("Gerente Comercial", as_index=False).agg(
        Lojas=("Filial", "nunique"),
        Supervisores=("Supervisor", "nunique"),
        **{
            "Meta Faturamento (R$)": ("Faturamento Total META", "sum"),
            "Realizado Faturamento (R$)": ("Faturamento Total Atual", "sum"),
            "Meta Margem Bruta (R$)": ("Margem Bruta META", "sum"),
            "Realizado Margem Bruta (R$)": ("Margem Bruta Atual", "sum"),
            "Prêmio máximo (R$)": ("Prêmio máximo da loja", "sum"),
            "Prêmio conquistado (R$)": ("Prêmio conquistado pelas metas", "sum"),
            "Saldo não conquistado (R$)": ("Saldo não conquistado", "sum"),
            "Prêmio oficial rateado (R$)": ("Prêmio oficial rateado", "sum"),
        }
    )

def _status_atingimento_premiacao(valor):
    valor = float(pd.to_numeric(valor, errors="coerce") or 0)
    if valor >= 100:
        return "Atingida", "green", "✅"
    if valor >= 90:
        return "Atenção", "yellow", "●"
    return "Crítico", "red", "●"


def _cor_barra_premiacao(valor):
    valor = float(pd.to_numeric(valor, errors="coerce") or 0)
    if valor >= 100:
        return "#20c66b"
    if valor >= 90:
        return "#22b96e"
    if valor >= 80:
        return "#ffc21c"
    return "#ff4d4f"


def _html_card_premiacao(icone, titulo, valor, nota, classe_icone, classe_valor=""):
    return f'<div class="hp-card"><div class="hp-card-top"><div class="hp-icon {classe_icone}">{icone}</div><div><div class="hp-card-label">{titulo}</div><div class="hp-card-value {classe_valor}">{valor}</div></div></div><div class="hp-card-note">{nota}</div></div>'


def _html_tabela_holerite(kpis, total_maximo, total_conquistado, total_saldo, ating_geral):
    linhas = []
    for item in kpis:
        ating = float(item["Atingimento (%)"])
        status, cor_status, simbolo = _status_atingimento_premiacao(ating)
        cor = _cor_barra_premiacao(ating)
        largura = min(max(ating, 0), 100)
        tipo = item.get("Tipo", "Valor (R$)")
        meta = percentual(item["Meta"]) if "Percentual" in tipo else moeda_real(item["Meta"])
        realizado = percentual(item["Realizado"]) if "Percentual" in tipo else moeda_real(item["Realizado"])
        linhas.append(f'<tr><td><b>{item["KPI"]}</b></td><td>{tipo}</td><td>{meta}</td><td>{realizado}</td><td>{percentual(ating)}<span class="hp-progress"><span style="width:{largura:.2f}%;background:{cor}"></span></span></td><td>{percentual(item["Peso (%)"])}</td><td>{moeda_real(item["Parcela máxima (R$)"])}</td><td class="hp-green">{moeda_real(item["Parcela conquistada (R$)"])}</td><td class="hp-red">{moeda_real(item["Valor perdido (R$)"])}</td><td><span class="hp-badge hp-badge-{cor_status}">{simbolo} {status}</span></td></tr>')
    status, cor_status, simbolo = _status_atingimento_premiacao(ating_geral)
    cor = _cor_barra_premiacao(ating_geral)
    largura = min(max(ating_geral, 0), 100)
    linhas.append(f'<tr><td class="hp-green">TOTAL GERAL</td><td>-</td><td>-</td><td>-</td><td>{percentual(ating_geral)}<span class="hp-progress"><span style="width:{largura:.2f}%;background:{cor}"></span></span></td><td>{percentual(100)}</td><td class="hp-green">{moeda_real(total_maximo)}</td><td class="hp-green">{moeda_real(total_conquistado)}</td><td class="hp-red">{moeda_real(total_saldo)}</td><td><span class="hp-badge hp-badge-{cor_status}">{simbolo} {status}</span></td></tr>')
    return '<div class="hp-table-wrap"><table class="hp-table"><thead><tr><th>KPI</th><th>Tipo</th><th>Meta</th><th>Realizado</th><th>Atingimento</th><th>Peso</th><th>Parcela Máxima</th><th>Parcela Conquistada</th><th>Valor Perdido</th><th>Status</th></tr></thead><tbody>' + ''.join(linhas) + '</tbody></table></div>'


def _evolucao_diaria_holerite_loja(periodo, loja, meta_faturamento):
    try:
        with conexao_cache() as con:
            cols = _colunas_tabela_cache(con, "base_vendas")
            col_loja = next((c for c in ["numero_loja", "loja", "unidade_negocio", "unidade", "filial"] if c in cols), None)
            col_data = next((c for c in ["datahora_venda_final", "data_venda", "datahora", "data"] if c in cols), None)
            col_valor = next((c for c in ["valortotal", "valor_total", "faturamento", "valor_venda"] if c in cols), None)
            if not col_loja or not col_data or not col_valor or "periodo_referencia" not in cols:
                return pd.DataFrame()
            consulta = f'''SELECT DATE("{col_data}") AS dia, COALESCE(SUM(CAST("{col_valor}" AS REAL)),0) AS venda
                           FROM base_vendas
                           WHERE periodo_referencia=? AND LOWER(TRIM(CAST("{col_loja}" AS TEXT)))=?
                           GROUP BY DATE("{col_data}") ORDER BY DATE("{col_data}")'''
            df = pd.read_sql_query(consulta, con, params=(str(periodo), str(loja).strip().lower()))
        if df.empty:
            return df
        df["dia"] = pd.to_datetime(df["dia"], errors="coerce")
        df = df.dropna(subset=["dia"]).sort_values("dia")
        df["venda"] = pd.to_numeric(df["venda"], errors="coerce").fillna(0)
        df["venda_acumulada"] = df["venda"].cumsum()
        df["Atingimento Real (%)"] = np.where(float(meta_faturamento or 0) > 0, df["venda_acumulada"] / float(meta_faturamento) * 100, 0)
        dias_mes = calendar.monthrange(int(str(periodo)[:4]), int(str(periodo)[5:7]))[1]
        df["Meta Projetada (%)"] = df["dia"].dt.day / dias_mes * 100
        return df
    except Exception:
        return pd.DataFrame()

def _listar_competencias_globais():
    """Lista competências para visualização sem depender de uma única fonte."""
    periodos = set()

    periodo_meta = str(METAS_GESTOR.get("periodo_referencia", "")).strip()[:7]
    if re.match(r"^\d{4}-\d{2}$", periodo_meta):
        periodos.add(periodo_meta)

    try:
        for item in carregar_historico():
            periodo = str(item.get("periodo_referencia", "")).strip()[:7]
            if re.match(r"^\d{4}-\d{2}$", periodo):
                periodos.add(periodo)
    except Exception:
        pass

    for arquivo in [
        DATA_DIR / "metas_lojas.json",
        DATA_DIR / "metas_por_comprador.json",
    ]:
        try:
            dados = json.loads(arquivo.read_text(encoding="utf-8"))
            lista = dados.get("metas", []) if isinstance(dados, dict) else dados
            for item in lista if isinstance(lista, list) else []:
                periodo = str(
                    item.get("periodo_referencia")
                    or item.get("competencia")
                    or ""
                ).strip()[:7]
                if re.match(r"^\d{4}-\d{2}$", periodo):
                    periodos.add(periodo)
        except Exception:
            pass

    # Disponibiliza os doze meses do ano vigente, mesmo antes da carga da base.
    ano_vigente = int(str(periodo_meta or date.today().strftime("%Y-%m"))[:4])
    for mes_numero in range(1, 13):
        periodos.add(f"{ano_vigente:04d}-{mes_numero:02d}")

    return sorted(periodos, reverse=True)


def _meta_para_competencia(periodo):
    """Retorna a meta correspondente ao mês selecionado."""
    periodo = str(periodo)[:7]
    candidatos = []
    try:
        candidatos.extend(carregar_historico())
    except Exception:
        pass
    candidatos.append(METAS_GESTOR)

    for item in candidatos:
        if str(item.get("periodo_referencia", ""))[:7] == periodo:
            return dict(item)

    # Se ainda não houver uma meta específica, usa os valores atuais como
    # modelo, alterando somente a vigência. O usuário poderá salvá-la na
    # Gestão de Metas.
    meta = dict(METAS_GESTOR)
    ano, mes = map(int, periodo.split("-"))
    import calendar as _calendar
    meta["periodo_referencia"] = periodo
    meta["data_inicio"] = f"{periodo}-01"
    meta["data_fim"] = f"{periodo}-{_calendar.monthrange(ano, mes)[1]:02d}"
    return meta


_COMPETENCIAS_GLOBAIS = _listar_competencias_globais()
_PERIODO_PADRAO_GLOBAL = str(
    METAS_GESTOR.get("periodo_referencia", date.today().strftime("%Y-%m"))
)[:7]

with st.sidebar:
    st.markdown("#### Mês de visualização")
    PERIODO_GLOBAL_SELECIONADO = st.selectbox(
        "Competência global",
        _COMPETENCIAS_GLOBAIS,
        index=(
            _COMPETENCIAS_GLOBAIS.index(_PERIODO_PADRAO_GLOBAL)
            if _PERIODO_PADRAO_GLOBAL in _COMPETENCIAS_GLOBAIS
            else 0
        ),
        label_visibility="collapsed",
        key="periodo_global_dashboard",
        help=(
            "Aplica-se às telas operacionais, metas e premiações. "
            "A Análise Comercial utiliza o próprio seletor anual."
        ),
    )
    st.caption(
        "Filtro global do projeto. Não altera a Análise Comercial."
    )

METAS_GESTOR = _meta_para_competencia(PERIODO_GLOBAL_SELECIONADO)
st.session_state["periodo_gestao_unificado_global"] = PERIODO_GLOBAL_SELECIONADO

PERIODO_DASHBOARD = PERIODO_GLOBAL_SELECIONADO
_TOKEN_VISOES = _arquivo_token(
    _arquivo_sqlite_leitura(),
    CACHE_DB_GZ_FILE,
    RUPTURA_AUTO_DB,
    Path(str(RUPTURA_AUTO_DB) + "-wal"),
    Path(str(RUPTURA_AUTO_DB) + "-shm"),
    DATA_DIR / "ultima_atualizacao_dados.json",
    DB_CONFIG_FILE,
    CACHE_DB_ORIGIN_FILE,
    RUPTURA_AUTO_CONTROLE,
    METAS_FILE,
    COMPRADORES_FILE,
    MAPA_COMPRADORES_FILE,
)
_CHAVE_VISOES = f"{PERIODO_DASHBOARD}|{_TOKEN_VISOES}"
_FORCAR_VISOES = bool(st.session_state.pop("_forcar_recalculo_visoes", False))
if _FORCAR_VISOES or st.session_state.get("_chave_visoes") != _CHAVE_VISOES:
    st.session_state["_dados_visoes"] = construir_visoes_dinamicas(
        PERIODO_DASHBOARD,
        _TOKEN_VISOES,
    )
    st.session_state["_chave_visoes"] = _CHAVE_VISOES
REALIZADOS, METAS, RESULTADO, PREMIO, PREMIO_KPI, STATUS_FONTES_DINAMICAS, PERIODO_REALIZADO_USADO = st.session_state["_dados_visoes"]

if PERIODO_REALIZADO_USADO != PERIODO_DASHBOARD:
    st.warning(
        f"A competência {PERIODO_DASHBOARD} ainda não possui Vendas, Entradas e Estoque "
        f"atualizados no cache. Os indicadores realizados abaixo usam temporariamente "
        f"{PERIODO_REALIZADO_USADO}. As metas continuam sendo as de {PERIODO_DASHBOARD}. "
        "Atualize o mês no módulo Banco de Dados para substituir o realizado provisório."
    )
# Segurança final: primeiro remove compradores inativos.
REALIZADOS = filtrar_dataframe_compradores_ativos(REALIZADOS)
METAS = filtrar_dataframe_compradores_ativos(METAS)
RESULTADO = filtrar_dataframe_compradores_ativos(RESULTADO)
PREMIO = filtrar_dataframe_compradores_ativos(PREMIO)

# SEGURANÇA CENTRALIZADA POR USUÁRIO.
# A partir daqui, perfis restritos não carregam dados de terceiros
# para telas, cards, gráficos ou exportações.
if _perfil_logado() != "Administrador":
    REALIZADOS = _filtrar_df_por_usuario_logado(REALIZADOS)
    METAS = _filtrar_df_por_usuario_logado(METAS)
    RESULTADO = _filtrar_df_por_usuario_logado(RESULTADO)
    PREMIO = _filtrar_df_por_usuario_logado(PREMIO)
    PREMIO_KPI = _filtrar_objeto_por_usuario_logado(PREMIO_KPI)

# Filtros e seletores seguem os compradores reconhecidos nas bases do período.
COMPRADORES = sorted(
    {
        str(x).strip() for x in REALIZADOS.get("Comprador", pd.Series(dtype=str)).tolist()
        if _nome_comprador_valido(x)
    },
    key=lambda x: x.casefold(),
)


@st.cache_data(ttl=600, show_spinner=False, max_entries=8)
def _ler_cache_analise_cached(tabela, token):
    try:
        with conexao_cache() as con:
            existe = con.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
                (tabela,),
            ).fetchone()
            if not existe:
                return pd.DataFrame()
            return pd.read_sql_query(f'SELECT * FROM "{tabela}"', con)
    except Exception:
        return pd.DataFrame()

# =========================================================
# EIROX DESIGN SYSTEM
# =========================================================

st.markdown("""
<style>
:root{
    --bg:#050b13;
    --panel:#0c1724;
    --panel2:#101f30;
    --line:#1f3850;
    --text:#f3f7fb;
    --muted:#8da2b8;
    --cyan:#22d3ee;
    --blue:#2f80ed;
    --green:#31d07f;
    --gold:#f8c24e;
    --red:#ff6b74;
}
.stApp{
    background:
      radial-gradient(circle at 15% 0%,rgba(47,128,237,.18),transparent 28%),
      radial-gradient(circle at 85% 0%,rgba(34,211,238,.12),transparent 22%),
      linear-gradient(180deg,#07111c 0%,#050b13 100%);
    color:var(--text);
}
[data-testid="stSidebar"]{
    background:linear-gradient(180deg,#07121f,#0c1a2b 58%,#07111d);
    border-right:1px solid #18324a;
}
[data-testid="stSidebar"] *{color:#eef6ff!important;}
.block-container{max-width:1820px;padding:1rem 1.35rem 2rem;}
.eirox-shell{
    background:linear-gradient(135deg,rgba(15,38,60,.96),rgba(7,19,31,.98));
    border:1px solid #204663;
    border-radius:24px;
    padding:26px 28px;
    margin-bottom:16px;
    box-shadow:0 22px 55px rgba(0,0,0,.36);
    position:relative;
    overflow:hidden;
}
.eirox-shell:after{
    content:"";
    position:absolute;
    width:260px;height:260px;
    right:-80px;top:-90px;
    background:radial-gradient(circle,rgba(34,211,238,.25),transparent 66%);
}
.brand-row{display:flex;align-items:center;justify-content:space-between;gap:18px;}
.eirox-brand{display:flex;align-items:center;gap:14px;}
.eirox-mark{
    width:54px;height:54px;border-radius:16px;
    display:flex;align-items:center;justify-content:center;
    background:linear-gradient(135deg,#22d3ee,#2f80ed);
    box-shadow:0 10px 25px rgba(47,128,237,.35);
    color:white;font-weight:950;font-size:22px;letter-spacing:-1px;
}
.eirox-title small{display:block;color:#6fdff1;text-transform:uppercase;font-weight:800;letter-spacing:1.5px;font-size:11px}
.eirox-title h1{margin:2px 0 0;color:white!important;font-size:34px}
.eirox-badge{
    border:1px solid #2a5676;border-radius:999px;padding:8px 12px;
    color:#b9e9f4;background:rgba(14,44,66,.65);font-weight:800;font-size:12px;
}
.kpi-grid{display:grid;grid-template-columns:repeat(5,minmax(0,1fr));gap:12px;margin-bottom:16px}
.kpi-card{
    background:linear-gradient(180deg,rgba(16,31,48,.98),rgba(10,23,37,.98));
    border:1px solid #1d405d;border-radius:18px;padding:16px;
    box-shadow:0 12px 28px rgba(0,0,0,.22);
}
.kpi-card .label{color:#91a8bd;font-size:11px;text-transform:uppercase;letter-spacing:.7px;font-weight:800}
.kpi-card .value{font-size:25px;font-weight:950;color:white;margin-top:6px}
.kpi-card .sub{font-size:12px;color:#56d7ec;margin-top:4px}
.meta-card{
    background:linear-gradient(180deg,#0e1d2d,#0a1724);
    border:1px solid #1d3f5b;border-radius:18px;overflow:hidden;
    box-shadow:0 12px 28px rgba(0,0,0,.20);min-height:100%;
}
.meta-card.premium{border-color:#80661c;box-shadow:0 12px 28px rgba(248,194,78,.08)}
.meta-card-title{
    padding:12px 10px;text-align:center;font-weight:900;color:white;
    background:linear-gradient(180deg,#16324a,#102338);border-bottom:1px solid #244c69;
}
.meta-card.premium .meta-card-title{background:linear-gradient(180deg,#4e3d12,#30270f);border-color:#80661c}
.meta-line{display:grid;grid-template-columns:minmax(0,1fr) 112px;min-height:38px;align-items:center;border-bottom:1px solid #142d42}
.meta-line span{padding:8px 10px;color:#d8e4ef;font-weight:700;font-size:13px}
.meta-line strong{
    height:100%;display:flex;align-items:center;justify-content:flex-end;
    padding:8px 10px;color:#07111c;font-size:13px;
    background:linear-gradient(180deg,#8df4ff,#40d9ef);border-left:1px solid #2a7890;
}
.meta-card.premium .meta-line strong{background:linear-gradient(180deg,#ffe58f,#f8c24e);border-color:#9b7413}
.section-title{
    border-radius:16px 16px 0 0;padding:11px 16px;text-align:center;
    font-weight:950;letter-spacing:.3px;color:white;border:1px solid #284a65;border-bottom:none;
}
.sec-gray{background:linear-gradient(180deg,#26384a,#1b2a38)}
.sec-green{background:linear-gradient(180deg,#17613e,#0f412a)}
.sec-blue{background:linear-gradient(180deg,#185078,#123a58)}
.sec-gold{background:linear-gradient(180deg,#806117,#523e0f)}
[data-testid="stDataFrame"]{
    background:#0a1724;border:1px solid #25445f;border-radius:0 0 16px 16px;overflow:hidden;
    box-shadow:0 12px 28px rgba(0,0,0,.18);
}
[data-testid="stMetric"]{
    background:linear-gradient(180deg,#0f1f30,#0a1724);
    border:1px solid #22425f;padding:14px;border-radius:16px;
}
.premium-box{
    display:flex;align-items:center;justify-content:space-between;
    background:linear-gradient(135deg,#12263a,#0b1927);
    border:1px solid #2d526f;border-radius:18px;padding:16px 18px;margin-bottom:12px;
    box-shadow:0 12px 28px rgba(0,0,0,.20)
}
.premium-box .value{
    background:linear-gradient(180deg,#ffe58f,#f8c24e);
    color:#1d1605;border-radius:12px;padding:10px 28px;font-size:22px;font-weight:950
}
.eirox-footer{margin-top:18px;color:#7890a5;font-size:12px;text-align:center;letter-spacing:.4px}
@media(max-width:1200px){.kpi-grid{grid-template-columns:repeat(2,minmax(0,1fr))}}
@media(max-width:700px){.kpi-grid{grid-template-columns:1fr}.brand-row{align-items:flex-start;flex-direction:column}.eirox-title h1{font-size:28px}}

.sidebar-logo img { display:block; }
[data-testid="stSidebar"] [data-testid="stRadio"] label {
    padding: 7px 9px;
    border-radius: 10px;
    margin-bottom: 3px;
}
[data-testid="stSidebar"] [data-testid="stRadio"] label:hover {
    background: rgba(255,0,0,.08);
}

</style>
""", unsafe_allow_html=True)

with st.sidebar:
    if LOGO_ECONOMIZE_B64:
        st.markdown(
            f"""
            <div style="display:flex;justify-content:center;padding:4px 0 12px;">
                <img src="data:image/png;base64,{LOGO_ECONOMIZE_B64}"
                     style="width:235px;max-width:100%;object-fit:contain;">
            </div>
            """,
            unsafe_allow_html=True
        )

    st.caption("Gestão de Performance Comercial")
    st.caption("☁️ Viewer online • Somente leitura")
    st.divider()
    _renderizar_usuario_logado()

    perfil_logado = _perfil_logado()
    menu_restrito = _menu_permitido_por_perfil()
    if menu_restrito is not None:
        opcoes_navegacao = menu_restrito
    else:
        opcoes_navegacao = [
            "👔 Resumo CEO",
            "📌 Meu Resumo",
            "📚 Análise Comercial",
            "📊 Realizados",
            "🎯 Métricas Destaque",
            "📈 Resultado Métricas",
            "📋 Resultados dos KPI's",
            "🏆 Prêmio Comprador",
            "💰 Prêmio por KPI",
            "🌟 Portal de Premiação",
            "🧾 Holerite da Premiação",
            "🏪 Holerite da Loja",
            "👔 Holerite do Gerente Comercial",
            "🏬 Premiação por Loja",
            "👥 Premiação por Supervisor e Gerente",
            "🔎 Auditoria de Compradores",
        ]

    visao_label = st.radio(
        "Navegação",
        opcoes_navegacao,
        label_visibility="collapsed"
    )

    mapa_visoes = {
        "📌 Metas e Parâmetros": "Metas e Parâmetros",
        "👔 Resumo CEO": "Resumo CEO",
        "📌 Meu Resumo": "Meu Resumo",
        "📚 Análise Comercial": "Análise Comercial",
        "📊 Realizados": "Realizados",
        "🎯 Métricas Destaque": "Métricas Destaque",
        "📈 Resultado Métricas": "Resultado Métricas",
        "📋 Resultados dos KPI's": "Resultados dos KPI's",
        "🏆 Prêmio Comprador": "Prêmio Comprador",
        "💰 Prêmio por KPI": "Prêmio por KPI",
        "🌟 Portal de Premiação": "Portal de Premiação",
        "🧾 Holerite da Premiação": "Holerite da Premiação",
        "🏪 Holerite da Loja": "Holerite da Loja",
        "👔 Holerite do Gerente Comercial": "Holerite do Gerente Comercial",
        "🏬 Premiação por Loja": "Premiação por Loja",
        "👥 Premiação por Supervisor e Gerente": "Premiação por Supervisor e Gerente",
        "🏪 Metas de Loja": "Metas de Loja",
        "🧭 Gestão de Metas": "Gestão de Metas",
        "📥 Importar Ruptura": "Importar Ruptura",
        "🗄️ Banco de Dados": "Banco de Dados",
        "👥 Compradores por Classificação": "Compradores por Classificação",
        "🔎 Auditoria de Compradores": "Auditoria de Compradores",
        "🧑‍💼 Cadastro de Compradores": "Cadastro de Compradores",
    }
    visao = mapa_visoes[visao_label]

    if perfil_logado != "Administrador":
        visoes_permitidas = {mapa_visoes[x] for x in opcoes_navegacao}
        if visao not in visoes_permitidas:
            st.error("Acesso não autorizado para este perfil.")
            st.stop()

    st.markdown("### Filtros")
    perfil_atual = _perfil_logado()
    if perfil_atual == "Comprador":
        comprador = _escopo_usuario_logado()
        st.info(f"🔒 Comprador: {comprador}")
    elif perfil_atual == "Vendedor":
        comprador = "Todos"
        st.info(f"🔒 Vendedor: {_escopo_usuario_logado()}")
    elif perfil_atual == "Gerente":
        comprador = "Todos"
        st.info(f"🔒 Gerente: {_escopo_usuario_logado()}")
    else:
        comprador = st.selectbox("Comprador", ["Todos"] + COMPRADORES)

    st.markdown("#### Período ativo")
    st.info(
        f"{METAS_GESTOR.get('periodo_referencia','-')}\\n\\n"
        f"{data_br(METAS_GESTOR.get('data_inicio',''))} a "
        f"{data_br(METAS_GESTOR.get('data_fim',''))}"
    )

    with st.expander("Status da publicação", expanded=False):
        try:
            atualizado = datetime.fromtimestamp(CACHE_DB_FILE.stat().st_mtime).strftime("%d/%m/%Y %H:%M")
        except Exception:
            atualizado = "Não identificado"
        st.write(f"**Base publicada:** {atualizado}")
        st.write("**Origem:** Atualizador Local")
        st.write("**Modo:** somente leitura")

    st.divider()
    st.caption("Rede Economize • Enterprise")

iniciar_contexto_exportacao(visao, METAS_GESTOR.get("periodo_referencia", "-"))
st.session_state["_escopo_exportacao_usuario"] = {
    "perfil": _perfil_logado(),
    "escopo": _escopo_usuario_logado(),
}

if MODO_VIEWER and visao in {
    "Metas e Parâmetros", "Metas de Loja", "Gestão de Metas",
    "Importar Ruptura", "Banco de Dados",
    "Compradores por Classificação", "Cadastro de Compradores",
}:
    st.error("Esta função está disponível somente no Atualizador Local.")
    st.stop()

st.markdown(f"""
<div class="eirox-shell">
  <div class="brand-row">
    <div style="display:flex;align-items:center;gap:24px;min-width:0;">
      <img src="data:image/png;base64,{LOGO_ECONOMIZE_B64}"
           style="width:290px;max-width:38vw;object-fit:contain;filter:drop-shadow(0 10px 16px rgba(0,0,0,.28));">
      <div class="eirox-title">
        <small>Rede Economize</small>
        <h1>Performance Comercial</h1>
      </div>
    </div>
    <div style="display:flex;gap:8px;flex-wrap:wrap">
      <div class="eirox-badge">Enterprise Edition</div>
      <div class="eirox-badge">Período: {METAS_GESTOR.get("periodo_referencia","-")}</div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

if _perfil_logado() == "Administrador":
    _TOTAIS_CARDS = carregar_totais_cards_diretos(PERIODO_REALIZADO_USADO, _TOKEN_VISOES)
    fat = float(_TOTAIS_CARDS.get("faturamento", 0.0))
    cmv = float(_TOTAIS_CARDS.get("cmv", 0.0))
    estoque = float(_TOTAIS_CARDS.get("estoque", 0.0))
    ruptura = float(_TOTAIS_CARDS.get("ruptura", 0.0))
    reposicao = float(_TOTAIS_CARDS.get("reposicao", 0.0))
else:
    fat = float(pd.to_numeric(REALIZADOS.get("Faturamento Total Atual", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
    cmv = float(pd.to_numeric(REALIZADOS.get("CMV mês Atual", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
    estoque = float(pd.to_numeric(REALIZADOS.get("Estoque Total", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
    ruptura = float(pd.to_numeric(REALIZADOS.get("Ruptura Ativa", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
    reposicao = float(pd.to_numeric(REALIZADOS.get("Entradas CUSTO", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())

st.markdown(f"""
<div class="kpi-grid">
  <div class="kpi-card"><div class="label">Faturamento Total</div><div class="value">{moeda_real(fat)}</div><div class="sub">Meta: {moeda_real(METAS_GESTOR["meta_venda_total_mes"])}</div></div>
  <div class="kpi-card"><div class="label">CMV Atual</div><div class="value">R$ {moeda(cmv)}</div><div class="sub">68,9% do faturamento</div></div>
  <div class="kpi-card"><div class="label">Estoque Total</div><div class="value">R$ {moeda(estoque)}</div><div class="sub">Cobertura consolidada</div></div>
  <div class="kpi-card"><div class="label">Ruptura Ativa</div><div class="value">{moeda_real(ruptura)}</div><div class="sub">Meta operacional: {percentual(METAS_GESTOR["meta_ruptura"])}</div></div>
  <div class="kpi-card"><div class="label">Reposição CMV</div><div class="value">{percentual(reposicao)}</div><div class="sub">Meta: {percentual(METAS_GESTOR["meta_reposicao"])}</div></div>
</div>
""", unsafe_allow_html=True)

# =========================================================
# GRÁFICOS EXECUTIVOS
# =========================================================

dados_grafico = REALIZADOS.copy()
metas_grafico = METAS.copy()

if comprador != "Todos":
    dados_grafico = dados_grafico[dados_grafico["Comprador"] == comprador]
    metas_grafico = metas_grafico[metas_grafico["Comprador"] == comprador]

col_g1, col_g2 = st.columns([1.1, 0.9], gap="large")

with col_g1:
    st.markdown("### Faturamento realizado x meta")
    df_fat = dados_grafico[["Comprador", "Faturamento Total Atual"]].merge(
        metas_grafico[["Comprador", "Faturamento Total META"]],
        on="Comprador",
        how="left"
    )
    df_fat = df_fat.melt(
        id_vars="Comprador",
        var_name="Indicador",
        value_name="Valor"
    )
    fig_fat = px.bar(
        df_fat,
        x="Comprador",
        y="Valor",
        color="Indicador",
        barmode="group",
        text_auto=False
    )
    fig_fat.update_layout(
        height=350,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font_color="#f3f7fb",
        margin=dict(l=10, r=10, t=15, b=10),
        legend_title_text="",
        yaxis_title="R$",
        xaxis_title=""
    )
    fig_fat.update_yaxes(gridcolor="rgba(255,255,255,.08)")
    plotly_chart_br(fig_fat, use_container_width=True, config={"displayModeBar": False})

with col_g2:
    st.markdown("### Composição do estoque")
    df_curvas = pd.DataFrame({
        "Curva": ["Curva A", "Curva B", "Curva C", "Curva D"],
        "Valor": [
            dados_grafico["Estoque Curva A"].sum(),
            dados_grafico["Estoque Curva B"].sum(),
            dados_grafico["Estoque Curva C"].sum(),
            dados_grafico["Estoque Curva D"].sum(),
        ]
    })
    fig_curvas = px.pie(
        df_curvas,
        names="Curva",
        values="Valor",
        hole=0.58
    )
    fig_curvas.update_traces(textposition="inside", textinfo="percent+label")
    fig_curvas.update_layout(
        height=350,
        paper_bgcolor="rgba(0,0,0,0)",
        font_color="#f3f7fb",
        margin=dict(l=10, r=10, t=15, b=10),
        legend_title_text=""
    )
    plotly_chart_br(fig_curvas, use_container_width=True, config={"displayModeBar": False})

col_g3, col_g4 = st.columns(2, gap="large")

with col_g3:
    st.markdown("### Reposição CMV por comprador")
    fig_rep = px.bar(
        dados_grafico.sort_values("Reposição CMV %"),
        x="Reposição CMV %",
        y="Comprador",
        orientation="h",
        text="Reposição CMV %"
    )
    fig_rep.add_vline(
        x=METAS_GESTOR["meta_reposicao"],
        line_dash="dash",
        annotation_text="Meta"
    )
    fig_rep.update_traces(
        text=[
            percentual(valor)
            for valor in dados_grafico.sort_values(
                "Reposição CMV %"
            )["Reposição CMV %"]
        ],
        texttemplate="%{text}",
        textposition="outside",
        hovertemplate=(
            "Comprador=%{y}<br>"
            "Reposição CMV=%{text}<extra></extra>"
        ),
    )
    fig_rep.update_layout(
        height=300,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font_color="#f3f7fb",
        margin=dict(l=10, r=25, t=15, b=10),
        xaxis_title="Percentual",
        yaxis_title=""
    )
    fig_rep.update_xaxes(gridcolor="rgba(255,255,255,.08)")
    plotly_chart_br(
        fig_rep,
        use_container_width=True,
        config={"displayModeBar": False},
        tipo="percentual",
    )

with col_g4:
    st.markdown("### Ruptura ativa por comprador")
    if (
        not RUPTURA_IMPORTADA.empty
        and "Comprador" in RUPTURA_IMPORTADA.columns
        and RUPTURA_IMPORTADA["Comprador"].astype(str).str.strip().ne("").any()
    ):
        df_rup = RUPTURA_IMPORTADA.copy()
        df_rup["Comprador"] = df_rup["Comprador"].astype(str).str.strip()
        df_rup = (
            df_rup[df_rup["Comprador"] != ""]
            .groupby("Comprador", as_index=False)["Ruptura Ativa"]
            .sum()
        )
    else:
        df_rup = dados_grafico[["Comprador", "Ruptura Ativa"]].copy()

    fig_rup = px.bar(
        df_rup.sort_values("Ruptura Ativa"),
        x="Ruptura Ativa",
        y="Comprador",
        orientation="h",
        text_auto=False
    )
    fig_rup.update_layout(
        height=300,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font_color="#f3f7fb",
        margin=dict(l=10, r=25, t=15, b=10),
        xaxis_title="R$",
        yaxis_title=""
    )
    fig_rup.update_xaxes(gridcolor="rgba(255,255,255,.08)")
    plotly_chart_br(fig_rup, use_container_width=True, config={"displayModeBar": False})


st.markdown("### Status das bases do período")
status_cols = st.columns(5)
for coluna, (fonte, registros) in zip(
    status_cols,
    STATUS_FONTES_DINAMICAS.items(),
):
    with coluna:
        card_status_base(fonte, registros)

def listar_periodos_gestao_metas():
    periodos = set()
    atual = str(METAS_GESTOR.get("periodo_referencia", "")).strip()
    if atual:
        periodos.add(atual)
    for item in carregar_historico():
        valor = str(item.get("periodo_referencia", "")).strip()
        if valor:
            periodos.add(valor)
    for item in carregar_metas_por_comprador():
        valor = str(item.get("periodo_referencia", "")).strip()
        if valor:
            periodos.add(valor)
    df_lojas = dataframe_metas_lojas()
    if not df_lojas.empty and "periodo_referencia" in df_lojas.columns:
        periodos.update(df_lojas["periodo_referencia"].dropna().astype(str).str.strip())
    return sorted([p for p in periodos if p], reverse=True)




def _render_premium_header(icone, titulo, subtitulo, periodo=None):
    periodo = str(periodo or globals().get("PERIODO_REALIZADO_USADO", ""))
    st.markdown(
        f'<div class="hp-title-wrap"><div><div class="hp-title">{icone} {titulo}</div>'
        f'<div class="hp-subtitle">{subtitulo}</div></div>'
        f'<div class="hp-update">Competência: {periodo} &nbsp;|&nbsp; Atualizado em: {datetime.now():%d/%m/%Y %H:%M:%S} ⟳</div></div>',
        unsafe_allow_html=True,
    )


def _render_premium_cards(itens):
    """Renderiza cards no mesmo padrão dos holerites."""
    if not itens:
        return
    cols = st.columns(len(itens))
    for col, item in zip(cols, itens):
        col.markdown(
            _html_card_premiacao(
                item.get("icone", "🎯"),
                item.get("titulo", "Indicador"),
                item.get("valor", ""),
                item.get("subtitulo", ""),
                item.get("classe_icone", "hp-blue-icon"),
                item.get("classe_valor", ""),
            ),
            unsafe_allow_html=True,
        )

if visao == "Metas e Parâmetros":
    c1, c2, c3, c4, c5 = st.columns([1.35,.9,1.15,1.15,1.65], gap="small")
    with c1:
        card_meta("Metas por Tipo", [
            ("Meta Venda Total Mês", moeda_real(METAS_GESTOR["meta_venda_total_mes"])),
            ("Meta CMV Mês", percentual(METAS_GESTOR["meta_cmv_mes"])),
            ("Fator Redução CMV", br_num(METAS_GESTOR["fator_reducao_cmv"], 2)),
            ("Fator Cobertura", br_num(METAS_GESTOR["fator_cobertura"], 2)),
        ])
    with c2:
        card_meta("Metas Operacionais", [
            ("Meta Ruptura", percentual(METAS_GESTOR["meta_ruptura"])),
            ("Meta Reposição", percentual(METAS_GESTOR["meta_reposicao"])),
        ])
    with c3:
        card_meta("Curvas de Estoque", [
            ("Curva A", percentual(METAS_GESTOR["curva_a"])),
            ("Curva B", percentual(METAS_GESTOR["curva_b"])),
            ("Curva C", percentual(METAS_GESTOR["curva_c"])),
            ("Curva D", percentual(METAS_GESTOR["curva_d"])),
        ])
    with c4:
        metas_part_periodo = [
            item for item in carregar_metas_por_comprador()
            if str(item.get("periodo_referencia", "")) == str(PERIODO_DASHBOARD)
            and str(item.get("comprador", "")).strip().casefold() in _conjunto_compradores_ativos()
            and _nome_comprador_valido(item.get("comprador", ""))
        ]
        linhas_rep = [
            (str(item.get("comprador", "")), percentual(float(item.get("participacao_venda_pct", 0))))
            for item in sorted(metas_part_periodo, key=lambda x: str(x.get("comprador", "")).casefold())
        ]
        linhas_rep.append(("Total", percentual(sum(float(x.get("participacao_venda_pct", 0)) for x in metas_part_periodo))))
        card_meta("Rep. Venda Comprador", linhas_rep)
    with c5:
        card_meta("Pesos sobre Prêmio", [
            ("Faturamento", percentual(METAS_GESTOR["peso_faturamento"])),
            ("CMV", percentual(METAS_GESTOR["peso_cmv"])),
            ("Fator Cobertura", percentual(METAS_GESTOR["peso_fator_cobertura"])),
            ("Estoque Curva A", percentual(METAS_GESTOR["peso_curva_a"])),
            ("Estoque Curva B", percentual(METAS_GESTOR["peso_curva_b"])),
            ("Estoque Curva C", percentual(METAS_GESTOR["peso_curva_c"])),
            ("Estoque Curva D", percentual(METAS_GESTOR["peso_curva_d"])),
            ("Ruptura Ativa", percentual(METAS_GESTOR["peso_ruptura"])),
            ("Reposição CMV", percentual(METAS_GESTOR["peso_reposicao"])),
            ("Peso Total", percentual(
                METAS_GESTOR["peso_faturamento"] + METAS_GESTOR["peso_cmv"] +
                METAS_GESTOR["peso_fator_cobertura"] + METAS_GESTOR["peso_curva_a"] +
                METAS_GESTOR["peso_curva_b"] + METAS_GESTOR["peso_curva_c"] +
                METAS_GESTOR["peso_curva_d"] + METAS_GESTOR["peso_ruptura"] +
                METAS_GESTOR["peso_reposicao"]
            )),
        ], destaque=True)

elif visao == "Resumo CEO":
    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff;font-size:22px">Resumo Executivo CEO — Meta x Realizado</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:4px">Visão consolidada de todas as metas do período ativo.</div>
      </div>
      <div style="color:#56d7ec;font-weight:800">Executive Overview</div>
    </div>
    """, unsafe_allow_html=True)

    real_ceo = filtrar_dataframe_compradores_ativos(REALIZADOS)
    meta_ceo = filtrar_dataframe_compradores_ativos(METAS)
    def _soma_ceo(df, coluna):
        return float(pd.to_numeric(df.get(coluna, pd.Series(dtype=float)), errors="coerce").fillna(0).sum())

    entrada_financeira_ceo = 0.0
    try:
        with conexao_cache() as con_ceo_entrada:
            linha_entrada_ceo = con_ceo_entrada.execute(
                "SELECT COALESCE(SUM(CAST(valor_nf_total AS REAL)), 0) "
                "FROM base_entradas_financeira WHERE periodo_referencia = ?",
                (str(PERIODO_DASHBOARD),),
            ).fetchone()
            entrada_financeira_ceo = float(linha_entrada_ceo[0] or 0)
    except Exception:
        entrada_financeira_ceo = _soma_ceo(real_ceo, "Entradas CUSTO")

    indicadores_ceo = [
        ("Faturamento", _soma_ceo(meta_ceo, "Faturamento Total META"), _soma_ceo(real_ceo, "Faturamento Total Atual"), "maior"),
        ("CMV", _soma_ceo(meta_ceo, "CMV mês META"), _soma_ceo(real_ceo, "CMV mês Atual"), "menor"),
        ("Estoque Total", _soma_ceo(meta_ceo, "Estoque Total META"), _soma_ceo(real_ceo, "Estoque Total"), "menor"),
        ("Ruptura Ativa", _soma_ceo(meta_ceo, "Ruptura Ativa"), _soma_ceo(real_ceo, "Ruptura Ativa"), "menor"),
        ("Reposição / Entradas", _soma_ceo(meta_ceo, "Entradas CUSTO"), entrada_financeira_ceo, "maior"),
    ]
    linhas_ceo = []
    for nome, meta_v, real_v, regra in indicadores_ceo:
        ating = ((real_v / meta_v) if regra == "maior" else (meta_v / real_v if real_v else 0)) * 100 if meta_v else 0
        linhas_ceo.append({"Indicador": nome, "Meta": meta_v, "Realizado": real_v, "Atingimento (%)": ating, "Status": "✅ Atingida" if ating >= 100 else ("🟡 Atenção" if ating >= 90 else "🔴 Abaixo")})
    df_ceo = pd.DataFrame(linhas_ceo)

    cards = st.columns(5)
    for col, linha in zip(cards, linhas_ceo):
        col.metric(linha["Indicador"], moeda_real(linha['Realizado']), f"{percentual(linha['Atingimento (%)'])} da meta")

    c1, c2 = st.columns([1.35, 1])
    with c1:
        fig = go.Figure()
        fig.add_bar(name="Meta", x=df_ceo["Indicador"], y=df_ceo["Meta"])
        fig.add_bar(name="Realizado", x=df_ceo["Indicador"], y=df_ceo["Realizado"])
        fig.update_layout(barmode="group", height=390, title="Meta x Realizado — Consolidado", margin=dict(l=10, r=10, t=55, b=10))
        plotly_chart_br(fig, use_container_width=True)
    with c2:
        st.markdown("### Semáforo das metas")
        dataframe_br(df_ceo[["Indicador", "Atingimento (%)", "Status"]], use_container_width=True, hide_index=True, height=390, column_config={"Atingimento (%)": st.column_config.ProgressColumn(min_value=0, max_value=120, format="%.1f%%")})

    st.markdown("### Meta x realizado por comprador ativo")
    comp_ceo = real_ceo.merge(meta_ceo, on="Comprador", how="outer").fillna(0)
    if not comp_ceo.empty:
        comp_ceo["Atingimento Faturamento (%)"] = comp_ceo.apply(lambda r: (r.get("Faturamento Total Atual", 0) / r.get("Faturamento Total META", 0) * 100) if r.get("Faturamento Total META", 0) else 0, axis=1)
        cols_ceo = [c for c in ["Comprador", "Faturamento Total META", "Faturamento Total Atual", "Atingimento Faturamento (%)", "CMV mês META", "CMV mês Atual", "Estoque Total META", "Estoque Total"] if c in comp_ceo.columns]
        dataframe_br(comp_ceo[cols_ceo], use_container_width=True, hide_index=True)

    st.markdown("### Meta x realizado por filial")
    st.caption(
        "Faturamento e margem bruta por filial, considerando as metas "
        "cadastradas no período ativo e o realizado da base de vendas."
    )
    quadro_filiais_ceo = montar_quadro_filiais_ceo(PERIODO_DASHBOARD)
    if quadro_filiais_ceo.empty:
        st.info("Não há metas de filial cadastradas para o período ativo.")
    else:
        dataframe_br(
            quadro_filiais_ceo,
            use_container_width=True,
            hide_index=True,
            height=min(430, 82 + 36 * max(len(quadro_filiais_ceo), 1)),
        )

elif visao == "Análise Comercial":
    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff;font-size:22px">Análise de Desempenho Comercial</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:4px">Venda, custo, entrada, CMV, fornecedores, estoque e resultado por área e competência.</div>
      </div>
      <div style="color:#56d7ec;font-weight:800">Análise Agrupada</div>
    </div>
    """, unsafe_allow_html=True)

    cfg_analise = carregar_config_analise_comercial()
    st.caption("⚡ Modo rápido: a tela utiliza resumos persistentes, sem carregar os movimentos brutos.")

    token_analise = _arquivo_token(_arquivo_sqlite_leitura(), CACHE_DB_GZ_FILE)

    @st.cache_data(ttl=3600, show_spinner=False, max_entries=12)
    def _carregar_resumos_analise(token):
        with conexao_cache() as con:
            tabelas = {r[0] for r in con.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()}
            obrigatorias = {
                "analise_vendas_resumo",
                "analise_entradas_resumo",
                "analise_entradas_financeiro_resumo",
                "analise_contas_resumo",
                "analise_estoque_resumo",
                "analise_posicao_resumo",
            }
            if not obrigatorias.issubset(tabelas):
                faltantes = sorted(obrigatorias - tabelas)
                raise RuntimeError(
                    "O SQLite publicado não contém todos os resumos da Análise Comercial: "
                    + ", ".join(faltantes)
                    + ". Atualize/publice a base pelo Atualizador Local."
                )

            vendas = pd.read_sql_query(
                "SELECT periodo_referencia, classificacao, venda, custo FROM analise_vendas_resumo",
                con,
            )
            entradas = pd.read_sql_query(
                "SELECT periodo_referencia, classificacao, compra FROM analise_entradas_resumo",
                con,
            )
            entradas_financeiras = pd.read_sql_query(
                "SELECT periodo_referencia, compra, notas "
                "FROM analise_entradas_financeiro_resumo",
                con,
            )
            contas = pd.read_sql_query(
                "SELECT periodo_referencia, plano_contas, pagamento FROM analise_contas_resumo",
                con,
            )
            estoque = pd.read_sql_query(
                "SELECT periodo_referencia, estoque FROM analise_estoque_resumo",
                con,
            )
            posicao = pd.read_sql_query(
                "SELECT periodo_referencia, contas_pagar, estoque "
                "FROM analise_posicao_resumo",
                con,
            )
        return vendas, entradas, entradas_financeiras, contas, estoque, posicao

    # Normalmente retorna poucas dezenas de linhas, mesmo quando a base bruta
    # possui milhões de movimentos.
    vendas_an, entradas_an, entradas_fin_an, contas_an, estoque_an, posicao_an = _carregar_resumos_analise(token_analise)

    def _normalizar_texto_analise(valor):
        texto = "" if pd.isna(valor) else str(valor)
        texto = unicodedata.normalize("NFKD", texto)
        texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
        return re.sub(r"\s+", " ", texto).strip().upper()

    def _area_analise(valor):
        normal = _normalizar_texto_analise(valor)
        mapa = [
            ("1 - ETICOS", "01 - Med Propagados"),
            ("ETICOS", "01 - Med Propagados"),
            ("2 - GENERICOS", "02 - Genericos/Similares"),
            ("GENERICOS", "02 - Genericos/Similares"),
            ("SIMILARES", "02 - Genericos/Similares"),
            ("5 - PERFUMARIA", "04 - Perfumaria / Dermocosméticos"),
            ("PERFUMARIA", "04 - Perfumaria / Dermocosméticos"),
            ("DERMOCOSMET", "04 - Perfumaria / Dermocosméticos"),
            ("FRALDAS", "05 - Fraldas"),
            ("LEITES", "06 - Leite"),
            ("LEITE", "06 - Leite"),
            ("HOSPITALARES", "07 - Varejos"),
            ("VAREJO", "07 - Varejos"),
            ("6 - CONVENIENCIA", "08 - Conveniencias"),
            ("CONVENIENCIA", "08 - Conveniencias"),
            ("SUPLEMENT", "09 - Nutracêuticos, Suplementos, Vitaminas"),
            ("NUTRACE", "09 - Nutracêuticos, Suplementos, Vitaminas"),
            ("VITAMIN", "09 - Nutracêuticos, Suplementos, Vitaminas"),
        ]
        for termo, area in mapa:
            if termo in normal:
                return area

        # Preserva a classificação real em vez de agrupar como "Outros".
        nome_original = "" if pd.isna(valor) else str(valor).strip()
        nome_original = re.sub(r"^\s*\d+\s*[-–—.]\s*", "", nome_original)
        nome_original = re.sub(r"\s*>{1,}\s*", " / ", nome_original)
        nome_original = re.sub(r"\s+", " ", nome_original).strip(" -/")
        return nome_original.title() if nome_original else "Classificação não identificada"

    if not vendas_an.empty:
        vendas_an["Área"] = vendas_an["classificacao"].map(_area_analise)
        vendas_an["Competência"] = vendas_an["periodo_referencia"].astype(str).str[:7]
        vendas_an["Venda"] = pd.to_numeric(vendas_an["venda"], errors="coerce").fillna(0)
        vendas_an["Custo"] = pd.to_numeric(vendas_an["custo"], errors="coerce").fillna(0)

    if not entradas_an.empty:
        entradas_an["Área"] = entradas_an["classificacao"].map(_area_analise)
        entradas_an["Competência"] = entradas_an["periodo_referencia"].astype(str).str[:7]
        entradas_an["Compra"] = pd.to_numeric(entradas_an["compra"], errors="coerce").fillna(0)

    if not entradas_fin_an.empty:
        entradas_fin_an["Competência"] = entradas_fin_an["periodo_referencia"].astype(str).str[:7]
        entradas_fin_an["Compra"] = pd.to_numeric(
            entradas_fin_an["compra"], errors="coerce"
        ).fillna(0)
        entradas_fin_an["Notas"] = pd.to_numeric(
            entradas_fin_an.get("notas", 0), errors="coerce"
        ).fillna(0)

    if not contas_an.empty:
        contas_an["Competência"] = contas_an["periodo_referencia"].astype(str).str[:7]
        contas_an["Valor"] = pd.to_numeric(contas_an["pagamento"], errors="coerce").fillna(0)

    if not estoque_an.empty:
        estoque_an["Competência"] = estoque_an["periodo_referencia"].astype(str).str[:7]
        estoque_an["Estoque"] = pd.to_numeric(
            estoque_an["estoque"], errors="coerce"
        ).fillna(0)

    if not posicao_an.empty:
        posicao_an["Competência"] = posicao_an["periodo_referencia"].astype(str).str[:7]
        posicao_an["Contas a Pagar"] = pd.to_numeric(
            posicao_an["contas_pagar"], errors="coerce"
        ).fillna(0)
        posicao_an["Estoque"] = pd.to_numeric(
            posicao_an["estoque"], errors="coerce"
        ).fillna(0)

    periodos_disponiveis = sorted(set(
        ([x for x in vendas_an.get("Competência", pd.Series(dtype=str)).dropna().astype(str).tolist() if re.match(r"^\\d{4}-\\d{2}$", x)] +
         [x for x in entradas_an.get("Competência", pd.Series(dtype=str)).dropna().astype(str).tolist() if re.match(r"^\\d{4}-\\d{2}$", x)] +
         [x for x in entradas_fin_an.get("Competência", pd.Series(dtype=str)).dropna().astype(str).tolist() if re.match(r"^\\d{4}-\\d{2}$", x)] +
         [x for x in contas_an.get("Competência", pd.Series(dtype=str)).dropna().astype(str).tolist() if re.match(r"^\\d{4}-\\d{2}$", x)] +
         [x for x in estoque_an.get("Competência", pd.Series(dtype=str)).dropna().astype(str).tolist() if re.match(r"^\\d{4}-\\d{2}$", x)] +
         [x for x in posicao_an.get("Competência", pd.Series(dtype=str)).dropna().astype(str).tolist() if re.match(r"^\\d{4}-\\d{2}$", x)])
    ))
    # A Evolução Comercial possui contexto anual próprio e não obedece
    # ao período ativo utilizado nas demais telas.
    ano_atual = date.today().year
    anos_encontrados = sorted(
        {
            int(str(periodo)[:4])
            for periodo in periodos_disponiveis
            if re.match(r"^\d{4}-\d{2}$", str(periodo))
        },
        reverse=True,
    )
    anos_disponiveis = sorted(
        set(anos_encontrados + [ano_atual]),
        reverse=True,
    )

    with st.expander("Filtros e plano de contas", expanded=True):
        f1, f2 = st.columns([1.2, 2])
        with f1:
            ano_selecionado = st.selectbox(
                "Ano da Evolução Comercial",
                anos_disponiveis,
                index=(
                    anos_disponiveis.index(ano_atual)
                    if ano_atual in anos_disponiveis
                    else 0
                ),
                help=(
                    "Este filtro é independente do período ativo das demais telas. "
                    "A análise sempre mostra os 12 meses do ano escolhido."
                ),
            )

            meses_selecionados = [
                f"{int(ano_selecionado):04d}-{mes:02d}"
                for mes in range(1, 13)
            ]

            st.caption(
                f"Exibindo janeiro a dezembro de {ano_selecionado}. "
                "Meses sem movimento permanecem visíveis com valor zero."
            )
        planos_base = []
        if not contas_an.empty and "plano_contas" in contas_an.columns:
            planos_base = sorted(
                contas_an["plano_contas"].dropna().astype(str).str.strip()
                .loc[lambda x: x.ne("")].unique().tolist()
            )

        planos_catalogo = carregar_catalogo_planos_contas()
        planos_salvos = (
            list(cfg_analise.get("planos_contas_selecionados", []))
            + [cfg_analise.get("plano_contas_padrao", PLANO_CONTAS_PAGAMENTO_PADRAO)]
            + list(cfg_analise.get("planos_adicionais", []))
        )
        opcoes_planos = list(dict.fromkeys(
            [
                plano for plano in
                planos_salvos + planos_base + planos_catalogo
                if str(plano).strip()
            ]
        ))

        selecionados_salvos = [
            plano for plano in cfg_analise.get(
                "planos_contas_selecionados",
                [cfg_analise.get(
                    "plano_contas_padrao",
                    PLANO_CONTAS_PAGAMENTO_PADRAO,
                )],
            )
            if plano in opcoes_planos
        ]
        if not selecionados_salvos:
            selecionados_salvos = [PLANO_CONTAS_PAGAMENTO_PADRAO]

        with f2:
            planos_selecionados = st.multiselect(
                "Planos de contas usados em Pagamento de Fornecedor",
                opcoes_planos,
                default=selecionados_salvos,
                help=(
                    "É possível selecionar um ou vários planos. A configuração "
                    "fica salva para as próximas análises."
                ),
            )

        st.markdown("**Planos selecionados nesta análise:**")
        if planos_selecionados:
            for plano_ativo in planos_selecionados:
                st.caption(f"✓ {plano_ativo}")
        else:
            st.warning("Selecione pelo menos um plano de contas.")

        cadd1, cadd2, cadd3 = st.columns([2, 1, 1])
        with cadd1:
            novo_plano = st.text_input(
                "Adicionar outro plano de contas",
                placeholder="Cole o caminho completo do plano",
            )
        with cadd2:
            if st.button("Adicionar plano", use_container_width=True):
                if novo_plano.strip():
                    adicionais = list(cfg_analise.get("planos_adicionais", []))
                    if novo_plano.strip() not in adicionais:
                        adicionais.append(novo_plano.strip())
                    cfg_analise["planos_adicionais"] = adicionais
                    salvar_config_analise_comercial(cfg_analise)
                    st.success("Plano adicionado à lista.")
                    st.rerun()
        with cadd3:
            if st.button("Salvar configuração", use_container_width=True):
                if not planos_selecionados:
                    st.error("Selecione pelo menos um plano.")
                else:
                    cfg_analise["planos_contas_selecionados"] = planos_selecionados
                    cfg_analise["plano_contas_padrao"] = planos_selecionados[0]
                    salvar_config_analise_comercial(cfg_analise)
                    st.success("Configuração de planos salva.")
                    st.rerun()

    if meses_selecionados:
        if not vendas_an.empty:
            vendas_an = vendas_an[vendas_an["Competência"].isin(meses_selecionados)]
        if not entradas_an.empty:
            entradas_an = entradas_an[entradas_an["Competência"].isin(meses_selecionados)]
        if not entradas_fin_an.empty:
            entradas_fin_an = entradas_fin_an[
                entradas_fin_an["Competência"].isin(meses_selecionados)
            ]
        if not contas_an.empty:
            contas_an = contas_an[contas_an["Competência"].isin(meses_selecionados)]
        if not estoque_an.empty:
            estoque_an = estoque_an[estoque_an["Competência"].isin(meses_selecionados)]
        if not posicao_an.empty:
            posicao_an = posicao_an[posicao_an["Competência"].isin(meses_selecionados)]

    contas_filtradas = contas_an.copy()
    if not contas_filtradas.empty and "plano_contas" in contas_filtradas.columns:
        alvos = {
            _normalizar_texto_analise(plano)
            for plano in planos_selecionados
        }
        contas_filtradas = contas_filtradas[
            contas_filtradas["plano_contas"]
            .map(_normalizar_texto_analise)
            .isin(alvos)
        ]

    venda_total = float(vendas_an.get("Venda", pd.Series(dtype=float)).sum()) if not vendas_an.empty else 0.0
    custo_total = float(vendas_an.get("Custo", pd.Series(dtype=float)).sum()) if not vendas_an.empty else 0.0
    entrada_total = float(entradas_fin_an.get("Compra", pd.Series(dtype=float)).sum()) if not entradas_fin_an.empty else 0.0
    pagamento_total = float(contas_filtradas.get("Valor", pd.Series(dtype=float)).sum()) if not contas_filtradas.empty else 0.0
    cmv_total = (custo_total / venda_total * 100) if venda_total else 0.0
    lucro_total = venda_total - custo_total
    margem_total = (lucro_total / venda_total * 100) if venda_total else 0.0

    cards_an = st.columns(6)
    valores_cards = [
        ("Venda Geral", moeda_real(venda_total)),
        ("Custo Médio", moeda_real(custo_total)),
        ("Entrada Geral", moeda_real(entrada_total)),
        ("CMV Geral", percentual(cmv_total)),
        ("Lucro Bruto", moeda_real(lucro_total)),
        ("Pag. Fornecedor", moeda_real(pagamento_total)),
    ]
    for col, (rotulo, valor) in zip(cards_an, valores_cards):
        col.metric(rotulo, valor)

    areas = sorted(set(vendas_an.get("Área", pd.Series(dtype=str)).dropna().tolist()) | set(entradas_an.get("Área", pd.Series(dtype=str)).dropna().tolist()))
    meses = list(meses_selecionados)
    registros = []
    for area in areas:
        for mes in meses:
            vv = vendas_an[(vendas_an.get("Área") == area) & (vendas_an.get("Competência") == mes)] if not vendas_an.empty else pd.DataFrame()
            ee = entradas_an[(entradas_an.get("Área") == area) & (entradas_an.get("Competência") == mes)] if not entradas_an.empty else pd.DataFrame()
            venda = float(vv.get("Venda", pd.Series(dtype=float)).sum()) if not vv.empty else 0.0
            custo = float(vv.get("Custo", pd.Series(dtype=float)).sum()) if not vv.empty else 0.0
            compra = float(ee.get("Compra", pd.Series(dtype=float)).sum()) if not ee.empty else 0.0
            registros.append({
                "Área": area,
                "Competência": mes,
                "Venda": venda,
                "Custo": custo,
                "Compra": compra,
                "CMV (%)": (custo / venda * 100) if venda else 0,
                "Entrada - CMV": compra - custo,
                "Lucro Bruto": venda - custo,
                "Margem (%)": ((venda - custo) / venda * 100) if venda else 0,
            })
    analise_df = pd.DataFrame(registros)

    st.markdown("### Análise agrupada por área")
    indicador = st.selectbox("Indicador da matriz", ["Venda", "Custo", "Compra", "CMV (%)", "Entrada - CMV", "Lucro Bruto", "Margem (%)"])
    if not analise_df.empty:
        matriz = analise_df.pivot_table(
            index="Área",
            columns="Competência",
            values=indicador,
            aggfunc="sum",
            fill_value=0,
        )
        matriz["Total Geral"] = (
            matriz.sum(axis=1)
            if indicador not in ["CMV (%)", "Margem (%)"]
            else analise_df.groupby("Área")[indicador].mean()
        )
        matriz_exibicao = matriz.copy().astype(object)
        for coluna_matriz in matriz_exibicao.columns:
            if indicador in ["CMV (%)", "Margem (%)"]:
                matriz_exibicao[coluna_matriz] = matriz[coluna_matriz].map(percentual)
            else:
                matriz_exibicao[coluna_matriz] = matriz[coluna_matriz].map(moeda_real)
        dataframe_br(matriz_exibicao, use_container_width=True, height=390)
    else:
        st.info("Ainda não existem dados suficientes no cache para montar a matriz.")

    ano_numero_status = int(ano_selecionado)
    ultimo_mes_status = (
        12 if ano_numero_status < date.today().year
        else date.today().month if ano_numero_status == date.today().year
        else 0
    )
    competencias_realizadas_status = [
        f"{ano_numero_status:04d}-{mes_numero:02d}"
        for mes_numero in range(1, ultimo_mes_status + 1)
    ]
    competencias_com_dados_status = set(
        vendas_an.get("Competência", pd.Series(dtype=str))
        .dropna().astype(str).tolist()
    ) | set(
        entradas_an.get("Competência", pd.Series(dtype=str))
        .dropna().astype(str).tolist()
    )
    faltantes_status = [
        competencia
        for competencia in competencias_realizadas_status
        if competencia not in competencias_com_dados_status
    ]
    if faltantes_status:
        if MODO_VIEWER:
            st.warning(
                "Competências ainda não presentes nos resumos publicados: "
                + ", ".join(faltantes_status)
                + ". Se já foram atualizadas localmente, use Reconstruir análise; caso contrário, atualize no aplicativo local e publique novamente."
            )
        else:
            st.warning(
                "Competências realizadas ainda não atualizadas: "
                + ", ".join(faltantes_status)
                + ". Use o botão Atualizar análise."
            )

    ac_btn, ac_msg = st.columns([1.1, 3])
    with ac_btn:
        if MODO_VIEWER:
            atualizar_analise_clicado = st.button(
                "🔄 Recarregar base publicada",
                key=f"reconstruir_evolucao_{ano_selecionado}",
                use_container_width=True,
                help="Recarrega o .gz publicado, valida com quick_check e substitui atomicamente a cópia runtime. Não acessa o PostgreSQL.",
            )
            if atualizar_analise_clicado:
                try:
                    with st.spinner("Recarregando e validando a base publicada..."):
                        _preparar_sqlite_publicado(forcar=True)
                        try:
                            _carregar_resumos_analise.clear()
                        except Exception:
                            pass
                        try:
                            _ler_cache_analise_cached.clear()
                        except Exception:
                            pass
                        st.cache_data.clear()
                    st.success("Base publicada recarregada e validada no Viewer.")
                    st.rerun()
                except Exception as erro:
                    st.error("Não foi possível recarregar a base SQLite publicada.")
                    st.code(str(erro), language=None)
        else:
            if st.button(
                "🔄 Atualizar análise",
                key=f"atualizar_evolucao_{ano_selecionado}",
                use_container_width=True,
                help="Atualiza cada competência diretamente no PostgreSQL.",
            ):
                cfg_atualizacao = carregar_config_banco()
                identidade_banco_atualizacao = validar_identidade_banco(cfg_atualizacao)
                ano_numero = int(ano_selecionado)
                ultimo_mes = (
                    12 if ano_numero < date.today().year
                    else date.today().month if ano_numero == date.today().year
                    else 0
                )

                if ultimo_mes == 0:
                    st.info("O ano selecionado ainda não possui competências realizadas.")
                else:
                    competencias_existentes = set(
                        vendas_an.get("Competência", pd.Series(dtype=str))
                        .dropna().astype(str).tolist()
                    ) | set(
                        entradas_an.get("Competência", pd.Series(dtype=str))
                        .dropna().astype(str).tolist()
                    )

                    competencias_realizadas = [
                        f"{ano_numero:04d}-{numero_mes:02d}"
                        for numero_mes in range(1, ultimo_mes + 1)
                    ]
                    competencias_faltantes = [
                        competencia
                        for competencia in competencias_realizadas
                        if competencia not in competencias_existentes
                    ]

                    competencias_atualizar = (
                        competencias_faltantes
                        if competencias_faltantes
                        else competencias_realizadas
                    )

                    total_etapas = len(competencias_atualizar) * 3 + 1
                    etapa = 0
                    mensagens_ano = []
                    progresso_ano = st.progress(
                        0,
                        text=(
                            "Atualizando competências faltantes..."
                            if competencias_faltantes
                            else "Atualizando competências realizadas..."
                        ),
                    )

                    for competencia_atualizacao in competencias_atualizar:
                        numero_mes = int(competencia_atualizacao[-2:])
                        ultimo_dia = calendar.monthrange(ano_numero, numero_mes)[1]
                        data_inicio_atualizacao = f"{competencia_atualizacao}-01"
                        data_fim_atualizacao = f"{competencia_atualizacao}-{ultimo_dia:02d} 23:59:59"

                        for fonte_atualizacao in ["vendas", "entradas", "contas_pagar"]:
                            etapa += 1
                            titulo = FONTES_BANCO[fonte_atualizacao]["titulo"]
                            progresso_ano.progress(
                                etapa / total_etapas,
                                text=f"{competencia_atualizacao} • {titulo}",
                            )
                            try:
                                qtd = executar_atualizacao_fonte(
                                    fonte_atualizacao,
                                    cfg_atualizacao,
                                    competencia_atualizacao,
                                    data_inicio_atualizacao,
                                    data_fim_atualizacao,
                                )
                                mensagens_ano.append(
                                    f"{competencia_atualizacao} • {titulo}: {qtd:,} registros"
                                )
                            except Exception as erro:
                                mensagens_ano.append(
                                    f"{competencia_atualizacao} • {titulo}: ERRO — {erro}"
                                )

                    competencia_estoque = f"{ano_numero:04d}-{ultimo_mes:02d}"
                    ultimo_dia = calendar.monthrange(ano_numero, ultimo_mes)[1]
                    etapa += 1
                    progresso_ano.progress(
                        etapa / total_etapas,
                        text=f"{competencia_estoque} • Estoque",
                    )
                    try:
                        executar_atualizacao_fonte(
                            "estoque",
                            cfg_atualizacao,
                            competencia_estoque,
                            f"{competencia_estoque}-01",
                            f"{competencia_estoque}-{ultimo_dia:02d} 23:59:59",
                        )
                    except Exception as erro:
                        mensagens_ano.append(
                            f"{competencia_estoque} • Estoque: ERRO — {erro}"
                        )

                    st.cache_data.clear()
                    progresso_ano.progress(1.0, text="Atualização concluída.")
                    st.success("Evolução Comercial atualizada mês a mês.")
                    with st.expander("Resultado da atualização"):
                        st.code("\n".join(mensagens_ano[-60:]), language=None)
                    st.rerun()
    with ac_msg:
        if MODO_VIEWER:
            st.caption(
                "No Viewer, este botão reconstrói os resumos usando somente os dados já publicados. "
                "Para trazer novas competências do PostgreSQL, atualize-as no Atualizador Local e publique novamente."
            )
        else:
            st.caption(
                "A atualização é executada somente quando solicitada. "
                "Na navegação normal, a tela continua lendo os resumos rápidos."
            )

    st.markdown("### Evolução Comercial — mês a mês")
    st.caption(
        "Mesma dinâmica do relatório: indicadores nas linhas, competências nas colunas "
        "e Total Geral ao final."
    )

    resumo_mensal = []
    for mes in meses:
        vm = vendas_an[vendas_an.get("Competência") == mes] if not vendas_an.empty else pd.DataFrame()
        em = entradas_fin_an[
            entradas_fin_an.get("Competência") == mes
        ] if not entradas_fin_an.empty else pd.DataFrame()
        pm = contas_filtradas[contas_filtradas.get("Competência") == mes] if not contas_filtradas.empty else pd.DataFrame()
        psm = posicao_an[posicao_an.get("Competência") == mes] if not posicao_an.empty else pd.DataFrame()
        stm = estoque_an[estoque_an.get("Competência") == mes] if not estoque_an.empty else pd.DataFrame()

        venda = float(vm.get("Venda", pd.Series(dtype=float)).sum()) if not vm.empty else 0.0
        custo = float(vm.get("Custo", pd.Series(dtype=float)).sum()) if not vm.empty else 0.0
        entrada = float(em.get("Compra", pd.Series(dtype=float)).sum()) if not em.empty else 0.0
        pagamento = float(pm.get("Valor", pd.Series(dtype=float)).sum()) if not pm.empty else 0.0
        contas_total = float(psm.get("Contas a Pagar", pd.Series(dtype=float)).sum()) if not psm.empty else 0.0
        estoque_posicao = (
            float(psm.get("Estoque", pd.Series(dtype=float)).sum())
            if not psm.empty
            else 0.0
        )
        estoque_resumo = (
            float(stm.get("Estoque", pd.Series(dtype=float)).sum())
            if not stm.empty
            else 0.0
        )
        estoque_mes = (
            estoque_posicao
            if abs(estoque_posicao) > 0.000001
            else estoque_resumo
        )

        resumo_mensal.append({
            "Competência": mes,
            "Venda geral": venda,
            "Custo Médio Geral": custo,
            "Entrada geral": entrada,
            "CMV Geral": (custo / venda * 100) if venda else 0.0,
            "Custo Médio Geral - Entrada Geral (competência)": custo - entrada,
            "Lucro Bruto Geral": venda - custo,
            "Margem Contribuição": ((venda - custo) / venda * 100) if venda else 0.0,
            "Pagamento de Fornecedor": pagamento,
            "Custo médio - Pagamento de Fornecedor (caixa)": custo - pagamento,
            "Contas a Pagar Fornecedor Total": contas_total,
            "Estoque Mês": estoque_mes,
            "Estoque - Contas a Pagar": estoque_mes - contas_total,
        })

    # =========================================================
    # RESUMO EXECUTIVO AUTOMÁTICO DO FECHAMENTO MENSAL
    # =========================================================
    def _mes_extenso_pt(competencia):
        nomes = {
            1: "janeiro", 2: "fevereiro", 3: "março", 4: "abril",
            5: "maio", 6: "junho", 7: "julho", 8: "agosto",
            9: "setembro", 10: "outubro", 11: "novembro", 12: "dezembro",
        }
        try:
            ano, mes = map(int, str(competencia).split("-")[:2])
            return nomes.get(mes, str(competencia)), ano
        except Exception:
            return str(competencia), int(ano_selecionado)

    def _variacao(atual, anterior):
        diferenca = float(atual or 0) - float(anterior or 0)
        percentual_var = (diferenca / float(anterior) * 100) if anterior else 0.0
        return diferenca, percentual_var

    def _frase_variacao(rotulo, atual, anterior, referencia):
        dif, pct = _variacao(atual, anterior)
        if abs(dif) < 0.005:
            return f"{rotulo} permaneceu estável em relação a {referencia}."
        movimento = "crescimento" if dif > 0 else "redução"
        return (
            f"Em relação a {referencia}, houve {movimento} de "
            f"{moeda_real(abs(dif))}, equivalente a {percentual(abs(pct))}."
        )

    def _lista_texto(itens, positivo=True, limite=5):
        linhas = []
        for nome, valor in itens[:limite]:
            sinal = "+" if valor >= 0 else "-"
            linhas.append(f"* {nome}: {sinal}{moeda_real(abs(valor))};")
        if linhas:
            linhas[-1] = linhas[-1].rstrip(";") + "."
        return "\n".join(linhas)

    meses_com_movimento = [
        item for item in resumo_mensal
        if abs(float(item.get("Venda geral", 0) or 0)) > 0.000001
    ]

    if meses_com_movimento:
        fechamento_atual = meses_com_movimento[-1]
        competencia_fechamento = str(fechamento_atual["Competência"])
        mes_nome, ano_fechamento = _mes_extenso_pt(competencia_fechamento)
        indice_atual = meses.index(competencia_fechamento) if competencia_fechamento in meses else len(meses) - 1
        competencia_anterior = meses[indice_atual - 1] if indice_atual > 0 else None
        fechamento_anterior = next(
            (item for item in resumo_mensal if item["Competência"] == competencia_anterior),
            None,
        )
        fechamento_janeiro = next(
            (item for item in resumo_mensal if item["Competência"].endswith("-01")),
            None,
        )
        anteriores = [
            item for item in resumo_mensal
            if item["Competência"] < competencia_fechamento
            and float(item.get("Venda geral", 0) or 0) != 0
        ]

        venda_atual = float(fechamento_atual.get("Venda geral", 0) or 0)
        custo_atual = float(fechamento_atual.get("Custo Médio Geral", 0) or 0)
        lucro_atual = float(fechamento_atual.get("Lucro Bruto Geral", 0) or 0)
        margem_atual = float(fechamento_atual.get("Margem Contribuição", 0) or 0)
        entrada_atual = float(fechamento_atual.get("Entrada geral", 0) or 0)
        pagamento_atual = float(fechamento_atual.get("Pagamento de Fornecedor", 0) or 0)
        estoque_atual = float(fechamento_atual.get("Estoque Mês", 0) or 0)
        contas_atual = float(fechamento_atual.get("Contas a Pagar Fornecedor Total", 0) or 0)
        gap_atual = float(fechamento_atual.get("Estoque - Contas a Pagar", 0) or 0)

        venda_anterior = float((fechamento_anterior or {}).get("Venda geral", 0) or 0)
        lucro_anterior = float((fechamento_anterior or {}).get("Lucro Bruto Geral", 0) or 0)
        margem_anterior = float((fechamento_anterior or {}).get("Margem Contribuição", 0) or 0)
        estoque_anterior = float((fechamento_anterior or {}).get("Estoque Mês", 0) or 0)
        contas_anterior = float((fechamento_anterior or {}).get("Contas a Pagar Fornecedor Total", 0) or 0)
        gap_anterior = float((fechamento_anterior or {}).get("Estoque - Contas a Pagar", 0) or 0)
        venda_janeiro = float((fechamento_janeiro or {}).get("Venda geral", 0) or 0)

        venda_acumulada = sum(float(i.get("Venda geral", 0) or 0) for i in meses_com_movimento)
        lucro_acumulado = sum(float(i.get("Lucro Bruto Geral", 0) or 0) for i in meses_com_movimento)
        custo_acumulado = sum(float(i.get("Custo Médio Geral", 0) or 0) for i in meses_com_movimento)
        entrada_acumulada = sum(float(i.get("Entrada geral", 0) or 0) for i in meses_com_movimento)
        pagamento_acumulado = sum(float(i.get("Pagamento de Fornecedor", 0) or 0) for i in meses_com_movimento)
        margem_acumulada = (lucro_acumulado / venda_acumulada * 100) if venda_acumulada else 0.0
        compra_acima_cmv_acumulada = entrada_acumulada - custo_acumulado
        media_anteriores = (
            sum(float(i.get("Venda geral", 0) or 0) for i in anteriores) / len(anteriores)
            if anteriores else 0.0
        )

        dif_venda_ant, pct_venda_ant = _variacao(venda_atual, venda_anterior)
        dif_venda_jan, pct_venda_jan = _variacao(venda_atual, venda_janeiro)
        dif_media, pct_media = _variacao(venda_atual, media_anteriores)
        dif_lucro, pct_lucro = _variacao(lucro_atual, lucro_anterior)
        dif_margem_pp = margem_atual - margem_anterior
        entrada_acima_cmv = entrada_atual - custo_atual
        entrada_acima_pct = (entrada_acima_cmv / custo_atual * 100) if custo_atual else 0.0
        pagamento_acima_cmv = pagamento_atual - custo_atual
        estoque_variacao = estoque_atual - estoque_anterior
        contas_variacao = contas_atual - contas_anterior
        gap_variacao = gap_atual - gap_anterior

        # Crescimento por área em relação ao mês imediatamente anterior.
        variacoes_area = []
        excessos_area = []
        if not analise_df.empty:
            atual_area = analise_df[analise_df["Competência"] == competencia_fechamento]
            anterior_area = analise_df[analise_df["Competência"] == competencia_anterior] if competencia_anterior else pd.DataFrame()
            todas_areas = sorted(set(atual_area.get("Área", pd.Series(dtype=str))) | set(anterior_area.get("Área", pd.Series(dtype=str))))
            for area in todas_areas:
                venda_area_atual = float(atual_area.loc[atual_area["Área"] == area, "Venda"].sum()) if not atual_area.empty else 0.0
                venda_area_anterior = float(anterior_area.loc[anterior_area["Área"] == area, "Venda"].sum()) if not anterior_area.empty else 0.0
                variacoes_area.append((str(area), venda_area_atual - venda_area_anterior))
                compra_area = float(atual_area.loc[atual_area["Área"] == area, "Compra"].sum()) if not atual_area.empty else 0.0
                custo_area = float(atual_area.loc[atual_area["Área"] == area, "Custo"].sum()) if not atual_area.empty else 0.0
                excessos_area.append((str(area), compra_area - custo_area))
        crescimentos = sorted([x for x in variacoes_area if x[1] > 0], key=lambda x: x[1], reverse=True)
        recuos = sorted([x for x in variacoes_area if x[1] < 0], key=lambda x: x[1])
        excessos = sorted([x for x in excessos_area if x[1] > 0], key=lambda x: x[1], reverse=True)

        maior_resultado = venda_atual >= max(float(i.get("Venda geral", 0) or 0) for i in meses_com_movimento)

        # Índice de saúde comercial simples, transparente e auditável.
        score_saude = 50.0
        score_saude += 12 if pct_venda_ant > 0 else (-12 if pct_venda_ant < -3 else -4)
        score_saude += 12 if pct_lucro >= pct_venda_ant else (5 if pct_lucro > 0 else -10)
        score_saude += 10 if margem_atual >= margem_acumulada else (-6 if margem_atual < margem_acumulada - 1 else 2)
        score_saude += 8 if entrada_acima_cmv <= 0 else (-8 if entrada_acima_pct > 10 else -3)
        score_saude += 5 if contas_variacao <= 0 else -5
        score_saude += 3 if gap_variacao >= 0 else -3
        score_saude = max(0, min(100, round(score_saude)))
        if score_saude >= 90:
            nivel_saude, emoji_saude = "Excelente", "🟢"
        elif score_saude >= 75:
            nivel_saude, emoji_saude = "Bom", "🟡"
        elif score_saude >= 60:
            nivel_saude, emoji_saude = "Atenção", "🟠"
        else:
            nivel_saude, emoji_saude = "Crítico", "🔴"

        introducao = (
            f"{mes_nome.capitalize()} fechou com faturamento de {moeda_real(venda_atual)}"
            + (f", o maior resultado mensal de {ano_fechamento}." if maior_resultado else ".")
        )
        desempenho = [introducao]
        if fechamento_anterior:
            desempenho.append(_frase_variacao("O faturamento", venda_atual, venda_anterior, _mes_extenso_pt(competencia_anterior)[0]))
        if venda_janeiro and competencia_fechamento != f"{ano_fechamento}-01":
            desempenho.append(
                f"Comparado a janeiro, a variação alcança {percentual(pct_venda_jan)}, "
                f"equivalente a {moeda_real(abs(dif_venda_jan))}."
            )
        if media_anteriores:
            posicao_media = "acima" if dif_media >= 0 else "abaixo"
            desempenho.append(
                f"O mês ficou {percentual(abs(pct_media))} {posicao_media} da média dos meses anteriores do ano."
            )
        desempenho.append(
            f"O lucro bruto foi de {moeda_real(lucro_atual)}, com variação de {percentual(pct_lucro)} sobre o mês anterior. "
            f"A margem bruta encerrou em {percentual(margem_atual)}, "
            f"uma mudança de {numero_decimal(dif_margem_pp, 2)} ponto(s) percentual(is)."
        )
        if margem_atual < margem_acumulada:
            desempenho.append(
                f"Ainda há espaço para retornar à margem média acumulada do ano, atualmente em {percentual(margem_acumulada)}."
            )
        else:
            desempenho.append(
                f"A margem do mês ficou acima da média acumulada do ano, que está em {percentual(margem_acumulada)}."
            )

        compras_texto = (
            f"As entradas de mercadorias totalizaram {moeda_real(entrada_atual)}, ficando "
            f"{moeda_real(abs(entrada_acima_cmv))} — ou {percentual(abs(entrada_acima_pct))} — "
            f"{'acima' if entrada_acima_cmv >= 0 else 'abaixo'} do CMV do mês."
        )
        pagamentos_texto = (
            f"Os pagamentos a fornecedores totalizaram {moeda_real(pagamento_atual)}, ficando "
            f"{moeda_real(abs(pagamento_acima_cmv))} {'acima' if pagamento_acima_cmv >= 0 else 'abaixo'} do CMV do mês."
        )
        posicao_texto = (
            f"O estoque {'cresceu' if estoque_variacao >= 0 else 'reduziu'} {moeda_real(abs(estoque_variacao))}, "
            f"encerrando o período em {moeda_real(estoque_atual)}. As contas a pagar "
            f"{'cresceram' if contas_variacao >= 0 else 'caíram'} {moeda_real(abs(contas_variacao))}, "
            f"fechando em {moeda_real(contas_atual)}. A diferença entre estoque e contas a pagar "
            f"{'melhorou' if gap_variacao >= 0 else 'piorou'} {moeda_real(abs(gap_variacao))}, "
            f"encerrando em {moeda_real(gap_atual)}."
        )

        recomendacoes = []
        if margem_atual < margem_acumulada:
            recomendacoes.append(f"recuperar a margem para patamar igual ou superior a {percentual(margem_acumulada)}")
        if entrada_acima_cmv > 0:
            recomendacoes.append("alinhar as compras ao CMV e ao giro real das categorias")
        if excessos:
            recomendacoes.append(f"reduzir a formação de estoque principalmente em {', '.join(nome for nome, _ in excessos[:3])}")
        if pct_venda_ant > 0:
            recomendacoes.append("preservar o ritmo de crescimento comercial")
        if not recomendacoes:
            recomendacoes.append("manter o equilíbrio atual entre vendas, margem, compras e capital de giro")

        fatores_positivos = []
        pontos_atencao = []
        if pct_venda_ant > 0:
            fatores_positivos.append(f"Faturamento avançou {percentual(pct_venda_ant)} e adicionou {moeda_real(abs(dif_venda_ant))} à receita.")
        elif pct_venda_ant < 0:
            pontos_atencao.append(f"Faturamento recuou {percentual(abs(pct_venda_ant))}, com redução de {moeda_real(abs(dif_venda_ant))}.")
        if pct_lucro > pct_venda_ant and lucro_atual > 0:
            fatores_positivos.append("O lucro bruto cresceu acima do faturamento, indicando melhora da qualidade do resultado.")
        elif pct_lucro < pct_venda_ant:
            pontos_atencao.append("O lucro bruto evoluiu abaixo das vendas, sinalizando pressão sobre rentabilidade, mix ou precificação.")
        if dif_margem_pp > 0:
            fatores_positivos.append(f"A margem bruta recuperou {numero_decimal(abs(dif_margem_pp), 2)} ponto(s) percentual(is).")
        elif dif_margem_pp < 0:
            pontos_atencao.append(f"A margem bruta cedeu {numero_decimal(abs(dif_margem_pp), 2)} ponto(s) percentual(is).")
        if contas_variacao < 0:
            fatores_positivos.append(f"As contas a pagar foram reduzidas em {moeda_real(abs(contas_variacao))}.")
        elif contas_variacao > 0:
            pontos_atencao.append(f"As contas a pagar cresceram {moeda_real(abs(contas_variacao))}.")
        if entrada_acima_cmv > 0:
            pontos_atencao.append(f"As compras superaram o consumo em {moeda_real(entrada_acima_cmv)} ({percentual(entrada_acima_pct)}).")
        else:
            fatores_positivos.append("As compras permaneceram alinhadas ou abaixo do CMV.")
        if gap_variacao > 0:
            fatores_positivos.append(f"A relação entre estoque e contas a pagar melhorou {moeda_real(gap_variacao)}.")
        elif gap_variacao < 0:
            pontos_atencao.append(f"A relação entre estoque e contas a pagar piorou {moeda_real(abs(gap_variacao))}.")
        if crescimentos:
            fatores_positivos.append(f"O principal vetor de crescimento foi {crescimentos[0][0]}, com {moeda_real(crescimentos[0][1])}.")
        if recuos:
            pontos_atencao.append(f"O maior recuo ocorreu em {recuos[0][0]}, com impacto de {moeda_real(abs(recuos[0][1]))}.")
        if not fatores_positivos:
            fatores_positivos.append("O mês manteve estabilidade nos principais indicadores.")
        if not pontos_atencao:
            pontos_atencao.append("Não foram identificados desvios críticos no período.")

        def _status_area(cond_bom, cond_atencao=True):
            if cond_bom:
                return "🟢 Favorável"
            if cond_atencao:
                return "🟡 Atenção"
            return "🔴 Crítico"

        diagnostico_areas = [
            ("Comercial", _status_area(pct_venda_ant > 0, pct_venda_ant >= -3), f"Variação: {percentual(pct_venda_ant)}"),
            ("Rentabilidade", _status_area(margem_atual >= margem_acumulada, margem_atual >= margem_acumulada - 1), f"Margem: {percentual(margem_atual)}"),
            ("Compras", _status_area(entrada_acima_cmv <= 0, entrada_acima_pct <= 10), f"Compras × CMV: {percentual(entrada_acima_pct)}"),
            ("Estoque", _status_area(estoque_variacao <= 0, estoque_variacao <= max(custo_atual * .05, 1)), f"Variação: {moeda_real(estoque_variacao)}"),
            ("Financeiro", _status_area(contas_variacao <= 0, contas_variacao <= max(custo_atual * .05, 1)), f"Contas a pagar: {moeda_real(contas_variacao)}"),
            ("Capital de Giro", _status_area(gap_variacao >= 0, gap_variacao >= -max(custo_atual * .03, 1)), f"Evolução: {moeda_real(gap_variacao)}"),
        ]

        prioridades = []
        if entrada_acima_cmv > 0:
            tema = "Ajustar compras ao consumo"
            if excessos:
                tema += f" em {', '.join(nome for nome, _ in excessos[:3])}"
            prioridades.append(("🔴 Alta", tema, "Reduzir estoque excedente e liberar capital de giro."))
        if margem_atual < margem_acumulada:
            prioridades.append(("🔴 Alta", "Recuperar margem e revisar precificação/mix", f"Retornar ao patamar mínimo de {percentual(margem_acumulada)}."))
        if recuos:
            prioridades.append(("🟡 Média", f"Plano de recuperação para {recuos[0][0]}", "Reverter a maior contribuição negativa do mês."))
        if pct_venda_ant > 0:
            prioridades.append(("🟢 Manter", "Sustentar o crescimento comercial", "Preservar ações e categorias que impulsionaram a receita."))
        if not prioridades:
            prioridades.append(("🟢 Manter", "Consolidar o equilíbrio operacional", "Manter vendas, margem, compras e caixa sob controle."))

        ultimos_tres = meses_com_movimento[-3:]
        crescimentos_recentes = []
        for indice_proj in range(1, len(ultimos_tres)):
            base_proj = float(ultimos_tres[indice_proj - 1].get("Venda geral", 0) or 0)
            atual_proj = float(ultimos_tres[indice_proj].get("Venda geral", 0) or 0)
            if base_proj:
                crescimentos_recentes.append((atual_proj / base_proj) - 1)
        taxa_proj = sum(crescimentos_recentes) / len(crescimentos_recentes) if crescimentos_recentes else 0.0
        taxa_proj = max(-0.15, min(0.15, taxa_proj))
        faturamento_projetado = max(0.0, venda_atual * (1 + taxa_proj))
        faixa_inferior = faturamento_projetado * 0.98
        faixa_superior = faturamento_projetado * 1.02
        margem_projetada = max(0.0, margem_atual + (dif_margem_pp * 0.5))

        resumo_markdown = f"""### {emoji_saude} Saúde Comercial — {nivel_saude} ({score_saude}/100)

## DESEMPENHO DE {mes_nome.upper()}

{' '.join(desempenho)}

"""
        if crescimentos:
            resumo_markdown += "**Os principais crescimentos de faturamento em relação ao mês anterior vieram de:**\n\n" + _lista_texto(crescimentos) + "\n\n"
        if recuos:
            resumo_markdown += "**Os principais recuos ocorreram em:**\n\n" + _lista_texto(recuos) + "\n\n"

        resumo_markdown += f"""## COMPRAS, ESTOQUE E FORNECEDORES

{compras_texto}

"""
        if excessos:
            total_excesso_top = sum(v for _, v in excessos[:3])
            participacao_top = total_excesso_top / entrada_acima_cmv * 100 if entrada_acima_cmv > 0 else 0
            resumo_markdown += "**A maior formação de estoque concentrou-se em:**\n\n" + _lista_texto(excessos, limite=3) + "\n\n"
            resumo_markdown += f"As três principais áreas representam aproximadamente {percentual(participacao_top)} do volume comprado acima do consumo no mês.\n\n"
        resumo_markdown += f"{pagamentos_texto}\n\n{posicao_texto}\n\n"

        resumo_markdown += f"""## ACUMULADO DE JANEIRO A {mes_nome.upper()}

* Faturamento: {moeda_real(venda_acumulada)};
* Lucro bruto: {moeda_real(lucro_acumulado)};
* Margem bruta acumulada: {percentual(margem_acumulada)};
* Compras acima do CMV acumulado: {moeda_real(compra_acima_cmv_acumulada)};
* Pagamentos a fornecedores: {moeda_real(pagamento_acumulado)}.

## O QUE MAIS INFLUENCIOU O RESULTADO

**Principais fatores positivos:**

""" + "\n".join(f"* {item}" for item in fatores_positivos[:6]) + """

**Pontos de atenção:**

""" + "\n".join(f"* {item}" for item in pontos_atencao[:6]) + f"""

## TENDÊNCIA PARA O PRÓXIMO MÊS

Mantido o ritmo recente, o faturamento tende a ficar entre **{moeda_real(faixa_inferior)}** e **{moeda_real(faixa_superior)}**, com referência central de **{moeda_real(faturamento_projetado)}**. A margem indicativa é de aproximadamente **{percentual(margem_projetada)}**. Esta projeção é estatística e deve ser reavaliada com calendário comercial, sazonalidade e campanhas previstas.

## CONCLUSÃO E DIRECIONAMENTO

De forma geral, {mes_nome} apresentou saúde comercial classificada como **{nivel_saude}**. O foco do próximo mês deve ser {', '.join(recomendacoes[:-1]) + (' e ' if len(recomendacoes) > 1 else '') + recomendacoes[-1]}.
"""

        st.markdown("---")
        cor_saude = "good" if score_saude >= 90 else ("warn" if score_saude >= 60 else "bad")
        cor_venda = "good" if pct_venda_ant >= 0 else "bad"
        cor_margem = "good" if dif_margem_pp >= 0 else "warn"
        cor_compras = "good" if entrada_acima_cmv <= 0 else ("warn" if entrada_acima_pct <= 10 else "bad")
        data_geracao_resumo = datetime.now().strftime("%d/%m/%Y às %H:%M")

        st.markdown(
            f"""
            <div class="rec-wrap">
              <div class="rec-hero">
                <div class="rec-kicker">Inteligência Comercial • Fechamento Mensal</div>
                <div class="rec-title">Copiloto Executivo — Fechamento Mensal</div>
                <div class="rec-subtitle">Diagnóstico executivo de desempenho, causas, riscos, prioridades e tendência, calculado exclusivamente com os dados persistidos na Análise Comercial.</div>
                <div class="rec-meta">
                  <span class="rec-chip">📅 Competência: {mes_nome.capitalize()}/{ano_fechamento}</span>
                  <span class="rec-chip">🏢 Visão: Rede consolidada</span>
                  <span class="rec-chip">🕒 Gerado em: {data_geracao_resumo}</span>
                  <span class="rec-chip">🔒 Dados auditáveis</span>
                </div>
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        delta_venda_texto = percentual(pct_venda_ant) if fechamento_anterior else "Sem comparativo anterior"
        delta_margem_texto = f"{numero_decimal(dif_margem_pp, 2)} p.p. versus mês anterior"
        compras_sinal = "abaixo do CMV" if entrada_acima_cmv <= 0 else "acima do CMV"
        st.markdown(
            f"""
            <div class="rec-grid">
              <div class="rec-card {cor_saude}">
                <div class="rec-card-label">Saúde Comercial</div>
                <div class="rec-card-value rec-score"><span class="rec-score-dot"></span>{nivel_saude}</div>
                <div class="rec-card-note">Pontuação executiva: <strong>{score_saude}/100</strong></div>
              </div>
              <div class="rec-card {cor_venda}">
                <div class="rec-card-label">Faturamento do mês</div>
                <div class="rec-card-value">{moeda_real(venda_atual)}</div>
                <div class="rec-card-note">{delta_venda_texto} versus mês anterior</div>
              </div>
              <div class="rec-card {cor_margem}">
                <div class="rec-card-label">Margem Bruta</div>
                <div class="rec-card-value">{percentual(margem_atual)}</div>
                <div class="rec-card-note">{delta_margem_texto}</div>
              </div>
              <div class="rec-card {cor_compras}">
                <div class="rec-card-label">Compras × CMV</div>
                <div class="rec-card-value">{moeda_real(entrada_acima_cmv)}</div>
                <div class="rec-card-note">{percentual(abs(entrada_acima_pct))} {compras_sinal}</div>
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        positivos_html = "".join(f"<li>{item}</li>" for item in fatores_positivos[:6])
        atencao_html = "".join(f"<li>{item}</li>" for item in pontos_atencao[:6])
        st.markdown(
            f"""
            <div class="rec-section-label">Leitura executiva do período</div>
            <div class="rec-insight-grid">
              <div class="rec-insight good"><div class="rec-insight-title">🟢 Fatores que impulsionaram o resultado</div><ul>{positivos_html}</ul></div>
              <div class="rec-insight warn"><div class="rec-insight-title">🟠 Riscos e pontos de atenção</div><ul>{atencao_html}</ul></div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        diagnostico_html = "".join(
            f"<div class='rec-diagnosis-item'><div class='rec-diagnosis-name'>{nome}</div><div class='rec-diagnosis-status'>{status}</div><div class='rec-diagnosis-note'>{nota}</div></div>"
            for nome, status, nota in diagnostico_areas
        )
        st.markdown('<div class="rec-section-label">Diagnóstico por área</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="rec-diagnosis">{diagnostico_html}</div>', unsafe_allow_html=True)

        prioridades_html = "".join(
            f"<tr><td><strong>{prioridade}</strong></td><td>{tema}</td><td>{impacto}</td></tr>"
            for prioridade, tema, impacto in prioridades
        )
        st.markdown('<div class="rec-section-label">Plano de ação priorizado</div>', unsafe_allow_html=True)
        st.markdown(
            f"<table class='rec-priority-table'><thead><tr><th>Prioridade</th><th>Ação recomendada</th><th>Impacto esperado</th></tr></thead><tbody>{prioridades_html}</tbody></table>",
            unsafe_allow_html=True,
        )
        st.markdown(
            f"<div class='rec-projection'><div class='rec-section-label'>Tendência indicativa para o próximo mês</div><div class='rec-projection-value'>{moeda_real(faixa_inferior)} a {moeda_real(faixa_superior)}</div><div class='rec-card-note'>Referência central: {moeda_real(faturamento_projetado)} • Margem indicativa: {percentual(margem_projetada)} • Base: ritmo médio recente, limitado a ±15%.</div></div>",
            unsafe_allow_html=True,
        )

        st.markdown('<div class="rec-section-label">Relatório executivo consolidado</div>', unsafe_allow_html=True)
        st.markdown('<div class="rec-report">', unsafe_allow_html=True)
        st.markdown(_markdown_executivo_seguro(resumo_markdown))
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="rec-actions"><div class="rec-section-label">Distribuição e histórico do fechamento</div>', unsafe_allow_html=True)
        nome_responsavel = st.text_input(
            "Destinatário do fechamento",
            value=str(st.session_state.get("destinatario_resumo_comercial", "Maycon")),
            key="destinatario_resumo_comercial",
            help="Nome utilizado na abertura do texto para envio por e-mail ou WhatsApp.",
        )
        texto_copia = (
            f"{nome_responsavel.strip() or 'Gestor'}, segue o fechamento comercial de {mes_nome} "
            f"e o consolidado de janeiro a {mes_nome} de {ano_fechamento}, conforme demonstrado na aba Análise.\n\n"
            + re.sub(r"[#*]", "", resumo_markdown)
        )
        with st.expander("✉️ Texto pronto para copiar e enviar", expanded=False):
            st.text_area("Fechamento comercial", value=texto_copia, height=500)

        col_baixar, col_salvar, col_info = st.columns([1, 1, 2])
        with col_baixar:
            st.download_button(
                "⬇️ Baixar resumo em TXT",
                data=texto_copia.encode("utf-8-sig"),
                file_name=f"fechamento_comercial_{competencia_fechamento}.txt",
                mime="text/plain",
                use_container_width=True,
            )
        with col_salvar:
            if st.button("💾 Salvar fechamento", use_container_width=True, key=f"salvar_fechamento_{competencia_fechamento}"):
                pasta_fechamentos = Path("data") / "fechamentos_comerciais"
                pasta_fechamentos.mkdir(parents=True, exist_ok=True)
                arquivo_fechamento = pasta_fechamentos / f"fechamento_{competencia_fechamento}.json"
                arquivo_fechamento.write_text(
                    json.dumps({
                        "competencia": competencia_fechamento,
                        "gerado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "destinatario": nome_responsavel,
                        "saude_score": score_saude,
                        "saude_nivel": nivel_saude,
                        "texto": texto_copia,
                    }, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
                st.success(f"Fechamento de {competencia_fechamento} salvo no histórico.")
        with col_info:
            st.caption(
                "O fechamento considera automaticamente o último mês do ano selecionado que possui vendas. "
                "Ao atualizar o banco, o texto é recalculado com os novos dados."
            )
        st.markdown('</div>', unsafe_allow_html=True)

    if resumo_mensal:
        evolucao_df = pd.DataFrame(resumo_mensal).set_index("Competência").T.fillna(0)

        nomes_meses = {
            f"{int(ano_selecionado):04d}-01": "Jan",
            f"{int(ano_selecionado):04d}-02": "Fev",
            f"{int(ano_selecionado):04d}-03": "Mar",
            f"{int(ano_selecionado):04d}-04": "Abr",
            f"{int(ano_selecionado):04d}-05": "Mai",
            f"{int(ano_selecionado):04d}-06": "Jun",
            f"{int(ano_selecionado):04d}-07": "Jul",
            f"{int(ano_selecionado):04d}-08": "Ago",
            f"{int(ano_selecionado):04d}-09": "Set",
            f"{int(ano_selecionado):04d}-10": "Out",
            f"{int(ano_selecionado):04d}-11": "Nov",
            f"{int(ano_selecionado):04d}-12": "Dez",
        }

        indicadores_percentuais = {"CMV Geral", "Margem Contribuição"}
        colunas_meses = list(evolucao_df.columns)

        # Total Geral conforme a natureza de cada indicador.
        total_geral = {}
        for indicador in evolucao_df.index:
            if indicador == "CMV Geral":
                total_venda = sum(item["Venda geral"] for item in resumo_mensal)
                total_custo = sum(item["Custo Médio Geral"] for item in resumo_mensal)
                total_geral[indicador] = (
                    total_custo / total_venda * 100 if total_venda else 0.0
                )
            elif indicador == "Margem Contribuição":
                total_venda = sum(item["Venda geral"] for item in resumo_mensal)
                total_custo = sum(item["Custo Médio Geral"] for item in resumo_mensal)
                total_geral[indicador] = (
                    (total_venda - total_custo) / total_venda * 100
                    if total_venda else 0.0
                )
            elif indicador in {
                "Contas a Pagar Fornecedor Total",
                "Estoque Mês",
                "Estoque - Contas a Pagar",
            }:
                # Indicadores de posição: usa o último mês que realmente possui
                # posição, sem considerar meses futuros zerados.
                serie_posicao = pd.to_numeric(
                    evolucao_df.loc[indicador, colunas_meses],
                    errors="coerce",
                ).fillna(0)
                meses_com_posicao = serie_posicao[serie_posicao.ne(0)]
                total_geral[indicador] = (
                    float(meses_com_posicao.iloc[-1])
                    if not meses_com_posicao.empty
                    else 0.0
                )
            else:
                total_geral[indicador] = float(
                    pd.to_numeric(evolucao_df.loc[indicador], errors="coerce")
                    .fillna(0).sum()
                )

        evolucao_df["Total Geral"] = pd.Series(total_geral)

        def _formatar_evolucao(valor, indicador):
            numero = float(valor or 0)
            if indicador in indicadores_percentuais:
                return percentual(numero)
            return moeda_real(numero)

        evolucao_exibicao = evolucao_df.copy().astype(object)
        evolucao_exibicao = evolucao_exibicao.rename(columns=nomes_meses)
        for indicador in evolucao_exibicao.index:
            for coluna in evolucao_exibicao.columns:
                coluna_origem = next(
                    (
                        competencia
                        for competencia, nome_mes in nomes_meses.items()
                        if nome_mes == coluna
                    ),
                    coluna,
                )
                evolucao_exibicao.loc[indicador, coluna] = _formatar_evolucao(
                    evolucao_df.loc[indicador, coluna_origem], indicador
                )

        dataframe_br(
            evolucao_exibicao,
            use_container_width=True,
            height=500,
        )

        # A tabela mantém janeiro a dezembro, mas meses totalmente zerados
        # não participam dos gráficos nem das médias visuais.
        grafico_mensal = pd.DataFrame(resumo_mensal)
        grafico_mensal["Mês"] = grafico_mensal["Competência"].map(nomes_meses)

        colunas_movimento = [
            "Venda geral",
            "Custo Médio Geral",
            "Entrada geral",
            "Pagamento de Fornecedor",
            "Contas a Pagar Fornecedor Total",
            "Estoque Mês",
        ]
        mascara_movimento = (
            grafico_mensal[colunas_movimento]
            .apply(pd.to_numeric, errors="coerce")
            .fillna(0)
            .abs()
            .sum(axis=1)
            > 0
        )
        grafico_realizado = grafico_mensal.loc[mascara_movimento].copy()

        if not grafico_realizado.empty:
            fig_evol = go.Figure()
            for coluna in [
                "Venda geral",
                "Custo Médio Geral",
                "Entrada geral",
                "Pagamento de Fornecedor",
            ]:
                fig_evol.add_scatter(
                    x=grafico_realizado["Mês"],
                    y=grafico_realizado[coluna],
                    mode="lines+markers",
                    name=coluna,
                    connectgaps=False,
                )
            fig_evol.update_layout(
                height=380,
                title="Evolução dos principais indicadores",
                margin=dict(l=10, r=10, t=50, b=10),
                yaxis_title="R$",
                xaxis_title=f"Meses realizados de {ano_selecionado}",
            )
            plotly_chart_br(fig_evol, use_container_width=True)
        else:
            st.info("Ainda não existem meses realizados para exibir no gráfico.")

        st.markdown("### CMV por área")
        if not analise_df.empty:
            cmv_area = analise_df.groupby("Área", as_index=False).agg(
                Venda=("Venda", "sum"),
                Custo=("Custo", "sum"),
            )
            cmv_area["CMV (%)"] = cmv_area.apply(
                lambda r: r["Custo"] / r["Venda"] * 100 if r["Venda"] else 0,
                axis=1,
            )
            fig_cmv = px.bar(
                cmv_area.sort_values("CMV (%)"),
                x="CMV (%)",
                y="Área",
                orientation="h",
                text_auto=False,
            )
            fig_cmv.update_layout(
                height=400,
                margin=dict(l=10, r=10, t=30, b=10),
            )
            plotly_chart_br(fig_cmv, use_container_width=True)
    else:
        st.info("Ainda não existem competências suficientes para montar a evolução comercial.")

    st.caption("Planos de contas aplicados: " + (" | ".join(planos_selecionados) if planos_selecionados else "Nenhum"))

elif visao == "Meu Resumo":
    section("Meu Resumo", "sec-blue")
    _perfil_resumo = _perfil_logado()
    _escopo_resumo = _escopo_usuario_logado()

    st.caption(
        "Resumo consolidado e seguro do período. Compradores e gerentes visualizam "
        "exclusivamente os resultados vinculados ao próprio acesso."
    )

    if _perfil_resumo == "Comprador":
        _nome = _escopo_resumo
        _real = REALIZADOS.loc[
            REALIZADOS["Comprador"].map(_norm_escopo).eq(_norm_escopo(_nome))
        ].copy() if "Comprador" in REALIZADOS.columns else pd.DataFrame()
        _meta = METAS.loc[
            METAS["Comprador"].map(_norm_escopo).eq(_norm_escopo(_nome))
        ].copy() if "Comprador" in METAS.columns else pd.DataFrame()
        _prem = PREMIO.loc[
            PREMIO["Comprador"].map(_norm_escopo).eq(_norm_escopo(_nome))
        ].copy() if "Comprador" in PREMIO.columns else pd.DataFrame()
        _kpis = _kpis_holerite_comprador(_nome)

        st.markdown(f"### 👤 {_nome}")
        st.caption("🔒 Visualização individual. Nenhum total da empresa ou resultado de outro comprador é carregado nesta tela.")

        _r = _real.iloc[0] if not _real.empty else None
        _m = _meta.iloc[0] if not _meta.empty else None

        def _vlinha(linha, coluna, padrao=0.0):
            if linha is None or coluna not in linha.index:
                return padrao
            try:
                return float(pd.to_numeric(linha[coluna], errors="coerce"))
            except Exception:
                return padrao

        _fat = _vlinha(_r, "Faturamento Total Atual")
        _fat_meta = _vlinha(_m, "Faturamento Total META")
        _estoque = _vlinha(_r, "Estoque Total")
        _ruptura = _vlinha(_r, "Ruptura %")
        _reposicao = _vlinha(_r, "Reposição CMV %")
        _premio_total = 0.0
        if not _prem.empty:
            _cols_prem = [c for c in _prem.columns if str(c).endswith(" Prêmio")]
            if _cols_prem:
                _premio_total = float(_prem[_cols_prem].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1).sum())

        _ating_fat = (_fat / _fat_meta * 100) if _fat_meta else 0.0
        _cards = st.columns(5)
        _cards[0].metric("Faturamento", moeda_real(_fat), f"{percentual(_ating_fat)} da meta")
        _cards[1].metric("Meta de faturamento", moeda_real(_fat_meta))
        _cards[2].metric("Estoque total", moeda_real(_estoque))
        _cards[3].metric("Ruptura", percentual(_ruptura))
        _cards[4].metric("Prêmio conquistado", moeda_real(_premio_total))

        st.markdown("### 🎯 Resultado dos KPI's")
        if _kpis is None or _kpis.empty:
            st.info("Ainda não existem KPI's calculados para este comprador nesta competência.")
        else:
            _tab = _kpis.copy()
            if "Atingimento (%)" in _tab.columns:
                _tab["Status"] = pd.to_numeric(_tab["Atingimento (%)"], errors="coerce").fillna(0).map(
                    lambda v: "✅ Atingida" if v >= 100 else ("🟡 Atenção" if v >= 90 else "🔴 Crítico")
                )
            dataframe_br(_tab, use_container_width=True, hide_index=True, export_title=f"Meu Resumo - {_nome}")

            _ating = pd.to_numeric(_tab.get("Atingimento (%)", pd.Series(dtype=float)), errors="coerce").fillna(0)
            _c = st.columns(4)
            _c[0].metric("KPI's avaliados", numero_inteiro(len(_tab)))
            _c[1].metric("KPI's atingidos", numero_inteiro((_ating >= 100).sum()))
            _c[2].metric("Em atenção", numero_inteiro(((_ating >= 90) & (_ating < 100)).sum()))
            _c[3].metric("Críticos", numero_inteiro((_ating < 90).sum()))

        st.markdown("### 📊 Indicadores operacionais")
        _ops = pd.DataFrame([
            {"Indicador": "Faturamento", "Resultado": moeda_real(_fat)},
            {"Indicador": "Atingimento do faturamento", "Resultado": percentual(_ating_fat)},
            {"Indicador": "Estoque total", "Resultado": moeda_real(_estoque)},
            {"Indicador": "Ruptura", "Resultado": percentual(_ruptura)},
            {"Indicador": "Reposição CMV", "Resultado": percentual(_reposicao)},
            {"Indicador": "Prêmio conquistado", "Resultado": moeda_real(_premio_total)},
        ])
        dataframe_br(_ops, use_container_width=True, hide_index=True, export_title=f"Indicadores - {_nome}")

    elif _perfil_resumo == "Gerente":
        _gerente = _escopo_resumo
        _hol_lojas = _holerite_lojas_por_meta(PERIODO_REALIZADO_USADO)
        _hol_ger = _holerite_gerentes_por_meta(PERIODO_REALIZADO_USADO)

        _hol_ger = _filtrar_gerente_privado(_hol_ger, "Gerente Comercial")
        _hol_lojas = _filtrar_gerente_privado(_hol_lojas, "Gerente Comercial")

        st.markdown(f"### 👔 {_gerente}")
        st.caption("🔒 Visualização gerencial privada. Somente as lojas vinculadas ao gerente conectado participam dos cálculos.")

        if _hol_ger.empty:
            st.info("Não existem resultados vinculados a este gerente na competência selecionada.")
        else:
            _g = _hol_ger.iloc[0]
            _meta_fat = float(pd.to_numeric(_g.get("Meta Faturamento (R$)", 0), errors="coerce") or 0)
            _fat = float(pd.to_numeric(_g.get("Realizado Faturamento (R$)", 0), errors="coerce") or 0)
            _meta_mb = float(pd.to_numeric(_g.get("Meta Margem Bruta (R$)", 0), errors="coerce") or 0)
            _mb = float(pd.to_numeric(_g.get("Realizado Margem Bruta (R$)", 0), errors="coerce") or 0)
            _premio = float(pd.to_numeric(_g.get("Prêmio conquistado (R$)", 0), errors="coerce") or 0)
            _premio_max = float(pd.to_numeric(_g.get("Prêmio máximo (R$)", 0), errors="coerce") or 0)
            _ating_fat = (_fat / _meta_fat * 100) if _meta_fat else 0.0
            _ating_mb = (_mb / _meta_mb * 100) if _meta_mb else 0.0
            _ating_geral = (_premio / _premio_max * 100) if _premio_max else 0.0
            _qtd_lojas = int(_hol_lojas["Filial"].nunique()) if not _hol_lojas.empty and "Filial" in _hol_lojas.columns else int(pd.to_numeric(_g.get("Lojas", 0), errors="coerce") or 0)

            _cards = st.columns(5)
            _cards[0].metric("Lojas sob gestão", numero_inteiro(_qtd_lojas))
            _cards[1].metric("Faturamento", moeda_real(_fat), f"{percentual(_ating_fat)} da meta")
            _cards[2].metric("Margem bruta", moeda_real(_mb), f"{percentual(_ating_mb)} da meta")
            _cards[3].metric("Atingimento geral", percentual(_ating_geral))
            _cards[4].metric("Prêmio conquistado", moeda_real(_premio))

            st.markdown("### 🎯 KPI's gerenciais")
            _kpi_g = pd.DataFrame([
                {"KPI": "Faturamento", "Meta": _meta_fat, "Realizado": _fat, "Atingimento (%)": _ating_fat},
                {"KPI": "Margem Bruta", "Meta": _meta_mb, "Realizado": _mb, "Atingimento (%)": _ating_mb},
            ])
            _kpi_g["Status"] = _kpi_g["Atingimento (%)"].map(
                lambda v: "✅ Atingida" if v >= 100 else ("🟡 Atenção" if v >= 90 else "🔴 Crítico")
            )
            dataframe_br(_kpi_g, use_container_width=True, hide_index=True, export_title=f"Meu Resumo Gerencial - {_gerente}")

            st.markdown("### 🏬 Resultado das minhas lojas")
            if _hol_lojas.empty:
                st.info("Não existem lojas vinculadas com resultado disponível.")
            else:
                _cols = [
                    c for c in [
                        "Filial", "Supervisor", "Faturamento Total META", "Faturamento Total Atual",
                        "Margem Bruta META", "Margem Bruta Atual", "Atingimento Geral (%)",
                        "Prêmio máximo da loja", "Prêmio conquistado pelas metas", "Saldo não conquistado"
                    ] if c in _hol_lojas.columns
                ]
                _lojas_view = _hol_lojas[_cols].copy()
                dataframe_br(
                    _lojas_view,
                    use_container_width=True,
                    hide_index=True,
                    height=min(430, 100 + 36 * max(len(_lojas_view), 1)),
                    export_title=f"Minhas Lojas - {_gerente}",
                )

    elif _perfil_resumo == "Administrador":
        st.info("Administrador: selecione abaixo o tipo de resumo individual que deseja consultar.")
        _tipo = st.radio("Tipo de acesso", ["Comprador", "Gerente"], horizontal=True, key="resumo_admin_tipo")
        if _tipo == "Comprador":
            _lista = sorted(REALIZADOS["Comprador"].dropna().astype(str).unique().tolist()) if "Comprador" in REALIZADOS.columns else []
            if not _lista:
                st.info("Não há compradores com resultado disponível.")
            else:
                _sel = st.selectbox("Comprador", _lista, key="resumo_admin_comprador")
                st.caption("Para conferir a visão exata desse usuário, acesse com o perfil do comprador. Esta seleção administrativa não altera permissões.")
                _real = REALIZADOS[REALIZADOS["Comprador"].astype(str) == str(_sel)].copy()
                _meta = METAS[METAS["Comprador"].astype(str) == str(_sel)].copy()
                _kpis = _kpis_holerite_comprador(_sel)
                if not _real.empty:
                    _r = _real.iloc[0]
                    _m = _meta.iloc[0] if not _meta.empty else None
                    _fat = float(pd.to_numeric(_r.get("Faturamento Total Atual", 0), errors="coerce") or 0)
                    _fat_meta = float(pd.to_numeric(_m.get("Faturamento Total META", 0), errors="coerce") or 0) if _m is not None else 0
                    _c = st.columns(3)
                    _c[0].metric("Faturamento", moeda_real(_fat))
                    _c[1].metric("Meta", moeda_real(_fat_meta))
                    _c[2].metric("Atingimento", percentual((_fat/_fat_meta*100) if _fat_meta else 0))
                if _kpis is not None and not _kpis.empty:
                    dataframe_br(_kpis, use_container_width=True, hide_index=True, export_title=f"Resumo Comprador - {_sel}")
        else:
            _hg = _holerite_gerentes_por_meta(PERIODO_REALIZADO_USADO)
            _lista = sorted(_hg["Gerente Comercial"].dropna().astype(str).unique().tolist()) if not _hg.empty and "Gerente Comercial" in _hg.columns else []
            if not _lista:
                st.info("Não há gerentes com resultado disponível.")
            else:
                _sel = st.selectbox("Gerente", _lista, key="resumo_admin_gerente")
                _g = _hg[_hg["Gerente Comercial"].astype(str) == str(_sel)].copy()
                if not _g.empty:
                    dataframe_br(_g, use_container_width=True, hide_index=True, export_title=f"Resumo Gerente - {_sel}")
    else:
        st.info("Este perfil ainda não possui um resumo executivo individual configurado.")

elif visao == "Realizados":
    section("Realizados", "sec-gray")
    df = REALIZADOS if comprador == "Todos" else REALIZADOS[REALIZADOS["Comprador"] == comprador]
    dataframe_br(preparar_tabela(df), use_container_width=True, hide_index=True, height=270)

elif visao == "Métricas Destaque":
    section("Métricas Destaque", "sec-green")
    df = METAS if comprador == "Todos" else METAS[METAS["Comprador"] == comprador]
    dataframe_br(preparar_tabela(df), use_container_width=True, hide_index=True, height=270)

elif visao == "Resultado Métricas":
    section("Resultado Métricas - Realizado", "sec-blue")
    df = RESULTADO if comprador == "Todos" else RESULTADO[RESULTADO["Comprador"] == comprador]
    dataframe_br(preparar_tabela(df), use_container_width=True, hide_index=True, height=270)

elif visao == "Resultados dos KPI's":
    section("Resultados dos KPI's", "sec-blue")
    st.caption("Visão exclusiva do resultado de cada KPI: Meta × Realizado × Atingimento × Status.")

    nomes_kpi = COMPRADORES if comprador == "Todos" else [comprador]
    partes_kpi = []
    for nome_kpi in nomes_kpi:
        detalhe_kpi = _kpis_holerite_comprador(nome_kpi)
        if detalhe_kpi is None or detalhe_kpi.empty:
            continue
        detalhe_kpi = detalhe_kpi.copy()
        detalhe_kpi.insert(0, "Comprador", nome_kpi)
        partes_kpi.append(detalhe_kpi)

    if not partes_kpi:
        st.info("Ainda não existem resultados de KPI disponíveis para o filtro selecionado.")
    else:
        resultado_kpis = pd.concat(partes_kpi, ignore_index=True)
        resultado_kpis["Atingimento (%)"] = pd.to_numeric(resultado_kpis["Atingimento (%)"], errors="coerce").fillna(0.0)
        resultado_kpis["Status"] = resultado_kpis["Atingimento (%)"].map(lambda v: _status_atingimento_premiacao(v)[0])
        resultado_kpis["Status"] = resultado_kpis.apply(
            lambda r: ("✅ Atingida" if r["Atingimento (%)"] >= 100 else ("🟡 Atenção" if r["Atingimento (%)"] >= 90 else "🔴 Crítico")),
            axis=1,
        )

        total_kpis = int(len(resultado_kpis))
        qtd_atingidos = int((resultado_kpis["Atingimento (%)"] >= 100).sum())
        qtd_atencao = int(((resultado_kpis["Atingimento (%)"] >= 90) & (resultado_kpis["Atingimento (%)"] < 100)).sum())
        qtd_criticos = int((resultado_kpis["Atingimento (%)"] < 90).sum())
        media_ating = float(resultado_kpis["Atingimento (%)"].mean()) if total_kpis else 0.0

        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("KPI's avaliados", numero_inteiro(total_kpis))
        c2.metric("Atingidos", numero_inteiro(qtd_atingidos))
        c3.metric("Em atenção", numero_inteiro(qtd_atencao))
        c4.metric("Críticos", numero_inteiro(qtd_criticos))
        c5.metric("Atingimento médio", percentual(media_ating))

        exib_kpis = resultado_kpis[["Comprador", "KPI", "Meta", "Realizado", "Atingimento (%)", "Status"]].copy()

        def _fmt_valor_kpi(nome, valor):
            nome = str(nome)
            valor = float(pd.to_numeric(valor, errors="coerce") or 0.0)
            if nome in {"Faturamento", "CMV", "Estoque Curva A", "Estoque Curva B", "Estoque Curva C", "Estoque Curva D", "Ruptura Ativa"}:
                return moeda_real(valor)
            if nome == "Reposição CMV":
                return percentual(valor)
            if nome == "Fator Cobertura":
                return _numero_base(valor, 2)
            return _numero_base(valor, 2)

        exib_kpis["Meta"] = [
            _fmt_valor_kpi(kpi, valor) for kpi, valor in zip(exib_kpis["KPI"], exib_kpis["Meta"])
        ]
        exib_kpis["Realizado"] = [
            _fmt_valor_kpi(kpi, valor) for kpi, valor in zip(exib_kpis["KPI"], exib_kpis["Realizado"])
        ]
        exib_kpis["Atingimento (%)"] = exib_kpis["Atingimento (%)"].map(percentual)

        dataframe_br(
            exib_kpis,
            use_container_width=True,
            hide_index=True,
            height=min(620, 80 + 38 * max(1, len(exib_kpis))),
            export_title="Resultados dos KPI's",
        )
        st.caption("Status: ✅ Atingida = 100% ou mais | 🟡 Atenção = 90% a 99,99% | 🔴 Crítico = abaixo de 90%.")

elif visao == "Prêmio Comprador":
    _render_premium_header("🏆", "Premiação por Comprador", "Resultado consolidado por comprador e por indicador, com valores oficiais e percentuais de atingimento.")
    df = PREMIO if comprador == "Todos" else PREMIO[PREMIO["Comprador"] == comprador]
    cols_premio = [c for c in df.columns if str(c).endswith(" Prêmio")]
    total_conquistado = float(pd.to_numeric(df[cols_premio].stack(), errors="coerce").fillna(0).sum()) if cols_premio and not df.empty else 0.0
    premio_maximo = float(len(df) * float(METAS_GESTOR.get("valor_premio_total", 3000.0))) if not df.empty else 0.0
    ating_geral = total_conquistado / premio_maximo * 100 if premio_maximo else 0.0
    _render_premium_cards([
        {"icone":"🏆","titulo":"Prêmio Máximo","valor":moeda_real(premio_maximo),"subtitulo":"Potencial dos compradores filtrados","classe_icone":"hp-blue-icon"},
        {"icone":"🥇","titulo":"Prêmio Conquistado","valor":moeda_real(total_conquistado),"subtitulo":"Soma das parcelas por KPI","classe_icone":"hp-green-icon"},
        {"icone":"🎯","titulo":"Atingimento Geral","valor":percentual(ating_geral),"subtitulo":"Percentual sobre o potencial","classe_icone":"hp-purple-icon"},
        {"icone":"⬇","titulo":"Saldo não Conquistado","valor":moeda_real(max(0,premio_maximo-total_conquistado)),"subtitulo":"Valor que deixou de ganhar","classe_icone":"hp-red-icon","classe_valor":"hp-red"},
        {"icone":"👥","titulo":"Compradores","valor":numero_inteiro(len(df)),"subtitulo":"Participantes no filtro","classe_icone":"hp-gold-icon"},
    ])
    st.markdown('<div class="hp-section-title">Como a premiação foi conquistada ⓘ</div>', unsafe_allow_html=True)
    dataframe_br(
        preparar_tabela_premio_comprador(df),
        use_container_width=True,
        hide_index=True,
        height=280,
        export_title="Prêmio Comprador x Métrica",
    )

    st.caption("As colunas Prêmio representam valores em reais; as colunas Realizado representam o percentual de atingimento de cada KPI.")
    if _perfil_logado() != "Administrador":
        st.caption("🔒 Esta tela está limitada exclusivamente à sua própria premiação.")
        st.stop()

    st.markdown("### Gerente Comercial")
    gerente = pd.DataFrame([[
        "Gerente Comercial", 14.95, 99.7, 15.00, 100.0, 48.96, 97.9,
        44.10, 58.8, 73.56, 98.1, 44.60, 89.2, 39.00, 78.0,
        57.26, 57.3, 67.07, 95.8
    ]], columns=PREMIO.columns)
    dataframe_br(
        preparar_tabela_premio_comprador(gerente),
        use_container_width=True,
        hide_index=True,
        height=110,
        export_title="Premiação do Gerente Comercial",
    )

elif visao == "Prêmio por KPI":
    if _perfil_logado() == "Comprador":
        _nome_kpi_privado = _escopo_usuario_logado()
        _detalhe_kpi_privado = _kpis_holerite_comprador(_nome_kpi_privado)
        st.markdown("### 🔒 Minha Premiação por KPI")
        if _detalhe_kpi_privado.empty:
            st.info("Não há premiação por KPI calculada para seu usuário nesta competência.")
        else:
            _cols_priv = [
                c for c in [
                    "KPI", "Meta", "Realizado", "Atingimento (%)",
                    "Peso (%)", "Prêmio máximo", "Prêmio conquistado",
                    "Saldo não conquistado"
                ] if c in _detalhe_kpi_privado.columns
            ]
            dataframe_br(
                _detalhe_kpi_privado[_cols_priv],
                use_container_width=True,
                hide_index=True,
                export_title=f"Minha Premiação por KPI - {_nome_kpi_privado}",
            )
        st.caption("🔒 Nenhum KPI ou valor de outro comprador é exibido.")
        st.stop()
    elif _perfil_logado() in {"Gerente", "Vendedor"}:
        st.info("Esta visão consolidada não está disponível para este perfil.")
        st.stop()

    _render_premium_header("💰", "Premiação por KPI", "Composição do prêmio máximo, atingimento e valor conquistado em cada indicador.")
    premio_maximo_kpi = float(pd.to_numeric(PREMIO_KPI["Prêmio por KPI atingível"], errors="coerce").fillna(0).sum())
    premio_atingido_kpi = float(pd.to_numeric(PREMIO_KPI["Prêmio Atingido"], errors="coerce").fillna(0).sum())
    ating_kpi = premio_atingido_kpi / premio_maximo_kpi * 100 if premio_maximo_kpi else 0.0
    _render_premium_cards([
        {"icone":"🏆","titulo":"Prêmio Máximo","valor":moeda_real(premio_maximo_kpi),"subtitulo":"Soma dos nove KPIs","classe_icone":"hp-blue-icon"},
        {"icone":"🥇","titulo":"Prêmio Conquistado","valor":moeda_real(premio_atingido_kpi),"subtitulo":"Valor efetivamente atingido","classe_icone":"hp-green-icon"},
        {"icone":"🎯","titulo":"Atingimento Geral","valor":percentual(ating_kpi),"subtitulo":"Percentual sobre o máximo","classe_icone":"hp-purple-icon"},
        {"icone":"⬇","titulo":"Saldo não Conquistado","valor":moeda_real(max(0,premio_maximo_kpi-premio_atingido_kpi)),"subtitulo":"Oportunidade restante","classe_icone":"hp-red-icon","classe_valor":"hp-red"},
        {"icone":"📊","titulo":"Indicadores","valor":numero_inteiro(len(PREMIO_KPI)),"subtitulo":"KPIs avaliados","classe_icone":"hp-gold-icon"},
    ])
    st.markdown('<div class="hp-section-title">Como o prêmio foi conquistado ⓘ</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff">Valor total de prêmio atingível</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:3px">Composição ponderada pelos nove KPIs</div>
      </div>
      <div class="value">R$ 3.000,00</div>
    </div>
    """, unsafe_allow_html=True)

    dataframe_br(preparar_tabela(PREMIO_KPI), use_container_width=True, hide_index=True, height=390)

    c1,c2,c3 = st.columns(3)
    with c1: st.metric("Prêmio atingido","R$ 2.638,34")
    with c2: st.metric("Atingimento geral","87,9%")
    with c3: st.metric("Saldo não atingido","R$ 361,66")

    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=87.9,
        number={"suffix":"%","font":{"color":"#f3f7fb"}},
        title={"text":"Atingimento total do prêmio","font":{"color":"#f3f7fb"}},
        gauge={
            "axis":{"range":[0,100],"tickcolor":"#8da2b8"},
            "bar":{"color":"#22d3ee"},
            "bgcolor":"#0c1724",
            "bordercolor":"#1f3850",
            "steps":[
                {"range":[0,70],"color":"#4a2024"},
                {"range":[70,90],"color":"#5a491a"},
                {"range":[90,100],"color":"#16452d"},
            ]
        }
    ))
    fig.update_layout(height=330,margin=dict(l=20,r=20,t=60,b=20),paper_bgcolor="rgba(0,0,0,0)")
    plotly_chart_br(fig,use_container_width=True,config={"displayModeBar":False})



elif visao == "Portal de Premiação":
    _render_premium_header("🌟", "Portal de Premiação Comercial", "Visão executiva consolidada por comprador, loja, supervisor e gerente.")
    st.caption(
        "Visão consolidada do valor conquistado. O prêmio oficial é calculado por comprador; "
        "as visões por loja, supervisor e gerente são rateios explicativos pela participação das vendas."
    )
    ranking_comp = _premio_total_por_comprador()
    rateio_h = _hierarquia_rateio(PERIODO_REALIZADO_USADO)

    _perfil_portal = _perfil_logado()
    _escopo_portal = _escopo_usuario_logado()

    if _perfil_portal == "Comprador":
        if "Comprador" in ranking_comp.columns:
            ranking_comp = ranking_comp.loc[
                ranking_comp["Comprador"].map(_norm_escopo).eq(_norm_escopo(_escopo_portal))
            ].copy()
        else:
            ranking_comp = ranking_comp.iloc[0:0].copy()

        _premio_individual = (
            float(pd.to_numeric(
                ranking_comp.get("Prêmio Total", pd.Series(dtype=float)),
                errors="coerce"
            ).fillna(0).sum())
            if not ranking_comp.empty else 0.0
        )

        st.markdown("### 🔒 Minha Premiação")
        c1, c2 = st.columns(2)
        c1.metric("Comprador", _escopo_portal or "-")
        c2.metric("Prêmio conquistado", moeda_real(_premio_individual))

        if ranking_comp.empty:
            st.info("Não há premiação calculada para seu usuário nesta competência.")
        else:
            _minha = ranking_comp.copy()
            for _col in ["Posição", "Pos.", "Ranking", "Rank"]:
                if _col in _minha.columns:
                    _minha = _minha.drop(columns=[_col])
            dataframe_br(
                _minha,
                use_container_width=True,
                hide_index=True,
                export_title=f"Minha Premiação - {_escopo_portal}",
            )
        st.caption("🔒 Esta área exibe exclusivamente sua própria premiação.")
        st.stop()

    if _perfil_portal == "Gerente":
        if not rateio_h.empty and "Gerente" in rateio_h.columns:
            _meu_rateio = rateio_h.loc[
                rateio_h["Gerente"].map(_norm_escopo).eq(_norm_escopo(_escopo_portal))
            ].copy()
        else:
            _meu_rateio = pd.DataFrame()

        _col_premio = next(
            (c for c in ["Prêmio rateado", "Prêmio Rateado", "Prêmio Total", "Premio rateado"]
             if c in _meu_rateio.columns),
            None,
        )
        _premio_gerente = (
            float(pd.to_numeric(_meu_rateio[_col_premio], errors="coerce").fillna(0).sum())
            if _col_premio and not _meu_rateio.empty else 0.0
        )

        st.markdown("### 🔒 Minha Premiação Gerencial")
        c1, c2 = st.columns(2)
        c1.metric("Gerente", _escopo_portal or "-")
        c2.metric("Prêmio conquistado", moeda_real(_premio_gerente))
        st.caption(
            "🔒 Não são exibidos compradores, outros gerentes, rankings ou valores de terceiros."
        )
        st.stop()

    if _perfil_portal == "Vendedor":
        st.markdown("### 🔒 Minha Premiação")
        st.info(
            "O Portal consolidado não exibe dados de terceiros para vendedores. "
            "Use o holerite individual quando houver premiação vinculada."
        )
        st.stop()
    total_premio = float(ranking_comp["Prêmio Total"].sum()) if not ranking_comp.empty else 0.0
    lojas_qtd = int(rateio_h["Loja"].nunique()) if not rateio_h.empty else 0
    supervisores_qtd = int(rateio_h.loc[rateio_h["Supervisor"] != "Não cadastrado", "Supervisor"].nunique()) if not rateio_h.empty else 0
    gerentes_qtd = int(rateio_h.loc[rateio_h["Gerente"] != "Não cadastrado", "Gerente"].nunique()) if not rateio_h.empty else 0
    _render_premium_cards([
        {"icone":"💰","titulo":"Premiação Total","valor":moeda_real(total_premio),"subtitulo":"Valor conquistado no período","classe_icone":"hp-green-icon"},
        {"icone":"👥","titulo":"Compradores Premiados","valor":numero_inteiro((ranking_comp["Prêmio Total"] > 0).sum() if not ranking_comp.empty else 0),"subtitulo":"Com valor conquistado","classe_icone":"hp-blue-icon"},
        {"icone":"🏬","titulo":"Lojas Analisadas","valor":numero_inteiro(lojas_qtd),"subtitulo":"Com participação apurada","classe_icone":"hp-purple-icon"},
        {"icone":"🧑‍💼","titulo":"Supervisores","valor":numero_inteiro(supervisores_qtd),"subtitulo":"Na hierarquia cadastrada","classe_icone":"hp-gold-icon"},
        {"icone":"👔","titulo":"Gerentes","valor":numero_inteiro(gerentes_qtd),"subtitulo":"Na hierarquia cadastrada","classe_icone":"hp-red-icon"},
    ])
    st.markdown('<div class="hp-section-title">Detalhamento da premiação ⓘ</div>', unsafe_allow_html=True)

    aba_c, aba_l, aba_s, aba_g = st.tabs(["Compradores", "Lojas", "Supervisores", "Gerentes"])
    with aba_c:
        if ranking_comp.empty:
            st.info("Sem premiação calculada para o período.")
        else:
            rank = ranking_comp.copy()
            rank.insert(0, "Posição", range(1, len(rank) + 1))
            dataframe_br(rank, use_container_width=True, hide_index=True, export_title="Ranking de Compradores")
            fig = px.bar(rank.sort_values("Prêmio Total"), x="Prêmio Total", y="Comprador", orientation="h", title="Premiação por comprador")
            plotly_chart_br(fig, use_container_width=True, tipo="moeda")
    with aba_l:
        if rateio_h.empty:
            st.info("Sem vendas por loja disponíveis para realizar o rateio explicativo.")
        else:
            loja = rateio_h.groupby("Loja", as_index=False).agg(Venda=("Venda", "sum"), **{"Prêmio rateado": ("Prêmio rateado", "sum")}).sort_values("Prêmio rateado", ascending=False)
            dataframe_br(loja, use_container_width=True, hide_index=True, export_title="Premiação por Loja")
    with aba_s:
        if rateio_h.empty:
            st.info("Sem dados de hierarquia disponíveis.")
        else:
            sup = rateio_h.groupby("Supervisor", as_index=False).agg(Lojas=("Loja", "nunique"), Compradores=("Comprador", "nunique"), Venda=("Venda", "sum"), **{"Prêmio rateado": ("Prêmio rateado", "sum")}).sort_values("Prêmio rateado", ascending=False)
            dataframe_br(sup, use_container_width=True, hide_index=True, export_title="Premiação por Supervisor")
    with aba_g:
        if rateio_h.empty:
            st.info("Sem dados de hierarquia disponíveis.")
        else:
            ger = rateio_h.groupby("Gerente", as_index=False).agg(Lojas=("Loja", "nunique"), Supervisores=("Supervisor", "nunique"), Compradores=("Comprador", "nunique"), Venda=("Venda", "sum"), **{"Prêmio rateado": ("Prêmio rateado", "sum")}).sort_values("Prêmio rateado", ascending=False)
            dataframe_br(ger, use_container_width=True, hide_index=True, export_title="Premiação por Gerente")

elif visao == "Holerite da Premiação":
    compradores_holerite = _premio_total_por_comprador()["Comprador"].tolist()
    if _perfil_logado() == "Comprador":
        _meu_comprador_holerite = _escopo_usuario_logado()
        compradores_holerite = [
            x for x in compradores_holerite
            if _norm_escopo(x) == _norm_escopo(_meu_comprador_holerite)
        ]
    elif _perfil_logado() != "Administrador":
        compradores_holerite = []
    if not compradores_holerite:
        section("Holerite Analítico da Premiação", "sec-gold")
        st.info("Sem compradores com premiação calculada.")
    else:
        st.markdown(
            f'<div class="hp-title-wrap"><div><div class="hp-title">🧾 Holerite Analítico da Premiação</div>'
            f'<div class="hp-subtitle">Detalhamento completo da premiação do comprador conforme as metas e pesos cadastrados</div></div>'
            f'<div class="hp-update">Atualizado em: {datetime.now():%d/%m/%Y %H:%M:%S} ⟳</div></div>',
            unsafe_allow_html=True,
        )

        padrao = comprador if comprador in compradores_holerite else compradores_holerite[0]
        f_comp, f_periodo = st.columns([1.35, 1])
        if _perfil_logado() == "Comprador":
            nome = compradores_holerite[0]
            f_comp.text_input(
                "Comprador",
                value=nome,
                disabled=True,
                key="comprador_holerite_privado",
            )
        else:
            nome = f_comp.selectbox(
                "Comprador",
                compradores_holerite,
                index=compradores_holerite.index(padrao),
                key="comprador_holerite",
            )
        f_periodo.text_input(
            "Competência",
            value=str(PERIODO_REALIZADO_USADO),
            disabled=True,
            key="hp_competencia_comprador",
        )

        detalhe = _kpis_holerite_comprador(nome)
        total = float(detalhe["Prêmio conquistado"].sum()) if not detalhe.empty else 0.0
        maximo = float(detalhe["Prêmio máximo"].sum()) if not detalhe.empty else 0.0
        saldo = max(0.0, maximo - total)
        geral = total / maximo * 100 if maximo else 0.0
        kpis_atingidos = int((pd.to_numeric(detalhe.get("Atingimento (%)", pd.Series(dtype=float)), errors="coerce").fillna(0) >= 100).sum()) if not detalhe.empty else 0

        cards = st.columns(5)
        cards[0].markdown(
            _html_card_premiacao("🏆", "Prêmio Máximo", moeda_real(maximo), "100,00% do potencial", "hp-blue-icon"),
            unsafe_allow_html=True,
        )
        cards[1].markdown(
            _html_card_premiacao("🥇", "Prêmio Conquistado", moeda_real(total), "Valor obtido pelo comprador", "hp-green-icon"),
            unsafe_allow_html=True,
        )
        cards[2].markdown(
            _html_card_premiacao("🎯", "Atingimento Geral", percentual(geral), "Percentual sobre o total", "hp-purple-icon"),
            unsafe_allow_html=True,
        )
        cards[3].markdown(
            _html_card_premiacao("⬇", "Saldo não Conquistado", moeda_real(saldo), "Valor que deixou de ganhar", "hp-red-icon", "hp-red"),
            unsafe_allow_html=True,
        )
        cards[4].markdown(
            _html_card_premiacao("✅", "KPIs em 100%", numero_inteiro(kpis_atingidos), f"de {numero_inteiro(len(detalhe))} indicadores", "hp-gold-icon"),
            unsafe_allow_html=True,
        )

        kpis_comprador = []
        tipos_percentuais = {"Reposição CMV"}
        tipos_decimais = {"Fator Cobertura"}
        for _, item in detalhe.iterrows():
            kpi = str(item.get("KPI", ""))
            if kpi in tipos_percentuais:
                tipo = "Percentual (%)"
            elif kpi in tipos_decimais:
                tipo = "Número decimal"
            else:
                tipo = "Valor (R$)"
            kpis_comprador.append({
                "KPI": kpi,
                "Tipo": tipo,
                "Meta": float(item.get("Meta", 0) or 0),
                "Realizado": float(item.get("Realizado", 0) or 0),
                "Atingimento (%)": float(item.get("Atingimento (%)", 0) or 0),
                "Peso (%)": float(item.get("Peso (%)", 0) or 0),
                "Parcela máxima (R$)": float(item.get("Prêmio máximo", 0) or 0),
                "Parcela conquistada (R$)": float(item.get("Prêmio conquistado", 0) or 0),
                "Valor perdido (R$)": float(item.get("Saldo não conquistado", 0) or 0),
            })

        st.markdown('<div class="hp-section-title">Como o prêmio foi conquistado ⓘ</div>', unsafe_allow_html=True)
        st.caption("Cada parcela corresponde ao prêmio máximo do KPI multiplicado pelo atingimento, limitado a 100%.")
        st.markdown(
            _html_tabela_holerite(kpis_comprador, maximo, total, saldo, geral),
            unsafe_allow_html=True,
        )
        registrar_tabela_exportacao(pd.DataFrame(kpis_comprador), f"Holerite Comprador - {nome}")

        rateio = _hierarquia_rateio(PERIODO_REALIZADO_USADO)
        rateio = rateio[rateio["Comprador"] == nome].copy() if not rateio.empty else rateio
        c_resumo, c_lojas, c_kpis = st.columns([0.9, 1.15, 1.35])

        with c_resumo:
            st.markdown('<div class="hp-panel"><div class="hp-panel-title">📈 Resumo Executivo</div>', unsafe_allow_html=True)
            fig_resumo = go.Figure(go.Pie(
                values=[total, max(saldo, 0)],
                labels=["Prêmio conquistado", "Saldo não conquistado"],
                hole=.70,
                marker=dict(colors=["#20c66b", "#ff4d4f"]),
                textinfo="none",
            ))
            fig_resumo.update_layout(
                height=255,
                margin=dict(l=5, r=5, t=5, b=5),
                showlegend=True,
                legend=dict(orientation="h", y=-.08),
                annotations=[dict(
                    text=f"<b>{percentual(geral)}</b><br><span style='font-size:11px'>Atingimento Geral</span>",
                    x=.5, y=.5, showarrow=False, font=dict(color="white", size=15),
                )],
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
            )
            st.plotly_chart(fig_resumo, use_container_width=True, config={"displayModeBar": False})
            st.markdown('</div>', unsafe_allow_html=True)

        with c_lojas:
            st.markdown('<div class="hp-panel"><div class="hp-panel-title">🏬 Contribuição Analítica por Loja</div>', unsafe_allow_html=True)
            if rateio.empty:
                st.info("Não há detalhamento de vendas por loja para este comprador.")
            else:
                lojas = rateio.groupby(["Loja", "Supervisor", "Gerente"], as_index=False).agg(
                    **{"Venda (R$)": ("Venda", "sum"), "Prêmio Rateado (R$)": ("Prêmio rateado", "sum")}
                )
                total_vendas = float(lojas["Venda (R$)"].sum())
                lojas["Participação (%)"] = np.where(total_vendas > 0, lojas["Venda (R$)"] / total_vendas * 100, 0)
                dataframe_br(
                    lojas.sort_values("Prêmio Rateado (R$)", ascending=False),
                    use_container_width=True,
                    hide_index=True,
                    height=260,
                    export_title=f"Lojas Comprador - {nome}",
                )
            st.markdown('</div>', unsafe_allow_html=True)

        with c_kpis:
            st.markdown('<div class="hp-panel"><div class="hp-panel-title">📊 Composição do Prêmio por KPI</div>', unsafe_allow_html=True)
            if detalhe.empty:
                st.info("Não há indicadores calculados para este comprador.")
            else:
                graf = detalhe[["KPI", "Prêmio conquistado", "Saldo não conquistado"]].copy()
                fig_kpi = go.Figure()
                fig_kpi.add_trace(go.Bar(
                    y=graf["KPI"], x=graf["Prêmio conquistado"], orientation="h",
                    name="Conquistado", marker_color="#20c66b",
                ))
                fig_kpi.add_trace(go.Bar(
                    y=graf["KPI"], x=graf["Saldo não conquistado"], orientation="h",
                    name="Saldo", marker_color="#ff4d4f",
                ))
                fig_kpi.update_layout(
                    barmode="stack", height=270, margin=dict(l=5, r=5, t=10, b=5),
                    xaxis_title="R$", yaxis_title="", paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)", legend=dict(orientation="h", y=1.12),
                )
                plotly_chart_br(fig_kpi, use_container_width=True, tipo="moeda")
            st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="hp-section-title">🏆 Ranking de KPIs do Comprador</div>', unsafe_allow_html=True)
        if not detalhe.empty:
            ranking_kpi = detalhe[["KPI", "Peso (%)", "Atingimento (%)", "Prêmio máximo", "Prêmio conquistado", "Saldo não conquistado"]].copy()
            ranking_kpi = ranking_kpi.sort_values("Prêmio conquistado", ascending=False).reset_index(drop=True)
            ranking_kpi.insert(0, "Pos.", np.arange(1, len(ranking_kpi) + 1))
            ranking_kpi.columns = ["Pos.", "KPI", "Peso (%)", "Atingimento (%)", "Prêmio Máximo (R$)", "Prêmio Conquistado (R$)", "Saldo (R$)"]
            ranking_kpi["Status"] = ranking_kpi["Atingimento (%)"].map(lambda v: _status_atingimento_premiacao(v)[0])
            dataframe_br(
                ranking_kpi,
                use_container_width=True,
                hide_index=True,
                height=min(430, 88 + 36 * len(ranking_kpi)),
                export_title=f"Ranking KPIs - {nome}",
            )

        st.markdown('<div class="hp-footnote">ⓘ O prêmio oficial é calculado por KPI, conforme pesos e metas cadastrados. O detalhamento por loja é apenas explicativo e não altera o valor oficial.</div>', unsafe_allow_html=True)


elif visao == "Holerite da Loja":
    hol_lojas = _holerite_lojas_por_meta(PERIODO_REALIZADO_USADO)
    if hol_lojas.empty:
        section("Holerite Analítico da Loja", "sec-gold")
        st.info("Não existem metas de loja e realizado suficientes para montar o holerite.")
    else:
        st.markdown(f'<div class="hp-title-wrap"><div><div class="hp-title">🏪 Holerite Analítico da Loja</div><div class="hp-subtitle">Detalhamento completo da premiação da loja conforme as metas cadastradas</div></div><div class="hp-update">Atualizado em: {datetime.now():%d/%m/%Y %H:%M:%S} ⟳</div></div>', unsafe_allow_html=True)
        f_loja, f_sup, f_ger, f_comp = st.columns([1.15, 1, 1, 1])
        loja_nome = f_loja.selectbox("Loja", hol_lojas["Filial"].astype(str).tolist(), key="holerite_loja_selecionada")
        linha = hol_lojas.loc[hol_lojas["Filial"].astype(str) == str(loja_nome)].iloc[0]
        f_sup.text_input("Supervisor", value=str(linha["Supervisor"]), disabled=True, key="hp_supervisor")
        f_ger.text_input("Gerente Comercial", value=str(linha["Gerente Comercial"]), disabled=True, key="hp_gerente")
        f_comp.text_input("Competência", value=str(PERIODO_REALIZADO_USADO), disabled=True, key="hp_competencia")
        premio_max = float(linha["Prêmio máximo da loja"]); premio_conq = float(linha["Prêmio conquistado pelas metas"]); ating_geral = float(linha["Atingimento Geral (%)"]); saldo = float(linha["Saldo não conquistado"])
        cards = st.columns(5)
        cards[0].markdown(_html_card_premiacao("🏆", "Prêmio Máximo", moeda_real(premio_max), "100,00% do potencial", "hp-blue-icon"), unsafe_allow_html=True)
        cards[1].markdown(_html_card_premiacao("🥇", "Prêmio Conquistado", moeda_real(premio_conq), "Valor obtido pela loja", "hp-green-icon"), unsafe_allow_html=True)
        cards[2].markdown(_html_card_premiacao("🎯", "Atingimento Geral", percentual(ating_geral), "Percentual sobre o total", "hp-purple-icon"), unsafe_allow_html=True)
        cards[3].markdown(_html_card_premiacao("⬇", "Saldo não Conquistado", moeda_real(saldo), "Valor que deixou de ganhar", "hp-red-icon", "hp-red"), unsafe_allow_html=True)
        cards[4].markdown(_html_card_premiacao("💰", "Meta Máxima da Loja", moeda_real(premio_max), "Valor total da premiação", "hp-gold-icon"), unsafe_allow_html=True)
        kpis_loja = [
            {"KPI":"Faturamento","Tipo":"Valor (R$)","Meta":float(linha["Faturamento Total META"]),"Realizado":float(linha["Faturamento Total Atual"]),"Atingimento (%)":float(linha["Atingimento Faturamento (%)"]),"Peso (%)":50.0,"Parcela máxima (R$)":premio_max*0.5,"Parcela conquistada (R$)":premio_max*0.5*min(max(float(linha["Atingimento Faturamento (%)"]),0),100)/100},
            {"KPI":"Margem Bruta","Tipo":"Valor (R$)","Meta":float(linha["Margem Bruta META"]),"Realizado":float(linha["Margem Bruta Atual"]),"Atingimento (%)":float(linha["Atingimento Margem Bruta (%)"]),"Peso (%)":50.0,"Parcela máxima (R$)":premio_max*0.5,"Parcela conquistada (R$)":premio_max*0.5*min(max(float(linha["Atingimento Margem Bruta (%)"]),0),100)/100},
        ]
        for item in kpis_loja: item["Valor perdido (R$)"] = item["Parcela máxima (R$)"] - item["Parcela conquistada (R$)"]
        st.markdown('<div class="hp-section-title">Como o prêmio foi conquistado ⓘ</div>', unsafe_allow_html=True)
        st.markdown(_html_tabela_holerite(kpis_loja, premio_max, premio_conq, saldo, ating_geral), unsafe_allow_html=True)
        registrar_tabela_exportacao(pd.DataFrame(kpis_loja), f"Holerite Loja - {loja_nome}")
        rateio_loja = _hierarquia_rateio(PERIODO_REALIZADO_USADO)
        rateio_loja = rateio_loja[rateio_loja["Loja"].map(_chave_loja_premiacao) == _chave_loja_premiacao(loja_nome)] if not rateio_loja.empty else rateio_loja
        c_resumo, c_compradores, c_evolucao = st.columns([0.9, 1.15, 1.35])
        with c_resumo:
            st.markdown('<div class="hp-panel"><div class="hp-panel-title">📈 Resumo Executivo</div>', unsafe_allow_html=True)
            fig_resumo = go.Figure(go.Pie(values=[premio_conq, max(saldo,0)], labels=["Prêmio conquistado","Saldo não conquistado"], hole=.70, marker=dict(colors=["#20c66b","#ff4d4f"]), textinfo="none"))
            fig_resumo.update_layout(height=255, margin=dict(l=5,r=5,t=5,b=5), showlegend=True, legend=dict(orientation="h", y=-.08), annotations=[dict(text=f"<b>{percentual(ating_geral)}</b><br><span style='font-size:11px'>Atingimento Geral</span>", x=.5,y=.5,showarrow=False,font=dict(color="white",size=15))], paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig_resumo, use_container_width=True, config={"displayModeBar":False})
            st.markdown('</div>', unsafe_allow_html=True)
        with c_compradores:
            st.markdown('<div class="hp-panel"><div class="hp-panel-title">👥 Participação dos Compradores na Loja</div>', unsafe_allow_html=True)
            if rateio_loja.empty: st.info("Não há detalhamento de compradores para esta loja.")
            else:
                comp = rateio_loja.groupby("Comprador", as_index=False).agg(**{"Venda (R$)":("Venda","sum"),"Prêmio Oficial Rateado (R$)":("Prêmio rateado","sum")})
                total_venda = float(comp["Venda (R$)"].sum()); comp["Participação na Venda (%)"] = np.where(total_venda>0,comp["Venda (R$)"]/total_venda*100,0)
                comp["% do Prêmio"] = np.where(comp["Prêmio Oficial Rateado (R$)"].sum()>0,comp["Prêmio Oficial Rateado (R$)"]/comp["Prêmio Oficial Rateado (R$)"].sum()*100,0)
                dataframe_br(comp.sort_values("Prêmio Oficial Rateado (R$)",ascending=False),use_container_width=True,hide_index=True,height=260,export_title=f"Compradores Loja - {loja_nome}")
            st.markdown('</div>', unsafe_allow_html=True)
        with c_evolucao:
            st.markdown('<div class="hp-panel"><div class="hp-panel-title">⏱️ Evolução do Atingimento no Mês</div>', unsafe_allow_html=True)
            evol = _evolucao_diaria_holerite_loja(PERIODO_REALIZADO_USADO, loja_nome, linha["Faturamento Total META"])
            if evol.empty: st.info("A base atual não possui data diária por loja para esta evolução.")
            else:
                fig_ev = go.Figure(); fig_ev.add_trace(go.Scatter(x=evol["dia"].dt.day,y=evol["Atingimento Real (%)"],mode="lines+markers",name="Atingimento Real (%)",line=dict(width=3,color="#348ce7"),marker=dict(size=5))); fig_ev.add_trace(go.Scatter(x=evol["dia"].dt.day,y=evol["Meta Projetada (%)"],mode="lines",name="Meta Projetada (%)",line=dict(width=2,dash="dash",color="#20c66b")))
                fig_ev.update_layout(height=270,margin=dict(l=5,r=5,t=10,b=5),xaxis_title="Dia",yaxis_title="%",yaxis=dict(ticksuffix="%"),paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",legend=dict(orientation="h",y=1.12),hovermode="x unified")
                st.plotly_chart(fig_ev,use_container_width=True,config={"displayModeBar":False})
            st.markdown('</div>', unsafe_allow_html=True)
        st.markdown('<div class="hp-section-title">🏆 Ranking de Lojas do Supervisor</div>', unsafe_allow_html=True)
        ranking = hol_lojas[hol_lojas["Supervisor"].astype(str) == str(linha["Supervisor"])].copy()
        if ranking.empty: ranking = hol_lojas.copy()
        ranking = ranking.sort_values("Prêmio conquistado pelas metas",ascending=False).reset_index(drop=True); ranking.insert(0,"Pos.",np.arange(1,len(ranking)+1))
        ranking_exib = ranking[["Pos.","Filial","Faturamento Total META","Faturamento Total Atual","Atingimento Geral (%)","Prêmio máximo da loja","Prêmio conquistado pelas metas","Saldo não conquistado"]].copy(); ranking_exib.columns=["Pos.","Loja","Meta (R$)","Realizado (R$)","Atingimento (%)","Prêmio Máximo (R$)","Prêmio Conquistado (R$)","Saldo (R$)"]; ranking_exib["Status"] = ranking_exib["Atingimento (%)"].map(lambda v:_status_atingimento_premiacao(v)[0])
        dataframe_br(ranking_exib,use_container_width=True,hide_index=True,height=min(430,88+36*len(ranking_exib)),export_title=f"Ranking Supervisor - {linha['Supervisor']}")
        st.markdown('<div class="hp-footnote">ⓘ Regras da premiação: 50% Faturamento + 50% Margem Bruta. Limitado a 100% por indicador.</div>', unsafe_allow_html=True)

elif visao == "Holerite do Gerente Comercial":
    hol_lojas = _holerite_lojas_por_meta(PERIODO_REALIZADO_USADO); hol_ger = _holerite_gerentes_por_meta(PERIODO_REALIZADO_USADO)
    if _perfil_logado() == "Gerente":
        _gerente_privado = _escopo_usuario_logado()
        hol_ger = _filtrar_gerente_privado(hol_ger, "Gerente Comercial")
        hol_lojas = _filtrar_gerente_privado(hol_lojas, "Gerente Comercial")
    elif _perfil_logado() != "Administrador":
        hol_ger = pd.DataFrame()
        hol_lojas = pd.DataFrame()
    st.markdown(f'<div class="hp-title-wrap"><div><div class="hp-title">👔 Holerite do Gerente Comercial</div><div class="hp-subtitle">Consolidação das metas e premiações das lojas sob sua responsabilidade</div></div><div class="hp-update">Atualizado em: {datetime.now():%d/%m/%Y %H:%M:%S} ⟳</div></div>',unsafe_allow_html=True)
    if hol_ger.empty: st.info("Não existem metas e realizado suficientes para montar o holerite gerencial.")
    else:
        f1,f2=st.columns([1.2,1])
        if _perfil_logado() == "Gerente":
            gerente_nome = str(hol_ger["Gerente Comercial"].iloc[0])
            f1.text_input("Gerente Comercial", value=gerente_nome, disabled=True, key="holerite_gerente_privado")
        else:
            gerente_nome=f1.selectbox("Gerente Comercial",hol_ger["Gerente Comercial"].astype(str).tolist(),key="holerite_gerente_selecionado")
        f2.text_input("Competência",value=str(PERIODO_REALIZADO_USADO),disabled=True,key="hp_comp_ger")
        g=hol_ger.loc[hol_ger["Gerente Comercial"].astype(str)==str(gerente_nome)].iloc[0]; ating_fat=g["Realizado Faturamento (R$)"]/g["Meta Faturamento (R$)"]*100 if g["Meta Faturamento (R$)"] else 0.0; ating_mb=g["Realizado Margem Bruta (R$)"]/g["Meta Margem Bruta (R$)"]*100 if g["Meta Margem Bruta (R$)"] else 0.0; ating_geral=g["Prêmio conquistado (R$)"]/g["Prêmio máximo (R$)"]*100 if g["Prêmio máximo (R$)"] else 0.0
        cards=st.columns(5); cards[0].markdown(_html_card_premiacao("🏆","Prêmio Máximo",moeda_real(g["Prêmio máximo (R$)"]),f"{numero_inteiro(g['Lojas'])} lojas", "hp-blue-icon"),unsafe_allow_html=True); cards[1].markdown(_html_card_premiacao("🥇","Prêmio Conquistado",moeda_real(g["Prêmio conquistado (R$)"]),"Valor consolidado", "hp-green-icon"),unsafe_allow_html=True); cards[2].markdown(_html_card_premiacao("🎯","Atingimento Geral",percentual(ating_geral),"Percentual sobre o total", "hp-purple-icon"),unsafe_allow_html=True); cards[3].markdown(_html_card_premiacao("⬇","Saldo não Conquistado",moeda_real(g["Saldo não conquistado (R$)"]),"Valor que deixou de ganhar", "hp-red-icon","hp-red"),unsafe_allow_html=True); cards[4].markdown(_html_card_premiacao("🏬","Lojas Gerenciadas",numero_inteiro(g["Lojas"]),f"{numero_inteiro(g['Supervisores'])} supervisores", "hp-gold-icon"),unsafe_allow_html=True)
        kpis_ger=[{"KPI":"Faturamento","Tipo":"Valor (R$)","Meta":float(g["Meta Faturamento (R$)"]),"Realizado":float(g["Realizado Faturamento (R$)"]),"Atingimento (%)":ating_fat,"Peso (%)":50.0,"Parcela máxima (R$)":float(g["Prêmio máximo (R$)"])*.5,"Parcela conquistada (R$)":float(g["Prêmio máximo (R$)"])*.5*min(max(ating_fat,0),100)/100},{"KPI":"Margem Bruta","Tipo":"Valor (R$)","Meta":float(g["Meta Margem Bruta (R$)"]),"Realizado":float(g["Realizado Margem Bruta (R$)"]),"Atingimento (%)":ating_mb,"Peso (%)":50.0,"Parcela máxima (R$)":float(g["Prêmio máximo (R$)"])*.5,"Parcela conquistada (R$)":float(g["Prêmio máximo (R$)"])*.5*min(max(ating_mb,0),100)/100}]
        for item in kpis_ger:item["Valor perdido (R$)"]=item["Parcela máxima (R$)"]-item["Parcela conquistada (R$)"]
        st.markdown('<div class="hp-section-title">Como o prêmio gerencial foi conquistado ⓘ</div>',unsafe_allow_html=True); st.markdown(_html_tabela_holerite(kpis_ger,float(g["Prêmio máximo (R$)"]),float(g["Prêmio conquistado (R$)"]),float(g["Saldo não conquistado (R$)"]),ating_geral),unsafe_allow_html=True); registrar_tabela_exportacao(pd.DataFrame(kpis_ger),f"Holerite Gerente - {gerente_nome}")
        lojas_g=hol_lojas[hol_lojas["Gerente Comercial"].astype(str)==str(gerente_nome)].copy(); st.markdown('<div class="hp-section-title">🏬 Resultado por Loja</div>',unsafe_allow_html=True)
        lojas_view=lojas_g[["Filial","Supervisor","Faturamento Total META","Faturamento Total Atual","Atingimento Geral (%)","Prêmio máximo da loja","Prêmio conquistado pelas metas","Saldo não conquistado"]].copy(); lojas_view.columns=["Loja","Supervisor","Meta (R$)","Realizado (R$)","Atingimento (%)","Prêmio Máximo (R$)","Prêmio Conquistado (R$)","Saldo (R$)"]; lojas_view["Status"]=lojas_view["Atingimento (%)"].map(lambda v:_status_atingimento_premiacao(v)[0]); dataframe_br(lojas_view.sort_values("Prêmio Conquistado (R$)",ascending=False),use_container_width=True,hide_index=True,height=390,export_title=f"Lojas Gerente - {gerente_nome}")
        st.markdown('<div class="hp-section-title">👥 Resultado por Supervisor</div>',unsafe_allow_html=True)
        sup_g=lojas_g.groupby("Supervisor",as_index=False).agg(Lojas=("Filial","nunique"),**{"Meta Faturamento (R$)":("Faturamento Total META","sum"),"Realizado Faturamento (R$)":("Faturamento Total Atual","sum"),"Prêmio máximo (R$)":("Prêmio máximo da loja","sum"),"Prêmio conquistado (R$)":("Prêmio conquistado pelas metas","sum"),"Saldo (R$)":("Saldo não conquistado","sum")}); sup_g["Atingimento do prêmio (%)"]=np.where(sup_g["Prêmio máximo (R$)"]>0,sup_g["Prêmio conquistado (R$)"]/sup_g["Prêmio máximo (R$)"]*100,0); sup_g["Status"]=sup_g["Atingimento do prêmio (%)"].map(lambda v:_status_atingimento_premiacao(v)[0]); dataframe_br(sup_g.sort_values("Prêmio conquistado (R$)",ascending=False),use_container_width=True,hide_index=True,export_title=f"Supervisores Gerente - {gerente_nome}")
        st.markdown('<div class="hp-footnote">ⓘ Regras da premiação: 50% Faturamento + 50% Margem Bruta. Limitado a 100% por indicador.</div>',unsafe_allow_html=True)

elif visao == "Premiação por Loja":
    _render_premium_header("🏬", "Premiação Analítica por Loja", "Contribuição das lojas na formação do prêmio oficial dos compradores.")
    st.warning("Esta visão não altera o prêmio oficial. Ela distribui o prêmio de cada comprador proporcionalmente às vendas de cada loja para explicar a contribuição operacional.")
    rateio = _hierarquia_rateio(PERIODO_REALIZADO_USADO)
    if rateio.empty:
        st.info("A base de Vendas não possui as colunas necessárias de loja e classificação para o rateio.")
    else:
        f1, f2, f3 = st.columns(3)
        lojas = sorted(rateio["Loja"].astype(str).unique())
        comps = sorted(rateio["Comprador"].astype(str).unique())
        sups = sorted(rateio["Supervisor"].astype(str).unique())
        loja_f = f1.selectbox("Loja", ["Todas"] + lojas, key="premio_loja_filtro")
        comp_f = f2.selectbox("Comprador", ["Todos"] + comps, key="premio_loja_comprador")
        sup_f = f3.selectbox("Supervisor", ["Todos"] + sups, key="premio_loja_supervisor")
        exib = rateio.copy()
        if loja_f != "Todas": exib = exib[exib["Loja"] == loja_f]
        if comp_f != "Todos": exib = exib[exib["Comprador"] == comp_f]
        if sup_f != "Todos": exib = exib[exib["Supervisor"] == sup_f]
        _render_premium_cards([
            {"icone":"💵","titulo":"Venda Considerada","valor":moeda_real(exib["Venda"].sum()),"subtitulo":"Base do rateio explicativo","classe_icone":"hp-blue-icon"},
            {"icone":"🥇","titulo":"Prêmio Rateado","valor":moeda_real(exib["Prêmio rateado"].sum()),"subtitulo":"Contribuição proporcional","classe_icone":"hp-green-icon"},
            {"icone":"👥","titulo":"Compradores Envolvidos","valor":numero_inteiro(exib["Comprador"].nunique()),"subtitulo":"Participantes no filtro","classe_icone":"hp-purple-icon"},
            {"icone":"🏬","titulo":"Lojas","valor":numero_inteiro(exib["Loja"].nunique()),"subtitulo":"Unidades consideradas","classe_icone":"hp-gold-icon"},
        ])
        st.markdown('<div class="hp-section-title">Detalhamento por loja ⓘ</div>', unsafe_allow_html=True)
        dataframe_br(exib, use_container_width=True, hide_index=True, height=430, export_title="Detalhamento da Premiação por Loja")

elif visao == "Premiação por Supervisor e Gerente":
    _render_premium_header("👥", "Premiação por Supervisor e Gerente", "Consolidação hierárquica da contribuição das lojas e compradores.")
    st.caption("Consolidação hierárquica do rateio explicativo por loja. O prêmio oficial continua pertencendo ao comprador.")
    rateio = _hierarquia_rateio(PERIODO_REALIZADO_USADO)
    aba_sup, aba_ger, aba_cad = st.tabs(["Supervisores", "Gerentes", "Configurar hierarquia"])
    with aba_sup:
        if rateio.empty:
            st.info("Sem dados para consolidar.")
        else:
            sup = rateio.groupby("Supervisor", as_index=False).agg(Lojas=("Loja", "nunique"), Compradores=("Comprador", "nunique"), Venda=("Venda", "sum"), **{"Prêmio rateado": ("Prêmio rateado", "sum")}).sort_values("Prêmio rateado", ascending=False)
            dataframe_br(sup, use_container_width=True, hide_index=True, export_title="Resumo por Supervisor")
            nome_sup = st.selectbox("Detalhar supervisor", sup["Supervisor"].tolist(), key="detalhe_supervisor")
            det = rateio[rateio["Supervisor"] == nome_sup]
            dataframe_br(det[["Loja", "Gerente", "Comprador", "Venda", "Prêmio rateado"]], use_container_width=True, hide_index=True, export_title=f"Supervisor - {nome_sup}")
    with aba_ger:
        if rateio.empty:
            st.info("Sem dados para consolidar.")
        else:
            ger = rateio.groupby("Gerente", as_index=False).agg(Lojas=("Loja", "nunique"), Supervisores=("Supervisor", "nunique"), Compradores=("Comprador", "nunique"), Venda=("Venda", "sum"), **{"Prêmio rateado": ("Prêmio rateado", "sum")}).sort_values("Prêmio rateado", ascending=False)
            dataframe_br(ger, use_container_width=True, hide_index=True, export_title="Resumo por Gerente")
            nome_ger = st.selectbox("Detalhar gerente", ger["Gerente"].tolist(), key="detalhe_gerente")
            detg = rateio[rateio["Gerente"] == nome_ger]
            dataframe_br(detg[["Loja", "Supervisor", "Comprador", "Venda", "Prêmio rateado"]], use_container_width=True, hide_index=True, export_title=f"Gerente - {nome_ger}")
    with aba_cad:
        st.info("Cadastre o supervisor e o gerente responsáveis por cada loja. Esta configuração afeta somente as visões analíticas de premiação.")
        hier = pd.DataFrame(carregar_hierarquia_premiacao())
        if hier.empty:
            lojas_base = _rateio_premiacao_loja(PERIODO_REALIZADO_USADO)
            hier = pd.DataFrame({"Loja": sorted(lojas_base["Loja"].unique()) if not lojas_base.empty else [], "Supervisor": "Não cadastrado", "Gerente": "Não cadastrado"})
        for col in ["Loja", "Supervisor", "Gerente"]:
            if col not in hier.columns: hier[col] = "Não cadastrado"
        editada = st.data_editor(hier[["Loja", "Supervisor", "Gerente"]], use_container_width=True, hide_index=True, num_rows="dynamic", key="editor_hierarquia_premio")
        if st.button("💾 Salvar hierarquia", type="primary", use_container_width=True):
            registros = editada.fillna("").to_dict("records")
            salvar_hierarquia_premiacao(registros)
            st.session_state["_flash_atualizacao_dados"] = "Hierarquia da premiação salva. As visões foram recalculadas."
            st.rerun()

elif visao == "Metas de Loja":
    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff;font-size:21px">Metas de Loja: Faturamento e Margem Bruta</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:4px">
          Acompanhamento executivo das metas financeiras por filial e gerente.
        </div>
      </div>
      <div style="color:#56d7ec;font-weight:800">Margem Bruta mínima: 34%</div>
    </div>
    """, unsafe_allow_html=True)

    df_todas_lojas = dataframe_metas_lojas()
    periodos_lojas = sorted(
        df_todas_lojas["periodo_referencia"].astype(str).dropna().unique().tolist(),
        reverse=True,
    ) if not df_todas_lojas.empty else []
    periodo_ativo_lojas = str(METAS_GESTOR.get("periodo_referencia", datetime.now().strftime("%Y-%m")))
    if periodo_ativo_lojas not in periodos_lojas:
        periodos_lojas = [periodo_ativo_lojas] + periodos_lojas

    periodo_painel_lojas = st.selectbox(
        "Período das metas de loja",
        periodos_lojas or [periodo_ativo_lojas],
        index=(periodos_lojas.index(periodo_ativo_lojas) if periodo_ativo_lojas in periodos_lojas else 0),
        key="periodo_painel_metas_lojas",
    )
    df_lojas = dataframe_metas_lojas(periodo_painel_lojas)
    if df_lojas.empty:
        st.warning("Nenhuma meta de loja cadastrada para o período selecionado.")
    else:
        total_meta = pd.to_numeric(df_lojas["meta_mes"], errors="coerce").fillna(0).sum()
        total_mb = pd.to_numeric(df_lojas["meta_margem_bruta_valor"], errors="coerce").fillna(0).sum()
        total_entrega = pd.to_numeric(df_lojas["representatividade_entrega_valor"], errors="coerce").fillna(0).sum()
        mb_pct_global = (total_mb / total_meta * 100) if total_meta else 0
        entrega_pct_global = (total_entrega / total_meta * 100) if total_meta else 0

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Meta global do mês", moeda_real(total_meta))
        c2.metric("Meta de margem bruta", moeda_real(total_mb))
        c3.metric("Margem bruta global", percentual(mb_pct_global))
        c4.metric("Representatividade da entrega", moeda_real(total_entrega), percentual(entrega_pct_global))

        tabela = df_lojas.rename(columns={
            "regional_loja": "Regional / Loja",
            "gerente": "Gerente",
            "meta_mes": "META Mês (R$)",
            "meta_margem_bruta_valor": "META Margem Bruta (R$)",
            "meta_margem_bruta_pct": "META MB (%)",
            "representatividade_entrega_pct": "Representatividade Entrega (%)",
            "representatividade_entrega_valor": "Representatividade Entrega (R$)",
            "status": "Status",
        })
        tabela = tabela[[
            "Regional / Loja", "Gerente", "META Mês (R$)",
            "META Margem Bruta (R$)", "META MB (%)",
            "Representatividade Entrega (%)", "Representatividade Entrega (R$)", "Status"
        ]]
        dataframe_br(
            tabela.style.format({
                "META Mês (R$)": lambda x: moeda_real(float(x)),
                "META Margem Bruta (R$)": lambda x: moeda_real(float(x)),
                "META MB (%)": lambda x: percentual(float(x)),
                "Representatividade Entrega (%)": lambda x: percentual(float(x)),
                "Representatividade Entrega (R$)": lambda x: moeda_real(float(x)),
            }),
            use_container_width=True, hide_index=True, height=360
        )

        graf = df_lojas.copy()
        graf["Meta do mês"] = pd.to_numeric(graf["meta_mes"], errors="coerce").fillna(0)
        graf["Margem bruta"] = pd.to_numeric(graf["meta_margem_bruta_valor"], errors="coerce").fillna(0)
        fig_loja = go.Figure()
        fig_loja.add_bar(name="Meta do mês", x=graf["regional_loja"], y=graf["Meta do mês"])
        fig_loja.add_bar(name="Meta margem bruta", x=graf["regional_loja"], y=graf["Margem bruta"])
        fig_loja.update_layout(
            barmode="group", height=390, paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)", font_color="#f3f7fb",
            margin=dict(l=20, r=20, t=45, b=20), yaxis_title="R$", xaxis_title=""
        )
        fig_loja.update_yaxes(gridcolor="rgba(255,255,255,.08)")
        plotly_chart_br(fig_loja, use_container_width=True, config={"displayModeBar": False})

elif visao == "Gestão de Metas de Loja":
    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff;font-size:21px">Gestão de Metas de Loja</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:4px">
          Cadastre, altere e acompanhe metas por filial, gerente e período.
        </div>
      </div>
      <div style="color:#56d7ec;font-weight:800">Dados persistentes</div>
    </div>
    """, unsafe_allow_html=True)

    df_edicao = dataframe_metas_lojas()
    periodos_existentes = sorted(df_edicao["periodo_referencia"].astype(str).unique().tolist(), reverse=True) if not df_edicao.empty else []
    periodo_padrao = METAS_GESTOR.get("periodo_referencia", datetime.now().strftime("%Y-%m"))
    if periodo_padrao not in periodos_existentes:
        periodos_existentes = [periodo_padrao] + periodos_existentes

    f1, f2 = st.columns([1, 2])
    with f1:
        periodo_gestao = st.selectbox("Período para gestão", periodos_existentes or [periodo_padrao])
    with f2:
        usuario_meta_loja = st.text_input("Responsável pela alteração", value="Gestor")

    with st.expander("📅 Alterar o período de uma meta já cadastrada", expanded=False):
        st.caption(
            "Transfira todas as metas do período selecionado para uma nova competência. "
            "O histórico anterior será preservado."
        )
        try:
            ano_atual, mes_atual = [int(x) for x in str(periodo_gestao).split("-")[:2]]
            data_periodo_atual = date(ano_atual, mes_atual, 1)
        except Exception:
            data_periodo_atual = date.today().replace(day=1)

        ap1, ap2 = st.columns([1, 2])
        with ap1:
            nova_competencia_data = st.date_input(
                "Nova competência",
                value=data_periodo_atual,
                key=f"nova_competencia_meta_loja_{periodo_gestao}",
            )
        with ap2:
            manter_periodo_original = st.checkbox(
                "Manter uma cópia no período original",
                value=False,
                help="Marcado: copia as metas. Desmarcado: move as metas para a nova competência.",
            )

        nova_competencia = nova_competencia_data.strftime("%Y-%m")
        alterar_periodo = st.button(
            "📅 Aplicar novo período",
            type="primary",
            use_container_width=True,
            key=f"alterar_periodo_meta_loja_{periodo_gestao}",
        )

        if alterar_periodo:
            if nova_competencia == str(periodo_gestao):
                st.warning("A nova competência deve ser diferente do período atual.")
            else:
                origem = df_edicao[
                    df_edicao["periodo_referencia"].astype(str) == str(periodo_gestao)
                ].copy()
                if origem.empty:
                    st.warning("Não existem metas cadastradas no período selecionado.")
                else:
                    transferidas = origem.copy()
                    transferidas["periodo_referencia"] = nova_competencia

                    if manter_periodo_original:
                        base_final = pd.concat([df_edicao, transferidas], ignore_index=True)
                    else:
                        base_sem_origem = df_edicao[
                            df_edicao["periodo_referencia"].astype(str) != str(periodo_gestao)
                        ].copy()
                        base_final = pd.concat([base_sem_origem, transferidas], ignore_index=True)

                    # Quando já houver a mesma filial no destino, prevalece a meta transferida.
                    base_final["periodo_referencia"] = base_final["periodo_referencia"].astype(str)
                    base_final["regional_loja"] = base_final["regional_loja"].astype(str)
                    base_final = base_final.drop_duplicates(
                        subset=["periodo_referencia", "regional_loja"],
                        keep="last",
                    )
                    salvar_metas_lojas(base_final.to_dict("records"), usuario_meta_loja)
                    acao = "copiadas" if manter_periodo_original else "transferidas"
                    st.success(
                        f"Metas {acao} de {periodo_gestao} para {nova_competencia}. "
                        "O histórico foi registrado."
                    )
                    st.session_state["periodo_painel_metas_lojas"] = nova_competencia
                    st.rerun()

    df_periodo = df_edicao[df_edicao["periodo_referencia"].astype(str) == str(periodo_gestao)].copy()
    if df_periodo.empty:
        df_periodo = pd.DataFrame([{
            "periodo_referencia": periodo_gestao, "regional_loja": "Nova filial", "gerente": "",
            "meta_mes": 0.0, "meta_margem_bruta_valor": 0.0, "meta_margem_bruta_pct": 34.0,
            "representatividade_entrega_pct": 0.0, "representatividade_entrega_valor": 0.0, "status": "Planejada"
        }])

    st.info("Você pode adicionar ou excluir linhas diretamente na tabela. A Margem Bruta em R$ pode ser recalculada automaticamente a partir da meta do mês e do percentual.")
    editado = st.data_editor(
        df_periodo, num_rows="dynamic", use_container_width=True, hide_index=True,
        column_config={
            "periodo_referencia": st.column_config.TextColumn("Período", disabled=True),
            "regional_loja": st.column_config.TextColumn("Regional / Loja", required=True),
            "gerente": st.column_config.TextColumn("Gerente", required=True),
            "meta_mes": st.column_config.NumberColumn("META Mês (R$)", min_value=0.0, format="R$ %.2f"),
            "meta_margem_bruta_valor": st.column_config.NumberColumn("META Margem Bruta (R$)", min_value=0.0, format="R$ %.2f"),
            "meta_margem_bruta_pct": st.column_config.NumberColumn("META MB (%)", min_value=0.0, max_value=100.0, format="%.2f%%"),
            "representatividade_entrega_pct": st.column_config.NumberColumn("Representatividade Entrega (%)", min_value=0.0, max_value=100.0, format="%.2f%%"),
            "representatividade_entrega_valor": st.column_config.NumberColumn("Representatividade Entrega (R$)", min_value=0.0, format="R$ %.2f"),
            "status": st.column_config.SelectboxColumn("Status", options=["Planejada", "Ativa", "Encerrada", "Cancelada"]),
        },
        key=f"editor_metas_lojas_{periodo_gestao}"
    )

    b1, b2, b3 = st.columns(3)
    recalcular = b1.button("🧮 Recalcular Margem em R$", use_container_width=True)
    salvar = b2.button("💾 Salvar metas de loja", type="primary", use_container_width=True)
    restaurar = b3.button("↩️ Restaurar exemplo inicial", use_container_width=True)

    if recalcular:
        temp = editado.copy()
        temp["periodo_referencia"] = periodo_gestao
        temp["meta_margem_bruta_valor"] = (
            pd.to_numeric(temp["meta_mes"], errors="coerce").fillna(0) *
            pd.to_numeric(temp["meta_margem_bruta_pct"], errors="coerce").fillna(0) / 100
        ).round(2)
        for c in ["meta_mes", "meta_margem_bruta_valor", "meta_margem_bruta_pct", "representatividade_entrega_pct", "representatividade_entrega_valor"]:
            temp[c] = pd.to_numeric(temp[c], errors="coerce").fillna(0.0)
        temp = temp[temp["regional_loja"].astype(str).str.strip() != ""].copy()
        restantes = df_edicao[df_edicao["periodo_referencia"].astype(str) != str(periodo_gestao)].copy()
        salvar_metas_lojas(pd.concat([restantes, temp], ignore_index=True).to_dict("records"), usuario_meta_loja)
        st.success("Margem bruta recalculada, salva e registrada no histórico.")
        st.rerun()

    if salvar:
        temp = editado.copy()
        temp["periodo_referencia"] = periodo_gestao
        for c in ["meta_mes", "meta_margem_bruta_valor", "meta_margem_bruta_pct", "representatividade_entrega_pct", "representatividade_entrega_valor"]:
            temp[c] = pd.to_numeric(temp[c], errors="coerce").fillna(0.0)
        temp = temp[temp["regional_loja"].astype(str).str.strip() != ""].copy()
        restantes = df_edicao[df_edicao["periodo_referencia"].astype(str) != str(periodo_gestao)].copy()
        consolidado = pd.concat([restantes, temp], ignore_index=True)
        salvar_metas_lojas(consolidado.to_dict("records"), usuario_meta_loja)
        st.success("Metas de loja salvas e histórico registrado.")
        st.rerun()

    if restaurar:
        outros = df_edicao[df_edicao["periodo_referencia"].astype(str) != str(periodo_gestao)].copy()
        padrao = pd.DataFrame([dict(x, periodo_referencia=periodo_gestao) for x in METAS_LOJAS_PADRAO])
        salvar_metas_lojas(pd.concat([outros, padrao], ignore_index=True).to_dict("records"), usuario_meta_loja)
        st.success("Exemplo inicial restaurado para o período selecionado.")
        st.rerun()

    st.markdown("### Resumo do período")
    resumo = editado.copy()
    total_meta = pd.to_numeric(resumo["meta_mes"], errors="coerce").fillna(0).sum()
    total_mb = pd.to_numeric(resumo["meta_margem_bruta_valor"], errors="coerce").fillna(0).sum()
    total_entrega = pd.to_numeric(resumo["representatividade_entrega_valor"], errors="coerce").fillna(0).sum()
    r1, r2, r3 = st.columns(3)
    r1.metric("Meta global", moeda_real(total_meta))
    r2.metric("Margem bruta global", moeda_real(total_mb), percentual(total_mb / total_meta * 100 if total_meta else 0))
    r3.metric("Entrega representativa", moeda_real(total_entrega), percentual(total_entrega / total_meta * 100 if total_meta else 0))

    with st.expander("📚 Histórico de alterações", expanded=False):
        hist_lojas = carregar_historico_metas_lojas()
        if hist_lojas:
            hist_resumo = pd.DataFrame([{
                "Data/Hora": h.get("data_hora", ""),
                "Usuário": h.get("usuario", ""),
                "Quantidade de lojas": len(h.get("registros", [])),
            } for h in reversed(hist_lojas)])
            dataframe_br(hist_resumo, use_container_width=True, hide_index=True)
        else:
            st.caption("Ainda não há alterações registradas.")

elif visao == "Gestão de Metas":
    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff;font-size:20px">Centro de Gestão de Metas</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:4px">
          Cadastre metas por período, preserve histórico e controle a vigência.
        </div>
      </div>
      <div style="color:#56d7ec;font-weight:800">Rede Economize Control Center</div>
    </div>
    """, unsafe_allow_html=True)

    periodos_unificados = listar_periodos_gestao_metas()
    periodo_atual_gestao = str(METAS_GESTOR.get("periodo_referencia", datetime.now().strftime("%Y-%m")))
    periodo_estado = str(st.session_state.get("periodo_gestao_unificado_global", "")).strip()
    for _periodo_extra in [periodo_atual_gestao, periodo_estado]:
        if _periodo_extra and _periodo_extra not in periodos_unificados:
            periodos_unificados.insert(0, _periodo_extra)

    def _abrir_competencia_global_callback():
        nova_data = st.session_state.get("nova_competencia_gestao_global")
        if nova_data:
            st.session_state["periodo_gestao_unificado_global"] = nova_data.strftime("%Y-%m")

    pg1, pg2, pg3 = st.columns([1.3, 1, 1.7])
    with pg1:
        periodo_gestao_unificado = st.selectbox("Período de trabalho para todas as metas", periodos_unificados or [periodo_atual_gestao], key="periodo_gestao_unificado_global")
    with pg2:
        nova_competencia_global = st.date_input("Nova competência", value=datetime.strptime(periodo_gestao_unificado + "-01", "%Y-%m-%d").date(), key="nova_competencia_gestao_global")
    with pg3:
        st.caption("Este período controla Meta Geral, Metas de Loja e Metas por Comprador.")
        st.button("➕ Abrir competência selecionada", key="abrir_competencia_global", use_container_width=True, on_click=_abrir_competencia_global_callback)

    aba1, aba_lojas, aba_compradores, aba2 = st.tabs([
        "📝 Meta Geral",
        "🏪 Metas de Loja",
        "🎯 Metas por Comprador",
        "📚 Histórico de Metas",
    ])

    with aba1:
        with st.form("form_metas_gestor"):
            st.markdown("### Identificação e período")
            p1, p2, p3, p4 = st.columns([1.2, 1, 1, 1])
            with p1:
                periodo_referencia = st.text_input(
                    "Período de referência",
                    value=periodo_gestao_unificado,
                    placeholder="Ex.: 2026-05",
                    key="periodo_referencia_meta_geral_unificado"
                )
            with p2:
                data_inicio = st.date_input(
                    "Data inicial",
                    value=datetime.strptime(
                        METAS_GESTOR.get("data_inicio", "2026-07-01"), "%Y-%m-%d"
                    ).date()
                )
            with p3:
                data_fim = st.date_input(
                    "Data final",
                    value=datetime.strptime(
                        METAS_GESTOR.get("data_fim", "2026-07-31"), "%Y-%m-%d"
                    ).date()
                )
            with p4:
                status_meta = st.selectbox(
                    "Status",
                    ["Planejada", "Ativa", "Encerrada", "Cancelada"],
                    index=["Planejada", "Ativa", "Encerrada", "Cancelada"].index(
                        METAS_GESTOR.get("status", "Ativa")
                    )
                )

            descricao = st.text_input(
                "Descrição da meta",
                value=METAS_GESTOR.get("descricao", "Metas comerciais do período")
            )
            usuario_cadastro = st.text_input(
                "Gestor responsável",
                value=METAS_GESTOR.get("usuario_cadastro", "Gestor")
            )

            st.markdown("### Metas gerais")
            g1, g2, g3, g4 = st.columns(4)
            with g1:
                meta_venda = st.number_input("Meta Venda Total Mês (R$)", min_value=0.0, value=float(METAS_GESTOR["meta_venda_total_mes"]), step=10000.0, format="%.2f")
            with g2:
                meta_cmv = st.number_input("Meta CMV Mês (%)", min_value=0.0, max_value=100.0, value=float(METAS_GESTOR["meta_cmv_mes"]), step=0.1, format="%.2f")
            with g3:
                fator_reducao = st.number_input("Fator Redução CMV", min_value=0.0, value=float(METAS_GESTOR["fator_reducao_cmv"]), step=0.01, format="%.2f")
            with g4:
                fator_cobertura = st.number_input("Fator Cobertura", min_value=0.0, value=float(METAS_GESTOR["fator_cobertura"]), step=0.05, format="%.2f")

            st.markdown("### Metas operacionais")
            o1, o2 = st.columns(2)
            with o1:
                meta_ruptura = st.number_input("Meta Ruptura (%)", min_value=0.0, max_value=100.0, value=float(METAS_GESTOR["meta_ruptura"]), step=0.1, format="%.2f")
            with o2:
                meta_reposicao = st.number_input("Meta Reposição (%)", min_value=0.0, max_value=200.0, value=float(METAS_GESTOR["meta_reposicao"]), step=0.1, format="%.2f")

            st.markdown("### Distribuição por curva de estoque")
            c1, c2, c3, c4 = st.columns(4)
            with c1: curva_a = st.number_input("Curva A (%)", 0.0, 100.0, float(METAS_GESTOR["curva_a"]), 1.0)
            with c2: curva_b = st.number_input("Curva B (%)", 0.0, 100.0, float(METAS_GESTOR["curva_b"]), 1.0)
            with c3: curva_c = st.number_input("Curva C (%)", 0.0, 100.0, float(METAS_GESTOR["curva_c"]), 1.0)
            with c4: curva_d = st.number_input("Curva D (%)", 0.0, 100.0, float(METAS_GESTOR["curva_d"]), 1.0)

            total_curvas = curva_a + curva_b + curva_c + curva_d
            st.caption(f"Total das curvas: {percentual(total_curvas)}")

            st.markdown("### Participação de venda por comprador")
            st.caption(
                "Os compradores são reconhecidos automaticamente pelas bases e pelo mapa "
                "de classificações. Os percentuais abaixo podem ser alterados."
            )
            compradores_reconhecidos = sorted(lista_compradores_ativos(), key=lambda x: x.casefold())
            metas_participacao = carregar_metas_por_comprador()
            linhas_participacao = []
            for nome in compradores_reconhecidos:
                atual = next((
                    item for item in metas_participacao
                    if str(item.get("periodo_referencia", "")) == str(periodo_referencia)
                    and str(item.get("comprador", "")).strip().casefold() == nome.casefold()
                ), {})
                realizado_nome = REALIZADOS.loc[
                    REALIZADOS["Comprador"].astype(str).str.strip().str.casefold() == nome.casefold(),
                    "Rep. Faturamento",
                ]
                participacao_real = float(realizado_nome.iloc[0]) if not realizado_nome.empty else 0.0
                linhas_participacao.append({
                    "Comprador": nome,
                    "Participação Meta (%)": float(atual.get("participacao_venda_pct", participacao_real)),
                    "Participação Real (%)": participacao_real,
                })
            editor_participacao = st.data_editor(
                pd.DataFrame(linhas_participacao),
                use_container_width=True,
                hide_index=True,
                disabled=["Comprador", "Participação Real (%)"],
                column_config={
                    "Comprador": st.column_config.TextColumn("Comprador reconhecido"),
                    "Participação Meta (%)": st.column_config.NumberColumn(
                        "Participação Meta (%)", min_value=0.0, max_value=100.0,
                        step=0.1, format="%.2f%%"
                    ),
                    "Participação Real (%)": st.column_config.NumberColumn(
                        "Participação Real (%)", format="%.2f%%"
                    ),
                },
                key="editor_participacao_compradores_dinamicos",
            )
            total_rep = pd.to_numeric(
                editor_participacao.get("Participação Meta (%)", pd.Series(dtype=float)),
                errors="coerce",
            ).fillna(0).sum()
            st.caption(f"Total da participação configurada: {percentual(total_rep)}")

            st.markdown("### Pesos da premiação")
            p1, p2, p3 = st.columns(3)
            with p1:
                peso_fat = st.number_input("Peso Faturamento (%)", 0.0, 100.0, float(METAS_GESTOR["peso_faturamento"]), 1.0)
                peso_cmv = st.number_input("Peso CMV (%)", 0.0, 100.0, float(METAS_GESTOR["peso_cmv"]), 1.0)
                peso_cob = st.number_input("Peso Fator Cobertura (%)", 0.0, 100.0, float(METAS_GESTOR["peso_fator_cobertura"]), 1.0)
            with p2:
                peso_a = st.number_input("Peso Curva A (%)", 0.0, 100.0, float(METAS_GESTOR["peso_curva_a"]), 1.0)
                peso_b = st.number_input("Peso Curva B (%)", 0.0, 100.0, float(METAS_GESTOR["peso_curva_b"]), 1.0)
                peso_c = st.number_input("Peso Curva C (%)", 0.0, 100.0, float(METAS_GESTOR["peso_curva_c"]), 1.0)
            with p3:
                peso_d = st.number_input("Peso Curva D (%)", 0.0, 100.0, float(METAS_GESTOR["peso_curva_d"]), 1.0)
                peso_ruptura = st.number_input("Peso Ruptura (%)", 0.0, 100.0, float(METAS_GESTOR["peso_ruptura"]), 1.0)
                peso_reposicao = st.number_input("Peso Reposição CMV (%)", 0.0, 100.0, float(METAS_GESTOR["peso_reposicao"]), 1.0)

            total_pesos = peso_fat + peso_cmv + peso_cob + peso_a + peso_b + peso_c + peso_d + peso_ruptura + peso_reposicao
            st.caption(f"Peso total configurado: {percentual(total_pesos)}")

            valor_premio = st.number_input("Valor total de prêmio atingível (R$)", min_value=0.0, value=float(METAS_GESTOR["valor_premio_total"]), step=100.0, format="%.2f")

            salvar = st.form_submit_button("💾 Salvar meta do período", use_container_width=True)

            if salvar:
                erros = []
                if not periodo_referencia.strip():
                    erros.append("Informe o período de referência.")
                if data_fim < data_inicio:
                    erros.append("A data final não pode ser anterior à data inicial.")
                if abs(total_curvas - 100.0) > 0.01:
                    erros.append("A soma das curvas deve ser igual a 100%.")
                if abs(total_rep - 100.0) > 0.01:
                    erros.append("A soma da participação dos compradores deve ser igual a 100%.")
                if abs(total_pesos - 100.0) > 0.01:
                    erros.append("A soma dos pesos da premiação deve ser igual a 100%.")

                if erros:
                    for erro in erros:
                        st.error(erro)
                else:
                    id_meta = f"META-{periodo_referencia.strip()}"
                    novas_metas = {
                        "id_meta": id_meta,
                        "periodo_referencia": periodo_referencia.strip(),
                        "data_inicio": data_inicio.strftime("%Y-%m-%d"),
                        "data_fim": data_fim.strftime("%Y-%m-%d"),
                        "descricao": descricao,
                        "status": status_meta,
                        "usuario_cadastro": usuario_cadastro,
                        "data_cadastro": METAS_GESTOR.get("data_cadastro", ""),
                        "meta_venda_total_mes": meta_venda,
                        "meta_cmv_mes": meta_cmv,
                        "fator_reducao_cmv": fator_reducao,
                        "fator_cobertura": fator_cobertura,
                        "meta_ruptura": meta_ruptura,
                        "meta_reposicao": meta_reposicao,
                        "curva_a": curva_a,
                        "curva_b": curva_b,
                        "curva_c": curva_c,
                        "curva_d": curva_d,
                        # Campos legados preservados apenas por compatibilidade.
                        # A participação atual é armazenada por comprador logo abaixo.
                        "rep_paulo": float(METAS_GESTOR.get("rep_paulo", 0)),
                        "rep_francieli": float(METAS_GESTOR.get("rep_francieli", 0)),
                        "rep_sebastiao": float(METAS_GESTOR.get("rep_sebastiao", 0)),
                        "peso_faturamento": peso_fat,
                        "peso_cmv": peso_cmv,
                        "peso_fator_cobertura": peso_cob,
                        "peso_curva_a": peso_a,
                        "peso_curva_b": peso_b,
                        "peso_curva_c": peso_c,
                        "peso_curva_d": peso_d,
                        "peso_ruptura": peso_ruptura,
                        "peso_reposicao": peso_reposicao,
                        "valor_premio_total": valor_premio,
                    }
                    # Salva a participação para cada comprador reconhecido.
                    metas_individuais = carregar_metas_por_comprador()
                    for _, linha_part in editor_participacao.iterrows():
                        nome_part = str(linha_part.get("Comprador", "")).strip()
                        pct_part = float(linha_part.get("Participação Meta (%)", 0) or 0)
                        encontrado = False
                        for item in metas_individuais:
                            if (
                                str(item.get("periodo_referencia", "")) == str(periodo_referencia.strip())
                                and str(item.get("comprador", "")).strip().casefold() == nome_part.casefold()
                            ):
                                item["participacao_venda_pct"] = pct_part
                                item["meta_venda"] = float(meta_venda) * pct_part / 100.0
                                item["ultima_atualizacao"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                encontrado = True
                                break
                        if not encontrado and nome_part:
                            novo_item = estrutura_meta_comprador_padrao(
                                nome_part, periodo_referencia.strip(), pct_part
                            )
                            novo_item["meta_venda"] = float(meta_venda) * pct_part / 100.0
                            novo_item["ultima_atualizacao"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            metas_individuais.append(novo_item)
                    salvar_metas_por_comprador(metas_individuais)
                    salvar_metas(novas_metas)
                    st.success("Meta do período salva com sucesso.")
                    st.rerun()

        if METAS_GESTOR.get("ultima_atualizacao"):
            st.caption(f"Última atualização: {METAS_GESTOR['ultima_atualizacao']}")

    with aba_lojas:
        st.markdown("### Gestão de Metas de Loja")
        st.caption(
            "Cadastre e altere as metas de faturamento e margem bruta por filial, "
            "usando o mesmo padrão de período, vigência, status e histórico da meta geral."
        )

        df_lojas_todas = dataframe_metas_lojas()
        periodos_lojas = sorted(
            df_lojas_todas["periodo_referencia"].astype(str).dropna().unique().tolist(),
            reverse=True,
        ) if not df_lojas_todas.empty else []
        periodo_padrao_lojas = str(METAS_GESTOR.get("periodo_referencia", datetime.now().strftime("%Y-%m")))
        if periodo_padrao_lojas not in periodos_lojas:
            periodos_lojas = [periodo_padrao_lojas] + periodos_lojas

        c_periodo_loja, c_inicio_loja, c_fim_loja, c_usuario_loja = st.columns([1, 1, 1, 1.3])
        with c_periodo_loja:
            periodo_gestao_lojas = periodo_gestao_unificado
            st.text_input("Período de referência", value=periodo_gestao_lojas, disabled=True, key="periodo_lojas_unificado_exibicao")
        try:
            ano_l, mes_l = [int(x) for x in str(periodo_gestao_lojas).split("-")[:2]]
            inicio_padrao_loja = date(ano_l, mes_l, 1)
            if mes_l == 12:
                fim_padrao_loja = date(ano_l, 12, 31)
            else:
                fim_padrao_loja = date(ano_l, mes_l + 1, 1) - timedelta(days=1)
        except Exception:
            inicio_padrao_loja = date.today().replace(day=1)
            fim_padrao_loja = date.today()
        with c_inicio_loja:
            data_inicio_lojas = st.date_input(
                "Data inicial",
                value=inicio_padrao_loja,
                key="gestao_unificada_inicio_lojas",
            )
        with c_fim_loja:
            data_fim_lojas = st.date_input(
                "Data final",
                value=fim_padrao_loja,
                key="gestao_unificada_fim_lojas",
            )
        with c_usuario_loja:
            usuario_meta_loja_unificado = st.text_input(
                "Gestor responsável",
                value="Gestor",
                key="gestao_unificada_usuario_lojas",
            )

        df_periodo_lojas = df_lojas_todas[
            df_lojas_todas["periodo_referencia"].astype(str) == str(periodo_gestao_lojas)
        ].copy() if not df_lojas_todas.empty else pd.DataFrame()

        colunas_lojas = [
            "periodo_referencia", "regional_loja", "gerente", "meta_mes",
            "meta_margem_bruta_valor", "meta_margem_bruta_pct",
            "representatividade_entrega_pct", "representatividade_entrega_valor", "status"
        ]
        if df_periodo_lojas.empty:
            df_periodo_lojas = pd.DataFrame(columns=colunas_lojas)
        for col in colunas_lojas:
            if col not in df_periodo_lojas.columns:
                df_periodo_lojas[col] = "" if col in {"periodo_referencia", "regional_loja", "gerente", "status"} else 0.0
        df_periodo_lojas["periodo_referencia"] = str(periodo_gestao_lojas)

        editor_lojas_unificado = st.data_editor(
            df_periodo_lojas[colunas_lojas],
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            column_config={
                "periodo_referencia": st.column_config.TextColumn("Período", disabled=True),
                "regional_loja": st.column_config.TextColumn("Regional / Loja", required=True),
                "gerente": st.column_config.TextColumn("Gerente", required=True),
                "meta_mes": st.column_config.NumberColumn("Meta Faturamento (R$)", min_value=0.0, format="R$ %.2f"),
                "meta_margem_bruta_valor": st.column_config.NumberColumn("Meta Margem Bruta (R$)", min_value=0.0, format="R$ %.2f"),
                "meta_margem_bruta_pct": st.column_config.NumberColumn("Meta MB (%)", min_value=0.0, max_value=100.0, format="%.2f%%"),
                "representatividade_entrega_pct": st.column_config.NumberColumn("Representatividade (%)", min_value=0.0, max_value=100.0, format="%.2f%%"),
                "representatividade_entrega_valor": st.column_config.NumberColumn("Representatividade (R$)", min_value=0.0, format="R$ %.2f"),
                "status": st.column_config.SelectboxColumn("Status", options=["Planejada", "Ativa", "Encerrada", "Cancelada"]),
            },
            key=f"gestao_unificada_editor_lojas_{periodo_gestao_lojas}",
        )

        bl1, bl2, bl3 = st.columns(3)
        recalcular_lojas = bl1.button("🧮 Recalcular margem", use_container_width=True, key="gestao_unificada_recalcular_lojas")
        salvar_lojas = bl2.button("💾 Salvar metas de loja", type="primary", use_container_width=True, key="gestao_unificada_salvar_lojas")
        duplicar_lojas = bl3.button("📑 Duplicar para outro período", use_container_width=True, key="gestao_unificada_duplicar_lojas")

        if recalcular_lojas or salvar_lojas:
            if data_fim_lojas < data_inicio_lojas:
                st.error("A data final não pode ser anterior à data inicial.")
            else:
                temp_lojas = editor_lojas_unificado.copy()
                temp_lojas["periodo_referencia"] = str(periodo_gestao_lojas)
                for c in ["meta_mes", "meta_margem_bruta_valor", "meta_margem_bruta_pct", "representatividade_entrega_pct", "representatividade_entrega_valor"]:
                    temp_lojas[c] = pd.to_numeric(temp_lojas[c], errors="coerce").fillna(0.0)
                if recalcular_lojas:
                    temp_lojas["meta_margem_bruta_valor"] = (
                        temp_lojas["meta_mes"] * temp_lojas["meta_margem_bruta_pct"] / 100
                    ).round(2)
                temp_lojas = temp_lojas[temp_lojas["regional_loja"].astype(str).str.strip() != ""].copy()
                restantes_lojas = df_lojas_todas[
                    df_lojas_todas["periodo_referencia"].astype(str) != str(periodo_gestao_lojas)
                ].copy() if not df_lojas_todas.empty else pd.DataFrame()
                consolidado_lojas = pd.concat([restantes_lojas, temp_lojas], ignore_index=True)
                salvar_metas_lojas(consolidado_lojas.to_dict("records"), usuario_meta_loja_unificado)
                st.success("Metas de loja salvas e registradas no histórico.")
                st.rerun()

        if duplicar_lojas:
            st.session_state["mostrar_duplicacao_lojas_unificada"] = True
        if st.session_state.get("mostrar_duplicacao_lojas_unificada"):
            d1, d2 = st.columns([1, 2])
            with d1:
                nova_comp_lojas = st.date_input(
                    "Nova competência",
                    value=inicio_padrao_loja,
                    key="gestao_unificada_nova_comp_lojas",
                ).strftime("%Y-%m")
            with d2:
                st.caption("A cópia preserva o período original e cria uma nova versão editável.")
            if st.button("Confirmar duplicação", type="primary", key="gestao_unificada_confirmar_dup_lojas"):
                origem_lojas = df_lojas_todas[
                    df_lojas_todas["periodo_referencia"].astype(str) == str(periodo_gestao_lojas)
                ].copy()
                if origem_lojas.empty:
                    st.warning("Não existem metas no período selecionado.")
                elif nova_comp_lojas == str(periodo_gestao_lojas):
                    st.warning("Escolha uma competência diferente.")
                else:
                    copia_lojas = origem_lojas.copy()
                    copia_lojas["periodo_referencia"] = nova_comp_lojas
                    base_lojas = pd.concat([df_lojas_todas, copia_lojas], ignore_index=True)
                    base_lojas = base_lojas.drop_duplicates(
                        subset=["periodo_referencia", "regional_loja"], keep="last"
                    )
                    salvar_metas_lojas(base_lojas.to_dict("records"), usuario_meta_loja_unificado)
                    st.session_state["mostrar_duplicacao_lojas_unificada"] = False
                    st.success(f"Metas duplicadas para {nova_comp_lojas}.")
                    st.rerun()

        if not editor_lojas_unificado.empty:
            total_meta_loja = pd.to_numeric(editor_lojas_unificado["meta_mes"], errors="coerce").fillna(0).sum()
            total_mb_loja = pd.to_numeric(editor_lojas_unificado["meta_margem_bruta_valor"], errors="coerce").fillna(0).sum()
            total_rep_loja = pd.to_numeric(editor_lojas_unificado["representatividade_entrega_valor"], errors="coerce").fillna(0).sum()
            rl1, rl2, rl3 = st.columns(3)
            rl1.metric("Meta global das lojas", moeda_real(total_meta_loja))
            rl2.metric("Margem bruta global", moeda_real(total_mb_loja), percentual(total_mb_loja / total_meta_loja * 100 if total_meta_loja else 0))
            rl3.metric("Representatividade", moeda_real(total_rep_loja), percentual(total_rep_loja / total_meta_loja * 100 if total_meta_loja else 0))

    with aba_compradores:
        st.markdown("### Gestão de Metas por Comprador")
        st.caption(
            "Cadastre metas individuais por comprador e período, mantendo o mesmo padrão "
            "de status, vigência, validação e histórico da meta geral."
        )

        historico_metas_unificado = carregar_historico()
        periodos_compradores = list(dict.fromkeys(
            [METAS_GESTOR.get("periodo_referencia", "")] +
            [h.get("periodo_referencia") for h in historico_metas_unificado if h.get("periodo_referencia")]
        ))
        cp1, cp2, cp3 = st.columns([1, 1, 1.3])
        with cp1:
            periodo_meta_comp_unificado = periodo_gestao_unificado
            st.text_input("Período de referência", value=periodo_meta_comp_unificado, disabled=True, key="periodo_compradores_unificado_exibicao")
        garantir_metas_compradores_periodo(periodo_meta_comp_unificado)
        compradores_ativos_unificados = lista_compradores_ativos()
        with cp2:
            comprador_meta_unificado = st.selectbox(
                "Comprador",
                compradores_ativos_unificados,
                key="gestao_unificada_comprador",
            )
        with cp3:
            usuario_meta_comprador = st.text_input(
                "Gestor responsável",
                value="Gestor",
                key="gestao_unificada_usuario_comprador",
            )

        meta_comp_atual = obter_meta_comprador(comprador_meta_unificado, periodo_meta_comp_unificado)
        with st.form("gestao_unificada_form_comprador"):
            gc1, gc2, gc3 = st.columns(3)
            with gc1:
                meta_venda_comp = st.number_input("Meta de Venda (R$)", min_value=0.0, value=float(meta_comp_atual.get("meta_venda", 0)), step=1000.0, format="%.2f")
                participacao_comp = st.number_input("Participação da Venda (%)", 0.0, 100.0, float(meta_comp_atual.get("participacao_venda_pct", 0)), 0.1, format="%.2f")
            with gc2:
                meta_cmv_comp = st.number_input("Meta CMV (%)", 0.0, 100.0, float(meta_comp_atual.get("meta_cmv_pct", 0)), 0.1, format="%.2f")
                fator_cob_comp = st.number_input("Fator de Cobertura", min_value=0.0, value=float(meta_comp_atual.get("fator_cobertura", 0)), step=0.05, format="%.2f")
            with gc3:
                meta_ruptura_comp = st.number_input("Meta de Ruptura (%)", 0.0, 100.0, float(meta_comp_atual.get("meta_ruptura_pct", 0)), 0.1, format="%.2f")
                meta_reposicao_comp = st.number_input("Meta de Reposição (%)", 0.0, 200.0, float(meta_comp_atual.get("meta_reposicao_pct", 0)), 0.1, format="%.2f")

            st.markdown("#### Distribuição do estoque por curva")
            cc1, cc2, cc3, cc4 = st.columns(4)
            with cc1: curva_a_comp = st.number_input("Curva A (%)", 0.0, 100.0, float(meta_comp_atual.get("meta_curva_a_pct", 0)), 1.0, key="unif_curva_a")
            with cc2: curva_b_comp = st.number_input("Curva B (%)", 0.0, 100.0, float(meta_comp_atual.get("meta_curva_b_pct", 0)), 1.0, key="unif_curva_b")
            with cc3: curva_c_comp = st.number_input("Curva C (%)", 0.0, 100.0, float(meta_comp_atual.get("meta_curva_c_pct", 0)), 1.0, key="unif_curva_c")
            with cc4: curva_d_comp = st.number_input("Curva D (%)", 0.0, 100.0, float(meta_comp_atual.get("meta_curva_d_pct", 0)), 1.0, key="unif_curva_d")

            pc1, pc2 = st.columns(2)
            with pc1:
                premio_comp = st.number_input("Valor de prêmio atingível (R$)", min_value=0.0, value=float(meta_comp_atual.get("valor_premio", 0)), step=100.0, format="%.2f")
            with pc2:
                status_comp = st.selectbox(
                    "Status",
                    ["Planejada", "Ativa", "Encerrada", "Cancelada"],
                    index=["Planejada", "Ativa", "Encerrada", "Cancelada"].index(meta_comp_atual.get("status", "Ativa")),
                    key="gestao_unificada_status_comprador",
                )
            salvar_comp = st.form_submit_button("💾 Salvar meta do comprador", use_container_width=True)

            if salvar_comp:
                total_curvas_comp = curva_a_comp + curva_b_comp + curva_c_comp + curva_d_comp
                if abs(total_curvas_comp - 100.0) > 0.01:
                    st.error("A soma das curvas deve ser igual a 100%.")
                else:
                    dados_comp = carregar_metas_por_comprador()
                    nova_meta_comp = {
                        "periodo_referencia": periodo_meta_comp_unificado,
                        "comprador": comprador_meta_unificado,
                        "meta_venda": meta_venda_comp,
                        "participacao_venda_pct": participacao_comp,
                        "meta_cmv_pct": meta_cmv_comp,
                        "meta_cmv_valor": meta_venda_comp * meta_cmv_comp / 100.0,
                        "fator_cobertura": fator_cob_comp,
                        "meta_estoque_total": meta_venda_comp * meta_cmv_comp / 100.0 * fator_cob_comp,
                        "meta_curva_a_pct": curva_a_comp,
                        "meta_curva_b_pct": curva_b_comp,
                        "meta_curva_c_pct": curva_c_comp,
                        "meta_curva_d_pct": curva_d_comp,
                        "meta_ruptura_pct": meta_ruptura_comp,
                        "meta_reposicao_pct": meta_reposicao_comp,
                        "valor_premio": premio_comp,
                        "status": status_comp,
                        "usuario_cadastro": usuario_meta_comprador,
                        "ultima_atualizacao": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                    }
                    substituiu_comp = False
                    for idx_comp, item_comp in enumerate(dados_comp):
                        if item_comp.get("periodo_referencia") == periodo_meta_comp_unificado and item_comp.get("comprador") == comprador_meta_unificado:
                            dados_comp[idx_comp] = nova_meta_comp
                            substituiu_comp = True
                            break
                    if not substituiu_comp:
                        dados_comp.append(nova_meta_comp)
                    salvar_metas_por_comprador(dados_comp)
                    st.success("Meta do comprador salva com sucesso.")
                    st.rerun()

        resumo_comp = pd.DataFrame([
            item for item in carregar_metas_por_comprador()
            if item.get("periodo_referencia") == periodo_meta_comp_unificado
            and str(item.get("comprador", "")).strip().casefold() in _conjunto_compradores_ativos()
        ])
        if not resumo_comp.empty:
            st.markdown("#### Resumo das metas do período")
            dataframe_br(resumo_comp, use_container_width=True, hide_index=True, height=310)
            soma_vendas_comp = pd.to_numeric(resumo_comp.get("meta_venda", 0), errors="coerce").fillna(0).sum()
            soma_part_comp = pd.to_numeric(resumo_comp.get("participacao_venda_pct", 0), errors="coerce").fillna(0).sum()
            rc1, rc2, rc3 = st.columns(3)
            rc1.metric("Soma das metas de venda", moeda_real(soma_vendas_comp))
            rc2.metric("Meta geral do período", moeda_real(METAS_GESTOR.get('meta_venda_total_mes', 0)))
            rc3.metric("Participação total", percentual(soma_part_comp))

    with aba2:
        historico = carregar_historico()
        if historico:
            hist_df = pd.DataFrame(historico)
            colunas = [
                "id_meta", "periodo_referencia", "data_inicio", "data_fim",
                "status", "descricao", "meta_venda_total_mes",
                "meta_cmv_mes", "valor_premio_total", "usuario_cadastro",
                "ultima_atualizacao"
            ]
            hist_df = hist_df[[c for c in colunas if c in hist_df.columns]].copy()
            if "data_inicio" in hist_df:
                hist_df["data_inicio"] = hist_df["data_inicio"].map(data_br)
            if "data_fim" in hist_df:
                hist_df["data_fim"] = hist_df["data_fim"].map(data_br)
            if "meta_venda_total_mes" in hist_df:
                hist_df["meta_venda_total_mes"] = hist_df["meta_venda_total_mes"].map(moeda)
            if "meta_cmv_mes" in hist_df:
                hist_df["meta_cmv_mes"] = hist_df["meta_cmv_mes"].map(percentual)
            if "valor_premio_total" in hist_df:
                hist_df["valor_premio_total"] = hist_df["valor_premio_total"].map(moeda)
            dataframe_br(hist_df, use_container_width=True, hide_index=True, height=360)

            opcoes = [h.get("id_meta") for h in historico]
            meta_escolhida = st.selectbox("Carregar meta do histórico", opcoes)
            c1, c2 = st.columns(2)
            with c1:
                if st.button("📥 Carregar como meta ativa", use_container_width=True):
                    registro = next(h for h in historico if h.get("id_meta") == meta_escolhida)
                    salvar_metas(registro.copy(), registrar_historico=False)
                    st.success("Meta carregada.")
                    st.rerun()
            with c2:
                if st.button("📑 Duplicar para novo período", use_container_width=True):
                    registro = next(h for h in historico if h.get("id_meta") == meta_escolhida).copy()
                    registro["id_meta"] = f"{registro.get('id_meta','META')}-COPIA"
                    registro["periodo_referencia"] = f"{registro.get('periodo_referencia','')}-copia"
                    registro["status"] = "Planejada"
                    registro["data_cadastro"] = ""
                    salvar_metas(registro)
                    st.success("Meta duplicada. Ajuste o novo período no cadastro.")
                    st.rerun()
        else:
            st.info("Ainda não existem metas no histórico.")


elif visao == "Importar Ruptura":
    _mensagem_atualizacao_pendente()
    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff;font-size:20px">Ruptura Automática</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:4px">
          Coloque o arquivo oficial na pasta IMPORTAR_RUPTURA. O sistema processa automaticamente.
        </div>
      </div>
      <div style="color:#56d7ec;font-weight:800">Automático</div>
    </div>
    """, unsafe_allow_html=True)

    arquivo = _arquivo_mais_recente()
    controle = {}
    if RUPTURA_AUTO_CONTROLE.exists():
        try:
            controle = json.loads(RUPTURA_AUTO_CONTROLE.read_text(encoding="utf-8"))
        except Exception:
            controle = {}
    ultima = controle.get("ultima_importacao", {})

    c1, c2, c3 = st.columns(3)
    with c1:
        st.metric("Período da meta", METAS_GESTOR.get("periodo_referencia", "-"))
    with c2:
        st.metric("Arquivo detectado", arquivo.name if arquivo else "Nenhum")
    with c3:
        st.metric("Última importação", ultima.get("data", "Não realizada"))

    st.caption("Pasta monitorada automaticamente:")
    st.code(str(PASTA_RUPTURA_AUTO.resolve()), language=None)

    b1, b2 = st.columns(2)
    with b1:
        if st.button("🔄 Verificar agora", use_container_width=True):
            try:
                r = processar_ruptura_automatica()
                if r["status"] == "importado":
                    st.session_state["_flash_atualizacao_dados"] = (
                        f"Ruptura atualizada: {r['registros']} registros em {r['periodo']}. "
                        "Cards, gráficos e tabelas foram recalculados."
                    )
                    st.rerun()
                else:
                    st.info(r["mensagem"])
            except Exception as e:
                st.error(str(e))
    with b2:
        if st.button("♻️ Reprocessar período", use_container_width=True):
            try:
                r = processar_ruptura_automatica(forcar=True)
                st.session_state["_flash_atualizacao_dados"] = (
                    f"Ruptura reprocessada: {r.get('registros', 0)} registros. "
                    "Cards, gráficos e tabelas foram recalculados."
                )
                st.rerun()
            except Exception as e:
                st.error(str(e))

    periodos = list(dict.fromkeys(
        [METAS_GESTOR.get("periodo_referencia", "")] +
        [h.get("periodo_referencia") for h in carregar_historico() if h.get("periodo_referencia")]
    ))
    periodo_rup = st.selectbox("Período para análise", periodos)
    base_rup = carregar_ruptura_auto(periodo_rup)

    if not base_rup.empty:
        k1, k2, k3, k4 = st.columns(4)
        with k1: st.metric("Itens", numero_inteiro(len(base_rup)))
        with k2: st.metric("Ruptura total", moeda_real(base_rup['Valor Ruptura'].sum()))
        with k3: st.metric("Necessidade a custo", moeda_real(base_rup['Valor Necessidade Custo'].sum()))
        with k4: st.metric("Não mapeados", numero_inteiro((base_rup["Comprador"] == "Não mapeado").sum()))

        g1, g2 = st.columns(2)
        with g1:
            por_comp = base_rup.groupby("Comprador", as_index=False)["Valor Ruptura"].sum().sort_values("Valor Ruptura")
            fig = px.bar(por_comp, x="Valor Ruptura", y="Comprador", orientation="h", title="Ruptura por comprador")
            fig.update_layout(height=330, paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", font_color="#f3f7fb")
            plotly_chart_br(fig, use_container_width=True, config={"displayModeBar": False})
        with g2:
            por_curva = base_rup.groupby("Curva Valor", as_index=False)["Valor Ruptura"].sum()
            fig2 = px.pie(por_curva, names="Curva Valor", values="Valor Ruptura", hole=.55, title="Ruptura por curva")
            fig2.update_layout(height=330, paper_bgcolor="rgba(0,0,0,0)", font_color="#f3f7fb")
            plotly_chart_br(fig2, use_container_width=True, config={"displayModeBar": False})

        dataframe_br(base_rup, use_container_width=True, hide_index=True, height=420)
    else:
        st.warning("Ainda não existem dados processados para esse período.")

    st.markdown("### Histórico")
    hist = historico_ruptura_auto()
    if not hist.empty:
        dataframe_br(hist, use_container_width=True, hide_index=True)
    else:
        st.info("Nenhuma importação registrada.")

elif visao == "Banco de Dados":
    _mensagem_atualizacao_pendente()
    status_sql_fontes = status_configuracao_fontes()
    if not status_sql_fontes.get("contas_pagar", {}).get("configurado", False):
        st.error(
            "Contas a Pagar não possui uma consulta SQL real configurada. "
            "A consulta atual é apenas um modelo e sempre retornaria zero. "
            "Abra a seção de edição de SQL e cole a consulta oficial."
        )

    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff;font-size:20px">Banco de Dados e Atualizações Mensais</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:4px">
          Configure o PostgreSQL, salve os scripts e atualize cada mês conforme o período da meta.
        </div>
      </div>
      <div style="color:#56d7ec;font-weight:800">Rede Economize Data Center</div>
    </div>
    """, unsafe_allow_html=True)

    aba_conexao, aba_sql, aba_atualizar, aba_historico = st.tabs([
        "🔐 Conexão",
        "🧾 Scripts SQL",
        "🔄 Atualização mensal",
        "📚 Histórico e dados salvos",
    ])

    with aba_conexao:
        CONFIG_BANCO = carregar_config_banco()
        st.markdown("### Configuração do PostgreSQL")
        st.caption(
            "Os dados são salvos localmente em `config/database.json`. "
            "Em ambiente online, a senha deverá ser migrada para variáveis de ambiente ou secrets."
        )

        with st.form("form_config_banco"):
            c1, c2 = st.columns([2, 1])
            with c1:
                host = st.text_input("Host", value=CONFIG_BANCO.get("host", ""))
                banco = st.text_input("Banco de dados", value=CONFIG_BANCO.get("banco", ""))
                usuario = st.text_input("Usuário", value=CONFIG_BANCO.get("usuario", ""))
            with c2:
                porta = st.number_input(
                    "Porta",
                    min_value=1,
                    max_value=65535,
                    value=int(CONFIG_BANCO.get("porta", 5432)),
                    step=1
                )
                sslmode = st.selectbox(
                    "SSL Mode",
                    ["prefer", "require", "disable"],
                    index=["prefer", "require", "disable"].index(
                        CONFIG_BANCO.get("sslmode", "prefer")
                    )
                )
                salvar_senha = st.checkbox(
                    "Salvar senha localmente",
                    value=bool(CONFIG_BANCO.get("salvar_senha", True))
                )

            senha = st.text_input(
                "Senha",
                value=CONFIG_BANCO.get("senha", "") if CONFIG_BANCO.get("salvar_senha", True) else "",
                type="password"
            )

            b1, b2 = st.columns(2)
            salvar_cfg = b1.form_submit_button("💾 Salvar configuração", use_container_width=True)
            testar_cfg = b2.form_submit_button("🔌 Testar conexão", use_container_width=True)

            cfg_digitada = {
                "tipo": "PostgreSQL",
                "host": host.strip(),
                "porta": int(porta),
                "banco": banco.strip(),
                "usuario": usuario.strip(),
                "senha": senha if salvar_senha else "",
                "sslmode": sslmode,
                "salvar_senha": salvar_senha,
                "ultima_validacao": CONFIG_BANCO.get("ultima_validacao", ""),
            }

            if salvar_cfg:
                salvar_config_banco(cfg_digitada)
                st.success("Configuração salva.")
                st.rerun()

            if testar_cfg:
                try:
                    cfg_teste = cfg_digitada.copy()
                    cfg_teste["senha"] = senha
                    identidade = validar_identidade_banco(cfg_teste)
                    cfg_digitada["ultima_validacao"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                    cfg_digitada["ultimo_banco_validado"] = identidade
                    if salvar_senha:
                        cfg_digitada["senha"] = senha
                    salvar_config_banco(cfg_digitada)
                    st.success(
                        "Conexão realizada com sucesso: "
                        f"{identidade.get('servidor') or cfg_teste.get('host')}:"
                        f"{identidade.get('porta') or cfg_teste.get('porta')} / "
                        f"Banco: {identidade.get('banco')} / Usuário: {identidade.get('usuario')}"
                    )
                except Exception as e:
                    st.error(f"Falha na conexão: {e}")

        if CONFIG_BANCO.get("ultima_validacao"):
            st.caption(f"Última conexão validada: {CONFIG_BANCO['ultima_validacao']}")
            identidade_salva = CONFIG_BANCO.get("ultimo_banco_validado") or {}
            if identidade_salva:
                st.info(
                    "Banco efetivamente validado: "
                    f"{identidade_salva.get('servidor') or CONFIG_BANCO.get('host')}:"
                    f"{identidade_salva.get('porta') or CONFIG_BANCO.get('porta')} · "
                    f"{identidade_salva.get('banco') or CONFIG_BANCO.get('banco')} · "
                    f"Usuário {identidade_salva.get('usuario') or CONFIG_BANCO.get('usuario')}"
                )

        st.markdown("#### Trocar de banco nesta máquina")
        st.caption(
            "Use este botão quando o projeto foi copiado de outro computador. "
            "Ele remove a conexão antiga, mas não apaga metas nem bases em cache."
        )
        if st.button("🧹 Apagar conexão antiga desta máquina", use_container_width=True):
            apagar_config_banco_local()
            st.success("Conexão antiga removida. Informe os dados do novo banco.")
            st.rerun()

    with aba_sql:
        cfg_planos_banco = carregar_config_analise_comercial()
        planos_ativos_banco = cfg_planos_banco.get(
            "planos_contas_selecionados",
            [PLANO_CONTAS_PAGAMENTO_PADRAO],
        )
        st.markdown("### Planos atualmente selecionados")
        for plano_banco in planos_ativos_banco:
            st.success(f"✓ {plano_banco}")
        st.caption(
            "A seleção é alterada e salva na tela Análise Comercial. "
            "Todos os planos retornados pela consulta permanecem armazenados no cache."
        )

        st.markdown("### Scripts utilizados nas atualizações")
        st.success(
            "Os scripts enviados anteriormente já estão carregados. "
            "Você pode alterar e salvar qualquer um deles nesta tela."
        )
        st.info(
            "Use os parâmetros `:data_inicio`, `:data_fim` e, quando necessário, "
            "`:periodo_referencia`. O sistema envia automaticamente as datas da meta selecionada."
        )

        fonte_sql = st.selectbox(
            "Fonte",
            list(FONTES_BANCO.keys()),
            format_func=lambda x: FONTES_BANCO[x]["titulo"]
        )
        sql_atual = ler_sql(FONTES_BANCO[fonte_sql]["arquivo_sql"])
        sql_configurado, sql_diagnostico = diagnosticar_sql_fonte(
            fonte_sql, sql_atual
        )
        if not sql_configurado:
            st.warning(sql_diagnostico)
        else:
            st.success("Consulta SQL configurada.")
        st.caption(
            f"Arquivo carregado: {FONTES_BANCO[fonte_sql]['arquivo_sql']} • "
            f"{len(sql_atual.splitlines())} linhas"
        )
        sql_editado = st.text_area(
            f"SQL de {FONTES_BANCO[fonte_sql]['titulo']}",
            value=sql_atual,
            height=420
        )

        if st.button("💾 Salvar script SQL", use_container_width=True):
            salvar_sql(FONTES_BANCO[fonte_sql]["arquivo_sql"], sql_editado)
            st.success("Script salvo.")

    with aba_atualizar:
        st.markdown("### Atualização por período da meta")

        historico_metas = carregar_historico()
        periodos_disponiveis = [
            h.get("periodo_referencia")
            for h in historico_metas
            if h.get("periodo_referencia")
        ]
        periodo_ativo = PERIODO_GLOBAL_SELECIONADO
        periodos_disponiveis = list(
            dict.fromkeys([periodo_ativo] + periodos_disponiveis)
        )

        periodo_selecionado = st.selectbox(
            "Período da meta",
            periodos_disponiveis,
            index=0
        )

        meta_periodo = next(
            (
                h for h in historico_metas
                if h.get("periodo_referencia") == periodo_selecionado
            ),
            METAS_GESTOR
        )

        data_inicio_periodo = meta_periodo.get("data_inicio", METAS_GESTOR.get("data_inicio"))
        data_fim_periodo = meta_periodo.get("data_fim", METAS_GESTOR.get("data_fim"))

        c1, c2, c3 = st.columns(3)
        with c1:
            st.metric("Período", periodo_selecionado)
        with c2:
            st.metric("Data inicial", data_br(data_inicio_periodo))
        with c3:
            st.metric("Data final", data_br(data_fim_periodo))

        st.caption(
            "Ao atualizar novamente o mesmo período, o sistema substitui somente os dados "
            "daquele mês e mantém os demais meses, metas, compradores, configurações e "
            "históricos já salvos."
        )

        fontes_selecionadas = st.multiselect(
            "Fontes a atualizar",
            list(FONTES_BANCO.keys()),
            default=list(FONTES_BANCO.keys()),
            format_func=lambda x: FONTES_BANCO[x]["titulo"]
        )

        confirmar = st.checkbox(
            f"Confirmo a atualização do período {periodo_selecionado}"
        )

        if st.button(
            "🧪 Testar somente a consulta de Contas a Pagar",
            use_container_width=True,
            key="testar_sql_contas_pagar_direto",
        ):
            cfg_teste = carregar_config_banco()
            cfg_teste["senha"] = (
                cfg_teste.get("senha")
                or st.session_state.get("senha_banco_execucao", "")
            )
            try:
                with st.spinner("Executando o SQL oficial diretamente no PostgreSQL..."):
                    sql_teste = ler_sql(FONTES_BANCO["contas_pagar"]["arquivo_sql"])
                    df_teste = executar_contas_pagar_psycopg2(
                        cfg_teste, sql_teste, data_inicio_periodo, data_fim_periodo, periodo_selecionado
                    )
                st.success(f"Consulta executada com sucesso: {len(df_teste):,} registros.".replace(",", "."))
                st.caption("Ambiente PostgreSQL: " + " | ".join(map(str, df_teste.attrs.get("ambiente_postgresql", ()))))
                dataframe_br(df_teste.head(20), use_container_width=True)
            except Exception as erro_teste:
                st.error("A consulta foi recusada pelo PostgreSQL.")
                st.code(str(erro_teste), language="text")

        if st.button(
            "🧪 Testar somente a consulta de Entradas",
            use_container_width=True,
            key="testar_sql_entradas_direto",
        ):
            cfg_teste = carregar_config_banco()
            cfg_teste["senha"] = (
                cfg_teste.get("senha")
                or st.session_state.get("senha_banco_execucao", "")
            )
            try:
                with st.spinner("Executando Entradas diretamente no PostgreSQL..."):
                    sql_teste = ler_sql(FONTES_BANCO["entradas"]["arquivo_sql"])
                    df_teste = executar_entradas_reconstruidas(
                        cfg_teste,
                        sql_teste,
                        data_inicio_periodo,
                        data_fim_periodo,
                        periodo_selecionado,
                    )
                tempo = float(df_teste.attrs.get("tempo_consulta_segundos", 0))
                st.success(
                    f"Consulta executada: {len(df_teste):,} registros em {tempo:.1f}s."
                    .replace(",", ".")
                )
                if df_teste.empty:
                    st.warning("A consulta não retornou registros; nada será apagado.")
                else:
                    total = float(pd.to_numeric(
                        df_teste.get("valor_nf_total", df_teste.get("entrada_custo_total", pd.Series(dtype=float))),
                        errors="coerce",
                    ).fillna(0).sum())
                    st.info(f"Total retornado para Entradas: {moeda_real(total)}")
                    dataframe_br(df_teste.head(20), use_container_width=True)
            except Exception as erro_teste:
                st.error("A consulta de Entradas foi recusada pelo PostgreSQL.")
                st.code(str(erro_teste), language="text")
                caminho_log = LOG_DIR / "entradas_erros.log"
                if caminho_log.exists():
                    st.download_button(
                        "⬇️ Baixar log completo de Entradas",
                        data=caminho_log.read_bytes(),
                        file_name="entradas_erros.log",
                        mime="text/plain",
                        use_container_width=True,
                    )


        st.markdown("### 🔎 Auditoria das regras de Entradas")
        st.caption(
            "Executa a consulta sem alterar o cache e compara as possíveis fórmulas "
            "com o total oficial do relatório A7."
        )
        col_ref_a7, col_auditar = st.columns([1, 2])
        with col_ref_a7:
            valor_referencia_a7 = st.number_input(
                "Compra oficial no A7",
                min_value=0.0,
                value=983238.25 if periodo_selecionado == "2026-07" else 0.0,
                step=0.01,
                format="%.2f",
                key=f"valor_referencia_a7_{periodo_selecionado}",
            )
        with col_auditar:
            auditar_entradas = st.button(
                "🔍 Auditar fórmulas e duplicidades",
                use_container_width=True,
                key="auditar_regras_entradas",
            )

        if auditar_entradas:
            cfg_auditoria = carregar_config_banco()
            cfg_auditoria["senha"] = (
                cfg_auditoria.get("senha")
                or st.session_state.get("senha_banco_execucao", "")
            )
            try:
                with st.spinner("Consultando Entradas e comparando as regras..."):
                    sql_auditoria = ler_sql(FONTES_BANCO["entradas"]["arquivo_sql"])
                    df_auditoria = executar_entradas_reconstruidas(
                        cfg_auditoria,
                        sql_auditoria,
                        data_inicio_periodo,
                        data_fim_periodo,
                        periodo_selecionado,
                    )
                    resumo_auditoria, duplicidades_auditoria, cfop_auditoria, diag_auditoria = gerar_auditoria_regras_entradas(
                        df_auditoria,
                        valor_referencia_a7,
                    )

                st.success(
                    f"Auditoria concluída: {numero_inteiro(diag_auditoria['linhas'])} linhas; "
                    f"{numero_inteiro(diag_auditoria['itens_unicos'])} itens únicos; "
                    f"{numero_inteiro(diag_auditoria['linhas_duplicadas'])} repetições."
                )
                melhor = resumo_auditoria.iloc[0]
                st.info(
                    f"Regra mais próxima do A7: **{melhor['Regra']}** "
                    f"({melhor['Tipo']}) — {moeda_real(melhor['Total'])}; "
                    f"diferença {moeda_real(melhor['Diferença para A7'])}."
                )

                exibir = resumo_auditoria.copy()
                for coluna in ["Total", "Diferença para A7", "Distância absoluta"]:
                    exibir[coluna] = exibir[coluna].map(moeda_real)
                dataframe_br(exibir, use_container_width=True, hide_index=True)

                csv_resumo = resumo_auditoria.to_csv(index=False, sep=";", decimal=",").encode("utf-8-sig")
                st.download_button(
                    "⬇️ Baixar resultado da auditoria",
                    data=csv_resumo,
                    file_name=f"auditoria_entradas_{periodo_selecionado}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )

                if not duplicidades_auditoria.empty:
                    with st.expander(
                        f"⚠️ Itens repetidos por classificação ({len(duplicidades_auditoria):,} linhas)".replace(",", "."),
                        expanded=False,
                    ):
                        dataframe_br(duplicidades_auditoria.head(1000), use_container_width=True, hide_index=True)
                else:
                    st.success("Não foram encontrados itens repetidos pelo identificador do item da nota fiscal.")

                if not cfop_auditoria.empty:
                    with st.expander("📊 Totais por CFOP", expanded=False):
                        cfop_exibir = cfop_auditoria.copy()
                        cfop_exibir["Total_custo"] = cfop_exibir["Total_custo"].map(moeda_real)
                        cfop_exibir["Total_valor_unitario"] = cfop_exibir["Total_valor_unitario"].map(moeda_real)
                        dataframe_br(cfop_exibir, use_container_width=True, hide_index=True)
            except Exception as erro_auditoria:
                st.error("Não foi possível concluir a auditoria das Entradas.")
                st.code(str(erro_auditoria), language="text")

        if st.button("🔄 Atualizar fontes selecionadas", use_container_width=True):
            if not confirmar:
                st.warning("Marque a confirmação antes de atualizar.")
            elif not fontes_selecionadas:
                st.warning("Selecione ao menos uma fonte.")
            else:
                cfg_execucao = carregar_config_banco()
                if not cfg_execucao.get("senha"):
                    senha_execucao = st.session_state.get("senha_banco_execucao", "")
                else:
                    senha_execucao = cfg_execucao.get("senha", "")

                if not cfg_execucao.get("host") or not cfg_execucao.get("banco"):
                    st.error("Configure o banco de dados antes da atualização.")
                elif not senha_execucao:
                    st.error(
                        "A senha não está salva. Marque 'Salvar senha localmente' "
                        "ou salve novamente a configuração."
                    )
                else:
                    cfg_execucao["senha"] = senha_execucao

                    try:
                        identidade_execucao, cache_limpo_troca = preparar_cache_para_banco_atual(cfg_execucao)
                        if cache_limpo_troca:
                            st.warning(
                                "Foi detectada troca de banco. O cache operacional da máquina anterior "
                                "foi preservado em backup e limpo antes desta atualização."
                            )
                    except Exception as erro_identidade:
                        st.error(f"Não foi possível validar o banco antes da atualização: {erro_identidade}")
                        st.stop()

                    total_fontes = len(fontes_selecionadas)
                    total_etapas = total_fontes + 2
                    resultados = []
                    logs = []
                    inicio_geral = time.perf_counter()

                    progresso = st.progress(
                        0,
                        text="Preparando a atualização..."
                    )
                    status_principal = st.empty()
                    metricas_progresso = st.columns(4)
                    log_area = st.empty()

                    def atualizar_painel(etapa, texto, fonte_atual="-", registros=0):
                        decorrido = time.perf_counter() - inicio_geral
                        percentual_etapa = min(etapa / total_etapas, 1.0)
                        progresso.progress(
                            percentual_etapa,
                            text=f"{percentual_etapa * 100:.0f}% • {texto}"
                        )
                        status_principal.info(texto)
                        metricas_progresso[0].metric(
                            "Etapa",
                            f"{etapa}/{total_etapas}"
                        )
                        metricas_progresso[1].metric(
                            "Fonte atual",
                            fonte_atual
                        )
                        metricas_progresso[2].metric(
                            "Registros",
                            numero_inteiro(registros)
                        )
                        metricas_progresso[3].metric(
                            "Tempo decorrido",
                            time.strftime("%H:%M:%S", time.gmtime(decorrido))
                        )
                        log_area.code(
                            "\n".join(logs[-15:]) or "Aguardando início...",
                            language=None
                        )

                    logs.append(
                        f"{datetime.now():%H:%M:%S} • Validando conexão e período {periodo_selecionado}"
                    )
                    atualizar_painel(
                        1,
                        "Validando conexão com o banco e preparando os scripts..."
                    )

                    for indice, fonte in enumerate(fontes_selecionadas, start=1):
                        titulo_fonte = FONTES_BANCO[fonte]["titulo"]
                        inicio_fonte = time.perf_counter()

                        logs.append(
                            f"{datetime.now():%H:%M:%S} • {titulo_fonte}: consulta iniciada"
                        )
                        atualizar_painel(
                            indice,
                            f"Consultando {titulo_fonte} no PostgreSQL...",
                            titulo_fonte,
                            0
                        )

                        try:
                            qtd = executar_atualizacao_fonte(
                                fonte,
                                cfg_execucao,
                                periodo_selecionado,
                                data_inicio_periodo,
                                data_fim_periodo,
                            )
                            tempo_fonte = time.perf_counter() - inicio_fonte
                            logs.append(
                                f"{datetime.now():%H:%M:%S} • {titulo_fonte}: "
                                f"{qtd:,} registros salvos em {tempo_fonte:.1f}s"
                                .replace(",", ".")
                            )
                            if fonte == "entradas":
                                validacao_entradas = st.session_state.get("_ultima_validacao_entradas", {})
                                if validacao_entradas:
                                    logs.append(
                                        f"{datetime.now():%H:%M:%S} • Entradas validadas: "
                                        f"{numero_inteiro(validacao_entradas.get('registros', 0))} itens analíticos • "
                                        f"{numero_inteiro(validacao_entradas.get('notas', 0))} notas • "
                                        f"Compra A7 {moeda_real(validacao_entradas.get('valor', 0))} • "
                                        f"{validacao_entradas.get('metodo', '')}"
                                    )
                            if fonte == "vendas":
                                logs.append(
                                    f"{datetime.now():%H:%M:%S} • Regra A7 aplicada: "
                                    "todos os itens finalizados, inclusive USO CONSUMO; "
                                    "movimentação usada somente para custo."
                                )
                                validacao_vendas = st.session_state.get(
                                    "_ultima_validacao_vendas", {}
                                )
                                if validacao_vendas:
                                    logs.append(
                                        f"{datetime.now():%H:%M:%S} • Vendas validada: "
                                        f"{moeda_real(validacao_vendas.get('faturamento', 0))} • "
                                        f"{numero_inteiro(validacao_vendas.get('quantidade', 0))} itens • "
                                        f"até {validacao_vendas.get('data_final', '')}"
                                    )
                            resultados.append(
                                {
                                    "Fonte": titulo_fonte,
                                    "Status": "Sucesso",
                                    "Registros": qtd,
                                    "Tempo": time.strftime(
                                        "%H:%M:%S",
                                        time.gmtime(tempo_fonte)
                                    ),
                                    "Mensagem": "Atualizada e salva",
                                }
                            )
                            atualizar_painel(
                                indice + 1,
                                f"{titulo_fonte} concluída e salva.",
                                titulo_fonte,
                                qtd
                            )

                        except Exception as e:
                            tempo_fonte = time.perf_counter() - inicio_fonte
                            logs.append(
                                f"{datetime.now():%H:%M:%S} • {titulo_fonte}: ERRO — "
                                f"{type(e).__name__}: {e}"
                            )
                            resultados.append(
                                {
                                    "Fonte": titulo_fonte,
                                    "Status": "Erro",
                                    "Registros": 0,
                                    "Tempo": time.strftime(
                                        "%H:%M:%S",
                                        time.gmtime(tempo_fonte)
                                    ),
                                    "Mensagem": f"{type(e).__name__}: {e}",
                                }
                            )
                            atualizar_painel(
                                indice + 1,
                                f"Erro ao atualizar {titulo_fonte}.",
                                titulo_fonte,
                                0
                            )

                    logs.append(
                        f"{datetime.now():%H:%M:%S} • Recalculando as visões e os KPIs"
                    )
                    atualizar_painel(
                        total_fontes + 1,
                        "Recalculando cards, métricas, gráficos e premiações..."
                    )

                    tempo_total = time.perf_counter() - inicio_geral
                    logs.append(
                        f"{datetime.now():%H:%M:%S} • Atualização finalizada em "
                        f"{time.strftime('%H:%M:%S', time.gmtime(tempo_total))}"
                    )
                    atualizar_painel(
                        total_etapas,
                        "Atualização finalizada.",
                        "Concluído",
                        sum(item["Registros"] for item in resultados)
                    )

                    sucessos = [
                        item for item in resultados
                        if item.get("Status") == "Sucesso"
                    ]
                    if any(item["Status"] == "Erro" for item in resultados):
                        status_principal.warning(
                            "Atualização concluída com uma ou mais falhas. "
                            "As fontes concluídas serão refletidas nas telas."
                        )
                    else:
                        status_principal.success(
                            "Todas as fontes selecionadas foram atualizadas e salvas."
                        )

                    if sucessos:
                        total_registros_atualizados = sum(
                            int(item.get("Registros", 0) or 0)
                            for item in sucessos
                        )
                        fontes_atualizadas = ", ".join(
                            str(item.get("Fonte", "")) for item in sucessos
                        )
                        _registrar_atualizacao_dados(
                            fontes_atualizadas,
                            periodo_selecionado,
                            total_registros_atualizados,
                        )
                        try:
                            identidade_atual = validar_identidade_banco(
                                carregar_config_banco()
                            )
                            _registrar_origem_cache_banco(
                                carregar_config_banco(), identidade_atual
                            )
                        except Exception:
                            pass
                        reconstruir_visoes_imediatamente(periodo_selecionado)
                        st.session_state.pop("_banco_alterado_pendente", None)
                        st.session_state["_flash_atualizacao_dados"] = (
                            f"Banco atualizado para {periodo_selecionado}: "
                            f"{fontes_atualizadas}. Cards, gráficos, tabelas e "
                            "premiações foram recalculados automaticamente."
                        )
                        st.rerun()

                    resultados_df = pd.DataFrame(resultados)
                    dataframe_br(
                        resultados_df,
                        use_container_width=True,
                        hide_index=True
                    )
                    erros_df = resultados_df[
                        resultados_df["Status"].astype(str).str.casefold().eq("erro")
                    ]
                    if not erros_df.empty:
                        with st.expander(
                            "🔎 Detalhes completos do erro", expanded=True
                        ):
                            for _, linha_erro in erros_df.iterrows():
                                st.code(
                                    f"{linha_erro['Fonte']}\n{linha_erro['Mensagem']}",
                                    language="text",
                                )

    with aba_historico:
        st.markdown("### Limpeza e reconstrução das bases")
        st.caption(
            "Remove somente Vendas, Estoque, Entradas, Contas a Pagar, resumos "
            "e histórico técnico. Metas, compradores, mapas, configurações e SQL "
            "permanecem preservados."
        )
        confirmar_limpeza = st.checkbox(
            "Confirmo que desejo apagar todas as bases e resumos salvos",
            key="confirmar_limpeza_total_bases",
        )
        if st.button(
            "🧹 Limpar tudo para nova atualização",
            type="primary",
            use_container_width=True,
            disabled=not confirmar_limpeza,
            key="limpar_todas_bases_operacionais",
        ):
            limpar_dados_operacionais()
            st.success(
                "Bases e resumos apagados. O projeto está pronto para uma nova carga."
            )
            st.rerun()

        st.markdown("### Importar base anual de Contas a Pagar")
        st.caption(
            "Use o CSV exportado pelo ERP. O sistema separa os registros por "
            "mês de vencimento e substitui apenas o cache de Contas a Pagar."
        )
        csv_contas_pagar = st.file_uploader(
            "Arquivo CSV anual",
            type=["csv", "txt"],
            key="upload_csv_anual_contas_pagar",
        )
        if csv_contas_pagar is not None:
            if st.button(
                "📥 Importar e reconstruir Contas a Pagar",
                use_container_width=True,
                key="importar_csv_anual_contas_pagar",
            ):
                try:
                    with st.spinner("Importando e separando as competências..."):
                        resultado_importacao, planos_importados = (
                            importar_csv_anual_contas_pagar(csv_contas_pagar)
                        )
                        st.cache_data.clear()
                    st.success(
                        "Contas a Pagar importado e salvo por competência."
                    )
                    dataframe_br(
                        resultado_importacao,
                        use_container_width=True,
                        hide_index=True,
                    )
                    if planos_importados:
                        st.caption(
                            f"Planos identificados: {len(planos_importados)}"
                        )
                except Exception as erro_importacao:
                    st.error(str(erro_importacao))

        arquivo_log_cp = LOG_DIR / "contas_pagar_erros.log"
        if arquivo_log_cp.exists():
            with st.expander(
                "🔎 Último erro técnico de Contas a Pagar",
                expanded=False,
            ):
                conteudo_log = arquivo_log_cp.read_text(
                    encoding="utf-8",
                    errors="replace",
                )
                st.code(conteudo_log[-8000:], language="text")

        sql_log_cp = LOG_DIR / "contas_pagar_sql_executado.sql"
        erro_log_cp = LOG_DIR / "contas_pagar_erro.txt"

        if erro_log_cp.exists():
            with st.expander(
                "🔎 Diagnóstico completo de Contas a Pagar",
                expanded=True,
            ):
                st.code(
                    erro_log_cp.read_text(
                        encoding="utf-8",
                        errors="replace",
                    )[-12000:],
                    language="text",
                )

        if sql_log_cp.exists():
            with st.expander(
                "🧾 SQL final executado em Contas a Pagar",
                expanded=False,
            ):
                st.code(
                    sql_log_cp.read_text(
                        encoding="utf-8",
                        errors="replace",
                    ),
                    language="sql",
                )

        st.markdown("### Histórico de atualizações")
        hist = historico_atualizacoes()
        if not hist.empty:
            dataframe_br(hist, use_container_width=True, hide_index=True, height=300)
        else:
            st.info("Nenhuma atualização de banco foi realizada.")

        st.markdown("### Consultar dados mensais salvos")
        c1, c2 = st.columns(2)
        with c1:
            fonte_consulta = st.selectbox(
                "Fonte salva",
                list(FONTES_BANCO.keys()),
                format_func=lambda x: FONTES_BANCO[x]["titulo"],
                key="fonte_consulta_cache"
            )
        with c2:
            periodo_consulta = st.selectbox(
                "Período salvo",
                list(dict.fromkeys(
                    [METAS_GESTOR.get("periodo_referencia", "")] +
                    [
                        h.get("periodo_referencia")
                        for h in carregar_historico()
                        if h.get("periodo_referencia")
                    ]
                )),
                key="periodo_consulta_cache"
            )

        snapshot = carregar_snapshot(fonte_consulta, periodo_consulta)
        if not snapshot.empty:
            st.metric("Registros encontrados", numero_inteiro(len(snapshot)))
            dataframe_br(snapshot, use_container_width=True, hide_index=True, height=400)

            csv_snapshot = snapshot.to_csv(
                index=False,
                sep=";",
                encoding="utf-8-sig"
            ).encode("utf-8-sig")
            st.download_button(
                "📥 Exportar dados salvos",
                data=csv_snapshot,
                file_name=f"{fonte_consulta}_{periodo_consulta}.csv",
                mime="text/csv",
                use_container_width=True
            )
        else:
            st.warning("Não existem dados salvos para essa fonte e período.")



elif visao == "Auditoria de Compradores":
    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff;font-size:20px">Auditoria de Categorias por Comprador</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:4px">
          Confere todas as categorias existentes em Vendas, Estoque e Entradas e identifica pendências de responsabilidade.
        </div>
      </div>
      <div style="color:#56d7ec;font-weight:800">Cobertura e Governança</div>
    </div>
    """, unsafe_allow_html=True)

    token_mapa = _arquivo_token(MAPA_COMPRADORES_FILE)
    auditoria = carregar_auditoria_categorias(
        PERIODO_DASHBOARD,
        _arquivo_token(CACHE_DB_FILE),
        token_mapa,
    )

    total_categorias = len(auditoria)
    total_ok = int((auditoria["Situação"] == "OK").sum()) if not auditoria.empty else 0
    total_sem = int(auditoria["Situação"].isin(["Sem comprador", "Sem classificação"]).sum()) if not auditoria.empty else 0
    total_dup = int((auditoria["Situação"] == "Duplicidade").sum()) if not auditoria.empty else 0
    total_inativo = int((auditoria["Situação"] == "Comprador inativo").sum()) if not auditoria.empty else 0
    cobertura = (total_ok / total_categorias * 100) if total_categorias else 0.0

    c1, c2, c3, c4, c5 = st.columns(5)
    with c1: card_status_base("Categorias na base", total_categorias)
    with c2: card_status_base("Com comprador", total_ok)
    with c3: card_status_base("Sem comprador", total_sem)
    with c4: card_status_base("Duplicidades/Inativos", total_dup + total_inativo)
    with c5:
        st.markdown(
            f"""<div class="status-card"><div class="status-label">Cobertura</div>
            <div class="status-value">{percentual(cobertura)}</div>
            <div class="status-pill">categorias auditadas</div></div>""",
            unsafe_allow_html=True,
        )

    if auditoria.empty:
        st.warning("Não existem categorias disponíveis nas bases do período selecionado.")
    else:
        st.markdown("### Pendências e cobertura")
        f1, f2, f3 = st.columns([1.2, 1.2, 2.0])
        with f1:
            situacoes = st.multiselect(
                "Situação",
                options=sorted(auditoria["Situação"].unique()),
                default=sorted(auditoria["Situação"].unique()),
                key="auditoria_situacoes",
            )
        with f2:
            compradores_filtro = sorted(
                [c for c in auditoria["Comprador atual"].dropna().unique() if str(c).strip()],
                key=str.casefold,
            )
            comprador_auditoria = st.selectbox(
                "Comprador atual",
                ["Todos"] + compradores_filtro,
                key="auditoria_comprador_filtro",
            )
        with f3:
            busca_auditoria = st.text_input(
                "Buscar categoria",
                placeholder="Digite parte da classificação...",
                key="auditoria_busca",
            )

        exibicao = auditoria[auditoria["Situação"].isin(situacoes)].copy()
        if comprador_auditoria != "Todos":
            exibicao = exibicao[exibicao["Comprador atual"] == comprador_auditoria]
        if busca_auditoria.strip():
            exibicao = exibicao[
                exibicao["Classificação Principal"].astype(str).str.contains(
                    busca_auditoria.strip(), case=False, na=False, regex=False
                )
            ]

        dataframe_br(
            exibicao,
            use_container_width=True,
            hide_index=True,
            height=min(520, 80 + 34 * max(len(exibicao), 1)),
            export_title="Auditoria de Categorias por Comprador",
        )

        st.markdown("### Corrigir vínculo em massa")
        pendentes = auditoria[auditoria["Situação"] != "OK"]["Classificação Principal"].tolist()
        categorias_selecionadas = st.multiselect(
            "Categorias para vincular",
            options=pendentes,
            key="auditoria_categorias_corrigir",
        )
        compradores_ativos_auditoria = sorted(lista_compradores_ativos(), key=str.casefold)
        comprador_destino = st.selectbox(
            "Comprador responsável",
            options=[""] + compradores_ativos_auditoria,
            key="auditoria_comprador_destino",
        )
        b1, b2 = st.columns([1, 2])
        with b1:
            if st.button("💾 Salvar vínculos", use_container_width=True, type="primary"):
                if not categorias_selecionadas:
                    st.warning("Selecione pelo menos uma categoria.")
                elif not comprador_destino:
                    st.warning("Selecione o comprador responsável.")
                else:
                    quantidade = salvar_vinculos_auditoria(
                        categorias_selecionadas,
                        comprador_destino,
                    )
                    st.success(
                        f"{numero_inteiro(quantidade)} categoria(s) vinculada(s) a {comprador_destino}. "
                        "As visões serão recalculadas."
                    )
                    st.rerun()
        with b2:
            st.caption(
                "A correção substitui somente o vínculo exato das categorias selecionadas, "
                "preservando os demais cadastros e compradores."
            )



elif visao == "Compradores por Classificação":
    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff;font-size:20px">Compradores por Classificação Principal</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:4px">
          Altere o comprador responsável por cada classificação sem modificar o código do projeto.
        </div>
      </div>
      <div style="color:#56d7ec;font-weight:800">Gestão de Responsabilidades</div>
    </div>
    """, unsafe_allow_html=True)

    st.info(
        "As alterações são salvas e passam a valer automaticamente nas visões, "
        "gráficos, metas, resultados e premiações."
    )

    mapa_df = pd.DataFrame(carregar_mapa_compradores_editavel())
    if mapa_df.empty:
        mapa_df = pd.DataFrame(columns=["Área", "Classificação Principal", "Comprador"])

    compradores_disponiveis = sorted(lista_compradores_ativos(), key=lambda x: x.casefold())

    editado = st.data_editor(
        mapa_df,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "Área": st.column_config.TextColumn("Área", width="medium"),
            "Classificação Principal": st.column_config.TextColumn(
                "Classificação Principal",
                width="large",
                required=True
            ),
            "Comprador": st.column_config.SelectboxColumn(
                "Comprador Responsável",
                options=compradores_disponiveis,
                required=True,
                width="medium"
            ),
        },
        key="editor_mapa_compradores"
    )

    c1, c2, c3 = st.columns(3)
    with c1:
        if st.button("💾 Salvar alterações", use_container_width=True):
            registros = editado.fillna("").to_dict(orient="records")
            registros = [
                {
                    "Área": str(r.get("Área", "")).strip(),
                    "Classificação Principal": str(r.get("Classificação Principal", "")).strip(),
                    "Comprador": str(r.get("Comprador", "")).strip(),
                }
                for r in registros
                if str(r.get("Classificação Principal", "")).strip()
            ]
            salvar_mapa_compradores_editavel(registros)
            st.success("Responsabilidades atualizadas com sucesso.")
            st.rerun()

    with c2:
        if st.button("↩️ Restaurar padrão inicial", use_container_width=True):
            salvar_mapa_compradores_editavel(MAPA_COMPRADORES_PADRAO)
            st.success("Mapa padrão restaurado.")
            st.rerun()

    with c3:
        csv_mapa = editado.to_csv(index=False, sep=";", encoding="utf-8-sig").encode("utf-8-sig")
        st.download_button(
            "📥 Exportar mapa",
            data=csv_mapa,
            file_name="mapa_compradores_classificacao.csv",
            mime="text/csv",
            use_container_width=True
        )

    st.markdown("### Resumo por comprador")
    resumo = (
        editado.groupby("Comprador", dropna=False)
        .size()
        .reset_index(name="Classificações")
        .sort_values("Classificações", ascending=False)
    )
    dataframe_br(resumo, use_container_width=True, hide_index=True)

    st.caption(
        "Após alterar o mapa, atualize ou reabra o painel para recalcular todas as visões "
        "com o novo responsável."
    )



elif visao == "Cadastro de Compradores":
    _renderizar_gestao_acessos_compradores()
    # Período seguro para cadastro e criação automática de metas.
    # Não depende de variáveis locais de outras telas.
    PERIODO_CADASTRO_COMPRADOR = str(
        globals().get("PERIODO_DASHBOARD")
        or globals().get("PERIODO_GLOBAL_SELECIONADO")
        or st.session_state.get("periodo_gestao_unificado_global")
        or datetime.now().strftime("%Y-%m")
    )

    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff;font-size:20px">Cadastro de Compradores</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:4px">
          Cadastre novos compradores e gerencie os compradores existentes.
        </div>
      </div>
      <div style="color:#56d7ec;font-weight:800">Gestão de Responsáveis</div>
    </div>
    """, unsafe_allow_html=True)

    st.info(
        "Use o formulário abaixo para incluir um comprador. A tabela serve para "
        "renomear, ativar, inativar ou excluir registros existentes."
    )

    st.markdown("### ➕ Novo comprador")
    with st.form("form_novo_comprador", clear_on_submit=True):
        f1, f2 = st.columns([3, 1])
        with f1:
            nome_novo = st.text_input(
                "Nome do comprador",
                placeholder="Digite o nome completo",
            )
        with f2:
            status_novo = st.selectbox(
                "Status inicial",
                ["Ativo", "Inativo"],
                index=0,
            )
        incluir = st.form_submit_button(
            "➕ Cadastrar comprador",
            type="primary",
            use_container_width=True,
        )

    if incluir:
        nome_limpo = str(nome_novo or "").strip()
        atuais = carregar_cadastro_compradores()
        nomes_existentes = {
            str(item.get("Comprador", "")).strip().casefold()
            for item in atuais
            if str(item.get("Comprador", "")).strip()
        }
        if not nome_limpo:
            st.error("Informe o nome do comprador.")
        elif nome_limpo.casefold() in nomes_existentes:
            st.error(f"O comprador {nome_limpo} já está cadastrado.")
        else:
            atuais.append({"Comprador": nome_limpo, "Status": status_novo})
            salvar_cadastro_compradores(atuais)
            garantir_meta_comprador_periodo(nome_limpo, PERIODO_CADASTRO_COMPRADOR, 0.0)
            _registrar_atualizacao_dados(
                fonte="Cadastro de Compradores",
                periodo=PERIODO_CADASTRO_COMPRADOR,
                registros=len(atuais),
            )
            st.session_state["_flash_atualizacao_dados"] = (
                f"Comprador {nome_limpo} cadastrado com sucesso."
            )
            st.rerun()

    st.markdown("### ✏️ Compradores cadastrados")
    cadastro_df = pd.DataFrame(carregar_cadastro_compradores())
    if cadastro_df.empty:
        cadastro_df = pd.DataFrame(columns=["Comprador", "Status"])
    for coluna in ["Comprador", "Status"]:
        if coluna not in cadastro_df.columns:
            cadastro_df[coluna] = ""
    cadastro_df = cadastro_df[["Comprador", "Status"]].copy()
    original_df = cadastro_df.copy()

    editado_compradores = st.data_editor(
        cadastro_df,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "Comprador": st.column_config.TextColumn(
                "Nome do Comprador", required=True, width="large"
            ),
            "Status": st.column_config.SelectboxColumn(
                "Status", options=["Ativo", "Inativo"],
                default="Ativo", required=True, width="medium"
            ),
        },
        key="editor_cadastro_compradores",
    )

    c1, c2, c3 = st.columns(3)
    with c1:
        salvar = st.button(
            "💾 Salvar alterações", use_container_width=True, type="primary"
        )
    with c2:
        limpar_todos = st.button(
            "🗑️ Excluir todos", use_container_width=True
        )
    with c3:
        restaurar = st.button(
            "↩️ Restaurar padrão", use_container_width=True
        )

    if salvar:
        novos = []
        nomes_usados = set()
        erro = ""
        for _, row in editado_compradores.fillna("").iterrows():
            nome = str(row.get("Comprador", "")).strip()
            status = str(row.get("Status", "Ativo")).strip() or "Ativo"
            if not nome:
                continue
            chave = nome.casefold()
            if chave in nomes_usados:
                erro = f"Comprador duplicado: {nome}"
                break
            nomes_usados.add(chave)
            novos.append({"Comprador": nome, "Status": status})

        if erro:
            st.error(erro)
        else:
            mapa = carregar_mapa_compradores_editavel()
            antigos = {
                str(x.get("Comprador", "")).strip().casefold():
                str(x.get("Comprador", "")).strip()
                for x in original_df.to_dict(orient="records")
                if str(x.get("Comprador", "")).strip()
            }
            atuais_nomes = {
                str(x.get("Comprador", "")).strip().casefold():
                str(x.get("Comprador", "")).strip()
                for x in novos
            }
            # Renomeações somente quando a posição correspondente permaneceu.
            limite = min(len(original_df), len(editado_compradores))
            for i in range(limite):
                antigo = str(original_df.iloc[i].get("Comprador", "")).strip()
                novo = str(editado_compradores.iloc[i].get("Comprador", "")).strip()
                if antigo and novo and antigo.casefold() != novo.casefold():
                    for item in mapa:
                        if str(item.get("Comprador", "")).strip().casefold() == antigo.casefold():
                            item["Comprador"] = novo
                    atualizar_nome_comprador_metas(antigo, novo)

            # Vínculos de compradores excluídos ficam como Não mapeado.
            excluidos = set(antigos) - set(atuais_nomes)
            if excluidos:
                for item in mapa:
                    if str(item.get("Comprador", "")).strip().casefold() in excluidos:
                        item["Comprador"] = "Não mapeado"

            salvar_mapa_compradores_editavel(mapa)
            salvar_cadastro_compradores(novos)
            for item in novos:
                garantir_meta_comprador_periodo(
                    item["Comprador"], PERIODO_CADASTRO_COMPRADOR, 0.0
                )
            _registrar_atualizacao_dados(
                fonte="Cadastro de Compradores",
                periodo=PERIODO_CADASTRO_COMPRADOR,
                registros=len(novos),
            )
            st.session_state["_flash_atualizacao_dados"] = (
                "Cadastro de compradores atualizado com sucesso."
            )
            st.rerun()

    if limpar_todos:
        salvar_cadastro_compradores([])
        mapa = carregar_mapa_compradores_editavel()
        for item in mapa:
            item["Comprador"] = "Não mapeado"
        salvar_mapa_compradores_editavel(mapa)
        _registrar_atualizacao_dados(
            fonte="Cadastro de Compradores",
            periodo=PERIODO_CADASTRO_COMPRADOR,
            registros=0,
        )
        st.session_state["_flash_atualizacao_dados"] = (
            "Todos os compradores foram excluídos. Você já pode cadastrar os novos."
        )
        st.rerun()

    if restaurar:
        salvar_cadastro_compradores([dict(x) for x in COMPRADORES_PADRAO])
        _registrar_atualizacao_dados(
            fonte="Cadastro de Compradores",
            periodo=PERIODO_CADASTRO_COMPRADOR,
            registros=len(COMPRADORES_PADRAO),
        )
        st.session_state["_flash_atualizacao_dados"] = (
            "Compradores padrão restaurados."
        )
        st.rerun()

    ativos = pd.DataFrame(
        [{"Comprador": nome} for nome in lista_compradores_ativos()]
    )
    st.markdown("### Compradores ativos")
    if ativos.empty:
        st.warning("Nenhum comprador ativo. Use o formulário acima para cadastrar.")
    else:
        dataframe_br(ativos, use_container_width=True, hide_index=True)

    st.caption(
        "Após cadastrar o novo comprador, acesse Compradores por Classificação "
        "para vincular as categorias sob responsabilidade dele."
    )

elif visao == "Metas por Comprador":
    st.markdown("""
    <div class="premium-box">
      <div>
        <div style="font-weight:950;color:#fff;font-size:20px">Metas por Comprador</div>
        <div style="color:#8da2b8;font-size:13px;margin-top:4px">
          Defina metas individuais para cada comprador e para cada período.
        </div>
      </div>
      <div style="color:#56d7ec;font-weight:800">Gestão Individual de Performance</div>
    </div>
    """, unsafe_allow_html=True)

    st.info(
        "Ao criar um novo comprador, o sistema gera automaticamente uma meta inicial "
        "para o período ativo. Depois, você pode ajustar os valores individualmente."
    )

    historico_metas = carregar_historico()
    periodos = list(dict.fromkeys(
        [METAS_GESTOR.get("periodo_referencia", "")] +
        [
            h.get("periodo_referencia")
            for h in historico_metas
            if h.get("periodo_referencia")
        ]
    ))

    c_periodo, c_comprador = st.columns(2)
    with c_periodo:
        periodo_meta_comprador = st.selectbox(
            "Período da meta",
            periodos,
            key="periodo_meta_comprador"
        )

    metas_individuais = garantir_metas_compradores_periodo(periodo_meta_comprador)
    compradores_ativos = lista_compradores_ativos()

    with c_comprador:
        comprador_meta = st.selectbox(
            "Comprador",
            compradores_ativos,
            key="comprador_meta_individual"
        )

    meta_atual = obter_meta_comprador(comprador_meta, periodo_meta_comprador)

    with st.form("form_meta_individual_comprador"):
        st.markdown(f"### Meta de {comprador_meta}")

        g1, g2, g3 = st.columns(3)
        with g1:
            meta_venda_ind = st.number_input(
                "Meta de Venda (R$)",
                min_value=0.0,
                value=float(meta_atual.get("meta_venda", 0)),
                step=1000.0,
                format="%.2f"
            )
            participacao_ind = st.number_input(
                "Participação da Venda (%)",
                min_value=0.0,
                max_value=100.0,
                value=float(meta_atual.get("participacao_venda_pct", 0)),
                step=0.1,
                format="%.2f"
            )
        with g2:
            meta_cmv_pct_ind = st.number_input(
                "Meta CMV (%)",
                min_value=0.0,
                max_value=100.0,
                value=float(meta_atual.get("meta_cmv_pct", 0)),
                step=0.1,
                format="%.2f"
            )
            fator_cob_ind = st.number_input(
                "Fator de Cobertura",
                min_value=0.0,
                value=float(meta_atual.get("fator_cobertura", 0)),
                step=0.05,
                format="%.2f"
            )
        with g3:
            meta_ruptura_ind = st.number_input(
                "Meta de Ruptura (%)",
                min_value=0.0,
                max_value=100.0,
                value=float(meta_atual.get("meta_ruptura_pct", 0)),
                step=0.1,
                format="%.2f"
            )
            meta_reposicao_ind = st.number_input(
                "Meta de Reposição (%)",
                min_value=0.0,
                max_value=200.0,
                value=float(meta_atual.get("meta_reposicao_pct", 0)),
                step=0.1,
                format="%.2f"
            )

        st.markdown("### Distribuição do estoque por curva")
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            curva_a_ind = st.number_input(
                "Curva A (%)", 0.0, 100.0,
                float(meta_atual.get("meta_curva_a_pct", 0)), 1.0
            )
        with c2:
            curva_b_ind = st.number_input(
                "Curva B (%)", 0.0, 100.0,
                float(meta_atual.get("meta_curva_b_pct", 0)), 1.0
            )
        with c3:
            curva_c_ind = st.number_input(
                "Curva C (%)", 0.0, 100.0,
                float(meta_atual.get("meta_curva_c_pct", 0)), 1.0
            )
        with c4:
            curva_d_ind = st.number_input(
                "Curva D (%)", 0.0, 100.0,
                float(meta_atual.get("meta_curva_d_pct", 0)), 1.0
            )

        valor_premio_ind = st.number_input(
            "Valor de Prêmio Atingível (R$)",
            min_value=0.0,
            value=float(meta_atual.get("valor_premio", 0)),
            step=100.0,
            format="%.2f"
        )

        status_ind = st.selectbox(
            "Status da meta individual",
            ["Planejada", "Ativa", "Encerrada", "Cancelada"],
            index=["Planejada", "Ativa", "Encerrada", "Cancelada"].index(
                meta_atual.get("status", "Ativa")
            )
        )

        salvar_meta_ind = st.form_submit_button(
            "💾 Salvar meta do comprador",
            use_container_width=True
        )

        if salvar_meta_ind:
            total_curvas = curva_a_ind + curva_b_ind + curva_c_ind + curva_d_ind
            if abs(total_curvas - 100.0) > 0.01:
                st.error("A soma das curvas deve ser igual a 100%.")
            else:
                dados = carregar_metas_por_comprador()
                nova_meta = {
                    "periodo_referencia": periodo_meta_comprador,
                    "comprador": comprador_meta,
                    "meta_venda": meta_venda_ind,
                    "participacao_venda_pct": participacao_ind,
                    "meta_cmv_pct": meta_cmv_pct_ind,
                    "meta_cmv_valor": meta_venda_ind * meta_cmv_pct_ind / 100.0,
                    "fator_cobertura": fator_cob_ind,
                    "meta_estoque_total": (
                        meta_venda_ind
                        * meta_cmv_pct_ind / 100.0
                        * fator_cob_ind
                    ),
                    "meta_curva_a_pct": curva_a_ind,
                    "meta_curva_b_pct": curva_b_ind,
                    "meta_curva_c_pct": curva_c_ind,
                    "meta_curva_d_pct": curva_d_ind,
                    "meta_ruptura_pct": meta_ruptura_ind,
                    "meta_reposicao_pct": meta_reposicao_ind,
                    "valor_premio": valor_premio_ind,
                    "status": status_ind,
                    "ultima_atualizacao": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                }

                atualizou = False
                for i, item in enumerate(dados):
                    if (
                        item.get("periodo_referencia") == periodo_meta_comprador
                        and item.get("comprador") == comprador_meta
                    ):
                        dados[i] = nova_meta
                        atualizou = True
                        break

                if not atualizou:
                    dados.append(nova_meta)

                salvar_metas_por_comprador(dados)
                st.success("Meta individual salva com sucesso.")
                st.rerun()

    st.markdown("### Resumo das metas do período")
    resumo_periodo = pd.DataFrame([
        item
        for item in carregar_metas_por_comprador()
        if item.get("periodo_referencia") == periodo_meta_comprador
    ])

    if not resumo_periodo.empty:
        colunas_resumo = [
            "comprador", "meta_venda", "participacao_venda_pct",
            "meta_cmv_pct", "fator_cobertura",
            "meta_ruptura_pct", "meta_reposicao_pct",
            "valor_premio", "status", "ultima_atualizacao"
        ]
        resumo_periodo = resumo_periodo[
            [c for c in colunas_resumo if c in resumo_periodo.columns]
        ].copy()

        if "meta_venda" in resumo_periodo:
            resumo_periodo["meta_venda"] = resumo_periodo["meta_venda"].map(moeda)
        if "valor_premio" in resumo_periodo:
            resumo_periodo["valor_premio"] = resumo_periodo["valor_premio"].map(moeda)
        for coluna in [
            "participacao_venda_pct", "meta_cmv_pct",
            "meta_ruptura_pct", "meta_reposicao_pct"
        ]:
            if coluna in resumo_periodo:
                resumo_periodo[coluna] = resumo_periodo[coluna].map(percentual)

        dataframe_br(
            resumo_periodo,
            use_container_width=True,
            hide_index=True
        )

        total_meta_venda = sum(
            float(item.get("meta_venda", 0))
            for item in carregar_metas_por_comprador()
            if item.get("periodo_referencia") == periodo_meta_comprador
            and item.get("status") != "Cancelada"
        )
        total_participacao = sum(
            float(item.get("participacao_venda_pct", 0))
            for item in carregar_metas_por_comprador()
            if item.get("periodo_referencia") == periodo_meta_comprador
            and item.get("status") != "Cancelada"
        )

        r1, r2, r3 = st.columns(3)
        with r1:
            st.metric("Soma das metas de venda", moeda_real(total_meta_venda))
        with r2:
            st.metric("Meta geral do período", moeda_real(METAS_GESTOR.get('meta_venda_total_mes', 0)))
        with r3:
            st.metric("Participação total", percentual(total_participacao))

        if abs(total_meta_venda - float(METAS_GESTOR.get("meta_venda_total_mes", 0))) > 0.01:
            st.warning(
                "A soma das metas individuais está diferente da meta geral do período."
            )
        if abs(total_participacao - 100.0) > 0.01:
            st.warning(
                "A soma das participações dos compradores está diferente de 100%."
            )


renderizar_exportacao_tela()

st.markdown("<div class='eirox-footer'>REDE ECONOMIZE • KPI COMERCIAL • ENTERPRISE PRO</div>", unsafe_allow_html=True)
